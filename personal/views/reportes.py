"""
Vistas para el módulo de Reportes RRHH.
Genera reportes Excel descargables: Plantilla de Personal, Asistencia Mensual, HE Detallado.
"""
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import Area, SubArea, Personal

# ─── Solo admin ───────────────────────────────────────────────────────────────
solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')

# ─── Estilos openpyxl ─────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="0D2B27")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL     = PatternFill("solid", fgColor="F0FDFA")
TEAL_FILL    = PatternFill("solid", fgColor="5EEAD4")
SUMMARY_FILL = PatternFill("solid", fgColor="134E4A")
SUMMARY_FONT = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL   = PatternFill("solid", fgColor="CCFBF1")
TOTAL_FONT   = Font(bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _auto_width(ws):
    """Ajusta el ancho de cada columna según el contenido más largo."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or '')) for cell in col),
            default=10
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)


def _header_row(ws, row_num, headers):
    """Escribe una fila de encabezados con el estilo teal oscuro."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_num].height = 30


def _data_cell(ws, row_num, col_num, value, is_even=False, number_format=None, align='left'):
    """Escribe una celda de datos con estilo alterno."""
    cell = ws.cell(row=row_num, column=col_num, value=value)
    cell.fill = ALT_FILL if is_even else PatternFill()
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def _calcular_antiguedad(fecha_alta):
    """Calcula años de antigüedad desde la fecha de alta hasta hoy."""
    if not fecha_alta:
        return 0
    hoy = date.today()
    anios = (hoy - fecha_alta).days // 365
    return anios


# ─────────────────────────────────────────────────────────────────────────────
# VISTA PANEL PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

@solo_admin
def reportes_panel(request):
    """Hub de reportes RRHH: acceso a plantilla, asistencia y HE."""
    hoy = timezone.localdate()
    areas = Area.objects.filter(activa=True).order_by('nombre')
    context = {
        'titulo': 'Reportes RRHH',
        'areas': areas,
        'anio_actual': hoy.year,
        'mes_actual': hoy.month,
        'meses': [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre'),
        ],
        'anios': list(range(hoy.year - 3, hoy.year + 1)),
    }
    return render(request, 'personal/reportes_panel.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# REPORTE 1: PLANTILLA DE PERSONAL
# ─────────────────────────────────────────────────────────────────────────────

@solo_admin
def reporte_plantilla(request):
    """
    Reporte de plantilla de personal en Excel.
    Filtros: estado, area (SubArea.area), grupo_tareo, tipo_contrato.
    Activa descarga si ?format=excel o si el método es POST.
    """
    # ── Leer filtros ──────────────────────────────────────────────────────────
    estado       = request.GET.get('estado', 'Activo')
    area_id      = request.GET.get('area', '')
    grupo_tareo  = request.GET.get('grupo_tareo', '')
    tipo_contrato = request.GET.get('tipo_contrato', '')

    # ── Queryset base ─────────────────────────────────────────────────────────
    qs = (
        Personal.objects
        .select_related('subarea', 'subarea__area')
        .order_by('subarea__area__nombre', 'apellidos_nombres')
    )

    if estado:
        qs = qs.filter(estado=estado)
    if area_id:
        qs = qs.filter(subarea__area_id=area_id)
    if grupo_tareo:
        qs = qs.filter(grupo_tareo=grupo_tareo)
    if tipo_contrato:
        qs = qs.filter(tipo_contrato=tipo_contrato)

    # Si no es solicitud de descarga, mostrar formulario (aunque aquí solo generamos Excel)
    if request.GET.get('format') != 'excel':
        # Redirigir al panel si se accede directamente sin ?format=excel
        areas = Area.objects.filter(activa=True).order_by('nombre')
        return render(request, 'personal/reportes_panel.html', {
            'titulo': 'Reportes RRHH',
            'areas': areas,
            'anio_actual': date.today().year,
            'mes_actual': date.today().month,
            'meses': [
                (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
                (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
                (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre'),
            ],
            'anios': list(range(date.today().year - 3, date.today().year + 1)),
        })

    # ── Generar Excel ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── HOJA 1: Plantilla de Personal ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Plantilla de Personal"

    headers = [
        'Nro', 'DNI', 'Apellidos y Nombres', 'Cargo', 'Área', 'SubÁrea',
        'Grupo', 'Tipo Trab.', 'Categoría', 'Rég. Pensión', 'AFP',
        'Tipo Contrato', 'Condición', 'Fecha Alta', 'Antigüedad (años)',
        'Sueldo Base', 'Estado', 'Email Corporativo', 'Celular',
    ]
    _header_row(ws1, 1, headers)
    ws1.freeze_panes = 'A2'

    masa_salarial = Decimal('0.00')

    for idx, p in enumerate(qs, 1):
        row_num = idx + 1
        is_even = idx % 2 == 0
        antiguedad = _calcular_antiguedad(p.fecha_alta)
        sueldo = p.sueldo_base or Decimal('0.00')
        masa_salarial += sueldo

        datos = [
            idx,
            p.nro_doc,
            p.apellidos_nombres,
            p.cargo,
            p.subarea.area.nombre if p.subarea and p.subarea.area else '',
            p.subarea.nombre if p.subarea else '',
            p.grupo_tareo,
            p.tipo_trab,
            p.get_categoria_display() if hasattr(p, 'get_categoria_display') else p.categoria,
            p.get_regimen_pension_display() if hasattr(p, 'get_regimen_pension_display') else p.regimen_pension,
            p.afp,
            p.tipo_contrato,
            p.condicion,
            p.fecha_alta,
            antiguedad,
            float(sueldo),
            p.estado,
            p.correo_corporativo,
            p.celular,
        ]

        for col_num, value in enumerate(datos, 1):
            align = 'right' if col_num in (1, 15, 16) else 'left'
            if col_num == 1:
                align = 'center'
            fmt = None
            if col_num == 2:          # DNI/CE — texto
                fmt = '@'
            elif col_num == 16:
                fmt = '#,##0.00'
            elif col_num == 14 and isinstance(value, date):
                fmt = 'DD/MM/YYYY'
            _data_cell(ws1, row_num, col_num, value, is_even, fmt, align)

    # Fila total masa salarial
    total_row = len(list(qs)) + 2
    total_cell = ws1.cell(row=total_row, column=15, value='MASA SALARIAL TOTAL')
    total_cell.font = TOTAL_FONT
    total_cell.fill = TOTAL_FILL
    total_cell.alignment = Alignment(horizontal='right')
    total_cell.border = THIN_BORDER

    masa_cell = ws1.cell(row=total_row, column=16, value=float(masa_salarial))
    masa_cell.font = TOTAL_FONT
    masa_cell.fill = TOTAL_FILL
    masa_cell.number_format = '#,##0.00'
    masa_cell.alignment = Alignment(horizontal='right')
    masa_cell.border = THIN_BORDER

    _auto_width(ws1)

    # ── HOJA 2: Resumen ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")

    # Total por área
    row = 1
    ws2.cell(row=row, column=1, value="RESUMEN — PLANTILLA DE PERSONAL").font = Font(bold=True, size=12, color="0D2B27")
    ws2.cell(row=row, column=1).fill = PatternFill("solid", fgColor="CCFBF1")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Sección 1: Total por área
    _header_row(ws2, row, ['Área', 'Total Empleados', 'Masa Salarial'])
    row += 1
    por_area = (
        qs.values('subarea__area__nombre')
        .annotate(total=Count('id'), masa=Sum('sueldo_base'))
        .order_by('subarea__area__nombre')
    )
    for i, item in enumerate(por_area):
        is_even = i % 2 == 0
        _data_cell(ws2, row, 1, item['subarea__area__nombre'] or 'Sin Área', is_even)
        _data_cell(ws2, row, 2, item['total'], is_even, align='center')
        _data_cell(ws2, row, 3, float(item['masa'] or 0), is_even, '#,##0.00', 'right')
        row += 1
    row += 1

    # Sección 2: STAFF vs RCO
    _header_row(ws2, row, ['Grupo', 'Total', '% del Total'])
    row += 1
    total_general = qs.count()
    for i, grp in enumerate(['STAFF', 'RCO', 'OTRO']):
        cnt = qs.filter(grupo_tareo=grp).count()
        pct = round(cnt / total_general * 100, 1) if total_general else 0
        is_even = i % 2 == 0
        _data_cell(ws2, row, 1, grp, is_even)
        _data_cell(ws2, row, 2, cnt, is_even, align='center')
        _data_cell(ws2, row, 3, f'{pct}%', is_even, align='center')
        row += 1
    row += 1

    # Sección 3: Por régimen pensión
    _header_row(ws2, row, ['Régimen Pensión', 'Total'])
    row += 1
    por_pension = (
        qs.values('regimen_pension')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    for i, item in enumerate(por_pension):
        is_even = i % 2 == 0
        _data_cell(ws2, row, 1, item['regimen_pension'], is_even)
        _data_cell(ws2, row, 2, item['total'], is_even, align='center')
        row += 1
    row += 1

    # Sección 4: Por tipo contrato
    _header_row(ws2, row, ['Tipo Contrato', 'Total'])
    row += 1
    por_contrato = (
        qs.values('tipo_contrato')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    for i, item in enumerate(por_contrato):
        is_even = i % 2 == 0
        _data_cell(ws2, row, 1, item['tipo_contrato'] or 'No especificado', is_even)
        _data_cell(ws2, row, 2, item['total'], is_even, align='center')
        row += 1
    row += 1

    # Sección 5: Masa salarial total
    _header_row(ws2, row, ['Concepto', 'Monto'])
    row += 1
    masa_total = qs.aggregate(total=Sum('sueldo_base'))['total'] or Decimal('0')
    _data_cell(ws2, row, 1, 'Masa Salarial Total', False)
    masa_c = ws2.cell(row=row, column=2, value=float(masa_total))
    masa_c.number_format = 'S/ #,##0.00'
    masa_c.font = TOTAL_FONT
    masa_c.fill = TOTAL_FILL
    masa_c.border = THIN_BORDER
    masa_c.alignment = Alignment(horizontal='right')

    _auto_width(ws2)

    # ── Respuesta HTTP ─────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"plantilla_personal_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# REPORTE 2: ASISTENCIA MENSUAL
# ─────────────────────────────────────────────────────────────────────────────

@solo_admin
def reporte_asistencia_mensual(request):
    """
    Reporte de asistencia mensual por empleado.
    Filtros: año, mes.
    Descarga si ?format=excel.
    """
    from asistencia.models import RegistroTareo

    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes  = int(request.GET.get('mes', hoy.month))

    if request.GET.get('format') != 'excel':
        areas = Area.objects.filter(activa=True).order_by('nombre')
        return render(request, 'personal/reportes_panel.html', {
            'titulo': 'Reportes RRHH',
            'areas': areas,
            'anio_actual': hoy.year,
            'mes_actual': hoy.month,
            'meses': [
                (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
                (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
                (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre'),
            ],
            'anios': list(range(hoy.year - 3, hoy.year + 1)),
        })

    # ── Registros del mes ─────────────────────────────────────────────────────
    registros = (
        RegistroTareo.objects
        .filter(fecha__year=anio, fecha__month=mes)
        .select_related('personal', 'personal__subarea', 'personal__subarea__area')
        .order_by('personal__apellidos_nombres', 'fecha')
    )

    # ── Agregar por empleado ───────────────────────────────────────────────────
    from collections import defaultdict
    empleados = defaultdict(lambda: {
        'apellidos_nombres': '',
        'area': '',
        'grupo': '',
        'dias_habiles': 0,
        'dias_asistio': 0,
        'dias_falta': 0,
        'he_25': Decimal('0'),
        'he_35': Decimal('0'),
        'he_100': Decimal('0'),
    })

    CODIGOS_ASISTENCIA = {'A', 'NOR', 'CHE', 'TE', 'FL', 'DM', 'VAC', 'SS', 'DL'}
    CODIGOS_FALTA = {'F'}
    CODIGOS_HABILES = {'A', 'NOR', 'CHE', 'F', 'TE', 'DM'}  # Días que cuentan como hábiles

    faltas_detalle = []

    for reg in registros:
        dni = reg.dni
        emp = empleados[dni]
        if not emp['apellidos_nombres']:
            if reg.personal:
                emp['apellidos_nombres'] = reg.personal.apellidos_nombres
                emp['area'] = (
                    reg.personal.subarea.area.nombre
                    if reg.personal.subarea and reg.personal.subarea.area
                    else ''
                )
            else:
                emp['apellidos_nombres'] = reg.nombre_archivo or dni
                emp['area'] = ''
        emp['grupo'] = reg.grupo

        codigo = (reg.codigo_dia or '').strip().upper()

        if codigo in CODIGOS_HABILES:
            emp['dias_habiles'] += 1
        if codigo in CODIGOS_ASISTENCIA:
            emp['dias_asistio'] += 1
        if codigo in CODIGOS_FALTA:
            emp['dias_falta'] += 1
            faltas_detalle.append({
                'fecha': reg.fecha,
                'dni': dni,
                'empleado': emp['apellidos_nombres'],
                'area': emp['area'],
                'grupo': reg.grupo,
            })

        emp['he_25']  += reg.he_25 or Decimal('0')
        emp['he_35']  += reg.he_35 or Decimal('0')
        emp['he_100'] += reg.he_100 or Decimal('0')

    # ── Generar Excel ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    MESES_ESP = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
    }

    # ── HOJA 1: Asistencia ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Asistencia"

    # Título del período
    titulo_periodo = f"Asistencia Mensual — {MESES_ESP[mes]} {anio}"
    ws1.cell(row=1, column=1, value=titulo_periodo).font = Font(bold=True, size=12, color="0D2B27")
    ws1.cell(row=1, column=1).fill = PatternFill("solid", fgColor="CCFBF1")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws1.row_dimensions[1].height = 25

    headers = [
        'DNI', 'Apellidos y Nombres', 'Área', 'Grupo',
        'Días Hábiles', 'Días Asistió', 'Días Falta', '% Asistencia',
        'HE 25% (hrs)', 'HE 35% (hrs)', 'HE 100% (hrs)',
    ]
    _header_row(ws1, 2, headers)
    ws1.freeze_panes = 'A3'

    sorted_emps = sorted(empleados.items(), key=lambda x: x[1]['apellidos_nombres'])
    for idx, (dni, emp) in enumerate(sorted_emps):
        row_num = idx + 3
        is_even = idx % 2 == 0
        habiles = emp['dias_habiles']
        asistio = emp['dias_asistio']
        pct = round(asistio / habiles * 100, 1) if habiles else 0

        datos = [
            dni,
            emp['apellidos_nombres'],
            emp['area'],
            emp['grupo'],
            habiles,
            asistio,
            emp['dias_falta'],
            f'{pct}%',
            float(emp['he_25']),
            float(emp['he_35']),
            float(emp['he_100']),
        ]
        for col_num, value in enumerate(datos, 1):
            align = 'center' if col_num in (4, 5, 6, 7, 8) else 'left'
            fmt = '#,##0.00' if col_num in (9, 10, 11) else None
            _data_cell(ws1, row_num, col_num, value, is_even, fmt, align)

    _auto_width(ws1)

    # ── HOJA 2: Faltas Detalle ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Faltas Detalle")
    ws2.cell(row=1, column=1, value=f"Detalle de Faltas — {MESES_ESP[mes]} {anio}").font = Font(bold=True, size=12, color="0D2B27")
    ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor="CCFBF1")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    _header_row(ws2, 2, ['Fecha', 'DNI', 'Empleado', 'Área', 'Grupo'])
    ws2.freeze_panes = 'A3'

    faltas_ord = sorted(faltas_detalle, key=lambda x: (x['fecha'], x['empleado']))
    for idx, falta in enumerate(faltas_ord):
        row_num = idx + 3
        is_even = idx % 2 == 0
        _data_cell(ws2, row_num, 1, falta['fecha'], is_even, 'DD/MM/YYYY')
        _data_cell(ws2, row_num, 2, falta['dni'], is_even, align='center')
        _data_cell(ws2, row_num, 3, falta['empleado'], is_even)
        _data_cell(ws2, row_num, 4, falta['area'], is_even)
        _data_cell(ws2, row_num, 5, falta['grupo'], is_even, align='center')

    _auto_width(ws2)

    # ── Respuesta HTTP ─────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"asistencia_{anio}{mes:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# REPORTE 3: HE DETALLADO
# ─────────────────────────────────────────────────────────────────────────────

@solo_admin
def reporte_he_detallado(request):
    """
    Reporte detallado de horas extras con costo por empleado.
    Filtros: año, mes, grupo (STAFF/RCO/todos).
    Descarga si ?format=excel.

    Costo HE:
      Valor hora = sueldo_base / 30 / 8
      HE 25%:  valor_hora * 1.25
      HE 35%:  valor_hora * 1.35
      HE 100%: valor_hora * 2.00
    """
    from asistencia.models import RegistroTareo

    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes  = int(request.GET.get('mes', hoy.month))
    grupo_filtro = request.GET.get('grupo', '')

    if request.GET.get('format') != 'excel':
        areas = Area.objects.filter(activa=True).order_by('nombre')
        return render(request, 'personal/reportes_panel.html', {
            'titulo': 'Reportes RRHH',
            'areas': areas,
            'anio_actual': hoy.year,
            'mes_actual': hoy.month,
            'meses': [
                (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
                (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
                (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre'),
            ],
            'anios': list(range(hoy.year - 3, hoy.year + 1)),
        })

    # ── Queryset ──────────────────────────────────────────────────────────────
    qs = (
        RegistroTareo.objects
        .filter(fecha__year=anio, fecha__month=mes)
        .filter(Q(he_25__gt=0) | Q(he_35__gt=0) | Q(he_100__gt=0))
        .select_related('personal', 'personal__subarea', 'personal__subarea__area')
    )
    if grupo_filtro:
        qs = qs.filter(grupo=grupo_filtro)

    # ── Agregar por empleado ───────────────────────────────────────────────────
    from collections import defaultdict
    empleados = defaultdict(lambda: {
        'apellidos_nombres': '',
        'area': '',
        'cargo': '',
        'grupo': '',
        'sueldo_base': Decimal('0'),
        'he_25': Decimal('0'),
        'he_35': Decimal('0'),
        'he_100': Decimal('0'),
    })

    # Detalle por día
    he_por_dia = []

    for reg in qs:
        dni = reg.dni
        emp = empleados[dni]
        if not emp['apellidos_nombres']:
            if reg.personal:
                emp['apellidos_nombres'] = reg.personal.apellidos_nombres
                emp['area'] = (
                    reg.personal.subarea.area.nombre
                    if reg.personal.subarea and reg.personal.subarea.area
                    else ''
                )
                emp['cargo'] = reg.personal.cargo
                emp['sueldo_base'] = reg.personal.sueldo_base or Decimal('0')
            else:
                emp['apellidos_nombres'] = reg.nombre_archivo or dni
        emp['grupo'] = reg.grupo
        emp['he_25']  += reg.he_25 or Decimal('0')
        emp['he_35']  += reg.he_35 or Decimal('0')
        emp['he_100'] += reg.he_100 or Decimal('0')

        if (reg.he_25 or 0) + (reg.he_35 or 0) + (reg.he_100 or 0) > 0:
            he_por_dia.append({
                'fecha': reg.fecha,
                'dni': dni,
                'empleado': emp['apellidos_nombres'],
                'he_25': reg.he_25 or Decimal('0'),
                'he_35': reg.he_35 or Decimal('0'),
                'he_100': reg.he_100 or Decimal('0'),
            })

    # ── Función de costo ──────────────────────────────────────────────────────
    def costo_he(sueldo, horas, multiplier):
        if not sueldo or not horas:
            return Decimal('0')
        valor_hora = sueldo / 30 / 8
        return round(valor_hora * horas * multiplier, 2)

    MULT_25  = Decimal('1.25')
    MULT_35  = Decimal('1.35')
    MULT_100 = Decimal('2.00')

    # ── Generar Excel ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    MESES_ESP = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
    }

    # ── HOJA 1: HE por Empleado ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "HE por Empleado"

    titulo_periodo = f"Horas Extras — {MESES_ESP[mes]} {anio}{' (' + grupo_filtro + ')' if grupo_filtro else ''}"
    ws1.cell(row=1, column=1, value=titulo_periodo).font = Font(bold=True, size=12, color="0D2B27")
    ws1.cell(row=1, column=1).fill = PatternFill("solid", fgColor="CCFBF1")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws1.row_dimensions[1].height = 25

    headers = [
        'DNI', 'Empleado', 'Área', 'Cargo',
        'HE 25% (hrs)', 'HE 25% (S/)',
        'HE 35% (hrs)', 'HE 35% (S/)',
        'HE 100% (hrs)', 'HE 100% (S/)',
        'Total HE (hrs)', 'Total HE (S/)',
    ]
    _header_row(ws1, 2, headers)
    ws1.freeze_panes = 'A3'

    sorted_emps = sorted(empleados.items(), key=lambda x: x[1]['apellidos_nombres'])

    total_he_25_hrs  = Decimal('0')
    total_he_35_hrs  = Decimal('0')
    total_he_100_hrs = Decimal('0')
    total_he_25_s    = Decimal('0')
    total_he_35_s    = Decimal('0')
    total_he_100_s   = Decimal('0')

    for idx, (dni, emp) in enumerate(sorted_emps):
        row_num = idx + 3
        is_even = idx % 2 == 0
        sueldo = emp['sueldo_base']

        c25  = costo_he(sueldo, emp['he_25'],  MULT_25)
        c35  = costo_he(sueldo, emp['he_35'],  MULT_35)
        c100 = costo_he(sueldo, emp['he_100'], MULT_100)
        total_hrs = emp['he_25'] + emp['he_35'] + emp['he_100']
        total_s   = c25 + c35 + c100

        total_he_25_hrs  += emp['he_25']
        total_he_35_hrs  += emp['he_35']
        total_he_100_hrs += emp['he_100']
        total_he_25_s    += c25
        total_he_35_s    += c35
        total_he_100_s   += c100

        datos = [
            dni,
            emp['apellidos_nombres'],
            emp['area'],
            emp['cargo'],
            float(emp['he_25']),
            float(c25),
            float(emp['he_35']),
            float(c35),
            float(emp['he_100']),
            float(c100),
            float(total_hrs),
            float(total_s),
        ]
        for col_num, value in enumerate(datos, 1):
            fmt = '#,##0.00' if col_num in (5, 6, 7, 8, 9, 10, 11, 12) else None
            align = 'right' if col_num > 4 else 'left'
            _data_cell(ws1, row_num, col_num, value, is_even, fmt, align)

    # Fila de gran total
    total_row = len(sorted_emps) + 3
    ws1.cell(row=total_row, column=4, value='GRAN TOTAL').font = TOTAL_FONT
    ws1.cell(row=total_row, column=4).fill = TOTAL_FILL
    ws1.cell(row=total_row, column=4).border = THIN_BORDER
    ws1.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')

    total_vals = [
        (5, float(total_he_25_hrs)),
        (6, float(total_he_25_s)),
        (7, float(total_he_35_hrs)),
        (8, float(total_he_35_s)),
        (9, float(total_he_100_hrs)),
        (10, float(total_he_100_s)),
        (11, float(total_he_25_hrs + total_he_35_hrs + total_he_100_hrs)),
        (12, float(total_he_25_s + total_he_35_s + total_he_100_s)),
    ]
    for col_num, val in total_vals:
        c = ws1.cell(row=total_row, column=col_num, value=val)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

    _auto_width(ws1)

    # ── HOJA 2: HE por Día ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("HE por Día")
    ws2.cell(row=1, column=1, value=f"HE por Día — {MESES_ESP[mes]} {anio}").font = Font(bold=True, size=12, color="0D2B27")
    ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor="CCFBF1")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    _header_row(ws2, 2, ['Fecha', 'DNI', 'Empleado', 'HE 25% (hrs)', 'HE 35% (hrs)', 'HE 100% (hrs)'])
    ws2.freeze_panes = 'A3'

    he_dia_ord = sorted(he_por_dia, key=lambda x: (x['fecha'], x['empleado']))
    for idx, item in enumerate(he_dia_ord):
        row_num = idx + 3
        is_even = idx % 2 == 0
        _data_cell(ws2, row_num, 1, item['fecha'], is_even, 'DD/MM/YYYY')
        _data_cell(ws2, row_num, 2, item['dni'], is_even, align='center')
        _data_cell(ws2, row_num, 3, item['empleado'], is_even)
        _data_cell(ws2, row_num, 4, float(item['he_25']),  is_even, '#,##0.00', 'right')
        _data_cell(ws2, row_num, 5, float(item['he_35']),  is_even, '#,##0.00', 'right')
        _data_cell(ws2, row_num, 6, float(item['he_100']), is_even, '#,##0.00', 'right')

    _auto_width(ws2)

    # ── Respuesta HTTP ─────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"he_detallado_{anio}{mes:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# REPORTE 4: VACACIONES Y PERMISOS
# ─────────────────────────────────────────────────────────────────────────────

@solo_admin
def reporte_vacaciones(request):
    """Reporte Excel de vacaciones y permisos del período seleccionado."""
    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes  = int(request.GET.get('mes', hoy.month))

    if request.GET.get('format') != 'excel':
        areas = Area.objects.filter(activa=True).order_by('nombre')
        return render(request, 'personal/reportes_panel.html', {
            'titulo': 'Reportes RRHH',
            'areas': areas,
            'anio_actual': hoy.year, 'mes_actual': hoy.month,
            'meses': [(1,'Enero'),(2,'Febrero'),(3,'Marzo'),(4,'Abril'),
                      (5,'Mayo'),(6,'Junio'),(7,'Julio'),(8,'Agosto'),
                      (9,'Septiembre'),(10,'Octubre'),(11,'Noviembre'),(12,'Diciembre')],
            'anios': list(range(hoy.year - 3, hoy.year + 1)),
        })

    wb = openpyxl.Workbook()
    # ── HOJA 1: Vacaciones ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Vacaciones'
    MESES_ES = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    titulo_str = f'Reporte de Vacaciones — {MESES_ES[mes]} {anio}'
    ws['A1'] = titulo_str
    ws['A1'].font = Font(size=13, bold=True, color='0D2B27')
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    headers = ['N°', 'Empleado', 'DNI', 'Área', 'Sub-Área', 'Grupo',
               'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado']
    _header_row(ws, 2, headers)

    try:
        from vacaciones.models import SolicitudVacacion
        qs_vac = SolicitudVacacion.objects.filter(
            fecha_inicio__year=anio, fecha_inicio__month=mes
        ).select_related('personal', 'personal__subarea', 'personal__subarea__area').order_by(
            'personal__subarea__area__nombre', 'personal__apellidos_nombres'
        )
        for i, v in enumerate(qs_vac, 1):
            even = (i % 2 == 0)
            p = v.personal
            area_nombre = (p.subarea.area.nombre if p and p.subarea and p.subarea.area else '—')
            sub_nombre  = (p.subarea.nombre if p and p.subarea else '—')
            row = i + 2
            _data_cell(ws, row, 1, i, even, align='center')
            _data_cell(ws, row, 2, p.apellidos_nombres if p else '—', even)
            _data_cell(ws, row, 3, p.nro_doc if p else '—', even, number_format='@', align='center')
            _data_cell(ws, row, 4, area_nombre, even)
            _data_cell(ws, row, 5, sub_nombre, even)
            _data_cell(ws, row, 6, p.grupo_tareo if p else '—', even, align='center')
            _data_cell(ws, row, 7, v.fecha_inicio, even, align='center')
            _data_cell(ws, row, 8, v.fecha_fin, even, align='center')
            _data_cell(ws, row, 9, getattr(v, 'dias_habiles', None) or getattr(v, 'dias_calendario', '—'), even, align='center')
            _data_cell(ws, row, 10, v.estado, even, align='center')
    except Exception:
        ws['A3'] = 'Módulo de vacaciones no disponible'

    _auto_width(ws)

    # ── HOJA 2: Permisos ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Permisos')
    ws2['A1'] = f'Reporte de Permisos — {MESES_ES[mes]} {anio}'
    ws2['A1'].font = Font(size=13, bold=True, color='0D2B27')
    ws2.merge_cells('A1:I1')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 28

    headers2 = ['N°', 'Empleado', 'DNI', 'Área', 'Tipo Permiso', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado']
    _header_row(ws2, 2, headers2)

    try:
        from vacaciones.models import SolicitudPermiso
        qs_per = SolicitudPermiso.objects.filter(
            fecha_inicio__year=anio, fecha_inicio__month=mes
        ).select_related('personal', 'personal__subarea', 'personal__subarea__area').order_by(
            'personal__apellidos_nombres'
        )
        for i, perm in enumerate(qs_per, 1):
            even = (i % 2 == 0)
            p = perm.personal
            area_nombre = (p.subarea.area.nombre if p and p.subarea and p.subarea.area else '—')
            row = i + 2
            _data_cell(ws2, row, 1, i, even, align='center')
            _data_cell(ws2, row, 2, p.apellidos_nombres if p else '—', even)
            _data_cell(ws2, row, 3, p.nro_doc if p else '—', even, number_format='@', align='center')
            _data_cell(ws2, row, 4, area_nombre, even)
            _data_cell(ws2, row, 5, getattr(perm, 'get_tipo_permiso_display', lambda: perm.tipo_permiso)(), even)
            _data_cell(ws2, row, 6, perm.fecha_inicio, even, align='center')
            _data_cell(ws2, row, 7, perm.fecha_fin, even, align='center')
            _data_cell(ws2, row, 8, getattr(perm, 'dias_solicitados', '—'), even, align='center')
            _data_cell(ws2, row, 9, perm.estado, even, align='center')
    except Exception:
        ws2['A3'] = 'Módulo de permisos no disponible'

    _auto_width(ws2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vacaciones_permisos_{anio}{mes:02d}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# REPORTE 5: CONTRATOS POR VENCER
# ─────────────────────────────────────────────────────────────────────────────

@solo_admin
def reporte_contratos(request):
    """Reporte Excel de contratos próximos a vencer."""
    hoy = date.today()
    dias = int(request.GET.get('dias', 60))

    if request.GET.get('format') != 'excel':
        areas = Area.objects.filter(activa=True).order_by('nombre')
        return render(request, 'personal/reportes_panel.html', {
            'titulo': 'Reportes RRHH',
            'areas': areas,
            'anio_actual': hoy.year, 'mes_actual': hoy.month,
            'meses': [(1,'Enero'),(2,'Febrero'),(3,'Marzo'),(4,'Abril'),
                      (5,'Mayo'),(6,'Junio'),(7,'Julio'),(8,'Agosto'),
                      (9,'Septiembre'),(10,'Octubre'),(11,'Noviembre'),(12,'Diciembre')],
            'anios': list(range(hoy.year - 3, hoy.year + 1)),
        })

    limite = hoy + timedelta(days=dias)
    qs = Personal.objects.filter(
        estado='Activo',
        fecha_fin_contrato__isnull=False,
        fecha_fin_contrato__gte=hoy,
        fecha_fin_contrato__lte=limite,
    ).select_related('subarea', 'subarea__area').order_by('fecha_fin_contrato')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contratos por Vencer'
    ws['A1'] = f'Contratos por Vencer — Próximos {dias} días ({hoy.strftime("%d/%m/%Y")})'
    ws['A1'].font = Font(size=13, bold=True, color='0D2B27')
    ws.merge_cells('A1:K1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    headers = ['N°', 'Empleado', 'DNI', 'Cargo', 'Área', 'Grupo', 'Tipo Contrato',
               'Inicio', 'Vencimiento', 'Días Restantes', 'Sueldo Base']
    _header_row(ws, 2, headers)

    RED_FILL   = PatternFill("solid", fgColor="FEE2E2")
    AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")

    for i, p in enumerate(qs, 1):
        row = i + 2
        dias_rest = (p.fecha_fin_contrato - hoy).days
        even = (i % 2 == 0)
        area_n = (p.subarea.area.nombre if p.subarea and p.subarea.area else '—')

        # Color urgencia
        urgencia_fill = RED_FILL if dias_rest <= 15 else (AMBER_FILL if dias_rest <= 30 else (ALT_FILL if even else PatternFill()))

        for col, val in enumerate([
            i,
            p.apellidos_nombres,
            p.nro_doc,
            p.cargo or '—',
            area_n,
            p.grupo_tareo,
            p.get_tipo_contrato_display() if hasattr(p, 'get_tipo_contrato_display') else '—',
            p.fecha_inicio_contrato,
            p.fecha_fin_contrato,
            dias_rest,
            float(p.sueldo_base or 0),
        ], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = urgencia_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal='center' if col in (1,3,6,7,8,9,10) else 'left',
                vertical='center'
            )
            if col == 3:  # DNI/CE
                cell.number_format = '@'
            elif col == 11:
                cell.number_format = '"S/ "#,##0.00'

    _auto_width(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="contratos_por_vencer_{hoy.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
