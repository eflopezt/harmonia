"""
Integraciones Perú — Exportadores.

Funciones para generar archivos en los formatos requeridos por SUNAT,
PLAME, AFP Net y bancos peruanos.
"""
import csv
import io
from datetime import date
from decimal import Decimal

from django.utils import timezone


# ──────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────

def _safe(value, default=''):
    """Convierte None a string vacío."""
    if value is None:
        return default
    return str(value).strip()


def _fecha(d, fmt='%d/%m/%Y'):
    """Formatea una fecha o retorna vacío."""
    if not d:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime(fmt)
    return ''


def _monto(value):
    """Formatea monto decimal sin símbolo."""
    if not value:
        return '0.00'
    try:
        return f'{Decimal(str(value)):.2f}'
    except Exception:
        return '0.00'


# ──────────────────────────────────────────────────────────────────────
# T-REGISTRO (SUNAT)
# ──────────────────────────────────────────────────────────────────────

# Mapeo de tipo_trab → código T-Registro
TIPO_TRAB_TREG = {
    'Empleado': '1',
    'Obrero': '2',
}

# Mapeo de tipo_contrato → código T-Registro (modalidad formativa / contrato)
MODALIDAD_TREG = {
    'INDEFINIDO': '1',
    'PLAZO_FIJO': '2',
    'INICIO_ACTIVIDAD': '3',
    'NECESIDAD_MERCADO': '4',
    'RECONVERSION_EMPRESARIAL': '5',
    'OBRA_SERVICIO': '6',
    'DISCONTINUO': '7',
    'TEMPORADA': '8',
    'SUPLENCIA': '9',
    'EMERGENCIA': '10',
    'SNP': '20',
    'PRACTICANTE': '30',
    'OTRO': '99',
    '': '99',
}

# Mapeo régimen pensionario → código T-Registro
PENSION_TREG = {
    'AFP': '1',
    'ONP': '2',
    'SIN_PENSION': '3',
}


def generar_t_registro_altas(queryset):
    """
    Genera archivo T-Registro de altas para SUNAT.
    Formato: pipe (|) delimitado, sin encabezado.

    Campos requeridos (simplificados):
    1. Tipo doc trabajador (1=DNI, 4=CE, 7=PAS)
    2. Nro doc
    3. Tipo trab (1=Empleado, 2=Obrero)
    4. Fecha ingreso (DD/MM/YYYY)
    5. Régimen pensionario (1=AFP, 2=ONP, 3=Sin)
    6. Modalidad contrato
    7. Remuneración básica
    8. Periodo (YYYY-MM)
    """
    from io import StringIO
    output = StringIO()

    TIPO_DOC_MAP = {'DNI': '1', 'CE': '4', 'Pasaporte': '7'}

    hoy = date.today()

    for p in queryset:
        row = [
            TIPO_DOC_MAP.get(p.tipo_doc, '1'),
            _safe(p.nro_doc),
            TIPO_TRAB_TREG.get(p.tipo_trab, '1'),
            _fecha(p.fecha_alta),
            PENSION_TREG.get(p.regimen_pension, '3'),
            _safe(p.afp) if p.regimen_pension == 'AFP' else '',
            _safe(p.cuspp),
            MODALIDAD_TREG.get(p.tipo_contrato, '99'),
            _monto(p.sueldo_base),
            _safe(p.cargo),
            _safe(p.correo_corporativo or p.correo_personal),
            hoy.strftime('%Y%m'),
        ]
        output.write('|'.join(row) + '\n')

    return output.getvalue(), queryset.count()


def generar_t_registro_bajas(queryset):
    """Genera archivo T-Registro de bajas (ceses) para SUNAT."""
    from io import StringIO
    output = StringIO()

    TIPO_DOC_MAP = {'DNI': '1', 'CE': '4', 'Pasaporte': '7'}

    for p in queryset:
        row = [
            TIPO_DOC_MAP.get(p.tipo_doc, '1'),
            _safe(p.nro_doc),
            _fecha(p.fecha_cese),
            '1',  # motivo cese: 01 = despido/renuncia (simplificado)
            hoy_str := date.today().strftime('%Y%m'),
        ]
        output.write('|'.join(row) + '\n')

    return output.getvalue(), queryset.count()


def generar_planilla_excel(queryset, periodo_str=''):
    """
    Genera planilla resumen en CSV compatible con Excel.
    Columnas: apellidos_nombres, DNI, cargo, área, sueldo,
              AFP/ONP, descuento AFP/ONP, neto calculado.
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=',')

    # Encabezado
    writer.writerow([
        'N°', 'APELLIDOS Y NOMBRES', 'DNI/DOC', 'CARGO',
        'AREA', 'SUBAREA', 'GRUPO', 'REGIMEN',
        'AFP/ONP', 'SUELDO BASE',
        'APORTE PENSION (13%)', 'ESSALUD (9% emp.)',
        'ASIG. FAMILIAR', 'NETO ESTIMADO',
        'BANCO', 'CTA. AHORROS',
        'TIPO CONTRATO', 'FIN CONTRATO',
    ])

    AFP_RATE = Decimal('0.13')      # ~13% AFP (variable por fondo)
    ONP_RATE = Decimal('0.13')      # 13% ONP fijo
    ESSALUD_RATE = Decimal('0.09')  # 9% ESSALUD (empleador)
    RMV = Decimal('1130')           # RMV vigente DS 006-2024-TR (ene-2025)
    ASIG_FAM = RMV * Decimal('0.10')  # 10% RMV

    for i, p in enumerate(queryset, 1):
        sueldo = p.sueldo_base or Decimal('0')
        asig = ASIG_FAM if p.asignacion_familiar else Decimal('0')
        bruto = sueldo + asig

        if p.regimen_pension == 'AFP':
            descuento_pension = (bruto * AFP_RATE).quantize(Decimal('0.01'))
        elif p.regimen_pension == 'ONP':
            descuento_pension = (bruto * ONP_RATE).quantize(Decimal('0.01'))
        else:
            descuento_pension = Decimal('0')

        essalud = (bruto * ESSALUD_RATE).quantize(Decimal('0.01'))
        neto = (bruto - descuento_pension).quantize(Decimal('0.01'))

        writer.writerow([
            i,
            p.apellidos_nombres,
            p.nro_doc,
            p.cargo,
            p.subarea.area.nombre if p.subarea and p.subarea.area else '—',
            p.subarea.nombre if p.subarea else '—',
            p.grupo_tareo,
            p.get_regimen_pension_display(),
            p.afp if p.regimen_pension == 'AFP' else 'ONP',
            _monto(sueldo),
            _monto(descuento_pension),
            _monto(essalud),
            _monto(asig),
            _monto(neto),
            p.banco,
            p.cuenta_ahorros,
            p.get_tipo_contrato_display() if p.tipo_contrato else '—',
            _fecha(p.fecha_fin_contrato),
        ])

    return output.getvalue(), queryset.count()


# ──────────────────────────────────────────────────────────────────────
# AFP NET
# ──────────────────────────────────────────────────────────────────────

def generar_afp_net(queryset, periodo_str=''):
    """
    Genera archivo AFP Net para declarar aportes.
    Formato simplificado: pipe delimitado.
    AFP Net acepta: CUSPP | Apellidos | Nombres | Remuneración | Aporte | Período
    """
    output = io.StringIO()

    AFP_RATE = Decimal('0.13')

    for p in queryset.filter(regimen_pension='AFP'):
        sueldo = p.sueldo_base or Decimal('0')
        aporte = (sueldo * AFP_RATE).quantize(Decimal('0.01'))

        # Separar apellidos y nombres del campo unificado
        nombre_full = p.apellidos_nombres
        partes = nombre_full.split(',', 1)
        apellidos = partes[0].strip() if len(partes) > 1 else nombre_full
        nombres = partes[1].strip() if len(partes) > 1 else ''

        row = [
            _safe(p.cuspp),
            apellidos,
            nombres,
            p.nro_doc,
            _monto(sueldo),
            _monto(aporte),
            periodo_str or date.today().strftime('%Y%m'),
            _safe(p.afp),
        ]
        output.write('|'.join(row) + '\n')

    return output.getvalue(), queryset.filter(regimen_pension='AFP').count()


# ──────────────────────────────────────────────────────────────────────
# BANCOS — Archivo de pago masivo
# ──────────────────────────────────────────────────────────────────────

# Formato simplificado compatible con Telecrédito BCP y similares
# Columnas: cuenta_destino, monto, nombre_beneficiario, nro_doc

def generar_pago_banco(queryset, banco_filtro='', monto_campo='sueldo_base'):
    """
    Genera archivo de pago masivo bancario.
    Formato CSV compatible con Telecrédito BCP / BBVA Pagos.

    Args:
        banco_filtro: si se especifica, solo incluye empleados de ese banco
        monto_campo: campo del modelo Personal a usar como monto
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Encabezado
    writer.writerow([
        'CUENTA_DESTINO', 'CUENTA_CCI',
        'NOMBRE_BENEFICIARIO', 'DOC_IDENTIDAD',
        'MONTO', 'MONEDA', 'DESCRIPCION',
    ])

    if banco_filtro:
        qs = queryset.filter(banco=banco_filtro)
    else:
        qs = queryset

    count = 0
    for p in qs:
        sueldo = getattr(p, monto_campo, None) or Decimal('0')
        if sueldo <= 0:
            continue

        writer.writerow([
            p.cuenta_ahorros or '',
            p.cuenta_cci or '',
            p.apellidos_nombres,
            p.nro_doc,
            _monto(sueldo),
            'PEN',
            f'Remuneracion {date.today().strftime("%m/%Y")}',
        ])
        count += 1

    return output.getvalue(), count


# ──────────────────────────────────────────────────────────────────────
# ESSALUD
# ──────────────────────────────────────────────────────────────────────

def generar_essalud(queryset, periodo_str=''):
    """Genera declaración ESSALUD (nómina de asegurados)."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'DNI', 'APELLIDOS Y NOMBRES', 'REMUNERACION',
        'APORTE_ESSALUD_9%', 'PERIODO',
    ])

    ESSALUD_RATE = Decimal('0.09')
    count = 0

    for p in queryset:
        sueldo = p.sueldo_base or Decimal('0')
        essalud = (sueldo * ESSALUD_RATE).quantize(Decimal('0.01'))

        writer.writerow([
            p.nro_doc,
            p.apellidos_nombres,
            _monto(sueldo),
            _monto(essalud),
            periodo_str or date.today().strftime('%Y%m'),
        ])
        count += 1

    return output.getvalue(), count


# ——————————————————————————————————————————————————————————————————————
# BANCOS — Formatos específicos por banco
# ——————————————————————————————————————————————————————————————————————


def generar_bcp_telecredito(queryset, descripcion='REMUNERACION'):
    '''Formato BCP Telecredito - TXT pipe-delimitado.'''
    output = io.StringIO()
    count = 0
    qs = queryset.filter(banco__icontains='BCP', cuenta_ahorros__gt='')
    for p in qs:
        sueldo = getattr(p, 'neto_a_pagar', None) or getattr(p, 'sueldo_base', None) or Decimal('0')
        if sueldo <= 0:
            continue
        row = [
            '0',
            _safe(p.cuenta_cci or p.cuenta_ahorros),
            _monto(sueldo),
            'PEN',
            descripcion[:40],
            _safe(p.nro_doc),
            p.apellidos_nombres[:60],
        ]
        output.write('|'.join(row) + '\r\n')
        count += 1
    return output.getvalue(), count


def generar_bbva_net_cash(queryset, descripcion='REMUNERACION'):
    '''Formato BBVA Net Cash - TXT tab-delimitado (CCI|Nombre|Doc|Monto|Moneda|Concepto|Fecha).'''
    output = io.StringIO()
    count = 0
    qs = queryset.filter(banco__icontains='BBVA', cuenta_ahorros__gt='')
    for p in qs:
        sueldo = getattr(p, 'neto_a_pagar', None) or getattr(p, 'sueldo_base', None) or Decimal('0')
        if sueldo <= 0:
            continue
        row = [
            _safe(p.cuenta_cci or p.cuenta_ahorros),
            p.apellidos_nombres[:60],
            _safe(p.nro_doc),
            _monto(sueldo),
            'PEN',
            descripcion[:30],
            date.today().strftime('%d/%m/%Y'),
        ]
        output.write('	'.join(row) + '\r\n')

        count += 1
    return output.getvalue(), count


def generar_interbank_masivo(queryset, descripcion='REMUNERACION'):
    '''Formato Interbank Pagos Masivos - pipe delimitado (A|CCI|Nombre|Doc|Monto|Moneda|Ref).'''
    output = io.StringIO()
    count = 0
    qs = queryset.filter(banco__icontains='INTERBANK', cuenta_ahorros__gt='')
    for p in qs:
        sueldo = getattr(p, 'neto_a_pagar', None) or getattr(p, 'sueldo_base', None) or Decimal('0')
        if sueldo <= 0:
            continue
        row = [
            'A',
            _safe(p.cuenta_cci or p.cuenta_ahorros),
            p.apellidos_nombres[:60],
            _safe(p.nro_doc),
            _monto(sueldo),
            'PEN',
            descripcion[:20],
        ]
        output.write('|'.join(row) + '\r\n')
        count += 1
    return output.getvalue(), count


def generar_scotiabank_masivo(queryset, descripcion='REMUNERACION'):
    '''Formato Scotiabank Pago Masivo - CSV con encabezado obligatorio.'''
    output = io.StringIO()
    writer = csv.writer(output)
    count = 0
    qs = queryset.filter(banco__icontains='SCOTIABANK', cuenta_ahorros__gt='')
    writer.writerow([
        'TIPO_TRANSACCION', 'CUENTA_DESTINO', 'CCI_DESTINO',
        'NOMBRE_BENEFICIARIO', 'DOCUMENTO_IDENTIDAD',
        'MONTO', 'MONEDA', 'DESCRIPCION', 'FECHA_PROCESO',
    ])
    for p in qs:
        sueldo = getattr(p, 'neto_a_pagar', None) or getattr(p, 'sueldo_base', None) or Decimal('0')
        if sueldo <= 0:
            continue
        writer.writerow([
            '01',
            _safe(p.cuenta_ahorros),
            _safe(p.cuenta_cci or ''),
            p.apellidos_nombres[:60],
            _safe(p.nro_doc),
            _monto(sueldo),
            'PEN',
            descripcion[:30],
            date.today().strftime('%d/%m/%Y'),
        ])
        count += 1
    return output.getvalue(), count


def generar_banco_nacion(queryset, descripcion='REMUNERACION'):
    '''Formato Banco de la Nacion - TXT coma-delimitado (DNI|Nombre|Cuenta|Monto|Glosa|Fecha).'''
    output = io.StringIO()
    count = 0
    qs = queryset.filter(banco__icontains='NACION', cuenta_ahorros__gt='')
    for p in qs:
        sueldo = getattr(p, 'neto_a_pagar', None) or getattr(p, 'sueldo_base', None) or Decimal('0')
        if sueldo <= 0:
            continue
        row = [
            _safe(p.nro_doc),
            p.apellidos_nombres[:60],
            _safe(p.cuenta_ahorros),
            _monto(sueldo),
            descripcion[:30],
            date.today().strftime('%d%m%Y'),
        ]
        output.write(','.join(row) + '\r\n')
        count += 1
    return output.getvalue(), count


# ——————————————————————————————————————————————————————————————————————
# PLAME — Planilla Mensual de Pagos (PDT 601 / T-PLAME)
# ——————————————————————————————————————————————————————————————————————


def generar_plame(queryset_personal, queryset_nomina=None, periodo_str=''):
    '''Genera TXT para importar en PDT 601 (T-PLAME / PLAME).

    Registro 06 por trabajador (23 campos pipe-delimitados):
    06|tipo_doc|nro_doc|ap_paterno|ap_materno|nombres|nacimiento|sexo|
    modalidad|ingreso|rem_comp|base_essalud|essalud|dias|horas|
    pension|cuspp|afp|aporte_afp|aporte_onp|ir_5ta|neto|periodo
    '''
    output = io.StringIO()
    hoy = date.today()
    periodo = periodo_str or hoy.strftime('%Y%m')

    TIPO_DOC_MAP = {'DNI': '1', 'CE': '4', 'Pasaporte': '7'}
    PENSION_MAP  = {'AFP': '1', 'ONP': '2', 'SIN_PENSION': '3'}
    MODALIDAD_MAP = {
        'INDEFINIDO': '1', 'PLAZO_FIJO': '2', 'INICIO_ACTIVIDAD': '3',
        'NECESIDAD_MERCADO': '4', 'OBRA_SERVICIO': '6', 'SNP': '20',
        'PRACTICANTE': '30', '': '1',
    }

    AFP_RATE     = Decimal('0.10')
    ESSALUD_RATE = Decimal('0.09')
    ONP_RATE     = Decimal('0.13')
    ASIG_FAM_MON = Decimal('102.50')

    count = 0
    for p in queryset_personal:
        sueldo   = p.sueldo_base or Decimal('0')
        asig     = ASIG_FAM_MON if p.asignacion_familiar else Decimal('0')
        rem_comp = (sueldo + asig).quantize(Decimal('0.01'))
        essalud  = (rem_comp * ESSALUD_RATE).quantize(Decimal('0.01'))

        if p.regimen_pension == 'AFP':
            aporte_afp = (rem_comp * AFP_RATE).quantize(Decimal('0.01'))
            aporte_onp = Decimal('0')
        elif p.regimen_pension == 'ONP':
            aporte_afp = Decimal('0')
            aporte_onp = (rem_comp * ONP_RATE).quantize(Decimal('0.01'))
        else:
            aporte_afp = Decimal('0')
            aporte_onp = Decimal('0')

        ir_5ta = Decimal('0')
        neto   = rem_comp - aporte_afp - aporte_onp

        if queryset_nomina:
            try:
                from django.db.models import Sum
                reg    = queryset_nomina.get(personal=p)
                neto   = reg.neto_a_pagar or neto
                ir_5ta = reg.lineas.filter(concepto__formula='IR_5TA').aggregate(
                    s=Sum('monto')
                )['s'] or Decimal('0')
            except Exception:
                pass

        partes      = p.apellidos_nombres.split(',', 1)
        ap_completo = partes[0].strip() if len(partes) > 1 else p.apellidos_nombres
        nombres     = partes[1].strip() if len(partes) > 1 else ''
        ap_partes   = ap_completo.split(' ', 1)
        ap_paterno  = ap_partes[0] if ap_partes else ap_completo
        ap_materno  = ap_partes[1] if len(ap_partes) > 1 else ''

        row = [
            '06',
            TIPO_DOC_MAP.get(p.tipo_doc, '1'),
            _safe(p.nro_doc),
            ap_paterno[:40],
            ap_materno[:40],
            nombres[:60],
            _fecha(p.fecha_nacimiento, '%d%m%Y') if hasattr(p, 'fecha_nacimiento') else '',
            _safe(getattr(p, 'sexo', 'M')),
            MODALIDAD_MAP.get(p.tipo_contrato or '', '1'),
            _fecha(p.fecha_alta, '%d%m%Y'),
            _monto(rem_comp),
            _monto(rem_comp),
            _monto(essalud),
            '30',
            '48',
            PENSION_MAP.get(p.regimen_pension, '3'),
            _safe(p.cuspp) if p.regimen_pension == 'AFP' else '',
            _safe(p.afp) if p.regimen_pension == 'AFP' else '',
            _monto(aporte_afp),
            _monto(aporte_onp),
            _monto(ir_5ta),
            _monto(neto),
            periodo,
        ]
        output.write('|'.join(row) + '\n')
        count += 1

    return output.getvalue(), count


# ──────────────────────────────────────────────────────────────────────
# CTS — DEPÓSITO BANCARIO (Mayo y Noviembre)
# ──────────────────────────────────────────────────────────────────────

def generar_cts_banco_excel(queryset_personal, montos_cts, banco_filtro=None):
    """
    Genera Excel para depósito de CTS en cuentas CTS de cada trabajador.

    Args:
        queryset_personal: QuerySet de Personal (activos)
        montos_cts: dict {personal_id: Decimal(monto_cts)}
        banco_filtro: 'BCP', 'BBVA', etc. o None para todos

    Returns:
        (bytes_excel, count, total_monto)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CTS Deposito'

    headers = ['N', 'DNI', 'NOMBRES', 'BANCO CTS', 'CUENTA CTS', 'MONTO CTS', 'MONEDA']
    hdr_fill = PatternFill('solid', fgColor='0F766E')
    hdr_font = Font(bold=True, color='FFFFFF')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font

    qs = queryset_personal.filter(cuenta_cts__gt='').exclude(cuenta_cts__isnull=True)
    if banco_filtro:
        qs = qs.filter(banco__icontains=banco_filtro)

    count = 0
    total = Decimal('0')
    for p in qs.order_by('apellidos_nombres'):
        monto = montos_cts.get(p.pk, Decimal('0'))
        if monto <= 0:
            continue
        count += 1
        total += monto
        ws.cell(row=count+1, column=1, value=count)
        ws.cell(row=count+1, column=2, value=p.nro_doc)
        ws.cell(row=count+1, column=3, value=p.apellidos_nombres)
        ws.cell(row=count+1, column=4, value=p.banco or '')
        ws.cell(row=count+1, column=5, value=p.cuenta_cts)
        ws.cell(row=count+1, column=6, value=float(monto))
        ws.cell(row=count+1, column=6).number_format = '#,##0.00'
        ws.cell(row=count+1, column=7, value='PEN')

    ws.cell(row=count+2, column=5, value='TOTAL').font = Font(bold=True)
    ws.cell(row=count+2, column=6, value=float(total))
    ws.cell(row=count+2, column=6).font = Font(bold=True)
    ws.cell(row=count+2, column=6).number_format = '#,##0.00'

    for c in range(1, 8):
        ws.column_dimensions[chr(64+c)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), count, float(total)

