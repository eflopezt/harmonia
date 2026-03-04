"""
Integraciones Contables - Generadores de asientos de planilla.

Formatos soportados:
  - Concar  (CSV - el mas usado en Peru para medianas empresas)
  - SIGO    (TXT pipe-delimitado - empresas publicas y mineras)
  - SAP/Generico (Excel - compatible con SAP FICO y otros ERPs contables)
  - SIRE SUNAT (PLE - Libros Electronicos Libro Diario, Registro 5.1)

Plan de cuentas PCGE 2020 (simplificado - configurable por empresa):
  6211  Sueldos y salarios
  6215  Gratificaciones (provision)
  6271  EsSalud (empleador)
  4111  Remuneraciones por pagar
  4031  AFP por pagar
  4011  ONP (SCN) por pagar
  4032  EsSalud por pagar
"""
import csv
import io
from datetime import date
from decimal import Decimal

# Plan de cuentas default (PCGE 2020)

PLAN_CUENTAS_DEFAULT = {
    "sueldos_debe":        ("6211", "Sueldos y salarios"),
    "gratif_debe":         ("6215", "Gratificaciones (provision)"),
    "essalud_debe":        ("6271", "EsSalud - aporte empleador"),
    "rem_pagar_haber":     ("4111", "Remuneraciones por pagar"),
    "afp_pagar_haber":     ("4031", "AFP por pagar"),
    "onp_pagar_haber":     ("4011", "ONP por pagar"),
    "essalud_pagar_haber": ("4032", "EsSalud por pagar"),
}


def _get_plan_cuentas():
    """Obtiene el plan de cuentas desde ConfiguracionSistema o usa el default."""
    try:
        from core.models import ConfiguracionSistema
        cfg = ConfiguracionSistema.objects.first()
        if cfg and hasattr(cfg, "plan_cuentas") and cfg.plan_cuentas:
            import json
            return {**PLAN_CUENTAS_DEFAULT, **json.loads(cfg.plan_cuentas)}
    except Exception:
        pass
    return PLAN_CUENTAS_DEFAULT


def _get_totales_periodo(periodo):
    """
    Calcula los totales del periodo desagregados por tipo de descuento/ingreso.
    Retorna dict: bruto, neto, afp, onp, essalud, gratif_prov.

    Prioriza los totales precalculados en PeriodoNomina.total_bruto / total_neto.
    Si estan en cero, agrega desde RegistroNomina y LineaNomina.
    """
    from django.db.models import Sum
    from nominas.models import RegistroNomina, LineaNomina

    bruto_pre = periodo.total_bruto or Decimal("0")
    neto_pre  = periodo.total_neto  or Decimal("0")

    if bruto_pre > 0:
        # Periodo ya calculado: usar totales del encabezado
        agg = RegistroNomina.objects.filter(periodo=periodo).aggregate(
            essalud=Sum("aporte_essalud"),
        )
        essalud = (agg["essalud"] or Decimal("0")).quantize(Decimal("0.01"))

        afp_total = LineaNomina.objects.filter(
            registro__periodo=periodo,
            concepto__formula__in=["AFP_APORTE", "AFP_COMISION", "AFP_SEGURO"],
        ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

        onp_total = LineaNomina.objects.filter(
            registro__periodo=periodo,
            concepto__formula="ONP",
        ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

        gratif_prov = (bruto_pre / Decimal("6")).quantize(Decimal("0.01"))

        return {
            "bruto":       bruto_pre.quantize(Decimal("0.01")),
            "neto":        neto_pre.quantize(Decimal("0.01")),
            "essalud":     essalud,
            "afp":         afp_total.quantize(Decimal("0.01")),
            "onp":         onp_total.quantize(Decimal("0.01")),
            "gratif_prov": gratif_prov,
        }

    # Periodo sin totales precalculados: agregar desde registros individuales
    registros = RegistroNomina.objects.filter(periodo=periodo)
    agg = registros.aggregate(
        bruto=Sum("total_ingresos"),
        neto=Sum("neto_a_pagar"),
        essalud=Sum("aporte_essalud"),
    )

    bruto   = agg["bruto"]   or Decimal("0")
    neto    = agg["neto"]    or Decimal("0")
    essalud = agg["essalud"] or Decimal("0")

    afp_total = LineaNomina.objects.filter(
        registro__periodo=periodo,
        concepto__formula__in=["AFP_APORTE", "AFP_COMISION", "AFP_SEGURO"],
    ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

    onp_total = LineaNomina.objects.filter(
        registro__periodo=periodo,
        concepto__formula="ONP",
    ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

    gratif_prov = (bruto / Decimal("6")).quantize(Decimal("0.01"))

    return {
        "bruto":       bruto.quantize(Decimal("0.01")),
        "neto":        neto.quantize(Decimal("0.01")),
        "essalud":     essalud.quantize(Decimal("0.01")),
        "afp":         afp_total.quantize(Decimal("0.01")),
        "onp":         onp_total.quantize(Decimal("0.01")),
        "gratif_prov": gratif_prov,
    }


def _monto_str(value):
    """Formatea Decimal con 2 decimales sin separador de miles."""
    if not value:
        return "0.00"
    return f"{Decimal(str(value)):.2f}"


# ====================================================================
# CONCAR - Formato CSV
# ====================================================================

def generar_asiento_concar(periodo):
    """
    Genera el asiento contable de planilla en formato CONCAR (CSV).

    Columnas Concar (SubDiario 01 = Libro Diario, TipoMov D=Debe H=Haber):
    SubDiario | NroAsiento | FechaAsiento | CodCuenta | CodCentroCosto |
    TipoDocumento | NroDocumento | FecRef1 | CodDocRef | NroDocRef |
    TipoMov | Glosa | TipoNota | MontoMN | MontoME | TipoCambio |
    Pendiente | Marcador | FecVcto | CodMoneda
    """
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "SubDiario", "NroAsiento", "FechaAsiento", "CodCuenta", "CodCentroCosto",
        "TipoDocumento", "NroDocumento", "FecRef1", "CodDocRef", "NroDocRef",
        "TipoMov", "Glosa", "TipoNota", "MontoMN", "MontoME", "TipoCambio",
        "Pendiente", "Marcador", "FecVcto", "CodMoneda",
    ])

    plan        = _get_plan_cuentas()
    tots        = _get_totales_periodo(periodo)
    fecha       = (periodo.fecha_fin or date.today()).strftime("%d/%m/%Y")
    nro_asiento = f"P{periodo.anio}{periodo.mes:02d}001"
    glosa       = f"Planilla {periodo.mes_nombre} {periodo.anio}"

    def row(cuenta_key, tipo_mov, monto):
        cod, _nombre = plan.get(cuenta_key, ("0000", ""))
        return [
            "01", nro_asiento, fecha, cod, "",
            "PL", nro_asiento, fecha, "", "",
            tipo_mov, glosa, "", _monto_str(monto), "0.00", "1.000",
            "N", "", "", "MN",
        ]

    # DEBE
    writer.writerow(row("sueldos_debe", "D", tots["bruto"]))
    writer.writerow(row("essalud_debe", "D", tots["essalud"]))
    writer.writerow(row("gratif_debe",  "D", tots["gratif_prov"]))

    # HABER
    writer.writerow(row("rem_pagar_haber",     "H", tots["neto"]))
    if tots["afp"] > 0:
        writer.writerow(row("afp_pagar_haber", "H", tots["afp"]))
    if tots["onp"] > 0:
        writer.writerow(row("onp_pagar_haber", "H", tots["onp"]))
    writer.writerow(row("essalud_pagar_haber", "H", tots["essalud"]))

    # Cuadre: diferencia de gratif_prov -> cuenta 4151 Gratificaciones por pagar
    diferencia = (
        tots["bruto"] + tots["essalud"] + tots["gratif_prov"]
        - tots["neto"] - tots["afp"] - tots["onp"] - tots["essalud"]
    )
    if abs(diferencia) > Decimal("0.02"):
        writer.writerow([
            "01", nro_asiento, fecha, "4151", "",
            "PL", nro_asiento, fecha, "", "",
            "H", f"{glosa} - Provision gratificaciones", "",
            _monto_str(diferencia), "0.00", "1.000",
            "N", "", "", "MN",
        ])

    return output.getvalue(), 1


# ====================================================================
# SIGO - Formato TXT pipe-delimitado
# ====================================================================

def generar_asiento_sigo(periodo):
    """
    Genera el asiento contable en formato SIGO (Sistema Integrado de Gestion).
    Usado por empresas publicas y sector minero. Sin encabezado, pipe-delimitado.

    Campos por linea:
    TipoReg|CodEmpresa|Periodo|NumAsiento|Fecha|CodCuenta|Glosa|
    Debe|Haber|CentroCosto|TipoDoc|NroDoc|CodMoneda|TipoCambio|Estado
    """
    output      = io.StringIO()
    plan        = _get_plan_cuentas()
    tots        = _get_totales_periodo(periodo)
    fecha       = (periodo.fecha_fin or date.today()).strftime("%d%m%Y")
    periodo_str = f"{periodo.anio}{periodo.mes:02d}"
    num_asiento = f"PL{periodo_str}001"
    glosa       = f"PLANILLA {periodo.mes_nombre.upper()} {periodo.anio}"

    def linea(cuenta_key, debe, haber):
        cod, _ = plan.get(cuenta_key, ("0000", ""))
        return [
            "6", "01", periodo_str, num_asiento, fecha, cod, glosa,
            _monto_str(debe), _monto_str(haber),
            "001", "PL", num_asiento, "MN", "1.000", "C",
        ]

    filas = [
        linea("sueldos_debe",        tots["bruto"],       Decimal("0")),
        linea("essalud_debe",        tots["essalud"],     Decimal("0")),
        linea("gratif_debe",         tots["gratif_prov"], Decimal("0")),
        linea("rem_pagar_haber",     Decimal("0"),        tots["neto"]),
        linea("essalud_pagar_haber", Decimal("0"),        tots["essalud"]),
    ]
    if tots["afp"] > 0:
        filas.append(linea("afp_pagar_haber", Decimal("0"), tots["afp"]))
    if tots["onp"] > 0:
        filas.append(linea("onp_pagar_haber", Decimal("0"), tots["onp"]))

    for f in filas:
        output.write("|".join(f) + "\n")

    return output.getvalue(), 1


# ====================================================================
# SAP / GENERICO - Excel (openpyxl)
# ====================================================================

def generar_asiento_sap_excel(periodo):
    """
    Genera el asiento contable en formato Excel generico compatible
    con SAP FICO, Oracle y otros ERPs contables.

    Hoja 1: Asiento resumen por cuenta contable.
    Hoja 2: Detalle por trabajador con columnas de nomina.
    Retorna bytes (Excel .xlsx).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    plan  = _get_plan_cuentas()
    tots  = _get_totales_periodo(periodo)
    fecha = periodo.fecha_fin or date.today()

    wb = openpyxl.Workbook()

    # Hoja 1: Asiento resumen
    ws1       = wb.active
    ws1.title = "Asiento Planilla"

    TEAL     = "FF0D2B27"
    HDR_FONT = Font(color="FFFFFFFF", bold=True, size=10)
    HDR_FILL = PatternFill(fill_type="solid", fgColor=TEAL)

    headers = [
        "Cuenta", "Descripcion", "Centro Costo", "Debe (S/)", "Haber (S/)",
        "Glosa", "Fecha", "Tipo Doc", "Referencia",
    ]
    for col, h in enumerate(headers, 1):
        cell           = ws1.cell(row=1, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center")

    glosa = f"Planilla {periodo.mes_nombre} {periodo.anio}"
    ref   = f"PL{periodo.anio}{periodo.mes:02d}"

    filas_asiento = [
        (plan["sueldos_debe"][0],        plan["sueldos_debe"][1],        tots["bruto"],        Decimal("0")),
        (plan["essalud_debe"][0],        plan["essalud_debe"][1],        tots["essalud"],      Decimal("0")),
        (plan["gratif_debe"][0],         plan["gratif_debe"][1],         tots["gratif_prov"],  Decimal("0")),
        (plan["rem_pagar_haber"][0],     plan["rem_pagar_haber"][1],     Decimal("0"),         tots["neto"]),
        (plan["afp_pagar_haber"][0],     plan["afp_pagar_haber"][1],     Decimal("0"),         tots["afp"]),
        (plan["onp_pagar_haber"][0],     plan["onp_pagar_haber"][1],     Decimal("0"),         tots["onp"]),
        (plan["essalud_pagar_haber"][0], plan["essalud_pagar_haber"][1], Decimal("0"),         tots["essalud"]),
    ]

    for i, (cuenta, desc, debe, haber) in enumerate(filas_asiento, 2):
        ws1.cell(row=i, column=1, value=cuenta)
        ws1.cell(row=i, column=2, value=desc)
        ws1.cell(row=i, column=3, value="001")
        ws1.cell(row=i, column=4, value=float(debe))
        ws1.cell(row=i, column=5, value=float(haber))
        ws1.cell(row=i, column=6, value=glosa)
        ws1.cell(row=i, column=7, value=fecha)
        ws1.cell(row=i, column=8, value="PL")
        ws1.cell(row=i, column=9, value=ref)

    last = len(filas_asiento) + 2
    ws1.cell(row=last, column=3, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=last, column=4, value=f"=SUM(D2:D{last-1})").font = Font(bold=True)
    ws1.cell(row=last, column=5, value=f"=SUM(E2:E{last-1})").font = Font(bold=True)

    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 30
    ws1.column_dimensions["G"].width = 14

    # Hoja 2: Detalle por trabajador
    ws2 = wb.create_sheet("Detalle por Trabajador")
    headers2 = [
        "DNI", "Apellidos y Nombres", "Cargo", "Area", "Grupo",
        "Sueldo Base", "Total Ingresos", "AFP/ONP", "EsSalud",
        "Total Descuentos", "Neto a Pagar",
    ]
    for col, h in enumerate(headers2, 1):
        cell      = ws2.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    from nominas.models import RegistroNomina
    registros = RegistroNomina.objects.filter(
        periodo=periodo
    ).select_related("personal__subarea__area").order_by("personal__apellidos_nombres")

    for i, reg in enumerate(registros, 2):
        p = reg.personal
        # AFP/ONP estimado: total descuentos menos prestamos y descuentos manuales
        afp_onp = (
            float(reg.total_descuentos or 0)
            - float(reg.descuento_prestamo or 0)
            - float(reg.otros_descuentos  or 0)
        )
        ws2.cell(row=i, column=1,  value=p.nro_doc)
        ws2.cell(row=i, column=2,  value=p.apellidos_nombres)
        ws2.cell(row=i, column=3,  value=p.cargo or "")
        ws2.cell(row=i, column=4,  value=p.subarea.area.nombre if p.subarea and p.subarea.area else "")
        ws2.cell(row=i, column=5,  value=reg.grupo or getattr(p, "grupo_tareo", ""))
        ws2.cell(row=i, column=6,  value=float(reg.sueldo_base or 0))
        ws2.cell(row=i, column=7,  value=float(reg.total_ingresos or 0))
        ws2.cell(row=i, column=8,  value=max(afp_onp, 0))
        ws2.cell(row=i, column=9,  value=float(reg.aporte_essalud or 0))
        ws2.cell(row=i, column=10, value=float(reg.total_descuentos or 0))
        ws2.cell(row=i, column=11, value=float(reg.neto_a_pagar or 0))

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), registros.count()


# ====================================================================
# SIRE SUNAT - PLE Libro Diario (Registro 5.1)
# ====================================================================

def generar_sire_libro_diario(periodo):
    """
    Genera el archivo PLE del Libro Diario en formato SUNAT SIRE.

    Formato: TXT pipe-delimitado, Registro 5.1 del PLE.
    Campos (13 por linea):
    01 Periodo (AAAAMM00)
    02 CUO - Codigo Unico de Operacion
    03 Correlativo del asiento (3 digitos)
    04 Fecha de la operacion (DD/MM/AAAA)
    05 Glosa
    06 Referencia / numero de documento
    07 Codigo libro (05=Libro Diario)
    08 Cuenta contable PCGE
    09 Descripcion de la cuenta
    10 Tipo de moneda (1=PEN)
    11 Monto Debe en soles
    12 Monto Haber en soles
    13 Indicador estado (1=activo, 2=anulado, 9=cierre)
    """
    output      = io.StringIO()
    plan        = _get_plan_cuentas()
    tots        = _get_totales_periodo(periodo)
    fecha       = (periodo.fecha_fin or date.today()).strftime("%d/%m/%Y")
    periodo_ple = f"{periodo.anio}{periodo.mes:02d}00"
    cuo         = f"PL{periodo.anio}{periodo.mes:02d}001"
    glosa       = f"Planilla {periodo.mes_nombre} {periodo.anio}"

    def linea(cuenta_key, debe, haber, correlativo):
        cod, nombre = plan.get(cuenta_key, ("0000", ""))
        return [
            periodo_ple,
            cuo,
            str(correlativo).zfill(3),
            fecha,
            glosa,
            cuo,
            "05",
            cod,
            nombre,
            "1",
            _monto_str(debe),
            _monto_str(haber),
            "1",
        ]

    asientos = [
        ("sueldos_debe",        tots["bruto"],       Decimal("0")),
        ("essalud_debe",        tots["essalud"],     Decimal("0")),
        ("gratif_debe",         tots["gratif_prov"], Decimal("0")),
        ("rem_pagar_haber",     Decimal("0"),        tots["neto"]),
        ("essalud_pagar_haber", Decimal("0"),        tots["essalud"]),
    ]
    if tots["afp"] > 0:
        asientos.append(("afp_pagar_haber", Decimal("0"), tots["afp"]))
    if tots["onp"] > 0:
        asientos.append(("onp_pagar_haber", Decimal("0"), tots["onp"]))

    for i, (cuenta_key, debe, haber) in enumerate(asientos, 1):
        output.write("|".join(linea(cuenta_key, debe, haber, i)) + "\n")

    return output.getvalue(), 1
