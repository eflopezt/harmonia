"""
Vistas del módulo de Amonestaciones y Proceso Disciplinario.
"""
from datetime import date, timedelta
import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST

from personal.models import Personal, Area
from .models import TipoFalta, MedidaDisciplinaria, Descargo


# ──────────────────────────────────────────────────────────────
# FERIADOS BÁSICOS PERÚ — para cálculo días hábiles en timeline
# ──────────────────────────────────────────────────────────────

def _feriados_peru(anio):
    """Retorna un set con las fechas feriadas nacionales del año dado (Perú)."""
    return {
        date(anio, 1, 1),   # Año Nuevo
        date(anio, 5, 1),   # Día del Trabajo
        date(anio, 6, 29),  # SS. Pedro y Pablo
        date(anio, 7, 28),  # Fiestas Patrias
        date(anio, 7, 29),  # Fiestas Patrias
        date(anio, 8, 30),  # Sta. Rosa de Lima
        date(anio, 10, 8),  # Combate de Angamos
        date(anio, 11, 1),  # Todos los Santos
        date(anio, 12, 8),  # Inmaculada Concepción
        date(anio, 12, 25), # Navidad
    }


def _dias_habiles_entre(fecha_inicio, fecha_fin):
    """
    Cuenta días hábiles (lun-vie, excluyendo feriados nacionales Perú)
    entre dos fechas, inclusivo en inicio, exclusivo en fin.
    Retorna int (puede ser negativo si fin < inicio).
    """
    if fecha_fin <= fecha_inicio:
        # Contar hacia atrás para detectar vencido
        delta = (fecha_fin - fecha_inicio).days
        return delta  # negativo o cero

    feriados = set()
    for y in range(fecha_inicio.year, fecha_fin.year + 1):
        feriados |= _feriados_peru(y)

    actual = fecha_inicio
    count = 0
    while actual < fecha_fin:
        if actual.weekday() < 5 and actual not in feriados:
            count += 1
        actual += timedelta(days=1)
    return count


def _sumar_dias_habiles(fecha_inicio, dias, feriados=None):
    """Suma N días hábiles a fecha_inicio (lun-vie, sin feriados Perú)."""
    if feriados is None:
        feriados = _feriados_peru(fecha_inicio.year) | _feriados_peru(fecha_inicio.year + 1)
    actual = fecha_inicio
    agregados = 0
    while agregados < dias:
        actual += timedelta(days=1)
        if actual.weekday() < 5 and actual not in feriados:
            agregados += 1
    return actual

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ══════════════════════════════════════════════════════════════
# ADMIN — PANEL PRINCIPAL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def disciplinaria_panel(request):
    """Panel principal de medidas disciplinarias."""
    qs = MedidaDisciplinaria.objects.select_related('personal', 'tipo_falta').all()

    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '')
    buscar = request.GET.get('q', '')
    anio = request.GET.get('anio', str(date.today().year))

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_medida=tipo)
    if buscar:
        qs = qs.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar)
        )
    if anio:
        qs = qs.filter(fecha_hechos__year=int(anio))

    # Stats
    en_proceso = qs.filter(estado__in=['NOTIFICADA', 'EN_DESCARGO', 'DESCARGO_RECIBIDO']).count()
    pendientes_descargo = qs.filter(estado='EN_DESCARGO').count()

    # Próximos a vencer: procesos EN_DESCARGO con fecha_limite_descargo en los próximos 3 días
    hoy = date.today()
    limite_3_dias = hoy + timedelta(days=3)
    proximos_vencer = (
        MedidaDisciplinaria.objects
        .select_related('personal', 'tipo_falta')
        .filter(
            estado='EN_DESCARGO',
            fecha_limite_descargo__isnull=False,
            fecha_limite_descargo__lte=limite_3_dias,
        )
        .order_by('fecha_limite_descargo')
    )

    # Enriquecer medidas con días hábiles transcurridos para la columna Días
    estados_activos_panel = {'BORRADOR', 'NOTIFICADA', 'EN_DESCARGO', 'DESCARGO_RECIBIDO'}
    medidas_raw = list(qs[:100])
    medidas_enriquecidas = []
    for m in medidas_raw:
        if m.estado in estados_activos_panel:
            dias_hab = _dias_habiles_entre(m.fecha_hechos, hoy)
            if dias_hab > 30:
                semaforo = 'rojo'
            elif dias_hab > 20:
                semaforo = 'amarillo'
            else:
                semaforo = 'verde'
        else:
            dias_hab = None
            semaforo = None
        medidas_enriquecidas.append({
            'medida': m,
            'dias_habiles': dias_hab,
            'semaforo': semaforo,
        })

    context = {
        'titulo': 'Proceso Disciplinario',
        'medidas': medidas_enriquecidas,
        'total': qs.count(),
        'filtro_estado': estado,
        'filtro_tipo': tipo,
        'buscar': buscar,
        'anio_filtro': anio,
        'hoy': hoy,
        'tipos_falta': TipoFalta.objects.filter(activo=True),
        'proximos_vencer': proximos_vencer,
        'stats': {
            'en_proceso': en_proceso,
            'pendientes_descargo': pendientes_descargo,
            'resueltas_anio': qs.filter(estado='RESUELTA').count(),
        },
    }
    return render(request, 'disciplinaria/panel.html', context)


@login_required
@solo_admin
def medida_crear(request):
    """Registrar nueva medida disciplinaria."""
    if request.method == 'POST':
        try:
            personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
            tipo_falta = get_object_or_404(TipoFalta, pk=request.POST['tipo_falta_id'])

            medida = MedidaDisciplinaria.objects.create(
                personal=personal,
                tipo_medida=request.POST['tipo_medida'],
                tipo_falta=tipo_falta,
                fecha_hechos=request.POST['fecha_hechos'],
                descripcion_hechos=request.POST['descripcion'],
                testigos=request.POST.get('testigos', ''),
                registrado_por=request.user,
                estado='BORRADOR',
            )
            if request.FILES.get('evidencias'):
                medida.evidencias = request.FILES['evidencias']
                medida.save(update_fields=['evidencias'])

            from core.audit import log_create
            log_create(request, medida,
                       f'{medida.get_tipo_medida_display()} registrada: {personal.apellidos_nombres}')
            messages.success(request, f'Medida disciplinaria registrada para {personal.apellidos_nombres}')
            return redirect('medida_detalle', pk=medida.pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': 'Nueva Medida Disciplinaria',
        'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres'),
        'tipos_falta': TipoFalta.objects.filter(activo=True),
    }
    return render(request, 'disciplinaria/crear.html', context)


@login_required
@solo_admin
def medida_detalle(request, pk):
    """Detalle de una medida disciplinaria con descargos."""
    medida = get_object_or_404(
        MedidaDisciplinaria.objects.select_related(
            'personal', 'tipo_falta', 'registrado_por', 'resuelto_por'
        ),
        pk=pk
    )
    descargos = medida.descargos.select_related('personal', 'revisado_por').all()

    # Historial del trabajador
    historial = MedidaDisciplinaria.objects.filter(
        personal=medida.personal, estado='RESUELTA',
    ).exclude(pk=pk).order_by('-fecha_hechos')[:10]

    context = {
        'titulo': f'{medida.get_tipo_medida_display()} — {medida.personal.apellidos_nombres}',
        'medida': medida,
        'descargos': descargos,
        'historial': historial,
    }
    return render(request, 'disciplinaria/detalle.html', context)


@login_required
@solo_admin
@require_POST
def medida_notificar(request, pk):
    """Notificar medida al trabajador (inicia plazo de descargo)."""
    medida = get_object_or_404(MedidaDisciplinaria, pk=pk)
    if medida.estado == 'BORRADOR':
        preaviso_raw = request.POST.get('fecha_preaviso')
        medida.fecha_carta_preaviso = date.fromisoformat(preaviso_raw) if preaviso_raw else date.today()
        if request.FILES.get('documento_preaviso'):
            medida.documento_preaviso = request.FILES['documento_preaviso']
        medida.notificar()

        from core.audit import log_update
        log_update(request, medida, {'estado': {'old': 'BORRADOR', 'new': medida.estado}},
                   f'Medida notificada: {medida.personal.apellidos_nombres}')
        return JsonResponse({
            'ok': True, 'estado': medida.estado,
            'fecha_limite': medida.fecha_limite_descargo.strftime('%d/%m/%Y') if medida.fecha_limite_descargo else '',
        })
    return JsonResponse({'ok': False, 'error': 'Solo se puede notificar medidas en borrador.'})


@login_required
@solo_admin
@require_POST
def medida_resolver(request, pk):
    """Resolver medida disciplinaria (decisión final)."""
    medida = get_object_or_404(MedidaDisciplinaria, pk=pk)
    if medida.estado in ('NOTIFICADA', 'EN_DESCARGO', 'DESCARGO_RECIBIDO'):
        resolucion = request.POST.get('resolucion', '')
        try:
            dias_suspension = int(request.POST.get('dias_suspension', 0) or 0)
        except (ValueError, TypeError):
            dias_suspension = 0
        fecha_cese_raw = request.POST.get('fecha_cese')

        medida.dias_suspension = dias_suspension
        if fecha_cese_raw:
            medida.fecha_cese = date.fromisoformat(fecha_cese_raw)
        if request.FILES.get('documento_resolucion'):
            medida.documento_resolucion = request.FILES['documento_resolucion']
        medida.resolver(request.user, resolucion)

        from core.audit import log_update
        log_update(request, medida, {'estado': {'old': 'EN_DESCARGO', 'new': 'RESUELTA'}},
                   f'Medida resuelta: {medida.personal.apellidos_nombres}')
        messages.success(request, 'Medida disciplinaria resuelta.')
        return redirect('medida_detalle', pk=pk)
    messages.error(request, 'No se puede resolver en este estado.')
    return redirect('medida_detalle', pk=pk)


@login_required
@solo_admin
@require_POST
def descargo_registrar(request, medida_pk):
    """Registrar descargo del trabajador."""
    medida = get_object_or_404(MedidaDisciplinaria, pk=medida_pk)

    descargo = Descargo.objects.create(
        medida=medida,
        personal=medida.personal,
        texto=request.POST.get('texto', ''),
        fecha_presentacion=date.fromisoformat(request.POST['fecha_presentacion']) if request.POST.get('fecha_presentacion') else date.today(),
    )
    if request.FILES.get('adjuntos'):
        descargo.archivos_adjuntos = request.FILES['adjuntos']
        descargo.save(update_fields=['archivos_adjuntos'])

    # Actualizar estado de medida
    medida.estado = 'DESCARGO_RECIBIDO'
    medida.save(update_fields=['estado'])

    from core.audit import log_create
    log_create(request, descargo, f'Descargo registrado: {medida.personal.apellidos_nombres}')
    return JsonResponse({'ok': True, 'pk': descargo.pk, 'a_tiempo': descargo.presentado_a_tiempo})


@login_required
@solo_admin
@require_POST
def descargo_evaluar(request, pk):
    """Evaluar un descargo."""
    descargo = get_object_or_404(Descargo, pk=pk)
    descargo.estado = request.POST.get('estado', 'EN_REVISION')
    descargo.evaluacion = request.POST.get('evaluacion', '')
    descargo.revisado_por = request.user
    descargo.fecha_revision = date.today()
    descargo.save()
    return JsonResponse({'ok': True, 'estado': descargo.get_estado_display()})


# ══════════════════════════════════════════════════════════════
# ADMIN — TIPOS DE FALTA (CONFIG)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def tipos_falta(request):
    """Configuración de tipos de falta."""
    tipos = TipoFalta.objects.all()
    context = {
        'titulo': 'Tipos de Falta',
        'tipos': tipos,
    }
    return render(request, 'disciplinaria/tipos_falta.html', context)


@login_required
@solo_admin
@require_POST
def tipo_falta_crear(request):
    """Crear tipo de falta."""
    try:
        t = TipoFalta.objects.create(
            nombre=request.POST['nombre'],
            codigo=request.POST['codigo'],
            descripcion=request.POST.get('descripcion', ''),
            gravedad=request.POST.get('gravedad', 'LEVE'),
            base_legal=request.POST.get('base_legal', ''),
        )
        return JsonResponse({'ok': True, 'pk': t.pk, 'nombre': str(t)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


# ══════════════════════════════════════════════════════════════
# REPORTE: HISTORIAL DISCIPLINARIO POR TRABAJADOR
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def historial_trabajador(request, personal_id):
    """Historial disciplinario completo de un trabajador."""
    empleado = get_object_or_404(Personal, pk=personal_id)
    medidas = MedidaDisciplinaria.objects.filter(
        personal=empleado,
    ).select_related('tipo_falta').order_by('-fecha_hechos')

    context = {
        'titulo': f'Historial — {empleado.apellidos_nombres}',
        'empleado': empleado,
        'medidas': medidas,
        'resumen': {
            'verbales': medidas.filter(tipo_medida='VERBAL', estado='RESUELTA').count(),
            'escritas': medidas.filter(tipo_medida='ESCRITA', estado='RESUELTA').count(),
            'suspensiones': medidas.filter(tipo_medida='SUSPENSION', estado='RESUELTA').count(),
        },
    }
    return render(request, 'disciplinaria/historial.html', context)


# ══════════════════════════════════════════════════════════════
# EXPORTACIÓN: REPORTE EXCEL DISCIPLINARIO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def exportar_reporte_disciplinario(request):
    """Exporta reporte disciplinario del año actual en formato Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl no instalado.', status=500)

    try:
        anio = int(request.GET.get('anio', date.today().year))
    except (ValueError, TypeError):
        anio = date.today().year

    medidas = (
        MedidaDisciplinaria.objects
        .select_related('personal', 'tipo_falta', 'registrado_por', 'resuelto_por')
        .filter(fecha_hechos__year=anio)
        .order_by('-fecha_hechos')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Disciplinario {anio}'

    # ── Estilos ──
    color_header = 'FF2E4057'   # Azul oscuro
    color_sub    = 'FF4A6FA5'   # Azul medio
    font_title   = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
    font_header  = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
    font_normal  = Font(name='Calibri', size=9)
    fill_header  = PatternFill('solid', fgColor=color_header)
    fill_sub     = PatternFill('solid', fgColor=color_sub)
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left         = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    estado_colors = {
        'BORRADOR':          'FFAAAAAA',
        'NOTIFICADA':        'FF4472C4',
        'EN_DESCARGO':       'FFFFB300',
        'DESCARGO_RECIBIDO': 'FF00ACC1',
        'RESUELTA':          'FF43A047',
        'ANULADA':           'FF9E9E9E',
    }

    # ── Título ──
    ws.merge_cells('A1:H1')
    ws['A1'] = f'Reporte Disciplinario — {anio}'
    ws['A1'].font = font_title
    ws['A1'].fill = PatternFill('solid', fgColor=color_header)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generado: {date.today().strftime("%d/%m/%Y")} | Total registros: {medidas.count()}'
    ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='FFFFFFFF')
    ws['A2'].fill = PatternFill('solid', fgColor=color_sub)
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 16

    # ── Encabezados ──
    headers = [
        'N°', 'Trabajador', 'DNI', 'Tipo Medida', 'Tipo Falta',
        'Fecha Hechos', 'Estado', 'Fecha Preaviso',
        'Limite Descargo', 'Fecha Resolución', 'Sanción / Resolución',
    ]
    ws.append([])  # fila 3 vacía para separar
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[4].height = 22

    # ── Datos ──
    for row_idx, m in enumerate(medidas, start=1):
        row_num = row_idx + 4

        # Color alternado de fila
        if row_idx % 2 == 0:
            row_fill = PatternFill('solid', fgColor='FFF5F5F5')
        else:
            row_fill = PatternFill('solid', fgColor='FFFFFFFF')

        valores = [
            row_idx,
            m.personal.apellidos_nombres,
            m.personal.nro_doc,
            m.get_tipo_medida_display(),
            m.tipo_falta.nombre,
            m.fecha_hechos.strftime('%d/%m/%Y') if m.fecha_hechos else '',
            m.get_estado_display(),
            m.fecha_carta_preaviso.strftime('%d/%m/%Y') if m.fecha_carta_preaviso else '',
            m.fecha_limite_descargo.strftime('%d/%m/%Y') if m.fecha_limite_descargo else '',
            m.fecha_resolucion.strftime('%d/%m/%Y') if m.fecha_resolucion else '',
            m.resolucion[:200] if m.resolucion else '',
        ]

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=valor)
            cell.font = font_normal
            cell.border = thin_border
            cell.fill = row_fill
            if col_idx in (1,):
                cell.alignment = center
            else:
                cell.alignment = left

        # Color especial para columna Estado (col 7)
        estado_cell = ws.cell(row=row_num, column=7)
        est_color = estado_colors.get(m.estado, 'FFAAAAAA')
        estado_cell.fill = PatternFill('solid', fgColor=est_color)
        estado_cell.font = Font(name='Calibri', size=9, bold=True,
                                color='FFFFFFFF' if m.estado != 'EN_DESCARGO' else 'FF000000')
        estado_cell.alignment = center

        ws.row_dimensions[row_num].height = 16

    # ── Anchos de columna ──
    col_widths = [5, 35, 12, 18, 28, 14, 18, 14, 16, 16, 50]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze y filtros ──
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}4'

    # ── Segunda hoja: Resumen por tipo ──
    ws2 = wb.create_sheet('Resumen')
    ws2['A1'] = f'Resumen por Tipo de Medida — {anio}'
    ws2['A1'].font = Font(name='Calibri', size=12, bold=True, color='FF2E4057')
    ws2.row_dimensions[1].height = 20

    ws2.append(['Tipo Medida', 'Total', 'Resueltas', 'En Proceso', 'Anuladas'])
    for col_idx in range(1, 6):
        cell = ws2.cell(row=2, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = thin_border

    for tipo_code, tipo_label in MedidaDisciplinaria.TIPO_CHOICES:
        sub = medidas.filter(tipo_medida=tipo_code)
        ws2.append([
            tipo_label,
            sub.count(),
            sub.filter(estado='RESUELTA').count(),
            sub.filter(estado__in=['NOTIFICADA', 'EN_DESCARGO', 'DESCARGO_RECIBIDO']).count(),
            sub.filter(estado='ANULADA').count(),
        ])

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 20 if col == 'A' else 12

    # ── Respuesta HTTP ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'reporte_disciplinario_{anio}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ══════════════════════════════════════════════════════════════
# DASHBOARD — KPIs Y ALERTAS LEGALES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def disciplinaria_dashboard(request):
    """
    Dashboard ejecutivo del módulo disciplinario.
    KPIs, alertas de plazos legales, top faltas, procesos por área.
    DS 003-97-TR: plazo máximo 30 días para imponer sanción desde detección.
    """
    import json
    from datetime import date as date_cls
    from dateutil.relativedelta import relativedelta

    hoy = date.today()
    anio_actual = hoy.year
    primer_dia_mes = hoy.replace(day=1)

    # ── KPI base ──
    qs_total = MedidaDisciplinaria.objects.all()
    estados_activos = ['BORRADOR', 'NOTIFICADA', 'EN_DESCARGO', 'DESCARGO_RECIBIDO']

    activos = qs_total.filter(estado__in=estados_activos).count()
    cerrados_mes = qs_total.filter(
        estado='RESUELTA',
        fecha_resolucion__gte=primer_dia_mes,
    ).count()
    total_anio = qs_total.filter(fecha_hechos__year=anio_actual).count()

    # ── Top 5 tipos de falta (año actual) ──
    por_tipo_falta = (
        MedidaDisciplinaria.objects
        .filter(fecha_hechos__year=anio_actual)
        .values('tipo_falta__nombre', 'tipo_falta__gravedad')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    # ── Por tipo de medida (año actual) ──
    por_tipo_medida = (
        MedidaDisciplinaria.objects
        .filter(fecha_hechos__year=anio_actual)
        .values('tipo_medida')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    tipo_medida_map = {t: 0 for t, _ in MedidaDisciplinaria.TIPO_CHOICES}
    for row in por_tipo_medida:
        tipo_medida_map[row['tipo_medida']] = row['total']

    # ── Por área (año actual) — top 8 ──
    por_area = (
        MedidaDisciplinaria.objects
        .filter(fecha_hechos__year=anio_actual)
        .values('personal__subarea__area__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')[:8]
    )

    # ── Procesos activos con días transcurridos y alerta de plazo ──
    procesos_activos = (
        MedidaDisciplinaria.objects
        .select_related('personal', 'personal__subarea__area', 'tipo_falta')
        .filter(estado__in=estados_activos)
        .order_by('fecha_hechos')
    )

    # Enriquecer con días hábiles transcurridos y semáforo
    feriados = _feriados_peru(anio_actual) | _feriados_peru(anio_actual + 1)
    procesos_con_dias = []
    alertas_vencimiento = []

    for p in procesos_activos:
        dias_hab = _dias_habiles_entre(p.fecha_hechos, hoy)
        # DS 003-97-TR: sanción debe imponerse dentro de los 30 días de conocida la falta.
        dias_restantes_legal = 30 - dias_hab

        if dias_restantes_legal <= 0:
            semaforo = 'rojo'
        elif dias_restantes_legal <= 5:
            semaforo = 'amarillo'
        else:
            semaforo = 'verde'

        entry = {
            'medida': p,
            'dias_habiles': dias_hab,
            'dias_restantes_legal': dias_restantes_legal,
            'semaforo': semaforo,
        }
        procesos_con_dias.append(entry)

        if semaforo in ('rojo', 'amarillo'):
            alertas_vencimiento.append(entry)

    # ── ANALYTICS: Top 8 tipos de falta — JSON para Chart.js doughnut ──
    # Paleta teal→red: 8 colores progresivos desde accent teal hasta rojo alerta
    _PALETTE_TIPO = [
        '#0f766e', '#0d9488', '#14b8a6', '#2dd4bf',
        '#f59e0b', '#ef4444', '#dc2626', '#991b1b',
    ]
    tipos_falta_json = '[]'
    try:
        top_tipos = (
            MedidaDisciplinaria.objects
            .filter(fecha_hechos__year=anio_actual)
            .values('tipo_falta__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')[:8]
        )
        tipos_falta_json = json.dumps([
            {
                'label': row['tipo_falta__nombre'] or 'Sin tipo',
                'value': row['total'],
                'color': _PALETTE_TIPO[i % len(_PALETTE_TIPO)],
            }
            for i, row in enumerate(top_tipos)
        ])
    except Exception:
        tipos_falta_json = '[]'

    # ── ANALYTICS: Tendencia 6 meses — JSON para Chart.js line ──
    tendencia_6m_json = '[]'
    try:
        # Calcular el primer día de hace 6 meses completos
        meses = []
        for i in range(5, -1, -1):
            # relativedelta: retroceder i meses desde hoy
            inicio_mes = (hoy.replace(day=1) - relativedelta(months=i))
            fin_mes = inicio_mes + relativedelta(months=1)
            label = inicio_mes.strftime('%m/%y')
            count = MedidaDisciplinaria.objects.filter(
                fecha_hechos__gte=inicio_mes,
                fecha_hechos__lt=fin_mes,
            ).count()
            meses.append({'label': label, 'value': count})
        tendencia_6m_json = json.dumps(meses)
    except Exception:
        tendencia_6m_json = '[]'

    # ── ANALYTICS: Distribución por estado — JSON para Chart.js doughnut ──
    _ESTADO_COLORS = {
        'BORRADOR':          '#9ca3af',
        'NOTIFICADA':        '#3b82f6',
        'EN_DESCARGO':       '#f59e0b',
        'DESCARGO_RECIBIDO': '#06b6d4',
        'RESUELTA':          '#10b981',
        'ANULADA':           '#6b7280',
    }
    _ESTADO_LABELS = dict(MedidaDisciplinaria.ESTADO_CHOICES)
    por_estado_json = '[]'
    try:
        estados_counts = (
            MedidaDisciplinaria.objects
            .values('estado')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        por_estado_json = json.dumps([
            {
                'label': _ESTADO_LABELS.get(row['estado'], row['estado']),
                'value': row['total'],
                'color': _ESTADO_COLORS.get(row['estado'], '#94a3b8'),
            }
            for row in estados_counts
            if row['total'] > 0
        ])
    except Exception:
        por_estado_json = '[]'

    # ── ANALYTICS: Procesos críticos (abiertos) — hasta 5, más recientes ──
    procesos_criticos = []
    try:
        procesos_criticos = list(
            MedidaDisciplinaria.objects
            .select_related('personal', 'tipo_falta')
            .filter(estado__in=estados_activos)
            .order_by('-fecha_hechos')[:5]
        )
    except Exception:
        procesos_criticos = []

    context = {
        'titulo': 'Dashboard Disciplinario',
        'hoy': hoy,
        'anio_actual': anio_actual,
        'kpi': {
            'activos': activos,
            'cerrados_mes': cerrados_mes,
            'total_anio': total_anio,
        },
        'por_tipo_falta': list(por_tipo_falta),
        'por_area': list(por_area),
        'tipo_medida_map': tipo_medida_map,
        'procesos_con_dias': procesos_con_dias,
        'alertas_vencimiento': alertas_vencimiento,
        'TIPO_CHOICES': MedidaDisciplinaria.TIPO_CHOICES,
        # Analytics — Chart.js JSON payloads
        'tipos_falta_json': tipos_falta_json,
        'tendencia_6m_json': tendencia_6m_json,
        'por_estado_json': por_estado_json,
        'procesos_criticos': procesos_criticos,
    }
    return render(request, 'disciplinaria/dashboard.html', context)


# ══════════════════════════════════════════════════════════════
# TIMELINE LEGAL DEL PROCESO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def proceso_timeline(request, pk):
    """
    Vista de timeline legal detallado para un proceso disciplinario.
    Muestra hitos reales y estimados con semáforo de cumplimiento.
    DS 003-97-TR: Art. 31 — 6 días hábiles para descargo; Art. 32 — resolución escrita.
    """
    medida = get_object_or_404(
        MedidaDisciplinaria.objects.select_related(
            'personal', 'personal__subarea__area',
            'tipo_falta', 'registrado_por', 'resuelto_por',
        ),
        pk=pk,
    )
    descargos = medida.descargos.select_related('personal', 'revisado_por').order_by('fecha_presentacion')
    primer_descargo = descargos.first()

    hoy = date.today()
    feriados = _feriados_peru(hoy.year) | _feriados_peru(hoy.year + 1)

    # ── Hito 1: Detección / Registro de hechos ──
    hito_hechos = {
        'nombre': 'Detección de Falta',
        'subtitulo': 'Fecha de los hechos registrada',
        'base_legal': '',
        'fecha_real': medida.fecha_hechos,
        'fecha_estimada': None,
        'completado': True,
        'estado': 'completado',
        'icono': 'fa-exclamation-circle',
    }

    # ── Hito 2: Carta de Preaviso ──
    fecha_estimada_preaviso = _sumar_dias_habiles(medida.fecha_hechos, 2, feriados)
    hito_preaviso = {
        'nombre': 'Carta de Preaviso / Imputación de Cargos',
        'subtitulo': 'Art. 31 DS 003-97-TR — Comunicación escrita de la falta',
        'base_legal': 'Art. 31 DS 003-97-TR',
        'fecha_real': medida.fecha_carta_preaviso,
        'fecha_estimada': fecha_estimada_preaviso if not medida.fecha_carta_preaviso else None,
        'completado': bool(medida.fecha_carta_preaviso),
        'estado': None,
        'icono': 'fa-envelope-open-text',
    }
    if medida.fecha_carta_preaviso:
        hito_preaviso['estado'] = 'completado'
    elif medida.estado == 'BORRADOR':
        hito_preaviso['estado'] = 'pendiente'
    else:
        hito_preaviso['estado'] = 'no_aplica'

    # ── Hito 3: Plazo de Descargo (6 días hábiles desde preaviso) ──
    fecha_base_descargo = medida.fecha_carta_preaviso or fecha_estimada_preaviso
    fecha_estimada_limite_descargo = _sumar_dias_habiles(fecha_base_descargo, 6, feriados)
    fecha_real_limite = medida.fecha_limite_descargo

    dias_restantes_descargo = None
    if fecha_real_limite:
        raw_delta = (fecha_real_limite - hoy).days
        dias_restantes_descargo = raw_delta  # positivo=queda tiempo, 0=hoy, negativo=vencido

    if medida.estado in ('EN_DESCARGO',):
        if dias_restantes_descargo is not None and dias_restantes_descargo < 0:
            estado_descargo = 'vencido'
        elif dias_restantes_descargo is not None and dias_restantes_descargo <= 2:
            estado_descargo = 'urgente'
        else:
            estado_descargo = 'en_curso'
    elif medida.estado in ('DESCARGO_RECIBIDO', 'RESUELTA', 'ANULADA'):
        estado_descargo = 'completado'
    else:
        estado_descargo = 'pendiente'

    hito_plazo_descargo = {
        'nombre': 'Plazo para Presentar Descargo',
        'subtitulo': '6 días hábiles desde la notificación de la carta de preaviso',
        'base_legal': 'Art. 31 DS 003-97-TR',
        'fecha_real': fecha_real_limite,
        'fecha_estimada': fecha_estimada_limite_descargo if not fecha_real_limite else None,
        'completado': medida.estado in ('DESCARGO_RECIBIDO', 'RESUELTA', 'ANULADA'),
        'estado': estado_descargo,
        'dias_restantes': dias_restantes_descargo,
        'icono': 'fa-hourglass-half',
    }

    # ── Hito 4: Recepción del Descargo ──
    if primer_descargo:
        estado_recepcion = 'completado'
        a_tiempo = primer_descargo.presentado_a_tiempo
    elif medida.estado in ('EN_DESCARGO', 'NOTIFICADA'):
        estado_recepcion = 'en_curso'
        a_tiempo = None
    else:
        estado_recepcion = 'pendiente'
        a_tiempo = None

    hito_descargo = {
        'nombre': 'Recepción del Descargo del Trabajador',
        'subtitulo': 'Derecho de defensa del trabajador (Art. 31 DS 003-97-TR)',
        'base_legal': 'Art. 31 DS 003-97-TR',
        'fecha_real': primer_descargo.fecha_presentacion if primer_descargo else None,
        'fecha_estimada': None,
        'completado': bool(primer_descargo),
        'estado': estado_recepcion,
        'a_tiempo': a_tiempo,
        'icono': 'fa-comment-dots',
    }

    # ── Hito 5: Resolución / Decisión Final ──
    # Estimado: 3 días hábiles tras vencer plazo descargo
    fecha_base_resolucion = fecha_real_limite or fecha_estimada_limite_descargo
    fecha_estimada_resolucion = _sumar_dias_habiles(fecha_base_resolucion, 3, feriados)

    if medida.fecha_resolucion:
        if medida.estado == 'ANULADA':
            estado_resolucion = 'anulado'
        else:
            estado_resolucion = 'completado'
    elif medida.estado == 'DESCARGO_RECIBIDO':
        estado_resolucion = 'en_curso'
    else:
        estado_resolucion = 'pendiente'

    hito_resolucion = {
        'nombre': 'Resolución / Decisión Final',
        'subtitulo': 'Sanción motivada y comunicada por escrito al trabajador',
        'base_legal': 'Art. 32 DS 003-97-TR',
        'fecha_real': medida.fecha_resolucion,
        'fecha_estimada': fecha_estimada_resolucion if not medida.fecha_resolucion else None,
        'completado': bool(medida.fecha_resolucion),
        'estado': estado_resolucion,
        'resuelto_por': medida.resuelto_por,
        'icono': 'fa-stamp',
    }

    # ── Hito 6: Notificación de Cese (solo DESPIDO) ──
    hito_cese = None
    if medida.tipo_medida == 'DESPIDO':
        hito_cese = {
            'nombre': 'Notificación de Cese',
            'subtitulo': 'Carta de despido con causa y fecha efectiva de cese',
            'base_legal': 'Art. 32 DS 003-97-TR',
            'fecha_real': medida.fecha_cese,
            'fecha_estimada': None,
            'completado': bool(medida.fecha_cese),
            'estado': 'completado' if medida.fecha_cese else (
                'en_curso' if medida.estado == 'RESUELTA' else 'pendiente'
            ),
            'icono': 'fa-door-open',
        }

    # ── Días hábiles totales transcurridos desde los hechos ──
    dias_habiles_total = _dias_habiles_entre(medida.fecha_hechos, hoy)

    hitos = [hito_hechos, hito_preaviso, hito_plazo_descargo, hito_descargo, hito_resolucion]
    if hito_cese:
        hitos.append(hito_cese)

    context = {
        'titulo': f'Timeline Legal — {medida.personal.apellidos_nombres}',
        'medida': medida,
        'descargos': descargos,
        'hitos': hitos,
        'hoy': hoy,
        'dias_habiles_total': dias_habiles_total,
        'plazo_legal_total': 30,  # días hábiles máximos DS 003-97-TR
        'plazo_superado': dias_habiles_total > 30,
    }
    return render(request, 'disciplinaria/timeline.html', context)


# ══════════════════════════════════════════════════════════════
# REPORTE AJAX — POR ÁREA (Chart.js)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def disciplinaria_reporte_area(request):
    """
    Endpoint AJAX para Chart.js.
    Retorna conteo de medidas disciplinarias por área del año actual.
    """
    try:
        anio = int(request.GET.get('anio', date.today().year))
    except (ValueError, TypeError):
        anio = date.today().year

    data = (
        MedidaDisciplinaria.objects
        .filter(fecha_hechos__year=anio)
        .values('personal__subarea__area__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    labels = []
    counts = []
    colors = [
        '#0f766e', '#0d9488', '#14b8a6', '#2dd4bf',
        '#5eead4', '#99f6e4', '#134e4a', '#042f2e',
    ]

    for i, row in enumerate(data):
        area_nombre = row['personal__subarea__area__nombre'] or 'Sin área'
        labels.append(area_nombre)
        counts.append(row['total'])

    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': f'Medidas {anio}',
            'data': counts,
            'backgroundColor': [colors[i % len(colors)] for i in range(len(counts))],
            'borderColor': '#ffffff',
            'borderWidth': 2,
        }],
        'anio': anio,
        'total': sum(counts),
    })
