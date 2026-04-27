"""
Vistas del módulo Tareo — Exportaciones.
"""
import calendar
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render

from asistencia.views._common import solo_admin, _qs_sin_papeleta


# ---------------------------------------------------------------------------
# EXPORTAR CARGA S10
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_carga_s10_view(request):
    """Genera y descarga el archivo CargaS10 para importar en el sistema S10."""
    from asistencia.services.exporters import CargaS10Exporter

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')

    try:
        exporter = CargaS10Exporter(anio, mes)
        buffer = exporter.generar()

        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        sufijo = '_corte' if tipo_periodo == 'corte' else ''
        filename = f'CargaS10_{MESES[mes-1]}_{anio}{sufijo}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        messages.error(request, f'Error generando CargaS10: {e}')
        return redirect('asistencia_dashboard')


# ---------------------------------------------------------------------------
# EXPORTAR REPORTE CIERRE
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_cierre_view(request):
    """Genera el reporte de cierre de mes."""
    from asistencia.services.exporters import ReporteCierreExporter

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')

    try:
        exporter = ReporteCierreExporter(anio, mes, tipo_periodo=tipo_periodo)
        buffer = exporter.generar()

        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        sufijo = '_corte' if tipo_periodo == 'corte' else ''
        filename = f'Cierre_{MESES[mes-1]}_{anio}{sufijo}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        messages.error(request, f'Error generando reporte de cierre: {e}')
        return redirect('asistencia_dashboard')


@login_required
@solo_admin
def reportes_exportar_panel(request):
    """Panel central de reportes y exportaciones."""
    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    tipo = request.GET.get('tipo_periodo', 'calendario')
    if tipo not in ('calendario', 'corte'):
        tipo = 'calendario'
    MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    # Calcular preview del rango de fechas
    d_ini, d_fin = _get_corte_config(request)
    ini, fin = _calcular_periodo(anio, mes, tipo, d_ini, d_fin)
    return render(request, 'asistencia/reportes_exportar.html', {
        'anio': anio, 'mes': mes, 'tipo_periodo': tipo,
        'fecha_ini': ini.strftime('%d/%m/%Y'),
        'fecha_fin': fin.strftime('%d/%m/%Y'),
        'anios': list(range(hoy.year - 2, hoy.year + 1)),
        'meses_list': [(i, MESES[i - 1]) for i in range(1, 13)],
    })


# ---------------------------------------------------------------------------
# EXPORTAR REPORTE DE HORAS RCO (Excel)
# ---------------------------------------------------------------------------

MESES_ES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']


def _calcular_periodo(anio: int, mes: int, tipo: str = 'calendario',
                      dia_inicio_corte: int = 22, dia_fin_corte: int = 21):
    """
    Calcula las fechas de inicio y fin del período.
    tipo='calendario' → del 1 al último día del mes seleccionado.
    tipo='corte'      → del dia_inicio_corte del mes anterior al dia_fin_corte del mes seleccionado.
    Los días de corte se leen de la configuración de la empresa (Empresa.dia_inicio_corte / dia_fin_corte).
    """
    import calendar as _cal
    if tipo == 'corte':
        if mes == 1:
            ini = date(anio - 1, 12, dia_inicio_corte)
        else:
            ini = date(anio, mes - 1, dia_inicio_corte)
        fin = date(anio, mes, dia_fin_corte)
    else:
        ini = date(anio, mes, 1)
        fin = date(anio, mes, _cal.monthrange(anio, mes)[1])
    return ini, fin


def _get_corte_config(request):
    """Obtiene los días de corte de ConfiguracionSistema.dia_corte_planilla.

    dia_corte=21 → ciclo (22 mes-ant → 21 mes). Retorna (dia_inicio, dia_fin).
    """
    from asistencia.models import ConfiguracionSistema
    cfg = ConfiguracionSistema.objects.first()
    if cfg:
        corte = cfg.dia_corte_planilla or 21
        return corte + 1, corte
    return 22, 21


def _label_periodo(ini, fin, tipo: str) -> str:
    """Genera etiqueta legible del período."""
    if tipo == 'corte':
        return f'Corte de Planilla: {ini.strftime("%d/%m/%Y")} al {fin.strftime("%d/%m/%Y")}'
    return f'Mes Calendario: {ini.strftime("%d/%m/%Y")} al {fin.strftime("%d/%m/%Y")}'


@login_required
@solo_admin
def exportar_horas_rco(request):
    """Excel con resumen de horas y HE por trabajador RCO del periodo."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from asistencia.models import RegistroTareo
    from personal.models import Personal

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')
    if tipo_periodo not in ('calendario', 'corte'):
        tipo_periodo = 'calendario'

    d_ini, d_fin = _get_corte_config(request)
    mes_ini, mes_fin = _calcular_periodo(anio, mes, tipo_periodo, d_ini, d_fin)
    label_periodo = _label_periodo(mes_ini, mes_fin, tipo_periodo)

    # Resumen por persona
    resumen = list(
        RegistroTareo.objects.filter(
            grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin
        ).values('dni', 'nombre_archivo', 'personal_id')
        .annotate(
            dias_trabajados=Count('id', filter=Q(
                codigo_dia__in=['T', 'NOR', 'TR', 'A', 'CDT', 'CPF', 'LCG', 'ATM', 'CHE', 'LIM', 'SS'])),
            dias_dl=Count('id', filter=Q(codigo_dia__in=['DL', 'DLA'])),
            dias_vac=Count('id', filter=Q(codigo_dia__in=['VAC', 'V'])),
            dias_dm=Count('id', filter=Q(codigo_dia='DM')),
            dias_sai=Count('id', filter=Q(codigo_dia='SAI')),
            total_horas=Sum('horas_marcadas'),
            total_hn=Sum('horas_normales'),
            total_he_25=Sum('he_25'),
            total_he_35=Sum('he_35'),
            total_he_100=Sum('he_100'),
        )
        .order_by('nombre_archivo')
    )

    # Faltas reales: excluir días cubiertos por papeleta APROBADA/EJECUTADA
    # (justificados aunque el codigo_dia haya quedado en FA por desincronía).
    faltas_por_pid = dict(
        _qs_sin_papeleta(
            RegistroTareo.objects.filter(
                grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin,
                codigo_dia__in=['FA', 'F'],
            ).exclude(dia_semana=6, condicion__in=['LOCAL', 'LIMA', ''])
        )
        .values('personal_id')
        .annotate(n=Count('id'))
        .values_list('personal_id', 'n')
    )
    for r in resumen:
        r['dias_falta'] = faltas_por_pid.get(r['personal_id'], 0)

    # Enriquecer con datos de Personal
    pids = [r['personal_id'] for r in resumen if r['personal_id']]
    personal_map = {p.id: p for p in Personal.objects.filter(id__in=pids)}

    for r in resumen:
        r['total_he_25'] = r['total_he_25'] or Decimal('0')
        r['total_he_35'] = r['total_he_35'] or Decimal('0')
        r['total_he_100'] = r['total_he_100'] or Decimal('0')
        r['total_he'] = r['total_he_25'] + r['total_he_35'] + r['total_he_100']
        r['total_horas'] = r['total_horas'] or Decimal('0')
        r['total_hn'] = r['total_hn'] or Decimal('0')
        p = personal_map.get(r['personal_id'])
        r['cargo'] = p.cargo if p else ''
        r['area'] = p.subarea.area.nombre if p and p.subarea else ''

    # ── Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Horas RCO {MESES_ES[mes - 1]} {anio}'

    # Estilos
    title_font = Font(bold=True, size=14, color='0f766e')
    sub_font = Font(size=9, color='64748b')
    header_font = Font(bold=True, size=9, color='FFFFFF')
    header_fill = PatternFill(start_color='0f766e', end_color='0f766e', fill_type='solid')
    data_font = Font(size=9)
    num_font = Font(size=9, bold=True)
    total_fill = PatternFill(start_color='134e4a', end_color='134e4a', fill_type='solid')
    total_font = Font(bold=True, size=9, color='FFFFFF')
    border = Border(bottom=Side(style='thin', color='e2e8f0'))
    center = Alignment(horizontal='center')

    # Titulo
    ws.cell(row=1, column=1, value=f'REPORTE DE HORAS — RCO — {MESES_ES[mes - 1].upper()} {anio}').font = title_font
    ws.cell(row=2, column=1, value=label_periodo).font = sub_font

    # Headers
    headers = ['N°', 'DNI', 'Apellidos y Nombres', 'Cargo', 'Area',
               'Dias Trab.', 'Faltas', 'DL', 'VAC', 'DM', 'SAI',
               'Hrs Marcadas', 'Hrs Normales', 'HE 25%', 'HE 35%', 'HE 100%', 'Total HE']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data
    for i, r in enumerate(resumen, 1):
        row = i + 4
        ws.cell(row=row, column=1, value=i).font = data_font
        c2 = ws.cell(row=row, column=2, value=r['dni'])
        c2.font = data_font
        c2.number_format = '@'  # DNI texto — preserva ceros iniciales
        ws.cell(row=row, column=3, value=r['nombre_archivo']).font = data_font
        ws.cell(row=row, column=4, value=r['cargo']).font = Font(size=8, color='64748b')
        ws.cell(row=row, column=5, value=r['area']).font = Font(size=8, color='64748b')
        ws.cell(row=row, column=6, value=r['dias_trabajados']).font = num_font
        ws.cell(row=row, column=7, value=r['dias_falta']).font = num_font
        ws.cell(row=row, column=8, value=r['dias_dl']).font = data_font
        ws.cell(row=row, column=9, value=r['dias_vac']).font = data_font
        ws.cell(row=row, column=10, value=r['dias_dm']).font = data_font
        ws.cell(row=row, column=11, value=r['dias_sai']).font = data_font
        ws.cell(row=row, column=12, value=float(r['total_horas'])).font = num_font
        ws.cell(row=row, column=13, value=float(r['total_hn'])).font = num_font
        ws.cell(row=row, column=14, value=float(r['total_he_25'])).font = num_font
        ws.cell(row=row, column=15, value=float(r['total_he_35'])).font = num_font
        ws.cell(row=row, column=16, value=float(r['total_he_100'])).font = num_font
        ws.cell(row=row, column=17, value=float(r['total_he'])).font = Font(bold=True, size=10, color='0f766e')
        for c in range(1, 18):
            ws.cell(row=row, column=c).border = border
            if c >= 6:
                ws.cell(row=row, column=c).alignment = center
                ws.cell(row=row, column=c).number_format = '#,##0.00' if c >= 12 else '0'

    # Totals row
    total_row = len(resumen) + 5
    for c in range(1, 18):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = center
    ws.cell(row=total_row, column=3, value='TOTALES')
    ws.cell(row=total_row, column=6, value=sum(r['dias_trabajados'] for r in resumen))
    ws.cell(row=total_row, column=7, value=sum(r['dias_falta'] for r in resumen))
    ws.cell(row=total_row, column=8, value=sum(r['dias_dl'] for r in resumen))
    ws.cell(row=total_row, column=9, value=sum(r['dias_vac'] for r in resumen))
    ws.cell(row=total_row, column=10, value=sum(r['dias_dm'] for r in resumen))
    ws.cell(row=total_row, column=11, value=sum(r['dias_sai'] for r in resumen))
    ws.cell(row=total_row, column=12, value=float(sum(r['total_horas'] for r in resumen)))
    ws.cell(row=total_row, column=13, value=float(sum(r['total_hn'] for r in resumen)))
    ws.cell(row=total_row, column=14, value=float(sum(r['total_he_25'] for r in resumen)))
    ws.cell(row=total_row, column=15, value=float(sum(r['total_he_35'] for r in resumen)))
    ws.cell(row=total_row, column=16, value=float(sum(r['total_he_100'] for r in resumen)))
    ws.cell(row=total_row, column=17, value=float(sum(r['total_he'] for r in resumen)))
    for c in [12, 13, 14, 15, 16, 17]:
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'

    # Conteo
    ws.cell(row=total_row + 2, column=1,
            value=f'Total trabajadores: {len(resumen)}').font = Font(size=9, color='64748b')

    # Column widths
    widths = [5, 12, 38, 25, 20, 9, 8, 6, 6, 6, 6, 12, 12, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    sufijo = '_corte' if tipo_periodo == 'corte' else ''
    response['Content-Disposition'] = f'attachment; filename="Horas_RCO_{MESES_ES[mes - 1]}_{anio}{sufijo}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# REPORTE DE FALTAS DEL MES (para cierre de planilla S10)
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_faltas_mes(request):
    """
    Excel con lista de trabajadores que tuvieron faltas (FA, F) y/o LSG en el mes.
    Incluye cálculo estimado de descuento por día.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from asistencia.models import RegistroTareo
    from personal.models import Personal

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')
    if tipo_periodo not in ('calendario', 'corte'):
        tipo_periodo = 'calendario'

    d_ini, d_fin = _get_corte_config(request)
    mes_ini, mes_fin = _calcular_periodo(anio, mes, tipo_periodo, d_ini, d_fin)
    label_periodo = _label_periodo(mes_ini, mes_fin, tipo_periodo)

    # Registros de FA/F/LSG del mes (excluir FA en domingos para LOCAL/LIMA)
    # y excluir días cubiertos por papeleta APROBADA/EJECUTADA — justificados.
    regs_detalle = list(
        _qs_sin_papeleta(
            RegistroTareo.objects.filter(
                fecha__gte=mes_ini, fecha__lte=mes_fin,
                codigo_dia__in=['FA', 'F', 'LSG'],
                personal__isnull=False,
            )
            .exclude(codigo_dia__in=['FA', 'F'], dia_semana=6, condicion__in=['LOCAL', 'LIMA', ''])
        )
        .select_related('personal__subarea__area')
        .order_by('fecha', 'personal__apellidos_nombres')
    )

    # Agrupar por personal
    faltas_map: dict[int, dict] = {}
    for r in regs_detalle:
        pid = r.personal_id
        if pid not in faltas_map:
            faltas_map[pid] = {'FA': 0, 'F': 0, 'LSG': 0}
        faltas_map[pid][r.codigo_dia] = faltas_map[pid].get(r.codigo_dia, 0) + 1

    if not faltas_map:
        # Sin faltas: crear 3 hojas igual que el caso con datos, solo con
        # mensaje informativo. Mantiene estructura consistente del archivo.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Faltas {MESES_ES[mes - 1]} {anio}'
        ws['A1'] = f'REPORTE DE FALTAS — {MESES_ES[mes - 1].upper()} {anio}'
        ws['A1'].font = Font(bold=True, size=14, color='991B1B')
        ws['A2'] = label_periodo
        ws['A2'].font = Font(size=9, color='64748b')
        ws['A4'] = 'Sin faltas registradas en el período.'
        ws['A4'].font = Font(size=10, italic=True, color='64748b')
        ws.column_dimensions['A'].width = 70

        ws_az = wb.create_sheet(title='Detalle A-Z')
        ws_az['A1'] = f'DETALLE POR TRABAJADOR (A-Z) — {MESES_ES[mes - 1].upper()} {anio}'
        ws_az['A1'].font = Font(bold=True, size=14, color='991B1B')
        ws_az['A2'] = label_periodo
        ws_az['A2'].font = Font(size=9, color='64748b')
        ws_az['A4'] = 'Sin faltas registradas en el período.'
        ws_az['A4'].font = Font(size=10, italic=True, color='64748b')
        ws_az.column_dimensions['A'].width = 70

        ws2 = wb.create_sheet(title='Detalle por fecha')
        ws2['A1'] = f'DETALLE DE FALTAS POR FECHA — {MESES_ES[mes - 1].upper()} {anio}'
        ws2['A1'].font = Font(bold=True, size=14, color='991B1B')
        ws2['A2'] = label_periodo
        ws2['A2'].font = Font(size=9, color='64748b')
        ws2['A4'] = 'Sin faltas registradas en el período.'
        ws2['A4'].font = Font(size=10, italic=True, color='64748b')
        ws2.column_dimensions['A'].width = 70
    else:
        pids = list(faltas_map.keys())
        personal_qs = Personal.objects.filter(id__in=pids).select_related('subarea__area').order_by('apellidos_nombres')

        # Días laborables del mes (approximación: días hábiles lun-sáb)
        dias_laborables = sum(
            1 for d in range(1, mes_fin.day + 1)
            if date(anio, mes, d).weekday() < 6  # lun-sáb
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Faltas {MESES_ES[mes - 1]} {anio}'

        # Estilos
        title_font = Font(bold=True, size=14, color='991B1B')
        sub_font = Font(size=9, color='64748b')
        header_font = Font(bold=True, size=9, color='FFFFFF')
        header_fill = PatternFill(start_color='991B1B', end_color='991B1B', fill_type='solid')
        data_font = Font(size=9)
        warn_font = Font(size=9, bold=True, color='991B1B')
        total_fill = PatternFill(start_color='7F1D1D', end_color='7F1D1D', fill_type='solid')
        total_font = Font(bold=True, size=9, color='FFFFFF')
        border = Border(bottom=Side(style='thin', color='e2e8f0'))
        center = Alignment(horizontal='center')
        alt_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')

        ws.cell(row=1, column=1, value=f'REPORTE DE FALTAS — {MESES_ES[mes - 1].upper()} {anio}').font = title_font
        ws.cell(row=2, column=1, value=f'{label_periodo} | Días laborables aprox.: {dias_laborables}').font = sub_font

        headers = ['N°', 'DNI', 'Apellidos y Nombres', 'Cargo', 'Área', 'Grupo',
                   'FA/F', 'LSG', 'Total Días Desc.',
                   'Sueldo Base', 'Valor Día', 'Descuento Estimado', 'Código S10']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        total_descuento = Decimal('0')
        for i, p in enumerate(personal_qs, 1):
            row = i + 4
            fd = faltas_map[p.pk]
            dias_fa = fd.get('FA', 0) + fd.get('F', 0)
            dias_lsg = fd.get('LSG', 0)
            total_dias_desc = dias_fa + dias_lsg
            sueldo = p.sueldo_base or Decimal('0')
            valor_dia = (sueldo / dias_laborables).quantize(Decimal('0.01')) if dias_laborables else Decimal('0')
            descuento = (valor_dia * total_dias_desc).quantize(Decimal('0.01'))
            total_descuento += descuento

            fill = alt_fill if i % 2 == 0 else None
            for c in range(1, 14):
                cell = ws.cell(row=row, column=c)
                if fill:
                    cell.fill = fill
                cell.border = border

            grupo_label = 'RCO' if p.grupo_tareo == 'RCO' else 'Staff'

            ws.cell(row=row, column=1, value=i).font = data_font
            c2 = ws.cell(row=row, column=2, value=p.nro_doc)
            c2.font = data_font
            c2.number_format = '@'  # DNI texto
            ws.cell(row=row, column=3, value=p.apellidos_nombres).font = data_font
            ws.cell(row=row, column=4, value=p.cargo or '').font = Font(size=8, color='64748b')
            ws.cell(row=row, column=5, value=p.subarea.area.nombre if p.subarea else '').font = Font(size=8, color='64748b')
            ws.cell(row=row, column=6, value=grupo_label).font = Font(size=8, bold=True)
            ws.cell(row=row, column=7, value=dias_fa).font = warn_font if dias_fa else data_font
            ws.cell(row=row, column=8, value=dias_lsg).font = warn_font if dias_lsg else data_font
            ws.cell(row=row, column=9, value=total_dias_desc).font = Font(bold=True, size=9)
            ws.cell(row=row, column=10, value=float(sueldo)).font = data_font
            ws.cell(row=row, column=10).number_format = '#,##0.00'
            ws.cell(row=row, column=11, value=float(valor_dia)).font = data_font
            ws.cell(row=row, column=11).number_format = '#,##0.00'
            ws.cell(row=row, column=12, value=float(descuento)).font = Font(bold=True, size=9, color='991B1B')
            ws.cell(row=row, column=12).number_format = '#,##0.00'
            cod_s10 = []
            if dias_fa:
                cod_s10.append('9800-FA')
            if dias_lsg:
                cod_s10.append('9810-LSG')
            ws.cell(row=row, column=13, value=' | '.join(cod_s10)).font = Font(size=8, color='64748b')

            for c in [6, 7, 8, 9, 10, 11, 12]:
                ws.cell(row=row, column=c).alignment = center

        # Fila totales
        total_row = len(pids) + 5
        for c in range(1, 14):
            ws.cell(row=total_row, column=c).fill = total_fill
            ws.cell(row=total_row, column=c).font = total_font
        ws.cell(row=total_row, column=3, value='TOTALES').alignment = center
        ws.cell(row=total_row, column=7, value=sum(fd.get('FA', 0) + fd.get('F', 0) for fd in faltas_map.values())).alignment = center
        ws.cell(row=total_row, column=8, value=sum(fd.get('LSG', 0) for fd in faltas_map.values())).alignment = center
        ws.cell(row=total_row, column=9, value=sum(fd.get('FA', 0) + fd.get('F', 0) + fd.get('LSG', 0) for fd in faltas_map.values())).alignment = center
        ws.cell(row=total_row, column=12, value=float(total_descuento)).number_format = '#,##0.00'
        ws.cell(row=total_row, column=12).alignment = center

        ws.cell(row=total_row + 2, column=1,
                value=f'Total trabajadores con descuentos: {len(pids)} | Descuento total estimado: S/ {total_descuento:,.2f}').font = Font(size=9, color='991B1B', bold=True)

        # Anchos de columna
        widths = [5, 12, 38, 25, 20, 8, 8, 8, 10, 14, 12, 18, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.freeze_panes = 'A5'

        # Constantes compartidas para hojas de detalle
        dias_es = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        cod_color = {
            'FA':  Font(size=9, bold=True, color='991B1B'),
            'F':   Font(size=9, bold=True, color='991B1B'),
            'LSG': Font(size=9, bold=True, color='7C2D12'),
        }
        # Resaltado tenue para alternar grupos por trabajador en la hoja A-Z
        worker_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2',
                                   fill_type='solid')

        # ──────────────────────────────────────────────────────────
        # HOJA 2 — Detalle por Trabajador (A-Z)
        # Cada falta del trabajador en filas consecutivas, ordenado
        # alfabéticamente por apellidos. Filtros automáticos en headers.
        # ──────────────────────────────────────────────────────────
        regs_az = sorted(
            regs_detalle,
            key=lambda r: (
                (r.personal.apellidos_nombres or '').upper(),
                r.fecha,
            ),
        )

        ws_az = wb.create_sheet(title='Detalle A-Z')
        ws_az.cell(row=1, column=1,
                   value=f'DETALLE POR TRABAJADOR (A-Z) — {MESES_ES[mes - 1].upper()} {anio}'
                   ).font = title_font
        ws_az.cell(row=2, column=1, value=label_periodo).font = sub_font

        headers_az = ['N°', 'Apellidos y Nombres', 'DNI', 'Cargo', 'Área',
                      'Grupo', 'Condición', 'Fecha', 'Día', 'Código',
                      'Observaciones']
        for c, h in enumerate(headers_az, 1):
            cell = ws_az.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Pintar filas alternando por trabajador (no por fila), para que
        # cada bloque de un mismo trabajador sea visualmente identificable.
        prev_pid = None
        bloque_alt = False
        for i, r in enumerate(regs_az, 1):
            row = i + 4
            p = r.personal
            if p.pk != prev_pid:
                bloque_alt = not bloque_alt
                prev_pid = p.pk
            fill = worker_fill if bloque_alt else None
            for c in range(1, 12):
                cell = ws_az.cell(row=row, column=c)
                if fill:
                    cell.fill = fill
                cell.border = border

            grupo_label = 'RCO' if p.grupo_tareo == 'RCO' else 'Staff'

            ws_az.cell(row=row, column=1, value=i).font = data_font
            ws_az.cell(row=row, column=2, value=p.apellidos_nombres).font = data_font
            c_dni = ws_az.cell(row=row, column=3, value=p.nro_doc)
            c_dni.font = data_font
            c_dni.number_format = '@'
            ws_az.cell(row=row, column=4, value=p.cargo or '').font = Font(size=8, color='64748b')
            ws_az.cell(row=row, column=5, value=p.subarea.area.nombre if p.subarea else '').font = Font(size=8, color='64748b')
            ws_az.cell(row=row, column=6, value=grupo_label).font = Font(size=8, bold=True)
            ws_az.cell(row=row, column=7, value=r.condicion or '').font = Font(size=8, color='64748b')
            c_fecha_az = ws_az.cell(row=row, column=8, value=r.fecha)
            c_fecha_az.font = data_font
            c_fecha_az.number_format = 'dd/mm/yyyy'
            ws_az.cell(row=row, column=9, value=dias_es[r.fecha.weekday()]).font = data_font
            ws_az.cell(row=row, column=10, value=r.codigo_dia).font = cod_color.get(r.codigo_dia, data_font)
            ws_az.cell(row=row, column=11, value=(r.observaciones or '')[:100]).font = Font(size=8, color='64748b')

            for c in [1, 3, 6, 7, 8, 9, 10]:
                ws_az.cell(row=row, column=c).alignment = center

        # Totales hoja A-Z
        total_row_az = len(regs_az) + 5
        for c in range(1, 12):
            ws_az.cell(row=total_row_az, column=c).fill = total_fill
            ws_az.cell(row=total_row_az, column=c).font = total_font
        ws_az.cell(row=total_row_az, column=2, value='TOTAL REGISTROS').alignment = center
        ws_az.cell(row=total_row_az, column=10, value=len(regs_az)).alignment = center

        # Anchos hoja A-Z
        widths_az = [5, 38, 12, 25, 20, 8, 10, 12, 6, 8, 40]
        for i, w in enumerate(widths_az, 1):
            ws_az.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # AutoFilter (clic en headers para filtrar/ordenar dentro de Excel)
        if regs_az:
            ws_az.auto_filter.ref = f'A4:K{len(regs_az) + 4}'
        ws_az.freeze_panes = 'A5'

        # ──────────────────────────────────────────────────────────
        # HOJA 3 — Detalle por Fecha
        # ──────────────────────────────────────────────────────────
        ws2 = wb.create_sheet(title='Detalle por fecha')

        ws2.cell(row=1, column=1,
                 value=f'DETALLE DE FALTAS POR FECHA — {MESES_ES[mes - 1].upper()} {anio}').font = title_font
        ws2.cell(row=2, column=1, value=label_periodo).font = sub_font

        headers2 = ['N°', 'Fecha', 'Día', 'DNI', 'Apellidos y Nombres',
                    'Cargo', 'Área', 'Grupo', 'Condición', 'Código', 'Observaciones']
        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for i, r in enumerate(regs_detalle, 1):
            row = i + 4
            p = r.personal
            fill = alt_fill if i % 2 == 0 else None
            for c in range(1, 12):
                cell = ws2.cell(row=row, column=c)
                if fill:
                    cell.fill = fill
                cell.border = border

            grupo_label = 'RCO' if p.grupo_tareo == 'RCO' else 'Staff'

            ws2.cell(row=row, column=1, value=i).font = data_font
            c_fecha = ws2.cell(row=row, column=2, value=r.fecha)
            c_fecha.font = data_font
            c_fecha.number_format = 'dd/mm/yyyy'
            ws2.cell(row=row, column=3, value=dias_es[r.fecha.weekday()]).font = data_font
            c_dni = ws2.cell(row=row, column=4, value=p.nro_doc)
            c_dni.font = data_font
            c_dni.number_format = '@'
            ws2.cell(row=row, column=5, value=p.apellidos_nombres).font = data_font
            ws2.cell(row=row, column=6, value=p.cargo or '').font = Font(size=8, color='64748b')
            ws2.cell(row=row, column=7, value=p.subarea.area.nombre if p.subarea else '').font = Font(size=8, color='64748b')
            ws2.cell(row=row, column=8, value=grupo_label).font = Font(size=8, bold=True)
            ws2.cell(row=row, column=9, value=r.condicion or '').font = Font(size=8, color='64748b')
            ws2.cell(row=row, column=10, value=r.codigo_dia).font = cod_color.get(r.codigo_dia, data_font)
            ws2.cell(row=row, column=11, value=(r.observaciones or '')[:100]).font = Font(size=8, color='64748b')

            for c in [1, 2, 3, 4, 8, 9, 10]:
                ws2.cell(row=row, column=c).alignment = center

        # Totales hoja 2
        total_row2 = len(regs_detalle) + 5
        for c in range(1, 12):
            ws2.cell(row=total_row2, column=c).fill = total_fill
            ws2.cell(row=total_row2, column=c).font = total_font
        ws2.cell(row=total_row2, column=5, value='TOTAL REGISTROS').alignment = center
        ws2.cell(row=total_row2, column=10, value=len(regs_detalle)).alignment = center

        # Anchos
        widths2 = [5, 12, 6, 12, 38, 25, 20, 8, 10, 8, 40]
        for i, w in enumerate(widths2, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws2.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    sufijo = '_corte' if tipo_periodo == 'corte' else ''
    response['Content-Disposition'] = f'attachment; filename="Faltas_{MESES_ES[mes - 1]}_{anio}{sufijo}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# PLANILLA CONSOLIDADA PARA S10 (todos los conceptos de nómina)
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_planilla_consolidada(request):
    """
    Excel consolidado con todos los conceptos por trabajador para importar en S10.
    Incluye: HE25/HE35/HE100, FA, LSG, DM, VAC, sueldo base.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from asistencia.models import RegistroTareo
    from personal.models import Personal

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')
    if tipo_periodo not in ('calendario', 'corte'):
        tipo_periodo = 'calendario'

    d_ini, d_fin = _get_corte_config(request)
    mes_ini, mes_fin = _calcular_periodo(anio, mes, tipo_periodo, d_ini, d_fin)
    label_periodo = _label_periodo(mes_ini, mes_fin, tipo_periodo)

    # Todos los registros del mes
    qs = (
        RegistroTareo.objects.filter(
            fecha__gte=mes_ini, fecha__lte=mes_fin,
            personal__isnull=False,
        )
        .values('personal_id', 'codigo_dia', 'personal__grupo_tareo')
        .annotate(
            cnt=Count('id'),
            sum_he25=Sum('he_25'),
            sum_he35=Sum('he_35'),
            sum_he100=Sum('he_100'),
        )
    )

    # Agrupar por personal
    datos: dict[int, dict] = {}
    for r in qs:
        pid = r['personal_id']
        if pid not in datos:
            datos[pid] = {
                'grupo': r['personal__grupo_tareo'] or '',
                'dias_trab': 0, 'fa': 0, 'lsg': 0, 'dm': 0,
                'vac': 0, 'dl': 0,
                'he25': Decimal('0'), 'he35': Decimal('0'), 'he100': Decimal('0'),
            }
        cod = (r['codigo_dia'] or '').upper()
        if cod in ('T', 'NOR', 'TR', 'A', 'SS', 'CDT', 'CPF', 'LCG', 'ATM', 'CHE', 'LIM'):
            datos[pid]['dias_trab'] += r['cnt']
        elif cod in ('FA', 'F'):
            datos[pid]['fa'] += r['cnt']
        elif cod == 'LSG':
            datos[pid]['lsg'] += r['cnt']
        elif cod == 'DM':
            datos[pid]['dm'] += r['cnt']
        elif cod in ('VAC', 'V'):
            datos[pid]['vac'] += r['cnt']
        elif cod in ('DL', 'DLA', 'B'):
            datos[pid]['dl'] += r['cnt']
        datos[pid]['he25'] += r['sum_he25'] or Decimal('0')
        datos[pid]['he35'] += r['sum_he35'] or Decimal('0')
        datos[pid]['he100'] += r['sum_he100'] or Decimal('0')

    pids = list(datos.keys())
    personal_qs = Personal.objects.filter(id__in=pids).select_related('subarea__area').order_by('apellidos_nombres')

    # Días laborables del mes
    dias_laborables = sum(
        1 for d in range(1, mes_fin.day + 1)
        if date(anio, mes, d).weekday() < 6
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Planilla {MESES_ES[mes - 1]} {anio}'

    title_font = Font(bold=True, size=14, color='1F4E79')
    sub_font = Font(size=9, color='64748b')
    header_font = Font(bold=True, size=9, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    data_font = Font(size=9)
    total_fill = PatternFill(start_color='17375E', end_color='17375E', fill_type='solid')
    total_font = Font(bold=True, size=9, color='FFFFFF')
    border = Border(bottom=Side(style='thin', color='e2e8f0'))
    center = Alignment(horizontal='center')
    alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    rco_fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')

    ws.cell(row=1, column=1, value=f'PLANILLA CONSOLIDADA — {MESES_ES[mes - 1].upper()} {anio}').font = title_font
    ws.cell(row=2, column=1, value=f'{label_periodo} | Días laborables: {dias_laborables}').font = sub_font
    ws.cell(row=3, column=1, value='INGRESOS: columnas amarillas | DESCUENTOS: columnas rojas | HE solo para grupo RCO').font = Font(size=8, italic=True, color='64748b')

    headers = [
        'N°', 'DNI/CE', 'Apellidos y Nombres', 'Cargo', 'Área', 'Grupo',
        'Sueldo Base', 'Días Trab.',
        # Descuentos (días)
        'Faltas (FA)', 'LSG', 'DM',
        # Ingresos/Neutros
        'Vacaciones', 'DL/Bajadas',
        # HE (solo RCO)
        'HE 25% (h)', 'HE 35% (h)', 'HE 100% (h)',
        # Cálculos estimados
        'Desc. FA+LSG', 'Total HE aprox.',
        'Concepto S10',
    ]

    # Colores de encabezado por tipo
    desc_fill = PatternFill(start_color='7F1D1D', end_color='7F1D1D', fill_type='solid')
    ing_fill = PatternFill(start_color='14532D', end_color='14532D', fill_type='solid')
    he_fill = PatternFill(start_color='78350F', end_color='78350F', fill_type='solid')
    calc_fill = PatternFill(start_color='312E81', end_color='312E81', fill_type='solid')

    col_fills = {
        9: desc_fill, 10: desc_fill, 11: desc_fill,  # FA, LSG, DM
        12: ing_fill, 13: ing_fill,                   # VAC, DL
        14: he_fill, 15: he_fill, 16: he_fill,        # HE
        17: calc_fill, 18: calc_fill, 19: calc_fill,  # cálculos
    }

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = header_font
        cell.fill = col_fills.get(c, header_fill)
        cell.alignment = center

    # Factor sueldo mensual → diario → hora (8h/día)
    for i, p in enumerate(personal_qs, 1):
        row = i + 5
        d = datos[p.pk]
        sueldo = p.sueldo_base or Decimal('0')
        valor_dia = (sueldo / dias_laborables).quantize(Decimal('0.01')) if dias_laborables else Decimal('0')
        valor_hora = (sueldo / (dias_laborables * 8)).quantize(Decimal('0.01')) if dias_laborables else Decimal('0')
        desc_fa_lsg = (valor_dia * (d['fa'] + d['lsg'])).quantize(Decimal('0.01'))
        # HE: 25%→+25%, 35%→+35%, 100%→+100% sobre valor_hora
        total_he = (
            d['he25'] * valor_hora * Decimal('1.25') +
            d['he35'] * valor_hora * Decimal('1.35') +
            d['he100'] * valor_hora * Decimal('2.00')
        ).quantize(Decimal('0.01'))

        row_fill = rco_fill if d['grupo'] == 'RCO' else (alt_fill if i % 2 == 0 else None)
        for c in range(1, 20):
            cell = ws.cell(row=row, column=c)
            if row_fill:
                cell.fill = row_fill
            cell.border = border

        ws.cell(row=row, column=1, value=i).font = data_font
        c2 = ws.cell(row=row, column=2, value=p.nro_doc)
        c2.font = data_font
        c2.number_format = '@'  # DNI/CE texto
        ws.cell(row=row, column=3, value=p.apellidos_nombres).font = data_font
        ws.cell(row=row, column=4, value=p.cargo or '').font = Font(size=8, color='64748b')
        ws.cell(row=row, column=5, value=p.subarea.area.nombre if p.subarea else '').font = Font(size=8, color='64748b')
        ws.cell(row=row, column=6, value=d['grupo']).font = Font(size=8, bold=True)
        ws.cell(row=row, column=7, value=float(sueldo)).number_format = '#,##0.00'
        ws.cell(row=row, column=7).font = data_font
        ws.cell(row=row, column=8, value=d['dias_trab']).font = data_font
        ws.cell(row=row, column=9, value=d['fa']).font = Font(size=9, bold=True, color='991B1B') if d['fa'] else data_font
        ws.cell(row=row, column=10, value=d['lsg']).font = Font(size=9, bold=True, color='991B1B') if d['lsg'] else data_font
        ws.cell(row=row, column=11, value=d['dm']).font = data_font
        ws.cell(row=row, column=12, value=d['vac']).font = data_font
        ws.cell(row=row, column=13, value=d['dl']).font = data_font
        ws.cell(row=row, column=14, value=float(d['he25'])).font = Font(size=9, color='92400E') if d['he25'] else data_font
        ws.cell(row=row, column=15, value=float(d['he35'])).font = Font(size=9, color='92400E') if d['he35'] else data_font
        ws.cell(row=row, column=16, value=float(d['he100'])).font = Font(size=9, color='92400E') if d['he100'] else data_font
        ws.cell(row=row, column=17, value=float(desc_fa_lsg)).number_format = '#,##0.00'
        ws.cell(row=row, column=17).font = Font(size=9, bold=True, color='991B1B') if desc_fa_lsg else data_font
        ws.cell(row=row, column=18, value=float(total_he)).number_format = '#,##0.00'
        ws.cell(row=row, column=18).font = Font(size=9, bold=True, color='78350F') if total_he else data_font

        # Conceptos S10 aplicables
        conceptos = []
        if d['fa']:
            conceptos.append(f'FA:{d["fa"]}d')
        if d['lsg']:
            conceptos.append(f'LSG:{d["lsg"]}d')
        if d['dm']:
            conceptos.append(f'DM:{d["dm"]}d')
        if d['vac']:
            conceptos.append(f'VAC:{d["vac"]}d')
        if d['he25']:
            conceptos.append(f'HE25:{float(d["he25"]):.1f}h')
        if d['he35']:
            conceptos.append(f'HE35:{float(d["he35"]):.1f}h')
        if d['he100']:
            conceptos.append(f'HE100:{float(d["he100"]):.1f}h')
        ws.cell(row=row, column=19, value=' | '.join(conceptos) if conceptos else 'Sin novedades').font = Font(size=8, color='64748b')

        for c in range(8, 19):
            ws.cell(row=row, column=c).alignment = center

    # Fila totales
    total_row = len(pids) + 6
    for c in range(1, 20):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).font = total_font
        ws.cell(row=total_row, column=c).alignment = center
    ws.cell(row=total_row, column=3, value='TOTALES')
    ws.cell(row=total_row, column=9, value=sum(d['fa'] for d in datos.values()))
    ws.cell(row=total_row, column=10, value=sum(d['lsg'] for d in datos.values()))
    ws.cell(row=total_row, column=11, value=sum(d['dm'] for d in datos.values()))
    ws.cell(row=total_row, column=12, value=sum(d['vac'] for d in datos.values()))
    ws.cell(row=total_row, column=14, value=float(sum(d['he25'] for d in datos.values())))
    ws.cell(row=total_row, column=15, value=float(sum(d['he35'] for d in datos.values())))
    ws.cell(row=total_row, column=16, value=float(sum(d['he100'] for d in datos.values())))
    for c in [14, 15, 16]:
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'

    ws.cell(row=total_row + 2, column=1,
            value=f'Total trabajadores: {len(pids)} | Fondo amarillo = RCO | * Montos estimados, verificar con S10').font = Font(size=9, color='64748b', italic=True)

    # Anchos de columna
    widths = [5, 12, 38, 25, 20, 8, 14, 9, 9, 8, 8, 10, 10, 12, 12, 12, 15, 15, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = 'A6'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    sufijo = '_corte' if tipo_periodo == 'corte' else ''
    response['Content-Disposition'] = f'attachment; filename="Planilla_{MESES_ES[mes - 1]}_{anio}{sufijo}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# REPORTE DE VALIDACIÓN / COHERENCIA DE DATOS (detección de errores)
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_validacion_datos(request):
    """
    Excel con errores y anomalías detectadas en los registros de asistencia del mes.
    Detecta: horas > 24, SS sin entrada, DNI duplicados, HE sospechosas, etc.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from asistencia.models import RegistroTareo
    from personal.models import Personal

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')
    if tipo_periodo not in ('calendario', 'corte'):
        tipo_periodo = 'calendario'

    d_ini, d_fin = _get_corte_config(request)
    mes_ini, mes_fin = _calcular_periodo(anio, mes, tipo_periodo, d_ini, d_fin)
    label_periodo = _label_periodo(mes_ini, mes_fin, tipo_periodo)

    errores = []  # lista de dicts: {tipo, severidad, dni, nombre, fecha, detalle}

    # ── 1. Horas marcadas > 24 en un día ──────────────────────────────────
    for r in RegistroTareo.objects.filter(
        fecha__gte=mes_ini, fecha__lte=mes_fin,
        horas_marcadas__gt=24,
    ).select_related('personal').values(
        'personal__nro_doc', 'personal__apellidos_nombres', 'fecha', 'horas_marcadas', 'codigo_dia'
    ):
        errores.append({
            'tipo': 'HORAS > 24h',
            'severidad': 'ERROR',
            'dni': r['personal__nro_doc'] or '',
            'nombre': r['personal__apellidos_nombres'] or '',
            'fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
            'detalle': f"Horas marcadas: {r['horas_marcadas']} h (código: {r['codigo_dia']})",
        })

    # ── 2. HE totales > 12h en un día ─────────────────────────────────────
    from django.db.models import F as DbF, ExpressionWrapper, DecimalField as DField
    for r in RegistroTareo.objects.filter(
        fecha__gte=mes_ini, fecha__lte=mes_fin,
    ).annotate(
        total_he=ExpressionWrapper(
            DbF('he_25') + DbF('he_35') + DbF('he_100'),
            output_field=DField(max_digits=6, decimal_places=2)
        )
    ).filter(total_he__gt=12).values(
        'personal__nro_doc', 'personal__apellidos_nombres', 'fecha', 'total_he', 'codigo_dia'
    ):
        errores.append({
            'tipo': 'HE > 12h/día',
            'severidad': 'ALERTA',
            'dni': r['personal__nro_doc'] or '',
            'nombre': r['personal__apellidos_nombres'] or '',
            'fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
            'detalle': f"HE totales: {r['total_he']} h (código: {r['codigo_dia']})",
        })

    # ── 3. Código SS (sin salida) ──────────────────────────────────────────
    for r in RegistroTareo.objects.filter(
        fecha__gte=mes_ini, fecha__lte=mes_fin,
        codigo_dia='SS',
    ).values('personal__nro_doc', 'personal__apellidos_nombres', 'fecha'):
        errores.append({
            'tipo': 'SIN SALIDA (SS)',
            'severidad': 'ALERTA',
            'dni': r['personal__nro_doc'] or '',
            'nombre': r['personal__apellidos_nombres'] or '',
            'fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
            'detalle': 'Registro con código SS — verificar marcación de salida',
        })

    # ── 4. DNI duplicados en Personal activo ──────────────────────────────
    from django.db.models import Count as DCount
    duplicados = (
        Personal.objects.values('nro_doc')
        .annotate(cnt=DCount('id'))
        .filter(cnt__gt=1, nro_doc__isnull=False)
        .exclude(nro_doc='')
    )
    for dup in duplicados:
        personas = Personal.objects.filter(nro_doc=dup['nro_doc']).values('apellidos_nombres', 'situacion')
        nombres = ' / '.join(f"{p['apellidos_nombres']} ({p['situacion']})" for p in personas)
        errores.append({
            'tipo': 'DNI DUPLICADO',
            'severidad': 'ERROR',
            'dni': dup['nro_doc'],
            'nombre': nombres,
            'fecha': '',
            'detalle': f"{dup['cnt']} trabajadores con el mismo DNI/CE",
        })

    # ── 5. Trabajadores sin DNI ───────────────────────────────────────────
    for p in Personal.objects.filter(
        Q(nro_doc__isnull=True) | Q(nro_doc=''),
        estado='Activo',
    ).values('apellidos_nombres', 'cargo'):
        errores.append({
            'tipo': 'SIN DNI/CE',
            'severidad': 'ERROR',
            'dni': '—',
            'nombre': p['apellidos_nombres'],
            'fecha': '',
            'detalle': f"Cargo: {p['cargo'] or 'No especificado'} — asignar DNI/CE antes del cierre",
        })

    # ── 6. FA en días de descanso (domingo) ───────────────────────────────
    for r in RegistroTareo.objects.filter(
        fecha__gte=mes_ini, fecha__lte=mes_fin,
        codigo_dia__in=['FA', 'F'],
        dia_semana=6,  # domingo
        condicion__in=['', 'LOCAL', 'LIMA'],
    ).values('personal__nro_doc', 'personal__apellidos_nombres', 'fecha'):
        errores.append({
            'tipo': 'FA EN DOMINGO',
            'severidad': 'ALERTA',
            'dni': r['personal__nro_doc'] or '',
            'nombre': r['personal__apellidos_nombres'] or '',
            'fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
            'detalle': 'Falta registrada en día de descanso semanal (domingo) — debería ser DS',
        })

    # ── 7. HE sin días trabajados (HE en día FA/DM/VAC) ───────────────────
    for r in RegistroTareo.objects.filter(
        fecha__gte=mes_ini, fecha__lte=mes_fin,
        codigo_dia__in=['FA', 'F', 'DM', 'LSG'],
    ).annotate(
        total_he=ExpressionWrapper(
            DbF('he_25') + DbF('he_35') + DbF('he_100'),
            output_field=DField(max_digits=6, decimal_places=2)
        )
    ).filter(total_he__gt=0).values(
        'personal__nro_doc', 'personal__apellidos_nombres', 'fecha', 'codigo_dia', 'total_he'
    ):
        errores.append({
            'tipo': 'HE EN DÍA NO LABORAL',
            'severidad': 'ERROR',
            'dni': r['personal__nro_doc'] or '',
            'nombre': r['personal__apellidos_nombres'] or '',
            'fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
            'detalle': f"HE={r['total_he']}h en día {r['codigo_dia']} — verificar importación",
        })

    # ── Generar Excel ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Validación {MESES_ES[mes - 1]} {anio}'

    title_font = Font(bold=True, size=14, color='7C3AED')
    sub_font = Font(size=9, color='64748b')
    header_font = Font(bold=True, size=9, color='FFFFFF')
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    error_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    alerta_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    error_font = Font(size=9, bold=True, color='991B1B')
    alerta_font = Font(size=9, bold=True, color='92400E')
    data_font = Font(size=9)
    border = Border(bottom=Side(style='thin', color='e2e8f0'))
    center = Alignment(horizontal='center')

    ws.cell(row=1, column=1, value=f'REPORTE DE VALIDACIÓN — {MESES_ES[mes - 1].upper()} {anio}').font = title_font
    ws.cell(row=2, column=1, value=f'{label_periodo} | Total anomalías detectadas: {len(errores)}').font = sub_font

    if not errores:
        ws.cell(row=4, column=1, value='✓ Sin anomalías detectadas en el período.').font = Font(bold=True, size=11, color='16A34A')
    else:
        headers = ['N°', 'Severidad', 'Tipo de Error', 'DNI/CE', 'Apellidos y Nombres', 'Fecha', 'Detalle']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for i, err in enumerate(errores, 1):
            row = i + 4
            es_error = err['severidad'] == 'ERROR'
            fill = error_fill if es_error else alerta_fill
            font_sev = error_font if es_error else alerta_font

            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = fill
                ws.cell(row=row, column=c).border = border

            ws.cell(row=row, column=1, value=i).font = data_font
            ws.cell(row=row, column=2, value=err['severidad']).font = font_sev
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=3, value=err['tipo']).font = font_sev
            c4 = ws.cell(row=row, column=4, value=err['dni'])
            c4.font = Font(size=9, name='Courier New')
            c4.number_format = '@'  # DNI texto
            ws.cell(row=row, column=5, value=err['nombre']).font = data_font
            ws.cell(row=row, column=6, value=err['fecha']).font = data_font
            ws.cell(row=row, column=6).alignment = center
            ws.cell(row=row, column=7, value=err['detalle']).font = Font(size=8, color='374151')

        # Resumen por tipo
        resumen_row = len(errores) + 6
        ws.cell(row=resumen_row, column=1, value='RESUMEN POR TIPO').font = Font(bold=True, size=10, color='7C3AED')
        tipos = {}
        for e in errores:
            tipos[e['tipo']] = tipos.get(e['tipo'], 0) + 1
        for j, (tipo, cnt) in enumerate(sorted(tipos.items()), 1):
            ws.cell(row=resumen_row + j, column=1, value=tipo).font = data_font
            ws.cell(row=resumen_row + j, column=2, value=cnt).font = Font(bold=True, size=9)

    # Anchos
    widths = [5, 10, 22, 14, 38, 12, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    sufijo = '_corte' if tipo_periodo == 'corte' else ''
    response['Content-Disposition'] = f'attachment; filename="Validacion_{MESES_ES[mes - 1]}_{anio}{sufijo}.xlsx"'
    return response
