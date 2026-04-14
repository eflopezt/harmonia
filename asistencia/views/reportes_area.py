"""
Módulo: Reportes de Asistencia por Área
========================================
Permite generar y enviar reportes de asistencia agrupados por área,
con destinos configurables (jefaturas + CC) y rango de fechas flexible.

Funcionalidades:
  - Panel principal con todas las áreas y su configuración
  - Guardar destinatarios por área (jefatura + CC)
  - Descargar ZIP con PDFs agrupados por área / subcarpeta
  - Enviar reportes por email a jefaturas con cuerpo personalizable
  - Rangos: semanal, quincenal, mensual, personalizado
"""
from __future__ import annotations

import io
import json
import logging
from datetime import date, time, timedelta
from zipfile import ZipFile

from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from asistencia.views._common import solo_admin
from personal.models import Area, Personal

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# Helpers internos de PDF (reutiliza lógica de reporte_individual)
# ─────────────────────────────────────────────────────────────

def _pdf_empleado_rango(personal, inicio: date, fin: date) -> bytes | None:
    """
    Genera el PDF de asistencia de un empleado para un rango de fechas
    arbitrario (semanal, quincenal, mensual, personalizado).
    """
    from asistencia.views.reporte_individual import (
        _build_staff_data, _build_rco_data,
        _render_staff_html, _render_rco_html,
        _get_papeletas, _render_pdf,
    )
    grupo = personal.grupo_tareo or 'STAFF'
    mes = inicio.month
    anio = inicio.year
    papeletas = _get_papeletas(personal, inicio, fin)
    try:
        if grupo == 'RCO':
            dias, totales = _build_rco_data(personal, inicio, fin)
            html = _render_rco_html(personal, dias, totales, papeletas, inicio, fin, mes, anio)
        else:
            dias, conteo = _build_staff_data(personal, inicio, fin)
            html = _render_staff_html(personal, dias, conteo, papeletas, inicio, fin, mes, anio)
        return _render_pdf(html)
    except Exception as e:
        logger.error(f'Error generando PDF para {personal.nro_doc}: {e}')
        return None


def _get_rango(tipo_periodo: str, fecha_inicio_str: str = '', fecha_fin_str: str = '') -> tuple[date, date]:
    """Calcula inicio/fin según el tipo de periodo seleccionado."""
    hoy = date.today()

    if tipo_periodo == 'SEMANAL':
        # Semana pasada: lun–dom
        lun = hoy - timedelta(days=hoy.weekday() + 7)
        return lun, lun + timedelta(days=6)

    if tipo_periodo == 'QUINCENAL':
        if hoy.day <= 15:
            # Primera quincena del mes anterior
            mes_ant = (hoy.replace(day=1) - timedelta(days=1))
            return mes_ant.replace(day=16), mes_ant.replace(day=mes_ant.day)
        else:
            return hoy.replace(day=1), hoy.replace(day=15)

    if tipo_periodo == 'MENSUAL':
        # Mes anterior completo
        primer_dia = hoy.replace(day=1)
        fin_mes = primer_dia - timedelta(days=1)
        return fin_mes.replace(day=1), fin_mes

    if tipo_periodo == 'ESTA_SEMANA':
        lun = hoy - timedelta(days=hoy.weekday())
        return lun, hoy

    if tipo_periodo == 'ESTE_MES':
        return hoy.replace(day=1), hoy

    # PERSONALIZADO — usa las fechas del request
    try:
        ini = date.fromisoformat(fecha_inicio_str)
        fin = date.fromisoformat(fecha_fin_str)
        return ini, fin
    except (ValueError, TypeError):
        # Default: últimos 7 días
        return hoy - timedelta(days=6), hoy


def _periodo_label(inicio: date, fin: date) -> str:
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if inicio.month == fin.month and inicio.year == fin.year:
        return f'{inicio.day}–{fin.day} de {meses[inicio.month]} {inicio.year}'
    return f'{inicio.strftime("%d/%m/%Y")} – {fin.strftime("%d/%m/%Y")}'


def _get_empresa() -> str:
    try:
        from asistencia.models import ConfiguracionSistema
        cfg = ConfiguracionSistema.get()
        return cfg.nombre_empresa or ''
    except Exception:
        return ''


# ─────────────────────────────────────────────────────────────
# VISTAS
# ─────────────────────────────────────────────────────────────

@login_required
@solo_admin
def panel_reportes_area(request):
    """Panel principal: muestra todas las áreas con sus configs y acciones."""
    from asistencia.models import ConfiguracionReporteArea, EnvioReporteArea
    from django.db.models import Max

    areas = Area.objects.filter(activa=True).order_by('nombre')

    # Cargar configuraciones existentes
    configs = {c.area_id: c for c in ConfiguracionReporteArea.objects.all()}

    # Último envío por área
    ultimos = {
        r['area_id']: r['ultimo']
        for r in EnvioReporteArea.objects.filter(
            estado__in=['ENVIADO', 'PARCIAL']
        ).values('area_id').annotate(ultimo=Max('creado_en'))
    }

    area_data = []
    for area in areas:
        n_empleados = Personal.objects.filter(
            subarea__area=area, estado='Activo'
        ).count()
        if n_empleados == 0:
            continue
        cfg = configs.get(area.pk)
        ultimo_envio = ultimos.get(area.pk)

        # Historial de emails (últimos 3)
        historial = []
        if cfg and cfg.historial_emails:
            historial = cfg.historial_emails[-3:][::-1]  # más recientes primero

        area_data.append({
            'area': area,
            'n_empleados': n_empleados,
            'cfg': cfg,
            'nombre_jefe': cfg.nombre_jefe if cfg else '',
            'emails_jefatura': cfg.get_emails_jefatura() if cfg else [],
            'emails_cc': cfg.get_emails_cc() if cfg else [],
            'asunto_template': cfg.asunto_template if cfg else 'Reporte de Asistencia - {area} - {periodo}',
            'cuerpo_template': cfg.cuerpo_template if cfg else '',
            'activo': cfg.activo if cfg else False,
            'tiene_config': bool(cfg),
            'ultimo_envio': ultimo_envio,
            'historial': historial,
        })

    context = {
        'area_data': area_data,
        'hoy': date.today(),
        'empresa': _get_empresa(),
        'titulo': 'Reportes de Asistencia por Área',
    }
    return render(request, 'asistencia/reportes_area_panel.html', context)


@login_required
@solo_admin
@require_POST
def configurar_area(request, area_id):
    """Guarda/actualiza la configuración de destinatarios de un área, con historial."""
    from asistencia.models import ConfiguracionReporteArea
    import re
    from django.utils import timezone

    area = get_object_or_404(Area, pk=area_id)

    def parse_emails(raw: str) -> list[str]:
        parts = re.split(r'[,;\n\r]+', raw or '')
        return [e.strip() for e in parts if '@' in e.strip()]

    emails_jefatura = parse_emails(request.POST.get('emails_jefatura', ''))
    emails_cc       = parse_emails(request.POST.get('emails_cc', ''))
    asunto_template = request.POST.get('asunto_template', '').strip() or \
                      'Reporte de Asistencia - {area} - {periodo}'
    cuerpo_template = request.POST.get('cuerpo_template', '').strip()
    nombre_jefe     = request.POST.get('nombre_jefe', '').strip()
    activo          = request.POST.get('activo', '1') == '1'

    cfg, created = ConfiguracionReporteArea.objects.get_or_create(area=area)

    # Registrar cambio en historial si los emails cambiaron
    emails_antes = cfg.emails_jefatura or []
    cc_antes     = cfg.emails_cc or []
    hubo_cambio  = sorted(emails_antes) != sorted(emails_jefatura) or \
                   sorted(cc_antes) != sorted(emails_cc)

    if hubo_cambio or created:
        entrada = {
            'fecha':          timezone.now().strftime('%d/%m/%Y %H:%M'),
            'accion':         'creacion' if created else 'actualizacion',
            'emails_antes':   emails_antes,
            'emails_despues': emails_jefatura,
            'cc_antes':       cc_antes,
            'cc_despues':     emails_cc,
            'usuario':        request.user.username if request.user.is_authenticated else 'sistema',
        }
        historial = list(cfg.historial_emails or [])
        historial.append(entrada)
        cfg.historial_emails = historial[-50:]  # máx 50 entradas

    cfg.emails_jefatura = emails_jefatura
    cfg.emails_cc       = emails_cc
    cfg.asunto_template = asunto_template
    cfg.cuerpo_template = cuerpo_template
    cfg.nombre_jefe     = nombre_jefe or cfg.nombre_jefe
    cfg.activo          = activo
    cfg.save()

    return JsonResponse({
        'ok':           True,
        'created':      created,
        'cambio':       hubo_cambio,
        'area':         area.nombre,
        'nombre_jefe':  cfg.nombre_jefe,
        'destinatarios': len(emails_jefatura),
        'cc':           len(emails_cc),
    })


@login_required
@solo_admin
def generar_zip_por_area(request):
    """
    Genera un ZIP con PDFs agrupados por área:
      reporte_asistencia/
        AREA_NOMBRE/
          12345678_APELLIDO_NOMBRE.pdf
    """
    tipo_periodo  = request.GET.get('tipo_periodo', 'PERSONALIZADO')
    fecha_ini_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    area_ids_str  = request.GET.getlist('areas')  # vacío = todas
    grupo_filter  = request.GET.get('grupo', 'TODOS')

    inicio, fin = _get_rango(tipo_periodo, fecha_ini_str, fecha_fin_str)
    periodo_lbl = _periodo_label(inicio, fin)

    # Empleados activos con area
    qs = Personal.objects.filter(
        estado='Activo',
        subarea__isnull=False,
    ).select_related('subarea__area').order_by(
        'subarea__area__nombre', 'apellidos_nombres'
    )

    if area_ids_str:
        qs = qs.filter(subarea__area_id__in=area_ids_str)
    if grupo_filter != 'TODOS':
        qs = qs.filter(grupo_tareo=grupo_filter)

    zip_buffer = io.BytesIO()
    total = 0
    with ZipFile(zip_buffer, 'w') as zf:
        for emp in qs:
            area_nombre = emp.subarea.area.nombre
            # Nombre de carpeta seguro
            carpeta = slugify(area_nombre, allow_unicode=False).upper().replace('-', '_')
            pdf = _pdf_empleado_rango(emp, inicio, fin)
            if pdf:
                nombre_pdf = (
                    f'{emp.nro_doc}_{emp.apellidos_nombres.replace(",","").replace(" ","_")}.pdf'
                )
                zf.writestr(f'reporte_asistencia/{carpeta}/{nombre_pdf}', pdf)
                total += 1

    if total == 0:
        return HttpResponse('No se encontraron registros para el período seleccionado.', status=404)

    zip_buffer.seek(0)
    fname = f'Asistencia_Areas_{inicio.strftime("%Y%m%d")}_{fin.strftime("%Y%m%d")}.zip'
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
@solo_admin
@require_POST
def enviar_reportes_por_area(request):
    """
    Envía los reportes por email a los destinatarios configurados por área.
    Acepta sobreescritura de destinatarios, CC, asunto y cuerpo desde el form.

    POST params:
      tipo_periodo, fecha_inicio, fecha_fin
      areas[]          → IDs de áreas a enviar (vacío = todas configuradas)
      grupo            → TODOS / STAFF / RCO
      override_emails  → emails adicionales separados por coma (van a todas las áreas)
      override_cc      → CC adicionales
      asunto_override  → Asunto personalizado (si no, usa config de área)
      cuerpo_override  → Cuerpo personalizado
    """
    from asistencia.models import ConfiguracionReporteArea, EnvioReporteArea
    from django.core.mail import get_connection, EmailMessage as DjEmail

    tipo_periodo  = request.POST.get('tipo_periodo', 'PERSONALIZADO')
    fecha_ini_str = request.POST.get('fecha_inicio', '')
    fecha_fin_str = request.POST.get('fecha_fin', '')
    area_ids      = request.POST.getlist('areas')
    grupo_filter  = request.POST.get('grupo', 'TODOS')
    asunto_override = request.POST.get('asunto_override', '').strip()
    cuerpo_override = request.POST.get('cuerpo_override', '').strip()

    def parse_emails_post(key):
        import re
        raw = request.POST.get(key, '')
        return [e.strip() for e in re.split(r'[,;\n\r]+', raw) if '@' in e.strip()]

    override_emails = parse_emails_post('override_emails')
    override_cc     = parse_emails_post('override_cc')

    inicio, fin = _get_rango(tipo_periodo, fecha_ini_str, fecha_fin_str)
    periodo_lbl = _periodo_label(inicio, fin)
    empresa = _get_empresa()

    # Cargar configuraciones
    configs_qs = ConfiguracionReporteArea.objects.filter(activo=True).select_related('area')
    if area_ids:
        configs_qs = configs_qs.filter(area_id__in=area_ids)

    if not configs_qs.exists():
        return JsonResponse({'ok': False, 'error': 'No hay áreas configuradas para envío.'}, status=400)

    resultados = []
    errores_globales = []

    for cfg in configs_qs:
        area = cfg.area

        # Empleados del área en el período
        qs_emp = Personal.objects.filter(
            subarea__area=area,
            estado='Activo',
        ).order_by('apellidos_nombres')
        if grupo_filter != 'TODOS':
            qs_emp = qs_emp.filter(grupo_tareo=grupo_filter)

        if not qs_emp.exists():
            resultados.append({
                'area': area.nombre,
                'estado': 'sin_empleados',
                'enviado': False,
            })
            continue

        # Generar ZIP del área
        zip_buf = io.BytesIO()
        n_pdfs = 0
        with ZipFile(zip_buf, 'w') as zf:
            for emp in qs_emp:
                pdf = _pdf_empleado_rango(emp, inicio, fin)
                if pdf:
                    nombre_pdf = (
                        f'{emp.nro_doc}_{emp.apellidos_nombres.replace(",","").replace(" ","_")}.pdf'
                    )
                    zf.writestr(nombre_pdf, pdf)
                    n_pdfs += 1

        if n_pdfs == 0:
            resultados.append({
                'area': area.nombre,
                'estado': 'sin_pdfs',
                'enviado': False,
            })
            continue

        # Destinatarios
        to_list = cfg.get_emails_jefatura() + override_emails
        cc_list = cfg.get_emails_cc() + override_cc

        if not to_list:
            resultados.append({
                'area': area.nombre,
                'estado': 'sin_destinatarios',
                'enviado': False,
            })
            continue

        # Asunto y cuerpo
        if asunto_override:
            asunto = asunto_override.format(area=area.nombre, periodo=periodo_lbl, empresa=empresa)
        else:
            asunto = cfg.render_asunto(area.nombre, periodo_lbl, empresa)

        if cuerpo_override:
            cuerpo = cuerpo_override.format(
                area=area.nombre, periodo=periodo_lbl,
                empresa=empresa, total_empleados=n_pdfs,
            )
        else:
            cuerpo = cfg.render_cuerpo(area.nombre, periodo_lbl, empresa, n_pdfs)

        # Enviar
        zip_buf.seek(0)
        zip_name = f'Asistencia_{slugify(area.nombre)}_{inicio.strftime("%Y%m%d")}_{fin.strftime("%Y%m%d")}.zip'

        try:
            msg = DjEmail(
                subject=asunto,
                body=cuerpo,
                to=to_list,
                cc=cc_list if cc_list else None,
            )
            msg.attach(zip_name, zip_buf.read(), 'application/zip')
            msg.send(fail_silently=False)
            enviado = True
            estado = 'enviado'
        except Exception as e:
            logger.error(f'Error enviando a {area.nombre}: {e}')
            enviado = False
            estado = f'error: {str(e)[:100]}'
            errores_globales.append({'area': area.nombre, 'error': str(e)})

        # Registrar en log
        try:
            EnvioReporteArea.objects.create(
                area=area,
                tipo_periodo=tipo_periodo,
                fecha_inicio=inicio,
                fecha_fin=fin,
                emails_destino=to_list,
                emails_cc=cc_list,
                asunto=asunto,
                cuerpo=cuerpo,
                estado='ENVIADO' if enviado else 'ERROR',
                empleados_total=qs_emp.count(),
                empleados_enviados=n_pdfs if enviado else 0,
                errores=[] if enviado else [estado],
                creado_por=request.user if request.user.is_authenticated else None,
            )
        except Exception as log_e:
            logger.warning(f'No se pudo crear log de envío: {log_e}')

        resultados.append({
            'area': area.nombre,
            'estado': estado,
            'enviado': enviado,
            'empleados': n_pdfs,
            'destinatarios': to_list,
        })

    total_enviados = sum(1 for r in resultados if r['enviado'])
    return JsonResponse({
        'ok': True,
        'total_areas': len(resultados),
        'enviados': total_enviados,
        'resultados': resultados,
        'errores': errores_globales,
        'periodo': periodo_lbl,
    })


@login_required
@solo_admin
def historial_envios(request):
    """Lista los últimos envíos registrados."""
    from asistencia.models import EnvioReporteArea
    envios = EnvioReporteArea.objects.select_related('area', 'creado_por').order_by('-creado_en')[:100]
    return render(request, 'asistencia/reportes_area_historial.html', {
        'envios': envios,
        'titulo': 'Historial de Envíos por Área',
    })


@login_required
@solo_admin
def reporte_excel_areas(request):
    """
    Genera ZIP con un Excel por área.
    Cada Excel:
      Sheet 1 "Asistencia": matriz empleado × fecha (entrada | salida por día)
      Sheet 2 "Resumen": totales por empleado
    Si se pasa solo un área, devuelve el Excel directo (sin ZIP).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    from asistencia.models import RegistroTareo, RegistroPapeleta

    tipo_periodo  = request.GET.get('tipo_periodo', 'SEMANAL')
    fecha_ini_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    area_ids      = request.GET.getlist('areas')

    inicio, fin = _get_rango(tipo_periodo, fecha_ini_str, fecha_fin_str)
    periodo_lbl = _periodo_label(inicio, fin)

    # Rango de fechas
    n_dias = (fin - inicio).days + 1
    fechas = [inicio + timedelta(days=i) for i in range(n_dias)]
    DIAS_ES = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

    # Colores
    C_FALTA    = PatternFill('solid', fgColor='C00000')
    C_TARDE    = PatternFill('solid', fgColor='FFCCCC')
    C_SALIDA_T = PatternFill('solid', fgColor='FFE0B2')
    C_DS       = PatternFill('solid', fgColor='D9D9D9')
    C_FERIADO  = PatternFill('solid', fgColor='D9B3FF')
    C_PAPELETA = PatternFill('solid', fgColor='BDD7EE')
    C_WEEKEND  = PatternFill('solid', fgColor='F5F5F5')
    C_HEADER   = PatternFill('solid', fgColor='1F4E79')
    C_SUBHDR   = PatternFill('solid', fgColor='2E75B6')
    C_TITLE    = PatternFill('solid', fgColor='D6E4F0')
    C_RESUMEN  = PatternFill('solid', fgColor='E2EFDA')

    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    # Borde derecho grueso para separar días
    sep_right = Side(style='medium', color='2E75B6')
    def _thin_sep(is_last_of_day=False):
        return Border(
            left=Side(style='thin', color='CCCCCC'),
            right=sep_right if is_last_of_day else Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )
    center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    center_w = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Hora referencia de entrada según condición (para detectar tardanza)
    HORA_REF = {
        'LOCAL':   time(8, 0),
        'FORANEO': time(7, 0),
        'FORÁNEO': time(7, 0),
        'LIMA':    time(8, 0),
    }
    MARGEN_TARDE = timedelta(minutes=15)

    def _es_tarde(hora_entrada, condicion):
        if hora_entrada is None:
            return False
        ref = HORA_REF.get(condicion or 'LOCAL', time(8, 0))
        from datetime import datetime as _dt
        ref_dt  = _dt.combine(date.today(), ref)
        entr_dt = _dt.combine(date.today(), hora_entrada)
        return (entr_dt - ref_dt) > MARGEN_TARDE

    def _es_salida_temp(hora_salida, horas_ef, condicion):
        """Salida temprana: menos de 80% de la jornada cumplida."""
        if hora_salida is None:
            return False
        return float(horas_ef or 0) < 6.0

    def _build_excel(area, empleados, tareos_map, papeletas_map, fechas, periodo_lbl):
        """Construye el workbook para un área."""
        wb = openpyxl.Workbook()

        # ──────────────────────────────────────────────────────────────
        # SHEET 1: ASISTENCIA
        # ──────────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Asistencia'

        N_FIJOS = 4  # Nombre, Cargo, Cond, Grupo
        # Cada fecha ocupa 2 columnas (E y S)
        # Luego 4 columnas resumen

        total_cols = N_FIJOS + len(fechas) * 2 + 4

        # ── Fila 1: Título ────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=total_cols)
        c = ws.cell(1, 1, f'REPORTE DE ASISTENCIA  |  ÁREA: {area.nombre.upper()}  |  {periodo_lbl}')
        c.font  = Font(bold=True, size=13, color='FFFFFF')
        c.fill  = C_HEADER
        c.alignment = center
        ws.row_dimensions[1].height = 22

        # ── Fila 2: Días de semana (encabezado de fechas) ─────────────
        for col_fijo in range(1, N_FIJOS + 1):
            c = ws.cell(2, col_fijo, '')
            c.fill = C_SUBHDR
        for i, f in enumerate(fechas):
            col = N_FIJOS + 1 + i * 2
            dia = DIAS_ES[f.weekday()]
            es_fin_semana = f.weekday() >= 5
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
            c = ws.cell(2, col, f'{dia} {f.day:02d}/{f.month:02d}')
            c.font = Font(bold=True, size=9, color='FFFFFF')
            c.fill = C_WEEKEND if es_fin_semana else C_SUBHDR
            c.alignment = center
        # Resumen cols header
        for i, lbl in enumerate(['H.NORM', 'H.EXTRA', 'FALTAS', 'TARDZ']):
            c = ws.cell(2, N_FIJOS + 1 + len(fechas)*2 + i, lbl)
            c.font = Font(bold=True, size=8, color='FFFFFF')
            c.fill = C_HEADER
            c.alignment = center
        ws.row_dimensions[2].height = 16

        # ── Fila 3: Etiquetas fijas + E/S por día ─────────────────────
        for col, lbl in enumerate(['APELLIDOS Y NOMBRES', 'CARGO', 'COND.', 'GRP.'], 1):
            c = ws.cell(3, col, lbl)
            c.font  = Font(bold=True, size=9, color='FFFFFF')
            c.fill  = C_HEADER
            c.alignment = center
        for i, f in enumerate(fechas):
            col = N_FIJOS + 1 + i * 2
            es_fin_semana = f.weekday() >= 5
            fill = C_WEEKEND if es_fin_semana else C_SUBHDR
            for j, lbl in enumerate(['E', 'S']):
                cc = ws.cell(3, col + j, lbl)
                cc.font = Font(bold=True, size=9, color='FFFFFF' if not es_fin_semana else '555555')
                cc.fill = fill
                cc.alignment = center
                cc.border = _thin_sep(j == 1)  # borde sep en col S
        for i in range(4):
            c = ws.cell(3, N_FIJOS + 1 + len(fechas)*2 + i, '')
            c.fill = C_HEADER
        ws.row_dimensions[3].height = 14

        # ── Anchos de columnas ────────────────────────────────────────
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 6
        for i in range(len(fechas) * 2):
            ws.column_dimensions[get_column_letter(N_FIJOS + 1 + i)].width = 7
        for i in range(4):
            ws.column_dimensions[get_column_letter(N_FIJOS + 1 + len(fechas)*2 + i)].width = 8

        ws.freeze_panes = ws.cell(4, N_FIJOS + 1)

        # ── Datos por empleado ────────────────────────────────────────
        # papeletas_map ya trae la etiqueta corta (iniciales o 4 primeros chars)
        resumen_rows = []

        for row_num, emp in enumerate(empleados, start=4):
            dni = emp.nro_doc
            condicion = emp.condicion or 'LOCAL'
            ws.row_dimensions[row_num].height = 14

            # Cols fijas
            for col, val in enumerate([
                emp.apellidos_nombres,
                emp.cargo or '—',
                condicion[:3].upper(),
                emp.grupo_tareo or 'RCO',
            ], 1):
                c = ws.cell(row_num, col, val)
                c.font = Font(size=9)
                c.border = thin
                c.alignment = Alignment(vertical='center',
                                        horizontal='left' if col == 1 else 'center')

            total_hnorm = 0.0
            total_hextra = 0.0
            total_faltas = 0
            total_tardanzas = 0

            for i, f in enumerate(fechas):
                col_e = N_FIJOS + 1 + i * 2
                col_s = col_e + 1
                es_fin_semana = f.weekday() >= 5

                reg = tareos_map.get((dni, f))
                pap = papeletas_map.get((dni, f))

                bg_e = C_WEEKEND if es_fin_semana else None
                bg_s = bg_e

                val_e = val_s = ''

                borde_e = _thin_sep(False)
                borde_s = _thin_sep(True)   # borde derecho grueso en col Salida

                if pap and not reg:
                    tipo_short = pap
                    ws.merge_cells(start_row=row_num, start_column=col_e,
                                   end_row=row_num, end_column=col_s)
                    c = ws.cell(row_num, col_e, tipo_short)
                    c.fill = C_PAPELETA
                    c.font = Font(bold=True, size=9, color='1F4E79')
                    c.alignment = center
                    c.border = borde_s
                    ws.cell(row_num, col_s).border = borde_s
                    continue

                if reg is None:
                    for col, brd in ((col_e, borde_e), (col_s, borde_s)):
                        cc = ws.cell(row_num, col, '')
                        cc.border = brd
                        if es_fin_semana:
                            cc.fill = C_WEEKEND
                    continue

                if reg.codigo_dia == 'SS':
                    # SS = marcación incompleta (presente, pero falta entrada o salida)
                    # NO es ausencia — mostrar la hora disponible y "SS" en la columna faltante
                    total_hnorm += float(reg.horas_normales or reg.horas_efectivas or 0)
                    entrada_ss = reg.hora_entrada_real
                    salida_ss  = reg.hora_salida_real
                    val_e = entrada_ss.strftime('%H:%M') if entrada_ss else 'SS'
                    val_s = salida_ss.strftime('%H:%M')  if salida_ss  else 'SS'
                    C_SS = PatternFill('solid', fgColor='FFE5CC')  # naranja claro
                    for col, val, brd in ((col_e, val_e, borde_e), (col_s, val_s, borde_s)):
                        c = ws.cell(row_num, col, val)
                        c.fill = C_SS
                        c.font = Font(size=9, bold=(val == 'SS'), color='8B4000' if val == 'SS' else '000000')
                        c.alignment = center
                        c.border = brd
                    continue

                if reg.codigo_dia == 'DS':
                    ws.merge_cells(start_row=row_num, start_column=col_e,
                                   end_row=row_num, end_column=col_s)
                    c = ws.cell(row_num, col_e, 'DESC')
                    c.fill = C_DS
                    c.font = Font(bold=True, size=9, color='555555')
                    c.alignment = center
                    c.border = borde_s
                    ws.cell(row_num, col_s).border = borde_s
                    continue

                # Código T (trabajado / feriado trabajado)
                entrada = reg.hora_entrada_real
                salida  = reg.hora_salida_real
                hef     = float(reg.horas_efectivas or 0)
                hextra  = float((reg.he_25 or 0) + (reg.he_35 or 0) + (reg.he_100 or 0))

                total_hnorm  += float(reg.horas_normales or 0)
                total_hextra += hextra

                val_e = entrada.strftime('%H:%M') if entrada else '—'
                val_s = salida.strftime('%H:%M')  if salida  else '—'

                es_tarde = _es_tarde(entrada, condicion)
                if es_tarde:
                    total_tardanzas += 1

                fill_e = C_FERIADO if reg.es_feriado else (C_TARDE if es_tarde else bg_e)
                fill_s = C_FERIADO if reg.es_feriado else (
                    C_SALIDA_T if _es_salida_temp(salida, hef, condicion) else bg_s
                )

                for col, val, fill, brd in (
                    (col_e, val_e, fill_e, borde_e),
                    (col_s, val_s, fill_s, borde_s),
                ):
                    c = ws.cell(row_num, col, val)
                    c.font = Font(size=9, bold=es_tarde and col == col_e,
                                  color='C00000' if es_tarde and col == col_e else '000000')
                    c.alignment = center
                    c.border = brd
                    if fill:
                        c.fill = fill

            # Columnas resumen
            base_col = N_FIJOS + 1 + len(fechas) * 2
            for j, (val, fmt) in enumerate([
                (round(total_hnorm, 1),  '0.0'),
                (round(total_hextra, 1), '0.0'),
                (total_faltas,           '0'),
                (total_tardanzas,        '0'),
            ]):
                c = ws.cell(row_num, base_col + j, val)
                c.font = Font(size=9, bold=True)
                c.alignment = center
                c.border = thin
                c.number_format = fmt
                if j == 2 and total_faltas > 0:
                    c.fill = PatternFill('solid', fgColor='FFCCCC')
                elif j == 3 and total_tardanzas > 0:
                    c.fill = PatternFill('solid', fgColor='FFE0B2')

            resumen_rows.append({
                'nombre':    emp.apellidos_nombres,
                'cargo':     emp.cargo or '—',
                'condicion': condicion,
                'grupo':     emp.grupo_tareo or 'RCO',
                'h_norm':    round(total_hnorm, 1),
                'h_extra':   round(total_hextra, 1),
                'faltas':    total_faltas,
                'tardanzas': total_tardanzas,
            })

        # ── Fila de totales diarios ───────────────────────────────────
        tot_row = 4 + len(empleados)
        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=N_FIJOS)
        c = ws.cell(tot_row, 1, 'TOTALES DEL DÍA →')
        c.font = Font(bold=True, size=9)
        c.fill = C_RESUMEN
        c.alignment = Alignment(horizontal='right', vertical='center')

        for i, f in enumerate(fechas):
            col_e = N_FIJOS + 1 + i * 2
            # Contar T del día
            t_count = sum(
                1 for emp in empleados
                if (reg := tareos_map.get((emp.nro_doc, f))) and reg.codigo_dia in ('T', 'SS')
            )
            abs_count = sum(
                1 for emp in empleados
                if tareos_map.get((emp.nro_doc, f)) is None
            )
            ws.merge_cells(start_row=tot_row, start_column=col_e, end_row=tot_row, end_column=col_e+1)
            label = f'✓{t_count}' + (f' ✗{abs_count}' if abs_count else '')
            c = ws.cell(tot_row, col_e, label)
            c.font = Font(size=8, bold=True)
            c.fill = C_RESUMEN
            c.alignment = center
            c.border = thin
        ws.row_dimensions[tot_row].height = 13

        # Leyenda
        leyenda_row = tot_row + 2
        ws.merge_cells(start_row=leyenda_row, start_column=1,
                       end_row=leyenda_row, end_column=total_cols)
        leyenda_items = [
            'ROJO=Tardanza entrada',
            'NARANJA=Salida temprana',
            'GRIS=DESC compensatorio',
            'AZUL=Papeleta/Permiso',
            'MORADO=Feriado trabajado',
            'NARANJA CLARO=SS (marcación incompleta, presente)',
            'FALTA=Ausencia sin justificar',
            'E>8:00 LOCAL / E>7:00 FORÁNEO = tardanza',
        ]
        c = ws.cell(leyenda_row, 1, '  LEYENDA:  ' + '   |   '.join(leyenda_items))
        c.font = Font(size=8, italic=True, color='555555')
        c.alignment = Alignment(vertical='center')

        # ──────────────────────────────────────────────────────────────
        # SHEET 2: RESUMEN
        # ──────────────────────────────────────────────────────────────
        ws2 = wb.create_sheet('Resumen')
        ws2.merge_cells('A1:H1')
        c = ws2.cell(1, 1, f'RESUMEN — {area.nombre}  |  {periodo_lbl}')
        c.font  = Font(bold=True, size=12, color='FFFFFF')
        c.fill  = C_HEADER
        c.alignment = center
        ws2.row_dimensions[1].height = 20

        hdrs2 = ['APELLIDOS Y NOMBRES', 'CARGO', 'COND.', 'GRUPO',
                 'H. NORM.', 'H. EXTRA', 'FALTAS', 'TARDANZAS']
        widths2 = [32, 22, 8, 7, 9, 9, 8, 9]
        for col, (h, w) in enumerate(zip(hdrs2, widths2), 1):
            c = ws2.cell(2, col, h)
            c.font = Font(bold=True, size=9, color='FFFFFF')
            c.fill = C_SUBHDR
            c.alignment = center
            c.border = thin
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.row_dimensions[2].height = 15

        # Ordenar por faltas desc, tardanzas desc
        resumen_rows.sort(key=lambda r: (-r['faltas'], -r['tardanzas'], r['nombre']))
        for row_num, r in enumerate(resumen_rows, start=3):
            vals = [r['nombre'], r['cargo'], r['condicion'][:3], r['grupo'],
                    r['h_norm'], r['h_extra'], r['faltas'], r['tardanzas']]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row_num, col, val)
                c.font = Font(size=9)
                c.border = thin
                c.alignment = Alignment(horizontal='left' if col <= 2 else 'center',
                                        vertical='center')
                if col == 7 and r['faltas'] > 0:
                    c.fill = PatternFill('solid', fgColor='FFCCCC')
                    c.font = Font(size=9, bold=True, color='C00000')
                elif col == 8 and r['tardanzas'] > 0:
                    c.fill = PatternFill('solid', fgColor='FFE0B2')
                    c.font = Font(size=9, bold=True, color='C06000')
            ws2.row_dimensions[row_num].height = 13

        ws2.freeze_panes = 'A3'
        ws2.auto_filter.ref = f'A2:H{2 + len(resumen_rows)}'

        return wb

    # ── Construir los Excels ──────────────────────────────────────────
    areas_qs = Area.objects.filter(activa=True).order_by('nombre')
    if area_ids:
        areas_qs = areas_qs.filter(pk__in=area_ids)

    # Pre-cargar todos los tareos del rango
    empleo_qs = Personal.objects.filter(
        estado='Activo', subarea__isnull=False,
    ).select_related('subarea__area').order_by('apellidos_nombres')
    if area_ids:
        empleo_qs = empleo_qs.filter(subarea__area_id__in=area_ids)

    all_tareos = RegistroTareo.objects.filter(
        fecha__gte=inicio, fecha__lte=fin,
        personal__in=empleo_qs,
    ).select_related('personal')

    all_papeletas = RegistroPapeleta.objects.filter(
        fecha_inicio__lte=fin, fecha_fin__gte=inicio,
        personal__in=empleo_qs,
        estado__in=['APROBADA', 'EJECUTADA', 'PENDIENTE'],
    ).select_related('personal')

    # Indexar tareos: (dni, fecha) → RegistroTareo
    tareos_idx = {}
    for t in all_tareos:
        tareos_idx[(t.dni, t.fecha)] = t

    # Indexar papeletas: (dni, fecha) → etiqueta corta
    papeletas_idx = defaultdict(dict)
    for p in all_papeletas:
        dni = p.personal.nro_doc
        # Usar iniciales si existen, si no las primeras 3 letras del tipo
        label = (p.iniciales or p.tipo_permiso or 'PER')[:4].upper()
        f = p.fecha_inicio
        while f <= p.fecha_fin:
            if inicio <= f <= fin:
                papeletas_idx[dni][f] = label
            f += timedelta(days=1)

    # Empleados por área
    emp_por_area = defaultdict(list)
    for emp in empleo_qs:
        emp_por_area[emp.subarea.area_id].append(emp)

    # Generar archivos
    archivos = []
    for area in areas_qs:
        empleados = emp_por_area.get(area.pk, [])
        if not empleados:
            continue

        tareos_map   = {(emp.nro_doc, f): tareos_idx.get((emp.nro_doc, f))
                        for emp in empleados for f in fechas}
        papeletas_map = {(emp.nro_doc, f): papeletas_idx.get(emp.nro_doc, {}).get(f)
                         for emp in empleados for f in fechas}

        wb = _build_excel(area, empleados, tareos_map, papeletas_map, fechas, periodo_lbl)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'Asistencia_{slugify(area.nombre)}_{inicio.strftime("%Y%m%d")}.xlsx'
        archivos.append((fname, buf.getvalue()))

    if not archivos:
        return HttpResponse('No hay registros para el período seleccionado.', status=404)

    # Devolver Excel directo si es una sola área
    if len(archivos) == 1:
        fname, data = archivos[0]
        resp = HttpResponse(data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    # ZIP con todos
    zip_buf = io.BytesIO()
    with ZipFile(zip_buf, 'w') as zf:
        for fname, data in archivos:
            zf.writestr(fname, data)
    zip_buf.seek(0)
    zip_name = f'Asistencia_Areas_{inicio.strftime("%Y%m%d")}_{fin.strftime("%Y%m%d")}.zip'
    resp = HttpResponse(zip_buf.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    return resp


@login_required
@solo_admin
def gestionar_emails(request):
    """
    Página de gestión rápida de emails por área.
    Tabla con todas las áreas y modal de edición completa.
    Muestra correo personal y corporativo del empleado si está en BD.
    """
    from asistencia.models import ConfiguracionReporteArea
    from personal.models import Personal

    areas = Area.objects.filter(activa=True).order_by('nombre')
    configs = {c.area_id: c for c in ConfiguracionReporteArea.objects.all()}

    rows = []
    for area in areas:
        n = Personal.objects.filter(subarea__area=area, estado='Activo').count()
        if n == 0:
            continue
        cfg = configs.get(area.pk)
        # Buscar correos del jefe si está como empleado en BD
        correos_jefe = {}
        if cfg and cfg.nombre_jefe:
            nombre_norm = cfg.nombre_jefe.upper().strip()
            match = Personal.objects.filter(
                apellidos_nombres__icontains=nombre_norm.split()[0]
            ).first() if nombre_norm else None
            if match:
                correos_jefe = {
                    'personal':     match.correo_personal or '',
                    'corporativo':  match.correo_corporativo or '',
                }
        rows.append({
            'area':            area,
            'n_empleados':     n,
            'cfg':             cfg,
            'nombre_jefe':     cfg.nombre_jefe if cfg else '',
            'emails_jefatura': cfg.get_emails_jefatura() if cfg else [],
            'emails_cc':       cfg.get_emails_cc() if cfg else [],
            'asunto':          cfg.asunto_template if cfg else 'Reporte de Asistencia - {area} - {periodo}',
            'cuerpo':          cfg.cuerpo_template if cfg else '',
            'activo':          cfg.activo if cfg else False,
            'correos_jefe':    correos_jefe,
        })

    return render(request, 'asistencia/gestionar_emails.html', {
        'rows':   rows,
        'titulo': 'Gestionar Emails por Área',
    })


@login_required
@solo_admin
def reporte_horario_simple(request):
    """
    Genera Excel con: ÁREA, DNI, NOMBRE, FECHA, DÍA, HORA ENTRADA, HORA SALIDA, CÓDIGO
    Un reporte liviano de marcaciones, sin cálculos complejos.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from asistencia.models import RegistroTareo

    tipo_periodo  = request.GET.get('tipo_periodo', 'SEMANAL')
    fecha_ini_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    area_ids      = request.GET.getlist('areas')
    grupo_filter  = request.GET.get('grupo', 'TODOS')

    inicio, fin = _get_rango(tipo_periodo, fecha_ini_str, fecha_fin_str)
    periodo_lbl = _periodo_label(inicio, fin)

    DIAS = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

    qs = RegistroTareo.objects.filter(
        fecha__gte=inicio, fecha__lte=fin,
    ).select_related('personal__subarea__area').order_by(
        'personal__subarea__area__nombre',
        'personal__apellidos_nombres',
        'fecha',
    )
    if area_ids:
        qs = qs.filter(personal__subarea__area_id__in=area_ids)
    if grupo_filter != 'TODOS':
        qs = qs.filter(grupo=grupo_filter)

    if not qs.exists():
        return HttpResponse('No hay registros para el período seleccionado.', status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Horario'

    # ── Estilos ──────────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    hdr_font  = Font(color='FFFFFF', bold=True, size=10)
    area_fill = PatternFill('solid', fgColor='D6E4F0')
    area_font = Font(bold=True, size=10)
    mono      = Font(name='Courier New', size=9)
    center    = Alignment(horizontal='center', vertical='center')
    thin      = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # ── Título ───────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'] = f'REPORTE DE HORARIOS — {periodo_lbl}'
    ws['A1'].font = Font(bold=True, size=12, color='1F4E79')
    ws.row_dimensions[1].height = 22

    # ── Cabecera ─────────────────────────────────────────────────────
    headers = ['ÁREA', 'DNI', 'APELLIDOS Y NOMBRES', 'FECHA', 'DÍA', 'ENTRADA', 'SALIDA', 'CÓD.']
    widths  = [28, 11, 36, 12, 6, 10, 10, 7]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 16
    ws.freeze_panes = 'A3'

    # ── Datos ────────────────────────────────────────────────────────
    row_num = 3
    prev_area = None
    for reg in qs:
        area_nombre = ''
        if reg.personal and reg.personal.subarea and reg.personal.subarea.area:
            area_nombre = reg.personal.subarea.area.nombre

        # Separador de área
        if area_nombre != prev_area:
            ws.merge_cells(f'A{row_num}:H{row_num}')
            c = ws.cell(row=row_num, column=1, value=f'▶  {area_nombre}')
            c.fill = area_fill
            c.font = area_font
            c.alignment = Alignment(vertical='center')
            ws.row_dimensions[row_num].height = 14
            row_num += 1
            prev_area = area_nombre

        nombre = reg.personal.apellidos_nombres if reg.personal else reg.dni
        entrada = reg.hora_entrada_real.strftime('%H:%M') if reg.hora_entrada_real else '—'
        salida  = reg.hora_salida_real.strftime('%H:%M') if reg.hora_salida_real else '—'
        dia_str = DIAS[reg.dia_semana] if reg.dia_semana is not None else ''

        vals = [area_nombre, reg.dni, nombre,
                reg.fecha.strftime('%d/%m/%Y'), dia_str,
                entrada, salida, reg.codigo_dia]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.border    = thin
            c.alignment = center if col in (1, 2, 4, 5, 6, 7, 8) else Alignment(vertical='center')
            if col in (6, 7):
                c.font = mono
            if reg.codigo_dia in ('DS', 'SS') and col == 8:
                c.font = Font(bold=True, color='E74C3C' if reg.codigo_dia == 'SS' else '27AE60')
            if reg.es_feriado and col == 4:
                c.fill = PatternFill('solid', fgColor='FDEBD0')

        ws.row_dimensions[row_num].height = 13
        row_num += 1

    # ── Autofilter ───────────────────────────────────────────────────
    ws.auto_filter.ref = f'A2:H{row_num - 1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'Horario_{inicio.strftime("%Y%m%d")}_{fin.strftime("%Y%m%d")}.xlsx'
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
@solo_admin
def reporte_pdf_area(request):
    """
    Genera un PDF de asistencia por área (matriz empleado × día).
    Formato: A4 horizontal, una columna por día con código/hora de entrada.
    Si se filtra por una sola área, devuelve el PDF directo.
    Si son varias, genera un ZIP con un PDF por área.
    """
    from collections import defaultdict
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from asistencia.models import RegistroTareo, RegistroPapeleta

    tipo_periodo  = request.GET.get('tipo_periodo', 'SEMANAL')
    fecha_ini_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    area_ids      = request.GET.getlist('areas')

    inicio, fin = _get_rango(tipo_periodo, fecha_ini_str, fecha_fin_str)
    periodo_lbl = _periodo_label(inicio, fin)

    n_dias = (fin - inicio).days + 1
    fechas = [inicio + timedelta(days=i) for i in range(n_dias)]
    DIAS_ES = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sá', 'Do']
    HORA_REF = {'LOCAL': time(8, 0), 'FORANEO': time(7, 0), 'FORÁNEO': time(7, 0), 'LIMA': time(8, 0)}
    MARGEN_TARDE = timedelta(minutes=15)

    def _es_tarde_pdf(hora_entrada, condicion):
        if hora_entrada is None:
            return False
        from datetime import datetime as _dt
        ref = HORA_REF.get(condicion or 'LOCAL', time(8, 0))
        return (_dt.combine(date.today(), hora_entrada) - _dt.combine(date.today(), ref)) > MARGEN_TARDE

    # Colores ReportLab
    COL_FALTA   = colors.HexColor('#C00000')
    COL_TARDE_E = colors.HexColor('#FFCCCC')
    COL_DS      = colors.HexColor('#D9D9D9')
    COL_FERIADO = colors.HexColor('#D9B3FF')
    COL_PAP     = colors.HexColor('#BDD7EE')
    COL_SS      = colors.HexColor('#FFE5CC')
    COL_HEADER  = colors.HexColor('#1F4E79')
    COL_SUBHDR  = colors.HexColor('#2E75B6')
    COL_WEEKEND = colors.HexColor('#E8E8E8')
    COL_ROWALT  = colors.HexColor('#F7FBFF')
    COL_SUMHDR  = colors.HexColor('#375623')
    COL_SUMALT  = colors.HexColor('#EBF1E6')

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle('t', parent=styles['Normal'],
                               fontSize=10, textColor=colors.white,
                               alignment=TA_CENTER, fontName='Helvetica-Bold')
    st_sub   = ParagraphStyle('s', parent=styles['Normal'],
                               fontSize=7, textColor=colors.white,
                               alignment=TA_CENTER, fontName='Helvetica-Bold')
    st_cell  = ParagraphStyle('c', parent=styles['Normal'],
                               fontSize=7, alignment=TA_LEFT, fontName='Helvetica')
    st_num   = ParagraphStyle('n', parent=styles['Normal'],
                               fontSize=7, alignment=TA_CENTER, fontName='Helvetica')
    st_bold  = ParagraphStyle('b', parent=styles['Normal'],
                               fontSize=7, alignment=TA_CENTER, fontName='Helvetica-Bold',
                               textColor=colors.white)

    PAGE_W, PAGE_H = landscape(A4)
    MARGIN = 15 * mm
    usable_w = PAGE_W - 2 * MARGIN

    # Anchos fijos (NOMBRE y CARGO se calculan dinámicamente en _build_pdf_area)
    W_COND   = 22
    W_NORM   = 28
    W_FALT   = 22
    W_TARD   = 22
    W_SUM    = W_NORM + W_FALT + W_TARD

    # Leyenda completa: code → (color_hex, descripción)
    LEYENDA_DEFS = {
        'TARD': ('#FFCCCC', 'Tardanza'),
        'SS':   ('#FFE5CC', 'SS – marcación incompleta'),
        'DS':   ('#D9D9D9', 'DS – descanso semanal'),
        'FER':  ('#D9B3FF', 'Feriado trabajado'),
        'F':    ('#C00000', 'F – Falta/Ausencia'),
        # Papeletas comunes
        'DL':   ('#BDD7EE', 'DL – Día libre'),
        'DLA':  ('#BDD7EE', 'DLA – Día libre adicional'),
        'VAC':  ('#BDD7EE', 'VAC – Vacaciones'),
        'LM':   ('#BDD7EE', 'LM – Licencia médica'),
        'LCG':  ('#BDD7EE', 'LCG – Licencia con goce'),
        'LSG':  ('#BDD7EE', 'LSG – Licencia sin goce'),
        'CHE':  ('#BDD7EE', 'CHE – Comisión de servicio'),
        'CDT':  ('#BDD7EE', 'CDT – Capacitación'),
        'CPF':  ('#BDD7EE', 'CPF – Permiso por fallecimiento'),
        'SUB':  ('#BDD7EE', 'SUB – Subsidio'),
        'ATM':  ('#BDD7EE', 'ATM – Atención médica'),
        'PER':  ('#BDD7EE', 'PER – Permiso'),
        'B':    ('#D9D9D9', 'B – Bajada/Descanso'),
    }

    def _build_pdf_area(area, empleados, tareos_map, papeletas_map):
        from reportlab.pdfbase import pdfmetrics

        # ── Auto-ajuste de anchos por contenido ──────────────────────
        FSIZE = 7
        max_n = max(
            (pdfmetrics.stringWidth(e.apellidos_nombres, 'Helvetica', FSIZE) for e in empleados),
            default=60,
        )
        max_c = max(
            (pdfmetrics.stringWidth(e.cargo or '—', 'Helvetica', FSIZE) for e in empleados),
            default=35,
        )
        w_nombre = min(max(max_n + 8, 60), 160)
        w_cargo  = min(max(max_c + 8, 35), 110)
        w_dias_avail = usable_w - w_nombre - w_cargo - W_COND - W_SUM
        w_d = max(14, w_dias_avail / max(n_dias, 1))

        col_widths = [w_nombre, w_cargo, W_COND] + [w_d] * n_dias + [W_NORM, W_FALT, W_TARD]

        story = []
        title_p = Paragraph(
            f'REPORTE DE ASISTENCIA  |  ÁREA: {area.nombre.upper()}  |  {periodo_lbl}',
            st_title,
        )
        story.append(title_p)
        story.append(Spacer(1, 3))

        hdr_dia = ['', '', ''] + [
            Paragraph(f'{DIAS_ES[f.weekday()]}<br/>{f.day:02d}/{f.month:02d}', st_sub)
            for f in fechas
        ] + ['NORM', 'FALT', 'TARD']

        table_data = [hdr_dia]
        table_styles = [
            ('BACKGROUND',    (0, 0), (-1, 0), COL_SUBHDR),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 6.5),
            ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING',   (0, 0), (-1, -1), 2),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
            ('WORDWRAP',      (0, 1), (1, -1), 1),  # wrap nombre y cargo
        ]

        for i, f in enumerate(fechas):
            if f.weekday() >= 5:
                c_idx = 3 + i
                table_styles.append(('BACKGROUND', (c_idx, 0), (c_idx, 0), COL_WEEKEND))
                table_styles.append(('TEXTCOLOR',  (c_idx, 0), (c_idx, 0), colors.HexColor('#555555')))

        resumen_rows = []
        used_codes   = set()   # para leyenda dinámica
        pap_codes    = set()   # códigos de papeleta que aparecen

        for row_idx, emp in enumerate(empleados):
            condicion = emp.condicion or 'LOCAL'
            total_hnorm = 0.0
            total_faltas = 0
            total_tardanzas = 0

            st_nombre = ParagraphStyle(
                f'nm{row_idx}', parent=styles['Normal'],
                fontSize=FSIZE, alignment=TA_LEFT, fontName='Helvetica',
                leading=8,
            )
            st_cargo_r = ParagraphStyle(
                f'cg{row_idx}', parent=styles['Normal'],
                fontSize=FSIZE, alignment=TA_LEFT, fontName='Helvetica',
                leading=8,
            )

            row_cells  = [
                Paragraph(emp.apellidos_nombres, st_nombre),
                Paragraph(emp.cargo or '—', st_cargo_r),
                Paragraph(condicion[:3].upper(), st_num),
            ]
            cell_styles = []

            for i, f in enumerate(fechas):
                col_idx      = 3 + i
                pap          = papeletas_map.get((emp.nro_doc, f))
                reg          = tareos_map.get((emp.nro_doc, f))
                es_fin_semana = f.weekday() >= 5

                # ① Papeleta tiene siempre prioridad sobre el código del tareo
                if pap:
                    row_cells.append(Paragraph(pap, st_num))
                    cell_styles.append((col_idx, COL_PAP, colors.HexColor('#1F4E79'), True))
                    pap_codes.add(pap)
                    continue

                # ② Sin registro
                if reg is None:
                    if es_fin_semana:
                        row_cells.append(Paragraph('', st_num))
                        cell_styles.append((col_idx, COL_WEEKEND, colors.black, False))
                    else:
                        total_faltas += 1
                        row_cells.append(Paragraph('F', st_bold))
                        cell_styles.append((col_idx, COL_FALTA, colors.white, True))
                        used_codes.add('F')
                    continue

                # ③ SS – marcación incompleta (presente)
                if reg.codigo_dia == 'SS':
                    total_hnorm += float(reg.horas_normales or reg.horas_efectivas or 0)
                    hora_disp = reg.hora_entrada_real or reg.hora_salida_real
                    txt = hora_disp.strftime('%H:%M') if hora_disp else 'SS'
                    row_cells.append(Paragraph(txt, st_num))
                    cell_styles.append((col_idx, COL_SS, colors.HexColor('#8B4000'), False))
                    used_codes.add('SS')
                    continue

                # ④ DS – descanso semanal
                if reg.codigo_dia == 'DS':
                    row_cells.append(Paragraph('DS', st_num))
                    cell_styles.append((col_idx, COL_DS, colors.HexColor('#555555'), True))
                    used_codes.add('DS')
                    continue

                # ⑤ Código T – trabajado
                total_hnorm += float(reg.horas_normales or 0)
                entrada   = reg.hora_entrada_real
                es_tarde  = _es_tarde_pdf(entrada, condicion)
                es_feriado = reg.es_feriado
                if es_tarde:
                    total_tardanzas += 1
                    used_codes.add('TARD')
                if es_feriado:
                    used_codes.add('FER')

                txt = entrada.strftime('%H:%M') if entrada else '—'
                if es_feriado:
                    bg, tc = COL_FERIADO, colors.black
                elif es_tarde:
                    bg, tc = COL_TARDE_E, colors.HexColor('#C00000')
                else:
                    bg, tc = None, colors.black

                row_cells.append(Paragraph(txt, ParagraphStyle(
                    f'dyn{row_idx}_{i}', parent=styles['Normal'],
                    fontSize=FSIZE, alignment=TA_CENTER,
                    fontName='Helvetica-Bold' if es_tarde else 'Helvetica',
                    textColor=tc,
                )))
                cell_styles.append((col_idx, bg, tc, es_tarde))

            row_cells += [
                Paragraph(f'{round(total_hnorm, 1)}', st_num),
                Paragraph(str(total_faltas)    if total_faltas    else '—', st_num),
                Paragraph(str(total_tardanzas) if total_tardanzas else '—', st_num),
            ]
            table_data.append(row_cells)

            data_row = row_idx + 1
            bg_fila  = COL_ROWALT if row_idx % 2 == 0 else colors.white
            table_styles.append(('BACKGROUND', (0, data_row), (2, data_row), bg_fila))
            table_styles.append(('BACKGROUND', (-3, data_row), (-1, data_row), bg_fila))
            for (col_i, bg, tc, bold) in cell_styles:
                if bg:
                    table_styles.append(('BACKGROUND', (col_i, data_row), (col_i, data_row), bg))
            if total_faltas > 0:
                table_styles.append(('BACKGROUND', (-2, data_row), (-2, data_row), colors.HexColor('#FFCCCC')))
                table_styles.append(('TEXTCOLOR',  (-2, data_row), (-2, data_row), colors.HexColor('#C00000')))
                table_styles.append(('FONTNAME',   (-2, data_row), (-2, data_row), 'Helvetica-Bold'))
            if total_tardanzas > 0:
                table_styles.append(('BACKGROUND', (-1, data_row), (-1, data_row), colors.HexColor('#FFE0B2')))
                table_styles.append(('FONTNAME',   (-1, data_row), (-1, data_row), 'Helvetica-Bold'))

            resumen_rows.append({
                'nombre': emp.apellidos_nombres, 'cargo': emp.cargo or '—',
                'condicion': condicion,
                'h_norm': round(total_hnorm, 1),
                'faltas': total_faltas, 'tardanzas': total_tardanzas,
            })

        # Separadores verticales entre semanas
        for i, f in enumerate(fechas):
            if f.weekday() == 6 and i < n_dias - 1:
                table_styles.append(('LINEAFTER', (3 + i, 0), (3 + i, -1), 1.5, COL_SUBHDR))

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(table_styles))
        story.append(t)

        # ── Leyenda dinámica (solo los códigos que aparecen) ─────────
        story.append(Spacer(1, 4))
        ley_parts = []
        # Códigos fijos con prioridad de aparición
        for code in ('TARD', 'SS', 'DS', 'FER', 'F'):
            if code in used_codes and code in LEYENDA_DEFS:
                col_hex, desc = LEYENDA_DEFS[code]
                txt_color = '#FFFFFF' if code == 'F' else col_hex
                bg_color  = col_hex
                ley_parts.append(
                    f'<font color="{col_hex}">■</font> {desc}'
                )
        # Papeletas que aparecen
        for pcode in sorted(pap_codes):
            col_hex, desc = LEYENDA_DEFS.get(pcode, ('#BDD7EE', f'{pcode} – Papeleta'))
            ley_parts.append(f'<font color="#1F4E79">■</font> {desc}')

        if ley_parts:
            story.append(Paragraph(
                '  <font color="#555555">LEYENDA:</font>  ' + '   ·   '.join(ley_parts),
                ParagraphStyle('ley', parent=styles['Normal'], fontSize=6.5,
                               textColor=colors.HexColor('#555555')),
            ))

        # ── Resumen ──────────────────────────────────────────────────
        story.append(Spacer(1, 8))
        story.append(Paragraph('RESUMEN POR EMPLEADO', ParagraphStyle(
            'rs_title', parent=styles['Normal'],
            fontSize=9, fontName='Helvetica-Bold',
            textColor=colors.HexColor('#375623'),
        )))
        story.append(Spacer(1, 3))

        resumen_rows.sort(key=lambda r: (-r['faltas'], -r['tardanzas'], r['nombre']))
        sum_hdr  = ['APELLIDOS Y NOMBRES', 'CARGO', 'COND.', 'H. NORM.', 'FALTAS', 'TARDANZAS']
        sum_data = [sum_hdr] + [
            [r['nombre'], r['cargo'], r['condicion'][:3], r['h_norm'],
             r['faltas'] or '—', r['tardanzas'] or '—']
            for r in resumen_rows
        ]
        sum_widths = [
            w_nombre, w_cargo,
            W_COND + 10, W_NORM + 10, W_FALT + 10, W_TARD + 10,
        ]
        sum_styles = [
            ('BACKGROUND',    (0, 0), (-1, 0), COL_SUMHDR),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 7),
            ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN',         (2, 1), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#AAAAAA')),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING',   (0, 0), (-1, -1), 3),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
            ('WORDWRAP',      (0, 1), (1, -1), 1),
        ]
        for i, r in enumerate(resumen_rows):
            row_i = i + 1
            bg = COL_SUMALT if i % 2 == 0 else colors.white
            sum_styles.append(('BACKGROUND', (0, row_i), (-1, row_i), bg))
            if r['faltas'] > 0:
                sum_styles.append(('BACKGROUND', (4, row_i), (4, row_i), colors.HexColor('#FFCCCC')))
                sum_styles.append(('TEXTCOLOR',  (4, row_i), (4, row_i), colors.HexColor('#C00000')))
                sum_styles.append(('FONTNAME',   (4, row_i), (4, row_i), 'Helvetica-Bold'))

        sum_t = Table(sum_data, colWidths=sum_widths)
        sum_t.setStyle(TableStyle(sum_styles))
        story.append(sum_t)

        return story

    # ── Pre-cargar datos ──────────────────────────────────────────────
    areas_qs = Area.objects.filter(activa=True).order_by('nombre')
    if area_ids:
        areas_qs = areas_qs.filter(pk__in=area_ids)

    empleo_qs = Personal.objects.filter(
        estado='Activo', subarea__isnull=False,
    ).select_related('subarea__area').order_by('apellidos_nombres')
    if area_ids:
        empleo_qs = empleo_qs.filter(subarea__area_id__in=area_ids)

    all_tareos = RegistroTareo.objects.filter(
        fecha__gte=inicio, fecha__lte=fin, personal__in=empleo_qs,
    ).select_related('personal')
    all_papeletas = RegistroPapeleta.objects.filter(
        fecha_inicio__lte=fin, fecha_fin__gte=inicio,
        personal__in=empleo_qs,
        estado__in=['APROBADA', 'EJECUTADA', 'PENDIENTE'],
    ).select_related('personal')

    tareos_idx = {}
    for t in all_tareos:
        tareos_idx[(t.dni, t.fecha)] = t

    from collections import defaultdict
    papeletas_idx = defaultdict(dict)
    for p in all_papeletas:
        dni = p.personal.nro_doc
        label = (p.iniciales or p.tipo_permiso or 'PER')[:4].upper()
        f = p.fecha_inicio
        while f <= p.fecha_fin:
            if inicio <= f <= fin:
                papeletas_idx[dni][f] = label
            f += timedelta(days=1)

    emp_por_area = defaultdict(list)
    for emp in empleo_qs:
        emp_por_area[emp.subarea.area_id].append(emp)

    # ── Generar PDFs ──────────────────────────────────────────────────
    archivos_pdf = []
    for area in areas_qs:
        empleados = emp_por_area.get(area.pk, [])
        if not empleados:
            continue
        tareos_map   = {(emp.nro_doc, f): tareos_idx.get((emp.nro_doc, f))
                        for emp in empleados for f in fechas}
        papeletas_map = {(emp.nro_doc, f): papeletas_idx.get(emp.nro_doc, {}).get(f)
                         for emp in empleados for f in fechas}

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=12 * mm, bottomMargin=12 * mm,
            title=f'Asistencia {area.nombre} {periodo_lbl}',
        )
        story = _build_pdf_area(area, empleados, tareos_map, papeletas_map)
        doc.build(story)
        buf.seek(0)
        fname = f'Asistencia_{slugify(area.nombre)}_{inicio.strftime("%Y%m%d")}.pdf'
        archivos_pdf.append((fname, buf.getvalue()))

    if not archivos_pdf:
        return HttpResponse('No hay registros para el período seleccionado.', status=404)

    if len(archivos_pdf) == 1:
        fname, data = archivos_pdf[0]
        resp = HttpResponse(data, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="{fname}"'
        return resp

    zip_buf = io.BytesIO()
    with ZipFile(zip_buf, 'w') as zf:
        for fname, data in archivos_pdf:
            zf.writestr(fname, data)
    zip_buf.seek(0)
    zip_name = f'Asistencia_Areas_{inicio.strftime("%Y%m%d")}_{fin.strftime("%Y%m%d")}.zip'
    resp = HttpResponse(zip_buf.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    return resp
