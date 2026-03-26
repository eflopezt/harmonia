"""
Vistas del módulo Asistencia — Papeletas (unificado).

Sistema unificado de papeletas: importadas desde Synkro/Excel O creadas
manualmente en el sistema. Cubre todos los tipos: VAC, LSG, LCG, CHE,
DM, LP, CT, bajadas, compensaciones, suspensiones, etc.

Base legal:
- D.Leg. 713 Art. 6: compensación por feriado/DSO trabajado
- DS 003-97-TR: licencias y permisos
- Ley 26644: licencia maternidad
- Ley 29409: licencia paternidad
"""
import calendar
from datetime import date, datetime

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_POST

from asistencia.views._common import solo_admin


@login_required
@solo_admin
def papeletas_view(request):
    """Lista unificada de papeletas con filtros."""
    from asistencia.models import RegistroPapeleta
    from personal.models import Personal

    # Filtros
    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '')
    origen = request.GET.get('origen', '')
    anio = int(request.GET.get('anio', date.today().year))
    personal_id = request.GET.get('personal', '')

    qs = RegistroPapeleta.objects.select_related(
        'personal', 'aprobado_por', 'creado_por'
    ).filter(fecha_inicio__year=anio)

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_permiso=tipo)
    if origen:
        qs = qs.filter(origen=origen)
    if personal_id:
        qs = qs.filter(personal_id=personal_id)

    hoy = date.today()

    from django.db.models import Count as _Count, Q as _Q
    pap_stats = qs.aggregate(
        pendientes_count=_Count('id', filter=_Q(estado='PENDIENTE')),
        aprobadas_count=_Count('id', filter=_Q(estado='APROBADA')),
        rechazadas_count=_Count('id', filter=_Q(estado='RECHAZADA')),
    )
    # Por tipo — top 3
    pap_por_tipo = list(
        qs.values('tipo_permiso')
        .annotate(total=_Count('id'))
        .order_by('-total')[:3]
    )

    context = {
        'titulo': 'Papeletas',
        'papeletas': qs.order_by('-fecha_inicio')[:500],
        'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres'),
        'anio_actual': anio,
        'anios': range(hoy.year - 1, hoy.year + 2),
        'estados': RegistroPapeleta.ESTADO_CHOICES,
        'tipos': RegistroPapeleta.TIPO_PERMISO_CHOICES,
        'origenes': RegistroPapeleta.ORIGEN_CHOICES,
        'filtro_estado': estado,
        'filtro_tipo': tipo,
        'filtro_origen': origen,
        'filtro_personal': personal_id,
        'total': qs.count(),
        'pendientes': qs.filter(estado='PENDIENTE').count(),
        'pap_stats': pap_stats,
        'pap_por_tipo': pap_por_tipo,
    }
    return render(request, 'asistencia/papeletas.html', context)


@login_required
@solo_admin
@require_POST
def papeleta_crear(request):
    """Crear papeleta manualmente (admin)."""
    from asistencia.models import RegistroPapeleta
    from personal.models import Personal
    try:
        personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
        tipo = request.POST['tipo_permiso']
        fecha_inicio = datetime.strptime(request.POST['fecha_inicio'], '%Y-%m-%d').date()
        fecha_fin_raw = request.POST.get('fecha_fin')
        fecha_fin = datetime.strptime(fecha_fin_raw, '%Y-%m-%d').date() if fecha_fin_raw else fecha_inicio
        estado = request.POST.get('estado', 'PENDIENTE')

        # Calcular días hábiles automáticamente si no se proporcionan
        dias_habiles_raw = int(request.POST.get('dias_habiles', 0) or 0)
        if not dias_habiles_raw:
            dias_habiles_raw = (fecha_fin - fecha_inicio).days + 1

        p = RegistroPapeleta.objects.create(
            personal=personal,
            dni=personal.nro_doc,
            tipo_permiso=tipo,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            fecha_referencia=datetime.strptime(request.POST['fecha_referencia'], '%Y-%m-%d').date() if request.POST.get('fecha_referencia') else None,
            detalle=request.POST.get('detalle', '').strip(),
            dias_habiles=dias_habiles_raw,
            origen='SISTEMA',
            estado=estado,
            creado_por=request.user,
            observaciones=request.POST.get('observaciones', '').strip(),
        )
        # Si se crea como aprobada directamente
        if estado in ('APROBADA', 'EJECUTADA'):
            p.aprobado_por = request.user
            p.fecha_aprobacion = date.today()
            p.save()

        return JsonResponse(_papeleta_dict(p))
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def papeleta_editar(request, pk):
    """Editar papeleta existente."""
    from asistencia.models import RegistroPapeleta
    p = get_object_or_404(RegistroPapeleta, pk=pk)
    try:
        p.tipo_permiso = request.POST.get('tipo_permiso', p.tipo_permiso)
        fi_raw = request.POST.get('fecha_inicio')
        p.fecha_inicio = datetime.strptime(fi_raw, '%Y-%m-%d').date() if fi_raw else p.fecha_inicio
        ff_raw = request.POST.get('fecha_fin')
        p.fecha_fin = datetime.strptime(ff_raw, '%Y-%m-%d').date() if ff_raw else p.fecha_fin
        fr_raw = request.POST.get('fecha_referencia')
        p.fecha_referencia = datetime.strptime(fr_raw, '%Y-%m-%d').date() if fr_raw else None
        p.detalle = request.POST.get('detalle', '').strip()
        dias_habiles_raw = int(request.POST.get('dias_habiles', 0) or 0)
        if not dias_habiles_raw:
            dias_habiles_raw = (p.fecha_fin - p.fecha_inicio).days + 1
        p.dias_habiles = dias_habiles_raw
        p.observaciones = request.POST.get('observaciones', '').strip()

        nuevo_estado = request.POST.get('estado', p.estado)
        if nuevo_estado != p.estado:
            p.estado = nuevo_estado
            if nuevo_estado in ('APROBADA', 'EJECUTADA'):
                p.aprobado_por = request.user
                p.fecha_aprobacion = date.today()

        p.save()
        return JsonResponse(_papeleta_dict(p))
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def papeleta_eliminar(request, pk):
    """Eliminar papeleta."""
    from asistencia.models import RegistroPapeleta
    p = get_object_or_404(RegistroPapeleta, pk=pk)
    # Solo permitir eliminar papeletas creadas en el sistema (no importadas)
    if p.origen == 'IMPORTACION':
        return JsonResponse(
            {'ok': False, 'error': 'No se pueden eliminar papeletas importadas. Use la gestión de importaciones.'},
            status=400)
    p.delete()
    return JsonResponse({'ok': True})


@login_required
@solo_admin
@require_POST
def papeleta_aprobar(request, pk):
    """Aprobar o rechazar una papeleta pendiente."""
    from asistencia.models import RegistroPapeleta
    p = get_object_or_404(RegistroPapeleta, pk=pk)
    accion = request.POST.get('accion', '')  # 'aprobar' o 'rechazar'

    if p.estado != 'PENDIENTE':
        return JsonResponse(
            {'ok': False, 'error': f'Solo se pueden revisar papeletas pendientes (actual: {p.get_estado_display()}).'},
            status=400)

    if accion == 'aprobar':
        p.estado = 'APROBADA'
    elif accion == 'rechazar':
        p.estado = 'RECHAZADA'
    else:
        return JsonResponse({'ok': False, 'error': 'Acción inválida.'}, status=400)

    old_estado = 'PENDIENTE'
    p.aprobado_por = request.user
    p.fecha_aprobacion = date.today()
    p.observaciones = request.POST.get('observaciones', '').strip()
    p.save()

    from core.audit import log_update
    log_update(request, p, {'estado': {'old': old_estado, 'new': p.estado}},
               f'Papeleta {p.get_estado_display().lower()}: {p.get_tipo_permiso_display()} de {p.personal}')

    return JsonResponse(_papeleta_dict(p))


@login_required
@solo_admin
def papeletas_exportar(request):
    """Exportar papeletas a Excel respetando los filtros activos."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from django.http import HttpResponse
    from asistencia.models import RegistroPapeleta

    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '')
    origen = request.GET.get('origen', '')
    anio = int(request.GET.get('anio', date.today().year))
    personal_id = request.GET.get('personal', '')

    qs = RegistroPapeleta.objects.select_related(
        'personal', 'aprobado_por', 'creado_por',
    ).filter(fecha_inicio__year=anio)

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_permiso=tipo)
    if origen:
        qs = qs.filter(origen=origen)
    if personal_id:
        qs = qs.filter(personal_id=personal_id)

    qs = qs.order_by('-fecha_inicio')[:2000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Papeletas'

    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='0F766E')
    hdr_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'DNI', 'Trabajador', 'Tipo', 'Fecha Inicio', 'Fecha Fin',
        'Días Háb.', 'Estado', 'Origen', 'Detalle', 'Observaciones',
        'Aprobado Por', 'Fecha Aprob.', 'Creado En',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for row_idx, p in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=p.dni or (p.personal.nro_doc if p.personal else ''))
        ws.cell(row=row_idx, column=2, value=p.personal.apellidos_nombres if p.personal else '')
        ws.cell(row=row_idx, column=3, value=p.get_tipo_permiso_display())
        ws.cell(row=row_idx, column=4, value=p.fecha_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=5, value=p.fecha_fin.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=6, value=p.dias_habiles)
        ws.cell(row=row_idx, column=7, value=p.get_estado_display())
        ws.cell(row=row_idx, column=8, value=p.get_origen_display())
        ws.cell(row=row_idx, column=9, value=p.detalle or '')
        ws.cell(row=row_idx, column=10, value=p.observaciones or '')
        ws.cell(row=row_idx, column=11, value=str(p.aprobado_por) if p.aprobado_por else '')
        ws.cell(row=row_idx, column=12, value=p.fecha_aprobacion.strftime('%d/%m/%Y') if p.fecha_aprobacion else '')
        ws.cell(row=row_idx, column=13, value=p.creado_en.strftime('%d/%m/%Y %H:%M') if p.creado_en else '')

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'papeletas_{anio}.xlsx'
    if estado:
        filename = f'papeletas_{anio}_{estado.lower()}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _papeleta_dict(p):
    """Serializa una papeleta a dict para JSON response."""
    return {
        'ok': True,
        'pk': p.pk,
        'personal_nombre': p.personal.apellidos_nombres if p.personal else p.dni,
        'personal_id': p.personal_id,
        'dni': p.dni,
        'tipo_permiso': p.tipo_permiso,
        'tipo_display': p.get_tipo_permiso_display(),
        'fecha_inicio': p.fecha_inicio.strftime('%Y-%m-%d'),
        'fecha_inicio_display': p.fecha_inicio.strftime('%d/%m/%Y'),
        'fecha_fin': p.fecha_fin.strftime('%Y-%m-%d'),
        'fecha_fin_display': p.fecha_fin.strftime('%d/%m/%Y'),
        'fecha_referencia': p.fecha_referencia.strftime('%Y-%m-%d') if p.fecha_referencia else '',
        'fecha_referencia_display': p.fecha_referencia.strftime('%d/%m/%Y') if p.fecha_referencia else '—',
        'detalle': p.detalle,
        'dias_habiles': p.dias_habiles,
        'estado': p.estado,
        'estado_display': p.get_estado_display(),
        'origen': p.origen,
        'origen_display': p.get_origen_display(),
        'observaciones': p.observaciones,
        'aprobado_por': str(p.aprobado_por) if p.aprobado_por else '',
        'fecha_aprobacion': p.fecha_aprobacion.strftime('%d/%m/%Y') if p.fecha_aprobacion else '',
        'es_compensacion': p.es_compensacion,
        'es_importada': p.es_importada,
    }


@login_required
@solo_admin
def papeletas_reporte_agrupado(request):
    """Reporte de papeletas agrupado por tipo y por trabajador."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from django.http import HttpResponse
    from asistencia.models import RegistroPapeleta
    from collections import defaultdict

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))

    _, num_dias = calendar.monthrange(anio, mes)
    mes_ini = date(anio, mes, 1)
    mes_fin = date(anio, mes, num_dias)

    MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    qs = RegistroPapeleta.objects.filter(
        fecha_inicio__lte=mes_fin, fecha_fin__gte=mes_ini,
        estado__in=['APROBADA', 'EJECUTADA'],
    ).select_related('personal').order_by('tipo_permiso', 'personal__apellidos_nombres', 'fecha_inicio')

    # Agrupar por tipo -> trabajador -> lista de rangos
    por_tipo = defaultdict(lambda: defaultdict(list))
    for p in qs:
        nombre = p.personal.apellidos_nombres if p.personal else p.dni
        dni = p.personal.nro_doc if p.personal else p.dni
        por_tipo[p.get_tipo_permiso_display()][f'{nombre} ({dni})'].append({
            'inicio': p.fecha_inicio,
            'fin': p.fecha_fin,
            'dias': p.dias_habiles,
            'obs': p.observaciones or '',
        })

    # Generar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Papeletas {MESES[mes-1]} {anio}'

    title_font = Font(bold=True, size=14, color='0f766e')
    tipo_font = Font(bold=True, size=11, color='FFFFFF')
    tipo_fill = PatternFill(start_color='0f766e', end_color='0f766e', fill_type='solid')
    header_font = Font(bold=True, size=9, color='FFFFFF')
    header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
    nombre_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    border = Border(
        bottom=Side(style='thin', color='e2e8f0'),
    )

    # Titulo
    ws.cell(row=1, column=1, value=f'REPORTE DE PAPELETAS — {MESES[mes-1].upper()} {anio}').font = title_font
    ws.cell(row=2, column=1, value=f'Periodo: {mes_ini.strftime("%d/%m/%Y")} al {mes_fin.strftime("%d/%m/%Y")}').font = Font(size=9, color='64748b')

    row = 4
    for tipo_nombre, trabajadores in sorted(por_tipo.items()):
        # Header de tipo
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = tipo_fill
            cell.font = tipo_font
        ws.cell(row=row, column=1, value=tipo_nombre)
        total_tipo = sum(sum(p['dias'] for p in papeletas) for papeletas in trabajadores.values())
        ws.cell(row=row, column=5, value=f'{total_tipo} días total')
        row += 1

        # Sub-headers
        for c, h in enumerate(['Trabajador', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Observación'], 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1

        for nombre_trab, papeletas in sorted(trabajadores.items()):
            for i, p in enumerate(papeletas):
                if i == 0:
                    ws.cell(row=row, column=1, value=nombre_trab).font = nombre_font
                else:
                    ws.cell(row=row, column=1, value='').font = data_font
                ws.cell(row=row, column=2, value=p['inicio'].strftime('%d/%m/%Y')).font = data_font
                ws.cell(row=row, column=3, value=p['fin'].strftime('%d/%m/%Y')).font = data_font
                ws.cell(row=row, column=4, value=p['dias']).font = Font(bold=True, size=9)
                ws.cell(row=row, column=5, value=p['obs'][:60]).font = Font(size=8, color='64748b')
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = border
                row += 1

        row += 1  # Espacio entre tipos

    # Anchos de columna
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 45

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Papeletas_{MESES[mes-1]}_{anio}.xlsx"'
    return response
