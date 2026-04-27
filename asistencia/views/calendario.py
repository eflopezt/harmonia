"""
Vista Calendario Grid — Asistencia diaria con justificación inline.
"""
import calendar
import json
from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_POST

from django.db.models import F as DbF

from asistencia.models import (
    RegistroTareo, HomologacionCodigo, CambioCodigoLog,
    FeriadoCalendario, ConfiguracionSistema,
)
from asistencia.views._common import solo_admin, _qs_staff_dedup, _papeletas_bulk
from personal.models import Personal, Area

# Códigos de presencia (días trabajados)
CODIGOS_PRESENCIA = {'T', 'NOR', 'TR', 'A', 'CDT', 'CPF', 'LCG', 'ATM', 'CHE', 'LIM', 'SS'}
CODIGOS_FALTA = {'F', 'FA', 'LSG'}
CODIGOS_DESCANSO = {'DL', 'DLA', 'DS'}
CODIGOS_VACACIONES = {'VAC', 'V'}
CODIGOS_MEDICO = {'DM', 'SUB'}

# Mapa de colores CSS por código
COLOR_MAP = {
    'NOR': 'present', 'T': 'present', 'A': 'present', 'TR': 'present',
    'SS': 'present',
    'F': 'falta', 'FA': 'falta',
    'VAC': 'vac', 'V': 'vac',
    'DL': 'descanso', 'DLA': 'descanso', 'DS': 'descanso',
    'DM': 'medico', 'SUB': 'medico',
    'CHE': 'comp', 'CDT': 'comp', 'CPF': 'comp',
    'LSG': 'lsg',
    'LCG': 'licencia', 'LF': 'licencia', 'LP': 'licencia',
    'FER': 'feriado', 'FL': 'feriado_laborado', 'DSL': 'descanso_laborado',
    'LIM': 'lima', 'ATM': 'lima',
    'NA': 'empty',
}

MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
DIAS_SEMANA = ['L', 'M', 'Mi', 'J', 'V', 'S', 'D']


@login_required
@solo_admin
def calendario_grid(request):
    """Grilla calendario mensual o ciclo de asistencia (fechas desde configuración)."""
    from datetime import timedelta
    from asistencia.models import ConfiguracionSistema
    config = ConfiguracionSistema.get()
    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    grupo = request.GET.get('grupo', 'TODOS')
    area_id = request.GET.get('area', '')
    condicion = request.GET.get('condicion', '')
    buscar = request.GET.get('q', '')
    modo = request.GET.get('modo', 'mes')

    # Corte desde configuración (ej: 21 → ciclo del 22 del mes anterior al 21 del mes actual)
    corte = config.dia_corte_planilla

    if modo == 'ciclo':
        mes_ini, mes_fin = config.get_ciclo_he(anio, mes)
    else:
        _, num_dias = calendar.monthrange(anio, mes)
        mes_ini = date(anio, mes, 1)
        mes_fin = date(anio, mes, num_dias)

    # Días del rango con día de semana
    dias_mes = []
    dt = mes_ini
    while dt <= mes_fin:
        dias_mes.append({
            'num': dt.day,
            'dow': DIAS_SEMANA[dt.weekday()],
            'es_finde': dt.weekday() >= 5,
            'fecha': dt,
        })
        dt += timedelta(days=1)

    # Feriados
    feriados = set(
        FeriadoCalendario.objects
        .filter(fecha__gte=mes_ini, fecha__lte=mes_fin)
        .values_list('fecha', flat=True)
    )
    for d in dias_mes:
        d['es_feriado'] = d['fecha'] in feriados

    # Query registros — excluir fuera de fecha_alta/fecha_cese (igual que RCO/STAFF)
    if grupo == 'STAFF':
        qs = _qs_staff_dedup(mes_ini, mes_fin)
    elif grupo == 'RCO':
        qs = RegistroTareo.objects.filter(
            grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin,
            personal__isnull=False)
    else:
        # TODOS — combinar STAFF dedup + RCO
        qs_staff = _qs_staff_dedup(mes_ini, mes_fin)
        qs_rco = RegistroTareo.objects.filter(
            grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin,
            personal__isnull=False)
        qs = qs_staff | qs_rco

    # Excluir registros fuera del periodo laboral del trabajador
    qs = qs.exclude(
        personal__fecha_cese__isnull=False, fecha__gt=DbF('personal__fecha_cese')
    ).exclude(
        personal__fecha_alta__isnull=False, fecha__lt=DbF('personal__fecha_alta')
    )

    # Filtros
    if area_id:
        qs = qs.filter(personal__subarea__area_id=area_id)
    if condicion:
        # Soportar FORANEO con o sin tilde
        if condicion.upper() in ('FORANEO', 'FORÁNEO'):
            qs = qs.filter(condicion__in=['FORANEO', 'FORÁNEO'])
        else:
            qs = qs.filter(condicion=condicion)
    if buscar:
        qs = qs.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(dni__icontains=buscar)
        )

    # Fetch all records — incluir fuente_codigo y horas_normales para dedup y detalle
    registros = list(qs.select_related('personal').values(
        'id', 'personal_id', 'personal__apellidos_nombres', 'personal__nro_doc',
        'personal__condicion', 'personal__fecha_alta', 'personal__fecha_cese',
        'fecha', 'codigo_dia', 'grupo', 'fuente_codigo',
        'hora_entrada_real', 'hora_salida_real',
        'horas_efectivas', 'horas_normales',
        'he_25', 'he_35', 'he_100', 'observaciones',
    ))

    # Pivot: personal_id -> {dia: registro}
    # Dedup: RELOJ > EXCEL; entre misma fuente, último pk gana (misma lógica que _build_rco_data)
    FUENTE_PRIORIDAD = {'PAPELETA': 0, 'MANUAL': 1, 'RELOJ': 2, 'FERIADO': 3, 'FALTA_AUTO': 4, 'EXCEL': 5}
    pivot = defaultdict(dict)
    personal_info = {}
    for r in registros:
        pid = r['personal_id']
        dia = r['fecha'].day
        existente = pivot[pid].get(dia)
        if existente:
            # Prioridad: menor número = mayor prioridad
            prio_new = FUENTE_PRIORIDAD.get(r['fuente_codigo'], 9)
            prio_old = FUENTE_PRIORIDAD.get(existente['fuente_codigo'], 9)
            if prio_new < prio_old:
                pivot[pid][dia] = r
            elif prio_new == prio_old:
                # Misma fuente: último pk gana (más reciente)
                if r['id'] > existente['id']:
                    pivot[pid][dia] = r
            # else: existente tiene mejor prioridad, no reemplazar
        else:
            pivot[pid][dia] = r
        if pid not in personal_info:
            personal_info[pid] = {
                'nombre': r['personal__apellidos_nombres'],
                'dni': r['personal__nro_doc'],
                'condicion': r['personal__condicion'] or '',
                'grupo': r['grupo'],
                'fecha_alta': r.get('personal__fecha_alta'),
                'fecha_cese': r.get('personal__fecha_cese'),
            }

    # Cargar papeletas aprobadas para todos los empleados visibles
    pap_bulk = _papeletas_bulk(list(personal_info.keys()), mes_ini, mes_fin)

    # Build rows sorted by name
    rows = []
    for pid in sorted(personal_info, key=lambda x: personal_info[x]['nombre']):
        info = personal_info[pid]
        celdas = []
        total_trab = 0
        total_falta = 0
        total_he = Decimal('0')
        fecha_alta = info.get('fecha_alta')
        fecha_cese = info.get('fecha_cese')
        for d in dias_mes:
            # N/A si fuera de periodo laboral
            if (fecha_alta and d['fecha'] < fecha_alta) or (fecha_cese and d['fecha'] > fecha_cese):
                celdas.append({'id': 0, 'codigo': 'NA', 'color': 'empty', 'editable': False})
                continue
            reg = pivot[pid].get(d['num'])
            if reg:
                codigo = reg['codigo_dia']
                # Auto DS para LOCAL domingos
                if d['fecha'].weekday() == 6 and info['condicion'].upper() in ('LOCAL', 'LIMA', '') and codigo in ('FA', 'F'):
                    codigo = 'DS'
                # Papeleta aprobada sobreescribe códigos genéricos (FA, F, DS)
                pap_reg = pap_bulk.get(pid, {}).get(d['fecha'])
                if pap_reg and codigo in ('FA', 'F', 'DS'):
                    codigo = pap_reg['codigo']
                color = COLOR_MAP.get(codigo, 'other')
                if codigo in CODIGOS_PRESENCIA:
                    total_trab += 1
                elif codigo in CODIGOS_FALTA:
                    total_falta += 1
                he = (reg['he_25'] or 0) + (reg['he_35'] or 0) + (reg['he_100'] or 0)
                total_he += Decimal(str(he))
                celda_reg = {
                    'id': reg['id'],
                    'codigo': codigo,
                    'color': color,
                    'entrada': str(reg['hora_entrada_real'])[:5] if reg['hora_entrada_real'] else '',
                    'salida': str(reg['hora_salida_real'])[:5] if reg['hora_salida_real'] else '',
                    'he': float(he) if he else 0,
                }
                if pap_reg:
                    celda_reg['papeleta'] = pap_reg
                celdas.append(celda_reg)
            else:
                # Fallback: papeleta → LIMA auto-A → DS domingo → FA
                pap_info = pap_bulk.get(pid, {}).get(d['fecha'])
                dow = d['fecha'].weekday()
                cond_upper = info['condicion'].upper()
                if pap_info:
                    # Papeleta siempre tiene prioridad (incluso domingos)
                    auto_cod = pap_info['codigo']
                elif dow == 6 and cond_upper in ('LOCAL', 'LIMA', ''):
                    auto_cod = 'DS'
                elif cond_upper == 'LIMA' and dow < 6:
                    # LIMA no marca asistencia, lun-sab = presente por defecto
                    auto_cod = 'A'
                else:
                    auto_cod = 'FA'
                color = COLOR_MAP.get(auto_cod, 'other' if auto_cod else 'empty')
                if auto_cod in CODIGOS_FALTA:
                    total_falta += 1
                celda_dict = {
                    'id': 0, 'codigo': auto_cod, 'color': color,
                    'editable': True, 'fecha_iso': d['fecha'].isoformat(),
                }
                if pap_info:
                    celda_dict['papeleta'] = pap_info
                celdas.append(celda_dict)
        rows.append({
            'personal_id': pid,
            'nombre': info['nombre'],
            'dni': info['dni'],
            'condicion': info['condicion'],
            'grupo': info['grupo'],
            'celdas': celdas,
            'total_trab': total_trab,
            'total_falta': total_falta,
            'total_he': float(total_he),
        })

    # Summary per day
    resumen_dias = []
    for i, d in enumerate(dias_mes):
        presentes = sum(1 for r in rows if r['celdas'][i]['codigo'] in CODIGOS_PRESENCIA)
        ausentes = sum(1 for r in rows if r['celdas'][i]['codigo'] in CODIGOS_FALTA)
        total = presentes + ausentes
        pct = round(presentes / total * 100) if total else 0
        resumen_dias.append({'presentes': presentes, 'ausentes': ausentes, 'pct': pct})

    # Stats globales
    total_presentes = sum(r['total_trab'] for r in rows)
    total_faltas = sum(r['total_falta'] for r in rows)
    total_registros = total_presentes + total_faltas
    pct_global = round(total_presentes / total_registros * 100, 1) if total_registros else 0

    # Areas para filtro
    areas = Area.objects.all().order_by('nombre')

    # Códigos disponibles para justificación
    codigos_disponibles = list(
        HomologacionCodigo.objects
        .filter(activo=True)
        .values('codigo_tareo', 'descripcion')
        .order_by('codigo_tareo')
    )

    context = {
        'titulo': f'{"Ciclo" if modo == "ciclo" else "Calendario"} de Asistencia — {MESES[mes]} {anio}',
        'anio': anio, 'mes': mes, 'mes_nombre': MESES[mes],
        'dias_mes': dias_mes, 'num_dias': len(dias_mes),
        'rows': rows, 'resumen_dias': resumen_dias,
        'total_empleados': len(rows),
        'total_presentes': total_presentes,
        'total_faltas': total_faltas,
        'pct_global': pct_global,
        'grupo': grupo, 'area_id': area_id, 'condicion': condicion, 'buscar': buscar,
        'modo': modo,
        # Para link de Exportar: 'corte' equivale a modo 'ciclo'.
        'tipo_periodo': 'corte' if modo == 'ciclo' else 'calendario',
        'ciclo_inicio_dia': corte + 1,
        'ciclo_fin_dia': corte,
        'areas': areas,
        'codigos_json': json.dumps(codigos_disponibles),
        'color_map_json': json.dumps(COLOR_MAP),
        'anios': list(range(hoy.year - 2, hoy.year + 1)),
        'meses_list': [(i, MESES[i]) for i in range(1, 13)],
    }
    return render(request, 'asistencia/calendario_grid.html', context)


@login_required
@solo_admin
def ajax_calendario_detalle(request, registro_id):
    """Detalle de celda para el modal."""
    reg = get_object_or_404(
        RegistroTareo.objects.select_related('personal'),
        pk=registro_id
    )
    cambios = list(
        CambioCodigoLog.objects
        .filter(registro=reg)
        .select_related('usuario')
        .values('codigo_anterior', 'codigo_nuevo', 'observacion', 'creado_en', 'usuario__username')
        .order_by('-creado_en')[:10]
    )
    for c in cambios:
        c['creado_en'] = c['creado_en'].strftime('%d/%m/%Y %H:%M')

    data = {
        'id': reg.id,
        'nombre': reg.personal.apellidos_nombres if reg.personal else reg.dni,
        'dni': reg.dni,
        'fecha': reg.fecha.strftime('%d/%m/%Y'),
        'fecha_dow': DIAS_SEMANA[reg.fecha.weekday()],
        'codigo': reg.codigo_dia,
        'color': COLOR_MAP.get(reg.codigo_dia, 'other'),
        'grupo': reg.grupo,
        'condicion': reg.condicion,
        'entrada': str(reg.hora_entrada_real)[:5] if reg.hora_entrada_real else '-',
        'salida': str(reg.hora_salida_real)[:5] if reg.hora_salida_real else '-',
        'horas_efectivas': float(reg.horas_efectivas or 0),
        'horas_normales': float(reg.horas_normales or 0),
        'he_25': float(reg.he_25 or 0),
        'he_35': float(reg.he_35 or 0),
        'he_100': float(reg.he_100 or 0),
        'observaciones': reg.observaciones or '',
        'fuente': reg.fuente_codigo,
        'almuerzo_manual': float(reg.almuerzo_manual) if reg.almuerzo_manual is not None else None,
        'cambios': cambios,
    }
    return JsonResponse(data)


def _recalcular_horas(reg):
    """Recalcular horas de un RegistroTareo usando la misma lógica del processor.

    Mantiene paridad con TareoProcessor._calcular_horas_raw del processor
    a través de helpers compartidos en asistencia.services._helpers.
    """
    import logging
    from asistencia.services._helpers import (
        calcular_almuerzo_h, redondear_media_hora, cap_horas_dia,
        calcular_horas_marcadas_desde_horarios,
    )
    config = ConfiguracionSistema.get()
    personal = reg.personal
    condicion = (reg.condicion or '').upper()
    CERO = Decimal('0')
    logger = logging.getLogger('asistencia')

    # Obtener jornada (misma lógica que processor._obtener_jornada)
    dia_semana = reg.fecha.weekday()
    # #5 fix: override personal aplica si está seteado (cualquier valor),
    # no solo si es != 8h.
    if personal and personal.jornada_horas is not None:
        jornada_h = Decimal(str(personal.jornada_horas))
    elif dia_semana == 6:  # domingo
        if condicion.replace('Á', 'A') == 'FORANEO':
            # FORÁNEO: domingo es parte del ciclo 21×7, jornada reducida 4h
            jornada_h = Decimal(str(config.jornada_domingo_horas))
        else:
            # LOCAL/LIMA: domingo es descanso semanal, si labora todo al 100%
            jornada_h = CERO
    elif condicion in ('FORANEO', 'FORÁNEO'):
        jornada_h = Decimal(str(config.jornada_foraneo_horas))
    elif dia_semana == 5:  # sábado
        jornada_h = Decimal(str(config.jornada_sabado_horas))
    else:
        jornada_h = Decimal(str(config.jornada_local_horas))

    # Calcular y capar horas marcadas
    horas_marcadas = calcular_horas_marcadas_desde_horarios(
        reg.fecha, reg.hora_entrada_real, reg.hora_salida_real,
        horas_marcadas_fallback=reg.horas_marcadas,
    )
    # #6 fix: tope sanitario 16h
    horas_marcadas = cap_horas_dia(
        horas_marcadas, dni=reg.dni, fecha=reg.fecha, logger=logger,
    )
    reg.horas_marcadas = horas_marcadas

    codigo = reg.codigo_dia
    # DS/FER aquí (= NO laborado, fuerza 0 horas).
    # Para "laborado en descanso/feriado" usar DSL/FL → 100% via codigo_fuerza_100.
    CODIGOS_SIN_HE = {'SS', 'DL', 'DLA', 'CHE', 'VAC', 'DM', 'LCG', 'LF', 'LP',
                      'LSG', 'FA', 'TR', 'CDT', 'CPF', 'ATM', 'SAI', 'F',
                      'V', 'SUB', 'B', 'LIM', 'NA', 'DS', 'FER'}

    def _set_horas(ef, norm, h25, h35, h100):
        """Asigna componentes redondeados a 0.5h (#3 fix)."""
        reg.horas_efectivas = redondear_media_hora(ef)
        reg.horas_normales = redondear_media_hora(norm)
        reg.he_25 = redondear_media_hora(h25)
        reg.he_35 = redondear_media_hora(h35)
        reg.he_100 = redondear_media_hora(h100)

    # SS: paga jornada completa
    # En LOCAL domingo o feriado: SS también va al 100% (D.Leg. 713)
    if codigo == 'SS':
        j = jornada_h if jornada_h > CERO else Decimal('8.5')
        _es_feriado_ss = reg.es_feriado or FeriadoCalendario.objects.filter(
            fecha=reg.fecha, activo=True).exists()
        _es_dom = reg.fecha.weekday() == 6
        if (_es_feriado_ss or _es_dom) and (_es_feriado_ss or jornada_h == CERO):
            _set_horas(j, CERO, CERO, CERO, j)
        else:
            _set_horas(j, j, CERO, CERO, CERO)
        return

    # Marcación incompleta: horas < jornada/2 → SS implícito
    if (horas_marcadas > CERO and horas_marcadas < jornada_h / 2
            and codigo not in CODIGOS_SIN_HE):
        _set_horas(jornada_h, jornada_h, CERO, CERO, CERO)
        return

    # Códigos sin horas
    if codigo in CODIGOS_SIN_HE or not horas_marcadas or horas_marcadas <= CERO:
        _set_horas(CERO, CERO, CERO, CERO, CERO)
        return

    # #2 fix: descuento almuerzo unificado (chequea jornada > 6h)
    almuerzo = calcular_almuerzo_h(horas_marcadas, jornada_h, reg.almuerzo_manual)
    horas_ef = max(CERO, horas_marcadas - almuerzo)

    # Feriado/Domingo trabajado → jornada normal + exceso HE 100%
    # EXCEPCIÓN: si el código fue cambiado manualmente a NOR/T/A → calcular como día normal
    # (el trabajador tiene descanso semanal en otro día, no el domingo)
    # Códigos LABORADOS (DSL/FL) marcan descanso/feriado trabajado en días que
    # el calendario no marca (p.ej. DSL miércoles = descanso rotativo trabajado).
    es_feriado = (reg.es_feriado
                  or FeriadoCalendario.objects.filter(fecha=reg.fecha, activo=True).exists()
                  or codigo == 'FL')
    es_descanso_semanal = reg.fecha.weekday() == 6 or codigo == 'DSL'
    codigo_fuerza_normal = codigo in ('NOR', 'T', 'A') and reg.fuente_codigo == 'MANUAL'
    codigo_fuerza_100 = codigo in ('DSL', 'FL')
    if (es_feriado or es_descanso_semanal) and not codigo_fuerza_normal:
        if es_feriado or codigo_fuerza_100 or jornada_h == CERO:
            # Feriado (toda condición), código DSL/FL o LOCAL domingo: TODAS al 100%
            _set_horas(horas_ef, CERO, CERO, CERO, horas_ef)
        else:
            # FORÁNEO domingo sin código de descanso (4h jornada):
            # normal hasta jornada, exceso HE 100%
            _set_horas(horas_ef, min(horas_ef, jornada_h), CERO, CERO,
                       max(CERO, horas_ef - jornada_h))
        return

    # #7 fix: he_bloqueado consulta ConfiguracionSistema + SolicitudHE
    # (misma lógica que el processor — sin solicitud aprobada cuando se requiere)
    he_bloqueado = False
    if personal and getattr(config, 'he_requiere_solicitud', False):
        from asistencia.models import SolicitudHE
        he_bloqueado = not SolicitudHE.objects.filter(
            personal=personal, fecha=reg.fecha, estado='APROBADA',
        ).exists()

    # Día normal: dentro de jornada
    if horas_ef <= jornada_h:
        _set_horas(horas_ef, horas_ef, CERO, CERO, CERO)
        return

    # Exceso sobre jornada
    if he_bloqueado:
        # HE bloqueadas para este empleado: el exceso se descarta
        _set_horas(jornada_h, jornada_h, CERO, CERO, CERO)
    # Horas extra (HE25 hasta 2h, resto HE35)
    exceso = horas_ef - jornada_h
    _set_horas(
        horas_ef, jornada_h,
        min(exceso, Decimal('2')),
        max(CERO, exceso - Decimal('2')),
        CERO,
    )


@login_required
@solo_admin
@require_POST
def ajax_calendario_cambiar(request, registro_id):
    """Cambiar código y/o entrada/salida de un registro (justificación)."""
    from datetime import time as dt_time
    from asistencia.services._helpers import fecha_en_periodo_cerrado
    reg = get_object_or_404(RegistroTareo, pk=registro_id)

    # #1 fix: bloquear edición si el período está CERRADO
    if fecha_en_periodo_cerrado(reg.fecha):
        from cierre.models import PeriodoCierre
        per = PeriodoCierre.objects.filter(
            anio=reg.fecha.year, mes=reg.fecha.month, estado='CERRADO'
        ).first()
        nombre = per.mes_nombre if per else f'{reg.fecha.month:02d}/{reg.fecha.year}'
        return JsonResponse({
            'error': f'Período {nombre} {reg.fecha.year} está CERRADO. '
                     'Reabrir el período antes de modificar.',
        }, status=409)

    nuevo_codigo = request.POST.get('codigo', '').strip().upper()
    observacion = request.POST.get('observacion', '').strip()
    sustento = request.FILES.get('sustento')
    nueva_entrada = request.POST.get('hora_entrada', '').strip()
    nueva_salida = request.POST.get('hora_salida', '').strip()

    if not nuevo_codigo and not nueva_entrada and not nueva_salida:
        return JsonResponse({'error': 'Debe indicar un código o cambiar entrada/salida'}, status=400)

    codigo_anterior = reg.codigo_dia
    entrada_anterior = str(reg.hora_entrada_real)[:5] if reg.hora_entrada_real else '-'
    salida_anterior = str(reg.hora_salida_real)[:5] if reg.hora_salida_real else '-'

    # Construir detalle del cambio para el log
    cambios_detalle = []

    if nuevo_codigo and nuevo_codigo != codigo_anterior:
        cambios_detalle.append(f'{codigo_anterior}→{nuevo_codigo}')
        reg.codigo_dia = nuevo_codigo
        reg.fuente_codigo = 'MANUAL'
    else:
        nuevo_codigo = codigo_anterior  # mantener el mismo

    if nueva_entrada:
        try:
            parts = nueva_entrada.split(':')
            t = dt_time(int(parts[0]), int(parts[1]))
            cambios_detalle.append(f'Entrada: {entrada_anterior}→{nueva_entrada}')
            reg.hora_entrada_real = t
        except (ValueError, IndexError):
            return JsonResponse({'error': 'Formato de entrada inválido (HH:MM)'}, status=400)

    if nueva_salida:
        try:
            parts = nueva_salida.split(':')
            t = dt_time(int(parts[0]), int(parts[1]))
            cambios_detalle.append(f'Salida: {salida_anterior}→{nueva_salida}')
            reg.hora_salida_real = t
        except (ValueError, IndexError):
            return JsonResponse({'error': 'Formato de salida inválido (HH:MM)'}, status=400)

    # Almuerzo manual
    almuerzo_val = request.POST.get('almuerzo', 'auto')
    if almuerzo_val == 'auto':
        reg.almuerzo_manual = None
    else:
        reg.almuerzo_manual = Decimal(almuerzo_val)

    # Si ahora tiene entrada Y salida pero el código era SS, cambiar a A
    if reg.hora_entrada_real and reg.hora_salida_real and reg.codigo_dia == 'SS':
        reg.codigo_dia = 'A'

    # Recalcular horas si cambió entrada/salida o almuerzo.
    # NO tocar fuente_codigo aquí: solo debe ser 'MANUAL' si el código se
    # cambió explícitamente (línea 507). Si solo se editan horas, el código
    # sigue viniendo del reloj/regla/feriado y codigo_fuerza_normal no debe
    # dispararse en domingos/feriados (LOCAL dom = 100%, no día normal).
    if nueva_entrada or nueva_salida or almuerzo_val != 'auto':
        _recalcular_horas(reg)

    # Log del cambio
    log_obs = ' | '.join(cambios_detalle)
    if observacion:
        log_obs = f'{log_obs} — {observacion}' if log_obs else observacion

    CambioCodigoLog.objects.create(
        registro=reg,
        codigo_anterior=codigo_anterior,
        codigo_nuevo=nuevo_codigo,
        observacion=log_obs,
        sustento=sustento,
        usuario=request.user,
    )

    # Actualizar observaciones del registro
    if log_obs:
        prev = reg.observaciones or ''
        reg.observaciones = f'{prev}\n[{request.user.username}] {log_obs}'.strip()
    reg.save()

    # Si el código nuevo es un permiso/licencia, crear papeleta automática
    # (si no hay ya una APROBADA/EJECUTADA cubriendo el día). Vincula el
    # registro a la papeleta para reflejarlo en reportes y descontar saldos.
    papeleta_info = None
    try:
        from asistencia.services.papeletas_sync import (
            CODIGO_A_TIPO, crear_papeleta_desde_codigo,
        )
        if reg.personal_id and nuevo_codigo in CODIGO_A_TIPO:
            obs_pap = (observacion or
                       f'Auto-creada desde matriz: {codigo_anterior}→{nuevo_codigo}')
            pap, creada = crear_papeleta_desde_codigo(
                reg.personal, reg.fecha, nuevo_codigo,
                usuario=request.user, observacion=obs_pap,
            )
            if pap is not None:
                reg.fuente_codigo = 'PAPELETA'
                reg.papeleta_ref = f'PAP#{pap.pk}'
                reg.save(update_fields=['fuente_codigo', 'papeleta_ref'])
                papeleta_info = {
                    'id': pap.pk,
                    'creada': creada,
                    'tipo': pap.get_tipo_permiso_display(),
                }
    except Exception as exc:
        import logging
        logging.getLogger('personal.business').warning(
            f'auto-papeleta error (reg={reg.pk}, cod={nuevo_codigo}): {exc}'
        )

    return JsonResponse({
        'ok': True,
        'codigo': nuevo_codigo,
        'color': COLOR_MAP.get(nuevo_codigo, 'other'),
        'anterior': codigo_anterior,
        'entrada': str(reg.hora_entrada_real)[:5] if reg.hora_entrada_real else '-',
        'salida': str(reg.hora_salida_real)[:5] if reg.hora_salida_real else '-',
        'horas_efectivas': float(reg.horas_efectivas or 0),
        'horas_normales': float(reg.horas_normales or 0),
        'he_25': float(reg.he_25 or 0),
        'he_35': float(reg.he_35 or 0),
        'he_100': float(reg.he_100 or 0),
        'papeleta': papeleta_info,
    })


@login_required
@solo_admin
@require_POST
def ajax_calendario_crear(request):
    """Crear un RegistroTareo nuevo para una celda vacía."""
    from datetime import time as dt_time
    personal_id = request.POST.get('personal_id')
    fecha_str = request.POST.get('fecha', '')
    nuevo_codigo = request.POST.get('codigo', '').strip().upper()
    observacion = request.POST.get('observacion', '').strip()
    nueva_entrada = request.POST.get('hora_entrada', '').strip()
    nueva_salida = request.POST.get('hora_salida', '').strip()

    if not personal_id or not fecha_str:
        return JsonResponse({'error': 'personal_id y fecha requeridos'}, status=400)
    if not nuevo_codigo:
        return JsonResponse({'error': 'Código requerido'}, status=400)

    try:
        fecha = date.fromisoformat(fecha_str)
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)

    personal = get_object_or_404(Personal, pk=personal_id)

    # Importación dummy para registros manuales (importacion_id es NOT NULL)
    from asistencia.models import TareoImportacion
    imp_manual, _ = TareoImportacion.objects.get_or_create(
        tipo='MANUAL', archivo_nombre='manual',
        defaults={
            'estado': 'COMPLETADO', 'total_registros': 0,
            'periodo_inicio': date(2025, 1, 1), 'periodo_fin': date(2030, 12, 31),
        },
    )

    # Crear registro
    reg = RegistroTareo(
        importacion=imp_manual,
        personal=personal,
        dni=personal.nro_doc,
        nombre_archivo=personal.apellidos_nombres,
        grupo=personal.grupo_tareo or 'STAFF',
        condicion=personal.condicion or 'LOCAL',
        fecha=fecha,
        dia_semana=fecha.weekday(),
        codigo_dia=nuevo_codigo,
        fuente_codigo='MANUAL',
        observaciones=f'[{request.user.username}] Creado manual: {nuevo_codigo}',
    )

    if nueva_entrada:
        try:
            parts = nueva_entrada.split(':')
            reg.hora_entrada_real = dt_time(int(parts[0]), int(parts[1]))
        except (ValueError, IndexError):
            return JsonResponse({'error': 'Formato de entrada inválido (HH:MM)'}, status=400)

    if nueva_salida:
        try:
            parts = nueva_salida.split(':')
            reg.hora_salida_real = dt_time(int(parts[0]), int(parts[1]))
        except (ValueError, IndexError):
            return JsonResponse({'error': 'Formato de salida inválido (HH:MM)'}, status=400)

    # Feriado
    reg.es_feriado = FeriadoCalendario.objects.filter(fecha=fecha, activo=True).exists()

    # Recalcular horas
    _recalcular_horas(reg)

    if observacion:
        reg.observaciones = f'{reg.observaciones}\n{observacion}'.strip()

    reg.save()

    # Log
    CambioCodigoLog.objects.create(
        registro=reg,
        codigo_anterior='',
        codigo_nuevo=nuevo_codigo,
        observacion=f'Registro creado manual — {observacion}' if observacion else 'Registro creado manual',
        usuario=request.user,
    )

    return JsonResponse({
        'ok': True,
        'id': reg.id,
        'codigo': nuevo_codigo,
        'color': COLOR_MAP.get(nuevo_codigo, 'other'),
        'entrada': str(reg.hora_entrada_real)[:5] if reg.hora_entrada_real else '-',
        'salida': str(reg.hora_salida_real)[:5] if reg.hora_salida_real else '-',
        'horas_efectivas': float(reg.horas_efectivas or 0),
    })


@login_required
@solo_admin
def calendario_exportar(request):
    """Exportar calendario a Excel respetando el tipo_periodo (calendario|corte)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from io import BytesIO
    from asistencia.views.exportaciones import _calcular_periodo, _get_corte_config

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))
    grupo = request.GET.get('grupo', 'TODOS')
    tipo_periodo = request.GET.get('tipo_periodo', 'calendario')
    if tipo_periodo not in ('calendario', 'corte'):
        tipo_periodo = 'calendario'

    d_ini_corte, d_fin_corte = _get_corte_config(request)
    mes_ini, mes_fin = _calcular_periodo(anio, mes, tipo_periodo, d_ini_corte, d_fin_corte)
    num_dias = (mes_fin - mes_ini).days + 1

    # Query
    if grupo == 'STAFF':
        qs = _qs_staff_dedup(mes_ini, mes_fin)
    elif grupo == 'RCO':
        qs = RegistroTareo.objects.filter(grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin, personal__isnull=False)
    else:
        qs = _qs_staff_dedup(mes_ini, mes_fin) | RegistroTareo.objects.filter(
            grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin, personal__isnull=False)

    registros = list(qs.select_related('personal').values(
        'personal_id', 'personal__apellidos_nombres', 'personal__nro_doc',
        'personal__condicion', 'fecha', 'codigo_dia', 'grupo'))

    # pivot keyed por r['fecha'] (date) en lugar de day, así soporta ciclo
    # 22→21 que cruza mes.
    pivot = defaultdict(dict)
    personal_info = {}
    for r in registros:
        pid = r['personal_id']
        pivot[pid][r['fecha']] = r['codigo_dia']
        if pid not in personal_info:
            personal_info[pid] = {
                'nombre': r['personal__apellidos_nombres'],
                'dni': r['personal__nro_doc'],
                'condicion': r['personal__condicion'] or '',
                'grupo': r['grupo'],
                'fecha_alta': r.get('personal__fecha_alta'),
                'fecha_cese': r.get('personal__fecha_cese'),
            }

    # Colors
    FILLS = {
        'present': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
        'falta': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        'vac': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
        'descanso': PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid'),
        'medico': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'comp': PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid'),
        'lsg': PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid'),
        'licencia': PatternFill(start_color='CFFAFE', end_color='CFFAFE', fill_type='solid'),
    }
    header_fill = PatternFill(start_color='1A2B47', end_color='1A2B47', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    wb = Workbook()
    ws = wb.active
    sufijo_titulo = ' Ciclo' if tipo_periodo == 'corte' else ''
    ws.title = f'{MESES[mes]} {anio}{sufijo_titulo}'[:31]

    # Lista ordenada de fechas del rango (soporta ciclo cross-mes)
    fechas_rango = []
    cur = mes_ini
    while cur <= mes_fin:
        fechas_rango.append(cur)
        from datetime import timedelta as _td
        cur += _td(days=1)

    # Header
    headers = ['N°', 'DNI', 'Empleado', 'Cond.', 'Grupo']
    for f in fechas_rango:
        headers.append(f'{f.day:02d}/{f.month:02d}\n{DIAS_SEMANA[f.weekday()]}')
    headers.extend(['Trab', 'Falta', 'HE'])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    row_num = 2
    for idx, pid in enumerate(sorted(personal_info, key=lambda x: personal_info[x]['nombre']), 1):
        info = personal_info[pid]
        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=2, value=info['dni']).border = thin_border
        ws.cell(row=row_num, column=3, value=info['nombre']).border = thin_border
        ws.cell(row=row_num, column=4, value=info['condicion']).border = thin_border
        ws.cell(row=row_num, column=5, value=info['grupo']).border = thin_border

        trab = falta = 0
        for i, f in enumerate(fechas_rango, start=1):
            codigo = pivot[pid].get(f, '')
            col = 5 + i
            cell = ws.cell(row=row_num, column=col, value=codigo)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.font = Font(size=8, bold=True)
            color_key = COLOR_MAP.get(codigo, '')
            if color_key in FILLS:
                cell.fill = FILLS[color_key]
            if codigo in CODIGOS_PRESENCIA:
                trab += 1
            elif codigo in CODIGOS_FALTA:
                falta += 1

        ws.cell(row=row_num, column=5 + num_dias + 1, value=trab).border = thin_border
        ws.cell(row=row_num, column=5 + num_dias + 2, value=falta).border = thin_border
        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 7
    for d in range(1, num_dias + 1):
        col_letter = ws.cell(row=1, column=5 + d).column_letter
        ws.column_dimensions[col_letter].width = 4.5

    ws.freeze_panes = 'F2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=calendario_{MESES[mes]}_{anio}.xlsx'
    return response
