"""
Exportadores del módulo Tareo.

CargaS10Exporter  → genera el archivo Excel para importar en S10
                     (horas extras y otros conceptos por empleado)
ReporteCierreExporter → reporte de cierre de mes con regularizaciones
"""
from __future__ import annotations

import calendar
import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.db.models import Q, Sum

logger = logging.getLogger('personal.business')

CERO = Decimal('0')

# Conceptos S10 estándar que genera el sistema
# Mapeados desde tareo → columna S10
CONCEPTOS_S10_DEFAULT = [
    ('HE25',  'HORAS EXTRAS 25%',                 'HORAS'),
    ('HE35',  'HORAS EXTRAS 35%',                 'HORAS'),
    ('HE100', 'HORAS EXTRAS 100%',                'HORAS'),
    ('DM',    'DIAS DESCANSO MEDICO',             'DIAS'),
    ('LF',    'DIAS LICENCIA FALLECIMIENTO',      'DIAS'),
    ('LSG',   'DIAS LICENCIA S/GOCE',             'DIAS'),
    ('LP',    'DIAS LICENCIA PATERNIDAD',         'DIAS'),
    ('FA',    'DIAS FALTA NO JUST.',              'DIAS'),
    ('SAI',   'DIAS SUSPENSION ACTO INSEGURO',    'DIAS'),
    ('VAC',   'DIAS DESCANSO VACACIONAL',         'DIAS'),
]


class CargaS10Exporter:
    """
    Genera el archivo CargaS10 en Excel.

    El CargaS10 tiene:
      Col A: Periodo (ej. "Febrero 2026 - 02")
      Col B: Código SAP
      Col C: DNI
      Col D: Nombre
      Cols E+: Una columna por concepto de planilla

    Solo incluye columnas con valores > 0 para el período,
    más las columnas HE que siempre se incluyen.
    """

    def __init__(self, anio: int, mes: int, config=None):
        from asistencia.models import ConfiguracionSistema, ConceptoMapeoS10
        self.anio = anio
        self.mes = mes
        self.config = config or ConfiguracionSistema.get()

        # Construir mapa de conceptos desde DB + defaults
        conceptos_db = {
            c.codigo_tareo: (c.nombre_concepto_s10, c.tipo_valor)
            for c in ConceptoMapeoS10.objects.filter(activo=True)
        }
        self.conceptos: list[tuple[str, str, str]] = []
        for cod, nombre, tipo in CONCEPTOS_S10_DEFAULT:
            nombre_real, tipo_real = conceptos_db.get(cod, (nombre, tipo))
            self.conceptos.append((cod, nombre_real, tipo_real))

        # Periodo label
        MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        self.periodo_label = f"{MESES[mes-1]} {anio} - {mes:02d}"

    def generar(self, importacion_ids: list[int] | None = None) -> BytesIO:
        """
        Genera el Excel CargaS10.

        importacion_ids: filtrar solo estas importaciones.
                         Si None, selecciona automáticamente la fuente
                         prioritaria: CSRT_RCO > RELOJ para el período.

        Prioridad de fuentes RCO:
          1. CSRT_RCO — hoja curada por RRHH (valores definitivos)
          2. RELOJ    — biométrico Synkro (fallback si no hay CSRT_RCO)
        Los registros RELOJ se excluyen cuando existen registros CSRT_RCO
        para el mismo período, evitando contabilizar HE duplicadas.
        """
        from personal.models import Personal
        from asistencia.models import RegistroTareo, TareoImportacion

        # ── Determinar ciclos ──────────────────────────────────
        inicio_he, fin_he = self.config.get_ciclo_he(self.anio, self.mes)
        inicio_asist, fin_asist = self.config.get_ciclo_asistencia(self.anio, self.mes)

        # ── Construir datos por empleado ───────────────────────
        qs_base = RegistroTareo.objects.filter(
            personal__isnull=False,
            grupo='RCO',    # RCO: HE se pagan
        )

        if importacion_ids:
            # El llamante especificó importaciones concretas → respetar
            qs_base = qs_base.filter(importacion_id__in=importacion_ids)
        else:
            # Auto-selección: preferir CSRT_RCO sobre RELOJ
            csrt_rco_ids = list(
                TareoImportacion.objects.filter(
                    tipo='CSRT_RCO',
                    periodo_fin__year=self.anio,
                    periodo_fin__month=self.mes,
                ).values_list('id', flat=True)
            )
            if csrt_rco_ids:
                # Hay datos curados: usar solo CSRT_RCO
                qs_base = qs_base.filter(importacion_id__in=csrt_rco_ids)
                logger.info(
                    'CargaS10 %d-%02d: usando CSRT_RCO (imp %s)',
                    self.anio, self.mes, csrt_rco_ids,
                )
            else:
                # Sin datos curados: usar todos los RELOJ del período
                logger.info(
                    'CargaS10 %d-%02d: sin CSRT_RCO, usando RELOJ',
                    self.anio, self.mes,
                )

        # Excluir registros post-cese y pre-ingreso
        from django.db.models import F as DbF
        qs_base = qs_base.exclude(
            personal__fecha_cese__isnull=False, fecha__gt=DbF('personal__fecha_cese')
        ).exclude(
            personal__fecha_alta__isnull=False, fecha__lt=DbF('personal__fecha_alta')
        )

        # HE: del ciclo HE (21 mes ant → 20 mes actual)
        qs_he = qs_base.filter(fecha__gte=inicio_he, fecha__lte=fin_he)

        # Asistencia: del mes completo
        qs_asist = qs_base.filter(fecha__gte=inicio_asist, fecha__lte=fin_asist)

        # Descuentos que NO van al mes actual (regularización)
        # = descuentos entre día 21 y fin de mes
        dia_regularizacion_inicio = date(self.anio, self.mes, self.config.dia_corte_planilla + 1)
        fin_mes = fin_asist

        # HE por persona
        he_por_persona: dict[int, dict] = {}
        for r in (qs_he.values('personal_id')
                  .annotate(sum_25=Sum('he_25'), sum_35=Sum('he_35'), sum_100=Sum('he_100'))):
            he_por_persona[r['personal_id']] = {
                'HE25': r['sum_25'] or CERO,
                'HE35': r['sum_35'] or CERO,
                'HE100': r['sum_100'] or CERO,
            }

        # Días especiales por persona (del mes de asistencia)
        # No regularizaciones (fecha < día corte+1)
        # Excluir FA en domingos para LOCAL/LIMA (son DS, no faltas)
        conceptos_dias: dict[int, dict[str, int]] = {}
        for r in qs_asist.filter(
            fecha__lt=dia_regularizacion_inicio
        ).values('personal_id', 'codigo_dia', 'dia_semana', 'condicion'):
            pid = r['personal_id']
            cod = r['codigo_dia']
            # Domingo + LOCAL/LIMA con FA → contar como DS, no FA
            if cod in ('FA', 'F') and r.get('dia_semana') == 6 and r.get('condicion', '').upper() in ('LOCAL', 'LIMA', ''):
                cod = 'DS'
            if pid not in conceptos_dias:
                conceptos_dias[pid] = {}
            conceptos_dias[pid][cod] = conceptos_dias[pid].get(cod, 0) + 1

        # Regularizaciones (descuentos del período posterior al corte → van al SIGUIENTE mes)
        regularizaciones: dict[int, dict[str, int]] = {}
        if self.config.regularizacion_activa:
            for r in qs_asist.filter(
                fecha__gte=dia_regularizacion_inicio,
                codigo_dia__in=['FA', 'LSG', 'SAI']
            ).values('personal_id', 'codigo_dia'):
                pid = r['personal_id']
                cod = r['codigo_dia']
                if pid not in regularizaciones:
                    regularizaciones[pid] = {}
                regularizaciones[pid][cod] = regularizaciones[pid].get(cod, 0) + 1

        # ── Personal involucrado ───────────────────────────────
        pids_todos = (set(he_por_persona.keys()) | set(conceptos_dias.keys()))
        personal_qs = (Personal.objects.filter(id__in=pids_todos)
                       .order_by('apellidos_nombres'))

        # ── Construir Excel ────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'CargaS10'

        # Estilo encabezados
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=9)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        reg_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

        # Fila 1: encabezados
        headers = ['Periodo', 'Código', 'DNI', 'Nombre']
        for _, nombre_col, _ in self.conceptos:
            headers.append(nombre_col)
        headers.append('REGULARIZACION_MES_SIGUIENTE')  # info de regularizaciones

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Filas de datos
        for row_idx, p in enumerate(personal_qs, 2):
            pid = p.pk
            he = he_por_persona.get(pid, {})
            dias = conceptos_dias.get(pid, {})
            reg_next = regularizaciones.get(pid, {})

            row_data = [
                self.periodo_label,
                p.codigo_sap or '',
                p.nro_doc,
                p.apellidos_nombres,
            ]

            for cod, _, tipo in self.conceptos:
                if cod == 'HE25':
                    val = float(he.get('HE25', CERO))
                elif cod == 'HE35':
                    val = float(he.get('HE35', CERO))
                elif cod == 'HE100':
                    val = float(he.get('HE100', CERO))
                elif cod == 'DM':
                    val = dias.get('DM', 0)
                elif cod == 'LF':
                    val = dias.get('LF', 0)
                elif cod == 'LSG':
                    val = dias.get('LSG', 0)
                elif cod == 'LP':
                    val = dias.get('LP', 0)
                elif cod == 'FA':
                    val = dias.get('FA', 0)
                elif cod == 'SAI':
                    val = dias.get('SAI', 0)
                elif cod == 'VAC':
                    val = dias.get('VAC', 0) + dias.get('V', 0)
                else:
                    val = 0

                row_data.append(val if val else 0)

            # Info regularización
            if reg_next:
                reg_str = '; '.join(f'{c}={n}' for c, n in reg_next.items())
                row_data.append(f'REGULARIZAR PRÓX MES: {reg_str}')
            else:
                row_data.append('')

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 3:  # DNI — forzar texto para preservar ceros iniciales
                    cell.number_format = '@'
                if reg_next and col_idx == len(row_data):
                    cell.fill = reg_fill
                    cell.font = Font(color='7F0000', bold=True, size=9)

        # Ajustar anchos
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.column_dimensions['A'].width = 22  # Periodo
        ws.column_dimensions['D'].width = 35  # Nombre

        # Segunda hoja: Regularizaciones del mes siguiente
        if any(regularizaciones.values()):
            ws2 = wb.create_sheet('Regularizaciones')
            ws2.append(['DNI', 'Nombre', 'Código', 'Días', 'Observación'])
            for p in personal_qs:
                if p.pk in regularizaciones:
                    for cod, dias_n in regularizaciones[p.pk].items():
                        ws2.append([
                            p.nro_doc,
                            p.apellidos_nombres,
                            cod,
                            dias_n,
                            f'Descuento diferido — aplica en planilla siguiente ({self.periodo_label})'
                        ])
                        ws2.cell(row=ws2.max_row, column=1).number_format = '@'  # DNI texto

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class ReporteCierreExporter:
    """
    Genera el reporte de cierre de mes en Excel con:
    - Resumen de asistencia por empleado
    - HE acumuladas (STAFF banco / RCO pago)
    - KPIs: tasa asistencia, faltas, tardanzas
    - Regularizaciones pendientes
    """

    def __init__(self, anio: int, mes: int, config=None, tipo_periodo: str = 'calendario'):
        from asistencia.models import ConfiguracionSistema
        self.anio = anio
        self.mes = mes
        self.config = config or ConfiguracionSistema.get()
        self.tipo_periodo = tipo_periodo if tipo_periodo in ('calendario', 'corte') else 'calendario'
        MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        self.mes_nombre = MESES[mes - 1]

    def _categorizar_codigo(self, cod: str, dia_semana: int | None = None,
                            condicion: str = '', datos_pid: dict | None = None):
        """
        Categoriza un código de día y lo suma al dict datos_pid.
        Retorna la categoría usada.
        """
        cod = (cod or '').upper()
        es_dom_local = dia_semana == 6 and condicion.upper() in ('LOCAL', 'LIMA', '')

        if cod == 'NA':
            cat = 'na'
        elif cod in ('T', 'NOR', 'TR', 'A', 'SS', 'CDT', 'CPF', 'LCG', 'ATM', 'CHE', 'LIM', 'CT', 'CAP'):
            cat = 'dias_trabajados'
        elif cod in ('FA', 'F'):
            # Domingo LOCAL/LIMA → DS, no falta
            cat = 'ds_feriado' if es_dom_local else 'faltas'
        elif cod in ('VAC', 'V'):
            cat = 'vacaciones'
        elif cod == 'DM':
            cat = 'dm'
        elif cod in ('DL', 'DLA', 'B'):
            cat = 'dl'
        elif cod == 'LSG':
            cat = 'lsg'
        elif cod == 'SAI':
            cat = 'sai'
        elif cod in ('LF', 'LP', 'LM'):
            cat = 'otros'
        elif cod in ('DS', 'FER', 'DOM'):
            cat = 'ds_feriado'
        else:
            cat = 'otros'

        if datos_pid is not None:
            datos_pid[cat] = datos_pid.get(cat, 0) + 1
        return cat

    def generar(self) -> BytesIO:
        from datetime import timedelta
        from personal.models import Personal
        from asistencia.models import BancoHoras, RegistroTareo, RegistroPapeleta

        if self.tipo_periodo == 'corte':
            inicio, fin = self.config.get_ciclo_he(self.anio, self.mes)
            label_periodo = f'Corte de Planilla: {inicio.strftime("%d/%m/%Y")} → {fin.strftime("%d/%m/%Y")}'
        else:
            inicio, fin = self.config.get_ciclo_asistencia(self.anio, self.mes)
            label_periodo = f'Mes Calendario: {inicio.strftime("%d/%m/%Y")} → {fin.strftime("%d/%m/%Y")}'
        total_dias = (fin - inicio).days + 1

        from django.db.models import F as DbF
        qs = (RegistroTareo.objects
              .filter(fecha__gte=inicio, fecha__lte=fin, personal__isnull=False)
              .exclude(personal__fecha_cese__isnull=False, fecha__gt=DbF('personal__fecha_cese'))
              .exclude(personal__fecha_alta__isnull=False, fecha__lt=DbF('personal__fecha_alta'))
              .values('personal_id', 'personal__apellidos_nombres',
                      'personal__nro_doc', 'personal__grupo_tareo',
                      'personal__codigo_sap', 'personal__condicion',
                      'codigo_dia', 'condicion', 'dia_semana', 'fecha')
              .annotate(
                  sum_he25=Sum('he_25'),
                  sum_he35=Sum('he_35'),
                  sum_he100=Sum('he_100'),
              ))

        # Recopilar info del personal y UN código por (pid, fecha)
        # Para evitar duplicados de múltiples importaciones
        info_personal: dict[int, dict] = {}  # pid → datos base del personal
        registro_por_fecha: dict[int, dict[date, str]] = {}  # pid → {fecha: codigo_dia}
        condicion_por_fecha: dict[int, dict[date, str]] = {}  # pid → {fecha: condicion}
        he_acum: dict[int, dict] = {}  # pid → {he25, he35, he100}

        for r in qs:
            pid = r['personal_id']
            if pid not in info_personal:
                info_personal[pid] = {
                    'nombre': r['personal__apellidos_nombres'],
                    'dni': r['personal__nro_doc'],
                    'grupo': r['personal__grupo_tareo'],
                    'sap': r['personal__codigo_sap'] or '',
                    'condicion_personal': (r['personal__condicion'] or '').upper(),
                }
                registro_por_fecha[pid] = {}
                condicion_por_fecha[pid] = {}
                he_acum[pid] = {'he25': CERO, 'he35': CERO, 'he100': CERO}

            fecha = r['fecha']
            cod = (r['codigo_dia'] or '').upper()
            cond_reg = r.get('condicion', '') or ''

            # Solo guardar un código por fecha (el primero que no sea vacío)
            if fecha not in registro_por_fecha[pid]:
                registro_por_fecha[pid][fecha] = cod
                condicion_por_fecha[pid][fecha] = cond_reg

            # HE siempre se acumulan (pueden venir de múltiples registros)
            he_acum[pid]['he25'] += (r['sum_he25'] or CERO)
            he_acum[pid]['he35'] += (r['sum_he35'] or CERO)
            he_acum[pid]['he100'] += (r['sum_he100'] or CERO)

        # ── Papeletas: cubrir días sin RegistroTareo ─────────────
        # Consultar papeletas activas (aprobadas/ejecutadas) que se solapan con el período
        papeletas_qs = (RegistroPapeleta.objects
                        .filter(
                            personal__isnull=False,
                            fecha_inicio__lte=fin,
                            fecha_fin__gte=inicio,
                            estado__in=('APROBADA', 'EJECUTADA'),
                        )
                        .select_related('personal')
                        .values('personal_id', 'personal__apellidos_nombres',
                                'personal__nro_doc', 'personal__grupo_tareo',
                                'personal__codigo_sap', 'personal__condicion',
                                'iniciales', 'tipo_permiso',
                                'fecha_inicio', 'fecha_fin'))

        # Mapeo tipo_permiso → código corto (fallback si iniciales está vacío)
        TIPO_A_CODIGO = {
            'DESCANSO_MEDICO': 'DM',
            'VACACIONES': 'VAC',
            'LICENCIA_SIN_GOCE': 'LSG',
            'LICENCIA_CON_GOCE': 'LCG',
            'LICENCIA_FALLECIMIENTO': 'LF',
            'LICENCIA_PATERNIDAD': 'LP',
            'LICENCIA_MATERNIDAD': 'LM',
            'BAJADAS': 'DL',
            'BAJADAS_ACUMULADAS': 'DLA',
            'COMPENSACION_HE': 'CHE',
            'COMPENSACION_FERIADO': 'CPF',
            'COMP_DIA_TRABAJO': 'CDT',
            'SUSPENSION': 'SUS',
            'SUSPENSION_ACTO_INSEGURO': 'SAI',
            'CAPACITACION': 'CAP',
            'TRABAJO_REMOTO': 'TR',
            'COMISION_TRABAJO': 'CT',
        }

        # Indexar papeletas por personal_id
        papeletas_por_pid: dict[int, list] = {}
        for pap in papeletas_qs:
            pid = pap['personal_id']
            if pid not in papeletas_por_pid:
                papeletas_por_pid[pid] = []
            papeletas_por_pid[pid].append(pap)

            # Si el personal no tiene datos de RegistroTareo, registrar info
            if pid not in info_personal:
                info_personal[pid] = {
                    'nombre': pap['personal__apellidos_nombres'],
                    'dni': pap['personal__nro_doc'],
                    'grupo': pap['personal__grupo_tareo'],
                    'sap': pap['personal__codigo_sap'] or '',
                    'condicion_personal': (pap['personal__condicion'] or '').upper(),
                }
                registro_por_fecha[pid] = {}
                condicion_por_fecha[pid] = {}
                he_acum[pid] = {'he25': CERO, 'he35': CERO, 'he100': CERO}

        # ── Fechas de vigencia del personal (alta/cese) ────────
        pids_todos = set(info_personal.keys()) | set(papeletas_por_pid.keys())
        vigencia: dict[int, tuple] = {}  # pid → (fecha_alta, fecha_cese)
        for p in Personal.objects.filter(id__in=pids_todos).values('id', 'fecha_alta', 'fecha_cese'):
            vigencia[p['id']] = (p['fecha_alta'], p['fecha_cese'])

        # ── Construir datos: un código por día por empleado ────
        todas_fechas = [inicio + timedelta(days=i) for i in range(total_dias)]
        datos: dict[int, dict] = {}

        for pid, info in info_personal.items():
            d = {
                **info,
                'dias_trabajados': 0, 'faltas': 0, 'lsg': 0, 'sai': 0,
                'vacaciones': 0, 'dm': 0, 'dl': 0,
                'ds_feriado': 0, 'na': 0, 'otros': 0,
                'he25': he_acum.get(pid, {}).get('he25', CERO),
                'he35': he_acum.get(pid, {}).get('he35', CERO),
                'he100': he_acum.get(pid, {}).get('he100', CERO),
            }
            datos[pid] = d

            paps = papeletas_por_pid.get(pid, [])
            cond = info['condicion_personal']
            f_alta, f_cese = vigencia.get(pid, (None, None))
            regs = registro_por_fecha.get(pid, {})
            conds = condicion_por_fecha.get(pid, {})

            for fecha in todas_fechas:
                # 1) Papeleta siempre tiene prioridad (override sobre RegistroTareo)
                cod_papeleta = None
                for pap in paps:
                    if pap['fecha_inicio'] <= fecha <= pap['fecha_fin']:
                        cod_papeleta = pap['iniciales'] or TIPO_A_CODIGO.get(pap['tipo_permiso'], '')
                        break

                if cod_papeleta:
                    self._categorizar_codigo(
                        cod_papeleta, fecha.weekday(), cond, d
                    )
                elif fecha in regs:
                    # 2) RegistroTareo — usar condicion del Personal (no del registro,
                    #    que puede estar mal por importaciones con condicion incorrecta)
                    self._categorizar_codigo(
                        regs[fecha], fecha.weekday(), cond, d
                    )
                elif (f_alta and fecha < f_alta) or (f_cese and fecha > f_cese):
                    # 3) Fuera de vigencia → NA
                    d['na'] += 1
                else:
                    # 4) Sin registro ni papeleta → aplicar misma lógica que Matricial
                    dia_semana = fecha.weekday()
                    if dia_semana == 6 and cond in ('LOCAL', 'LIMA', ''):
                        # Domingo LOCAL/LIMA → DS
                        d['ds_feriado'] += 1
                    elif cond == 'LIMA' and dia_semana < 6:
                        # LIMA lun-sáb sin registro → asistencia automática
                        d['dias_trabajados'] += 1
                    else:
                        # Resto → falta
                        d['faltas'] += 1

        # ── Generar Excel ────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Cierre {self.mes_nombre} {self.anio}'

        title_font = Font(bold=True, size=12, color='1F4E79')
        header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        header_font = Font(bold=True, size=9)

        tipo_label = 'CORTE' if self.tipo_periodo == 'corte' else 'MES CALENDARIO'
        ws.merge_cells('A1:S1')
        ws['A1'] = f'REPORTE DE CIERRE — {self.mes_nombre.upper()} {self.anio} | {tipo_label} | {label_periodo} | {total_dias} días'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Código SAP', 'DNI', 'Apellidos y Nombres', 'Grupo',
                   'Días Trabajados', 'Faltas', 'LSG', 'SAI', 'Vacaciones', 'DM', 'DL/Bajadas',
                   'DS/Feriado', 'NA', 'Otros', 'TOTAL',
                   'HE 25% (h)', 'HE 35% (h)', 'HE 100% (h)', '% Asistencia']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Días hábiles = total periodo menos descansos semanales/feriados menos NA
        for row_idx, d in enumerate(sorted(datos.values(), key=lambda x: x['nombre']), 3):
            dias_lab = total_dias - d['ds_feriado'] - d['na']
            pct_asist = (d['dias_trabajados'] / dias_lab * 100) if dias_lab else 0
            total_sum = (d['dias_trabajados'] + d['faltas'] + d['lsg'] + d['sai']
                         + d['vacaciones'] + d['dm'] + d['dl']
                         + d['ds_feriado'] + d['na'] + d['otros'])
            ws.append([
                d['sap'], d['dni'], d['nombre'], d['grupo'],
                d['dias_trabajados'], d['faltas'], d['lsg'], d['sai'], d['vacaciones'],
                d['dm'], d['dl'], d['ds_feriado'], d['na'], d['otros'], total_sum,
                float(d['he25']), float(d['he35']), float(d['he100']),
                round(pct_asist, 1),
            ])
            ws.cell(row=row_idx, column=2).number_format = '@'  # DNI texto

        for col_idx in range(1, 20):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.column_dimensions['C'].width = 35

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
