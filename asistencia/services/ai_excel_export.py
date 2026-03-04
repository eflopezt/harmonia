"""
Generador de Reportes Excel para Harmoni AI.

Genera reportes ejecutivos en formato .xlsx con datos reales de RRHH.
Usa openpyxl con estilo Harmoni (header dark teal, accent teal).
"""
from __future__ import annotations

import logging
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger('harmoni.ai')

# ── Estilos Harmoni ──
HEADER_FILL = PatternFill(start_color='0D2B27', end_color='0D2B27', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, color='0F766E', size=14)
SUBTITLE_FONT = Font(name='Calibri', bold=True, color='0F766E', size=11)
DATA_FONT = Font(name='Calibri', size=10)
NUMBER_FONT = Font(name='Calibri', size=10)
ACCENT_FILL = PatternFill(start_color='F0FDFA', end_color='F0FDFA', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


class ReporteGerenciaExporter:
    """
    Genera reporte ejecutivo Excel con 4 sheets:
    1. Resumen KPI — indicadores clave
    2. Headcount por Área — tabla área × STAFF/RCO/Total
    3. Tendencias — últimos 12 KPISnapshots
    4. Alertas Activas — alertas RRHH
    """

    def __init__(self, user):
        self.user = user
        self.hoy = date.today()

    def generate(self) -> openpyxl.Workbook:
        """Genera el workbook completo."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self._sheet_resumen_kpi(wb)
        self._sheet_headcount_area(wb)
        self._sheet_tendencias(wb)
        self._sheet_alertas(wb)

        return wb

    def _get_personal(self):
        """Obtiene queryset de personal activo filtrado por permisos."""
        from personal.permissions import filtrar_personal
        return filtrar_personal(self.user).filter(estado='Activo')

    def _apply_header_row(self, ws, row: int, headers: list[str]):
        """Aplica estilos de encabezado a una fila."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def _apply_data_cell(self, ws, row: int, col: int, value, is_number=False, accent=False):
        """Aplica estilo a una celda de datos."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = NUMBER_FONT if is_number else DATA_FONT
        cell.alignment = CENTER if is_number else LEFT
        cell.border = THIN_BORDER
        if accent:
            cell.fill = ACCENT_FILL
        return cell

    def _auto_width(self, ws, min_width=10, max_width=40):
        """Auto-ajusta ancho de columnas."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    # ─── Sheet 1: Resumen KPI ───

    def _sheet_resumen_kpi(self, wb):
        ws = wb.create_sheet('Resumen KPI')

        personal = self._get_personal()
        total = personal.count()
        staff = personal.filter(grupo_tareo='STAFF').count()
        rco = personal.filter(grupo_tareo='RCO').count()
        masculino = personal.filter(sexo='M').count()
        femenino = personal.filter(sexo='F').count()

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f'Reporte Ejecutivo RRHH — {self.hoy.strftime("%d/%m/%Y")}'
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells('A2:D2')
        ws['A2'].value = 'Harmoni — Sistema de Gestión de Recursos Humanos'
        ws['A2'].font = Font(name='Calibri', italic=True, color='6B7280', size=10)

        # KPI table
        row = 4
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = 'Indicadores Clave'
        ws[f'A{row}'].font = SUBTITLE_FONT

        row = 5
        self._apply_header_row(ws, row, ['Indicador', 'Valor', 'Detalle', 'Estado'])

        kpis = [
            ('Headcount Total', total, f'{staff} STAFF + {rco} RCO', '✓'),
            ('Personal STAFF', staff, f'{staff * 100 // max(total, 1)}% del total', ''),
            ('Personal RCO', rco, f'{rco * 100 // max(total, 1)}% del total', ''),
            ('Género Masculino', masculino, f'{masculino * 100 // max(total, 1)}%', ''),
            ('Género Femenino', femenino, f'{femenino * 100 // max(total, 1)}%', ''),
        ]

        # Contratos por vencer
        try:
            vencer_30 = personal.filter(
                fecha_fin_contrato__isnull=False,
                fecha_fin_contrato__lte=self.hoy + timedelta(days=30),
                fecha_fin_contrato__gte=self.hoy,
            ).count()
            kpis.append(('Contratos por vencer (30d)', vencer_30, 'Próximos 30 días',
                         '⚠' if vencer_30 > 0 else '✓'))
        except Exception:
            pass

        # Asistencia hoy
        try:
            from asistencia.models import RegistroTareo
            from django.db.models import Q, Count
            tareo_hoy = RegistroTareo.objects.filter(fecha=self.hoy, personal__in=personal)
            agg = tareo_hoy.aggregate(
                trabajando=Count('id', filter=Q(codigo_dia__in=['T', 'NOR', 'TR'])),
                faltas=Count('id', filter=Q(codigo_dia='FA')),
            )
            kpis.append(('Trabajando hoy', agg['trabajando'] or 0, '', ''))
            kpis.append(('Faltas hoy', agg['faltas'] or 0, '',
                         '⚠' if (agg['faltas'] or 0) > 5 else '✓'))
        except Exception:
            pass

        # Vacaciones pendientes
        try:
            from vacaciones.models import SolicitudVacacion
            vac_pend = SolicitudVacacion.objects.filter(
                estado='PENDIENTE', personal__in=personal).count()
            kpis.append(('Vacaciones pendientes', vac_pend, 'Solicitudes',
                         '⚠' if vac_pend > 5 else '✓'))
        except Exception:
            pass

        # Rotación
        try:
            from analytics.models import KPISnapshot
            last_snap = KPISnapshot.objects.order_by('-periodo').first()
            if last_snap:
                rot = float(last_snap.tasa_rotacion)
                kpis.append(('Rotación mensual', f'{rot:.1f}%',
                             last_snap.periodo.strftime('%b %Y'),
                             '⚠' if rot > 5 else '✓'))
        except Exception:
            pass

        # Capacitaciones
        try:
            from capacitaciones.models import Capacitacion
            cap_curso = Capacitacion.objects.filter(estado='EN_CURSO').count()
            cap_prog = Capacitacion.objects.filter(estado='PROGRAMADA').count()
            kpis.append(('Capacitaciones en curso', cap_curso, f'{cap_prog} programadas', ''))
        except Exception:
            pass

        # Evaluaciones
        try:
            from evaluaciones.models import CicloEvaluacion
            ciclo_activo = CicloEvaluacion.objects.filter(estado='ACTIVO').count()
            kpis.append(('Ciclos evaluación activos', ciclo_activo, '', ''))
        except Exception:
            pass

        for i, (indicador, valor, detalle, estado) in enumerate(kpis):
            r = row + 1 + i
            accent = i % 2 == 0
            self._apply_data_cell(ws, r, 1, indicador, accent=accent)
            self._apply_data_cell(ws, r, 2, valor, is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 3, detalle, accent=accent)
            self._apply_data_cell(ws, r, 4, estado, accent=accent)

        self._auto_width(ws)

    # ─── Sheet 2: Headcount por Área ───

    def _sheet_headcount_area(self, wb):
        ws = wb.create_sheet('Headcount por Área')

        personal = self._get_personal()

        from django.db.models import Count, Q

        areas_data = (
            personal
            .values('subarea__area__nombre')
            .annotate(
                total=Count('id'),
                staff=Count('id', filter=Q(grupo_tareo='STAFF')),
                rco=Count('id', filter=Q(grupo_tareo='RCO')),
                masc=Count('id', filter=Q(sexo='M')),
                fem=Count('id', filter=Q(sexo='F')),
            )
            .order_by('-total')
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'].value = 'Headcount por Área'
        ws['A1'].font = SUBTITLE_FONT

        row = 3
        headers = ['Área', 'STAFF', 'RCO', 'Total', 'Masculino', 'Femenino']
        self._apply_header_row(ws, row, headers)

        grand_total = [0, 0, 0, 0, 0]
        for i, area in enumerate(areas_data):
            r = row + 1 + i
            nombre = area['subarea__area__nombre'] or 'Sin Área'
            accent = i % 2 == 0
            self._apply_data_cell(ws, r, 1, nombre, accent=accent)
            self._apply_data_cell(ws, r, 2, area['staff'], is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 3, area['rco'], is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 4, area['total'], is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 5, area['masc'], is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 6, area['fem'], is_number=True, accent=accent)

            grand_total[0] += area['staff']
            grand_total[1] += area['rco']
            grand_total[2] += area['total']
            grand_total[3] += area['masc']
            grand_total[4] += area['fem']

        # Total row
        r = row + 1 + len(areas_data)
        for col_idx, val in enumerate([
            'TOTAL', grand_total[0], grand_total[1],
            grand_total[2], grand_total[3], grand_total[4],
        ], 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(name='Calibri', bold=True, size=10)
            cell.fill = PatternFill(start_color='134E4A', end_color='134E4A', fill_type='solid')
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        self._auto_width(ws)

    # ─── Sheet 3: Tendencias ───

    def _sheet_tendencias(self, wb):
        ws = wb.create_sheet('Tendencias')

        try:
            from analytics.models import KPISnapshot
            snaps = list(KPISnapshot.objects.order_by('-periodo')[:12][::-1])
        except Exception:
            snaps = []

        ws.merge_cells('A1:G1')
        ws['A1'].value = 'Tendencias Mensuales (Últimos 12 meses)'
        ws['A1'].font = SUBTITLE_FONT

        if not snaps:
            ws['A3'].value = 'No hay datos de KPI disponibles.'
            ws['A3'].font = DATA_FONT
            return

        row = 3
        headers = ['Período', 'Headcount', 'Rotación %', 'Asistencia %',
                    'Horas Extra', 'Altas', 'Bajas']
        self._apply_header_row(ws, row, headers)

        for i, snap in enumerate(snaps):
            r = row + 1 + i
            accent = i % 2 == 0
            self._apply_data_cell(ws, r, 1, snap.periodo.strftime('%b %Y'), accent=accent)
            self._apply_data_cell(ws, r, 2, snap.total_empleados, is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 3, f'{float(snap.tasa_rotacion):.1f}', is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 4, f'{float(snap.tasa_asistencia):.1f}', is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 5, f'{float(snap.total_horas_extra):.0f}', is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 6, snap.altas_mes, is_number=True, accent=accent)
            self._apply_data_cell(ws, r, 7, snap.bajas_mes, is_number=True, accent=accent)

        self._auto_width(ws)

    # ─── Sheet 4: Alertas Activas ───

    def _sheet_alertas(self, wb):
        ws = wb.create_sheet('Alertas Activas')

        try:
            from analytics.models import AlertaRRHH
            alertas = list(
                AlertaRRHH.objects.filter(activa=True)
                .order_by('-severidad', '-fecha_generada')[:50]
            )
        except Exception:
            alertas = []

        ws.merge_cells('A1:E1')
        ws['A1'].value = 'Alertas RRHH Activas'
        ws['A1'].font = SUBTITLE_FONT

        if not alertas:
            ws['A3'].value = 'No hay alertas activas. ✓'
            ws['A3'].font = DATA_FONT
            return

        row = 3
        headers = ['Severidad', 'Categoría', 'Título', 'Descripción', 'Fecha']
        self._apply_header_row(ws, row, headers)

        sev_colors = {
            'CRITICA': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
            'ALTA': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
            'MEDIA': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
            'BAJA': PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid'),
        }

        for i, alerta in enumerate(alertas):
            r = row + 1 + i
            fill = sev_colors.get(alerta.severidad, ACCENT_FILL)

            cell = ws.cell(row=r, column=1, value=alerta.severidad)
            cell.font = Font(name='Calibri', bold=True, size=10)
            cell.fill = fill
            cell.alignment = CENTER
            cell.border = THIN_BORDER

            self._apply_data_cell(ws, r, 2, alerta.categoria)
            self._apply_data_cell(ws, r, 3, alerta.titulo)
            self._apply_data_cell(ws, r, 4, str(alerta.descripcion or '')[:100])
            self._apply_data_cell(ws, r, 5, alerta.fecha_generada.strftime('%d/%m/%Y') if alerta.fecha_generada else '')

        self._auto_width(ws)
