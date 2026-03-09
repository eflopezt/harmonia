"""
Vistas para el asistente IA de Harmoni.

Endpoints:
  POST /asistencia/ia/chat/     → Chat streaming (SSE) — con fallback sin IA
  GET  /asistencia/ia/status/   → Estado de conexión IA
  GET  /asistencia/ia/context/  → Datos del sistema (JSON)
  GET  /asistencia/ia/insights/ → Insights generados por IA
  POST /asistencia/ia/analizar/ → Análisis de gráfico (SSE)
  POST /asistencia/ia/preguntar/→ Pregunta libre (SSE)
"""
from __future__ import annotations

import json
import logging
import time
import uuid

from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.views.decorators.http import require_GET, require_POST

from .services.ai_service import get_service
from .services.ai_context import (
    build_system_prompt,
    build_system_prompt_data,
    build_insights_prompt,
    detect_chart_request,
    detect_dashboard_request,
    detect_edit_request,
    detect_export_request,
    detect_individual_query,
    detect_module_context,
    detect_multiple_chart_requests,
    detect_pin_to_dashboard,
    generate_chart_data,
    generate_dashboard_data,
    get_individual_ranking,
    responder_sin_ia,
)

logger = logging.getLogger('harmoni.ai')


def _sse_text(text: str) -> str:
    """
    Formatea texto multilinea para SSE.
    En SSE cada línea del mensaje debe tener su propio 'data:' prefix.
    Newlines dentro del texto rompen el mensaje si no se manejan.
    Solución: reemplazar \\n por un marcador que el JS puede reconvertir.
    """
    # Reemplazar newlines reales por \\n literal para que SSE no se rompa
    return text.replace('\n', '\\n')

# Cache de conectividad IA (evita test_connection en cada request)
_ia_cache: dict = {'ok': None, 'ts': 0}
_IA_CACHE_TTL = 30  # segundos


def _is_ollama_reachable() -> bool:
    """
    Verifica conectividad con el proveedor IA configurado, con cache de 30s.
    Nombre mantenido por compatibilidad; funciona con cualquier provider.
    """
    now = time.time()
    if now - _ia_cache['ts'] < _IA_CACHE_TTL and _ia_cache['ok'] is not None:
        return _ia_cache['ok']

    svc = get_service()
    if svc is None:
        _ia_cache.update(ok=False, ts=now)
        return False

    try:
        result = svc.test_connection()
        # Solo chequeamos 'ok' — 'modelo_activo' es exclusivo de OllamaService
        ok = result.get('ok', False)
    except Exception:
        ok = False

    _ia_cache.update(ok=ok, ts=now)
    return ok


# ─────────────────────────────────────────────────
# ESTADO DE IA
# ─────────────────────────────────────────────────

@login_required
@require_GET
def ai_status(request):
    """
    Verifica si la IA está disponible.
    El frontend usa esto para mostrar/ocultar el widget de chat.
    Usa cache de conectividad para evitar test_connection en cada page load.
    """
    ollama_ok = _is_ollama_reachable()
    svc = get_service()

    if svc is None or not ollama_ok:
        return JsonResponse({
            'available': True,
            'provider': 'FALLBACK',
            'model': None,
            'fallback': True,
        })

    return JsonResponse({
        'available': True,
        'provider': getattr(svc, 'provider_name', 'IA'),
        'model': getattr(svc, 'modelo', None),
        'fallback': False,
    })


# ─────────────────────────────────────────────────
# CONTEXTO DEL SISTEMA (datos sin IA)
# ─────────────────────────────────────────────────

@login_required
@require_GET
def ai_context(request):
    """
    Retorna datos factuales del sistema como JSON.
    No requiere Ollama — responde con datos de la BD directamente.
    """
    data = build_system_prompt_data(request.user)
    return JsonResponse(data)


# ─────────────────────────────────────────────────
# HELPER: Edición de PDF con IA + PyMuPDF
# ─────────────────────────────────────────────────

def _apply_pdf_edit(message: str, file_context: dict, svc) -> str | None:
    """
    Usa IA para determinar los cambios de texto en el PDF y los aplica con PyMuPDF.
    Guarda el resultado en Django cache y retorna el edit_id (o None si falló).

    Flujo:
      1. Obtiene bytes del PDF original desde cache (usando file_id).
      2. Llama a la IA con prompt especializado para obtener replacements JSON.
      3. Aplica los reemplazos vía PyMuPDF (redacción + re-inserción).
      4. Guarda PDF editado en cache con nuevo edit_id (TTL 15 min).
    """
    import re as _re

    file_id = file_context.get('file_id', '')
    if not file_id:
        logger.warning('_apply_pdf_edit: no file_id en file_context')
        return None

    original_bytes = cache.get(f'ai_pdf_{file_id}')
    if not original_bytes:
        logger.warning(f'_apply_pdf_edit: PDF bytes expirados para file_id={file_id}')
        return None

    extracted_text = file_context.get('content', '')

    # ── 1. Pedir a la IA los reemplazos como JSON ──────────────────────────
    replacements = []
    if svc:
        replacement_system = (
            'Eres un asistente que genera JSON de reemplazos de texto. '
            'Siempre respondes ÚNICAMENTE con un JSON array válido, sin texto adicional. '
            'Ejemplo: [{"old": "TEXTO_ORIGINAL", "new": "TEXTO_NUEVO"}]'
        )
        replacement_prompt = (
            f'El documento PDF contiene este texto:\n{extracted_text[:3000]}\n\n'
            f'El usuario pide: "{message}"\n\n'
            f'Genera un JSON array con los reemplazos exactos necesarios. '
            f'Usa el texto EXACTO tal como aparece en el documento. '
            f'Responde SOLO el JSON array:'
        )
        try:
            ai_response = svc.generate(replacement_prompt, system=replacement_system) or ''
            json_match = _re.search(r'\[.*?\]', ai_response, _re.DOTALL)
            if json_match:
                replacements = json.loads(json_match.group())
            else:
                logger.warning(f'_apply_pdf_edit: sin JSON en respuesta IA: {ai_response[:200]}')
        except Exception as exc:
            logger.warning(f'_apply_pdf_edit IA call failed: {exc}')

    # Fallback: parsear reemplazos del mensaje si la IA no respondió
    if not replacements:
        patterns = [
            # "texto X" por "texto Y"  (con comillas dobles)
            _re.compile(r'"([^"]+)"\s+(?:por|con|a)\s+"([^"]+)"', _re.IGNORECASE),
            # 'texto X' por 'texto Y'  (con comillas simples)
            _re.compile(r"'([^']+)'\s+(?:por|con|a)\s+'([^']+)'", _re.IGNORECASE),
            # cambia/reemplaza EL? DNI/RUC/NÚMERO X por Y  (sin comillas, para números)
            _re.compile(
                r'(?:cambia|reemplaza|pon|coloca|actualiza)\s+(?:el\s+)?'
                r'(?:dni|ruc|cod|código|numero|número|nro\.?|n°)?\s*'
                r'([A-Z0-9][A-Z0-9 ]{0,30}?)\s+(?:por|con|a)\s+([A-Z0-9][A-Z0-9 ]{0,30})',
                _re.IGNORECASE,
            ),
            # Detectar pares de números tipo DNI/RUC que aparecen como "NNNNNNNN por NNNNNNNN"
            _re.compile(r'\b(\d{8,11})\b.*?\bpor\b.*?\b(\d{8,11})\b', _re.IGNORECASE),
        ]
        for pat in patterns:
            for m in pat.finditer(message):
                old_val = m.group(1).strip()
                new_val = m.group(2).strip()
                if old_val and new_val and old_val != new_val:
                    replacements.append({'old': old_val, 'new': new_val})
        # Deduplicar
        seen = set()
        unique_reps = []
        for r in replacements:
            key = (r['old'], r['new'])
            if key not in seen:
                seen.add(key)
                unique_reps.append(r)
        replacements = unique_reps

    if not replacements:
        logger.warning('_apply_pdf_edit: no se detectaron reemplazos')
        return None

    # ── 2. Aplicar reemplazos con PyMuPDF ─────────────────────────────────
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning('_apply_pdf_edit: PyMuPDF no instalado')
        return None

    try:
        doc = fitz.open(stream=original_bytes, filetype='pdf')
        changes_made = 0

        for rep in replacements:
            old_text = str(rep.get('old', '')).strip()
            new_text = str(rep.get('new', '')).strip()
            if not old_text or old_text == new_text:
                continue

            for page in doc:
                instances = page.search_for(old_text)
                if not instances:
                    continue

                # Estimar font size mirando el bloque de texto más cercano
                font_size = 9.0
                try:
                    for block in page.get_text('dict')['blocks']:
                        for line in block.get('lines', []):
                            for span in line.get('spans', []):
                                if old_text in span.get('text', ''):
                                    font_size = span.get('size', 9.0)
                                    raise StopIteration
                except StopIteration:
                    pass

                # Guardar posiciones ANTES de redactar (apply_redactions puede mover cosas)
                rects_copy = [fitz.Rect(r) for r in instances]

                # Redactar (rellenar con blanco)
                for rect in rects_copy:
                    page.add_redact_annot(rect, fill=(1, 1, 1))
                page.apply_redactions()

                # Insertar texto nuevo en misma posición
                for rect in rects_copy:
                    page.insert_text(
                        (rect.x0, rect.y1 - 1),
                        new_text,
                        fontsize=font_size,
                        color=(0.0, 0.0, 0.0),
                    )
                    changes_made += 1

        if changes_made == 0:
            logger.warning('_apply_pdf_edit: ningún texto encontrado para reemplazar')
            doc.close()
            return None

        edited_bytes = doc.tobytes()
        doc.close()

        # Guardar en cache con nuevo ID
        edit_id = uuid.uuid4().hex[:16]
        cache.set(f'ai_edit_{edit_id}', edited_bytes, 900)  # 15 min TTL
        logger.info(f'_apply_pdf_edit: {changes_made} cambios aplicados → edit_id={edit_id}')
        return edit_id

    except Exception as exc:
        logger.warning(f'_apply_pdf_edit PyMuPDF error: {exc}')
        return None


# ─────────────────────────────────────────────────
# CHAT STREAMING (SSE) — con fallback
# ─────────────────────────────────────────────────

@login_required
@require_POST
def ai_chat_stream(request):
    """
    Chat con streaming via Server-Sent Events.
    Body JSON: {"message": "...", "history": [...]}

    Si Ollama no está disponible, intenta responder directamente
    con datos de la BD (fallback sin IA).
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    message = body.get('message', '').strip()
    file_context = body.get('file_context')  # {'type': 'pdf|excel|image|text', 'name': ..., 'content': ...}
    if not message and not file_context:
        return JsonResponse({'error': 'Mensaje vacío'}, status=400)
    if not message:
        message = f'Analiza este archivo y dame un resumen ejecutivo con los puntos clave.'

    history = body.get('history', [])

    svc = get_service()
    ollama_ok = _is_ollama_reachable()

    # ── Detectar edición de PDF (antes que todo) ──
    if detect_edit_request(message, file_context):
        edit_id = _apply_pdf_edit(message, file_context, svc)
        if edit_id:
            def pdf_edit_stream():
                yield 'data: [FALLBACK]\n\n'
                resp_text = (
                    '✏️ He aplicado los cambios al documento. '
                    'Haz clic para descargar el PDF editado:\\n\\n'
                    f'[DOWNLOAD:pdf_edit:{edit_id}]'
                )
                yield f'data: {resp_text}\n\n'
                yield 'data: [DONE]\n\n'

            resp = StreamingHttpResponse(pdf_edit_stream(), content_type='text/event-stream')
            resp['Cache-Control'] = 'no-cache'
            resp['X-Accel-Buffering'] = 'no'
            return resp
        # Si falló la edición, continúa con flujo normal (IA explicará por qué)

    # ── Detectar export request (antes de todo) ──
    export_req = detect_export_request(message)
    if export_req:
        def export_stream():
            yield 'data: [FALLBACK]\n\n'
            resp_text = (
                '📊 He preparado tu **Reporte Ejecutivo** de RRHH. '
                'Haz clic para descargarlo:\\n\\n'
                f'[DOWNLOAD:{export_req["type"]}]'
            )
            yield f'data: {resp_text}\n\n'
            yield 'data: [DONE]\n\n'

        response = StreamingHttpResponse(
            export_stream(), content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response

    # ── Detectar dashboard multi-gráfico ──
    dashboard_req = detect_dashboard_request(message)
    if dashboard_req:
        dashboard = generate_dashboard_data(request.user)
        if dashboard and dashboard.get('charts'):
            def dashboard_stream():
                yield 'data: [FALLBACK]\n\n'
                yield 'data: [MAXIMIZE]\n\n'
                for chart in dashboard['charts']:
                    chart_json = json.dumps(chart, ensure_ascii=False)
                    yield f'data: [CHART]{chart_json}[/CHART]\n\n'
                # Summary text
                summary = (
                    f'📊 **{dashboard["title"]}** — '
                    f'{len(dashboard["charts"])} indicadores clave. '
                    f'Usa el botón de maximizar para ver los gráficos en detalle.'
                )
                yield f'data: {_sse_text(summary)}\n\n'
                yield 'data: [DONE]\n\n'

            response = StreamingHttpResponse(
                dashboard_stream(), content_type='text/event-stream')
            response['Cache-Control'] = 'no-cache'
            response['X-Accel-Buffering'] = 'no'
            return response

    # ── Detectar queries individuales (antes del fallback) ──
    indiv_type = detect_individual_query(message)
    indiv_data = None
    if indiv_type:
        indiv_data = get_individual_ranking(indiv_type, request.user)

    # ── Fallback sin IA si Ollama no está disponible ──
    if not ollama_ok:
        # Charts también funcionan sin Ollama (datos vienen de BD)
        chart_reqs = detect_multiple_chart_requests(message, history)
        chart_datas = []
        if chart_reqs:
            for cr in chart_reqs:
                cd = generate_chart_data(cr['type'], request.user, message)
                if cd:
                    chart_datas.append(cd)

        fallback = responder_sin_ia(message, request.user)

        def fallback_stream():
            yield 'data: [FALLBACK]\n\n'
            if chart_datas:
                for chart_data in chart_datas:
                    chart_json = json.dumps(chart_data, ensure_ascii=False)
                    yield f'data: [CHART]{chart_json}[/CHART]\n\n'
                # Breve análisis estático del primer chart
                first = chart_datas[0]
                total = sum(first.get('values', []))
                if total:
                    top = max(
                        zip(first.get('labels', []),
                            first.get('values', [])),
                        key=lambda x: x[1],
                    )
                    analysis = (
                        f'**{first.get("title", "Gráfico")}** — '
                        f'Total: {total}. '
                        f'Mayor: {top[0]} con {top[1]} '
                        f'({top[1] * 100 // total}%).'
                    )
                    if len(chart_datas) > 1:
                        analysis += f' (+{len(chart_datas) - 1} gráfico(s) adicional(es) mostrado(s))'
                    yield f'data: {_sse_text(analysis)}\n\n'
                if wants_pin:
                    pin_payload = json.dumps({
                        'titulo': first.get('title', 'Gráfico'),
                        'chart_type': first.get('chart', 'bar'),
                        'data_source': chart_reqs[0].get('type', 'custom') if chart_reqs else 'custom',
                        'config': first,
                    }, ensure_ascii=False)
                    yield f'data: [PIN_WIDGET]{pin_payload}[/PIN_WIDGET]\n\n'
            elif indiv_data is not None:
                if indiv_data:
                    ranking_text = '\n'.join(
                        f'{i + 1}. **{r["nombre"]}** ({r["area"]}): {r["valor"]} {r["unidad"]}'
                        for i, r in enumerate(indiv_data)
                    )
                    yield f'data: {_sse_text(ranking_text)}\n\n'
                else:
                    yield f'data: {_sse_text("No hay registros de " + indiv_type.replace("_", " ") + " este mes en la BD.")}\n\n'
            elif fallback:
                yield f'data: {_sse_text(fallback)}\n\n'
            else:
                no_match = (
                    'No tengo suficiente informacion para responder '
                    'esa consulta directamente. Prueba con preguntas como:\n'
                    '- "¿Cuántos empleados activos hay?"\n'
                    '- "¿Hay aprobaciones pendientes?"\n'
                    '- "¿Cómo va la asistencia hoy?"\n'
                    '- "Muéstrame un gráfico del personal por área"\n\n'
                    'Para consultas más complejas, configura un proveedor IA en '
                    '**Asistencia > Configuración > Pestaña IA**.'
                )
                yield f'data: {_sse_text(no_match)}\n\n'
            yield 'data: [DONE]\n\n'

        response = StreamingHttpResponse(
            fallback_stream(), content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response

    # ── Detectar si quiere fijar un gráfico en el dashboard ──
    wants_pin = detect_pin_to_dashboard(message)

    # ── Detectar si pide uno o más gráficos ──
    chart_reqs = detect_multiple_chart_requests(message, history)
    chart_datas = []
    if chart_reqs:
        for cr in chart_reqs:
            cd = generate_chart_data(cr['type'], request.user, message)
            if cd:
                chart_datas.append(cd)

    # ── Lazy loading: detectar módulos relevantes ──
    modules = detect_module_context(message)
    system_prompt = build_system_prompt(request.user, modules=modules)

    # ── RAG: inyectar conocimiento base relevante ──
    try:
        from core.knowledge_service import get_knowledge_context
        knowledge_ctx = get_knowledge_context(message, limit=4)
        if knowledge_ctx:
            system_prompt = system_prompt + '\n\n' + knowledge_ctx
    except Exception:
        pass  # knowledge_service es opcional, no bloquea el chat

    # Limitar historial a los últimos 10 turnos
    messages = []
    for msg in history[-10:]:
        role = msg.get('role', 'user')
        content = msg.get('content', '')
        if role in ('user', 'assistant') and content:
            messages.append({'role': role, 'content': content})

    # Construir mensaje enriquecido para el LLM
    ai_message = message

    # ── Archivo adjunto: inyectar contenido en el prompt ──
    if file_context:
        fc_type = file_context.get('type', '')
        fc_name = file_context.get('name', 'archivo')
        fc_content = file_context.get('content', '')

        if fc_type == 'image':
            # Imágenes: mensaje multimodal [texto + imagen] para Gemini vision
            # OpenAI/DeepSeek (text-only) ignoran la parte de imagen y solo ven el texto
            mime = file_context.get('mime', 'image/jpeg')
            # fc_content es "data:mime;base64,XXXX" — extraer solo el base64
            b64_data = fc_content.split(',', 1)[1] if fc_content and ',' in fc_content else (fc_content or '')
            default_prompt = (
                f'Analiza la imagen "{fc_name}". Si es un documento (DNI, pasaporte, contrato, '
                f'boleta, certificado, licencia), extrae TODOS los datos en formato estructurado. '
                f'Si es otra imagen, describe qué contiene y qué información relevante puedes identificar. '
                f'Pregunta del usuario: {message}'
            )
            ai_message = [
                {'type': 'text',  'text': default_prompt},
                {'type': 'image', 'mime_type': mime, 'data': b64_data},
            ]
        elif fc_content:
            trunc_note = ' [contenido truncado a 12000 chars]' if file_context.get('truncated') else ''
            ai_message = (
                f'[ARCHIVO ADJUNTO: {fc_name}{trunc_note}]\n'
                f'{fc_content}\n'
                f'[FIN ARCHIVO]\n\n'
                f'Pregunta del usuario: {message}\n\n'
                f'INSTRUCCION ADICIONAL: Si extraes datos de un trabajador (DNI, nombre), '
                f'al final de tu respuesta agrega una línea: '
                f'"¿Quieres que busque a este trabajador en el sistema para ver su expediente completo?"'
            )

    if indiv_data is not None:
        # Datos individuales exactos de BD — inyectar en el prompt
        if indiv_data:
            ranking_text = '\n'.join(
                f'{i + 1}. {r["nombre"]} ({r["area"]}): {r["valor"]} {r["unidad"]}'
                for i, r in enumerate(indiv_data)
            )
            ai_message = (
                f'{message}\n\n'
                f'[DATOS EXACTOS DE LA BD para responder esta pregunta]\n'
                f'{ranking_text}\n'
                f'[FIN DATOS]\n\n'
                f'Usa estos datos exactos para responder. Menciona los nombres completos.'
            )
        else:
            ai_message = (
                f'{message}\n\n'
                f'[NOTA: No hay registros de {indiv_type.replace("_", " ")} este mes en la BD]'
            )
    elif chart_datas:
        # Si hay charts, dar contexto al LLM para análisis
        summaries = []
        for cd in chart_datas:
            total = sum(cd.get('values', []))
            if total > 0:
                parts = [
                    f'{l}: {v} ({v * 100 // total}%)'
                    for l, v in zip(cd.get('labels', []), cd.get('values', []))
                ]
                summaries.append(f'{cd.get("title", "Gráfico")}: {", ".join(parts)} (Total: {total})')
            else:
                summaries.append(f'{cd.get("title", "Gráfico")}: Sin datos')
        ai_message = (
            f'El usuario pidio graficos y YA se los estoy mostrando visualmente. '
            f'Tu solo necesitas dar un breve analisis en texto (3-4 lineas). '
            f'Los datos reales son:\n' + '\n'.join(summaries) +
            '\nAnaliza brevemente que significan estos numeros para RRHH.'
        )

    messages.append({'role': 'user', 'content': ai_message})

    def event_stream():
        try:
            # Si hay charts, enviar marcadores PRIMERO
            if chart_datas:
                for chart_data in chart_datas:
                    chart_json = json.dumps(chart_data, ensure_ascii=False)
                    yield f'data: [CHART]{chart_json}[/CHART]\n\n'

            for chunk in svc.chat_stream(
                messages=messages,
                system=system_prompt,
                temperature=0.35,
                num_predict=900,
            ):
                yield f'data: {chunk}\n\n'

            # Si quiere fijar, enviar marker al final
            if wants_pin and chart_datas:
                pin_payload = json.dumps({
                    'titulo': chart_datas[0].get('title', 'Gráfico'),
                    'chart_type': chart_datas[0].get('chart', 'bar'),
                    'data_source': chart_reqs[0].get('type', 'custom') if chart_reqs else 'custom',
                    'config': chart_datas[0],
                }, ensure_ascii=False)
                yield f'data: [PIN_WIDGET]{pin_payload}[/PIN_WIDGET]\n\n'

            yield 'data: [DONE]\n\n'
        except Exception as e:
            logger.warning(f'ai_chat_stream error: {e}')
            # Intentar fallback en caso de error de streaming
            fallback = responder_sin_ia(message, request.user)
            if fallback:
                yield 'data: [FALLBACK]\n\n'
                yield f'data: {_sse_text(fallback)}\n\n'
            else:
                yield 'data: _Error de conexión con IA._\n\n'
            yield 'data: [DONE]\n\n'

    response = StreamingHttpResponse(
        event_stream(),
        content_type='text/event-stream',
    )
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


# ─────────────────────────────────────────────────
# INSIGHTS IA (para dashboard home)
# ─────────────────────────────────────────────────

@login_required
@require_GET
def ai_insights(request):
    """
    Genera insights IA basados en KPIs actuales.
    Cachea resultado por 6 horas para performance.
    Si Ollama no está disponible, genera insights estáticos básicos.
    """
    from django.core.cache import cache
    from datetime import date

    cache_key = f'harmoni_ai_insights_{request.user.id}_{date.today().isoformat()}'
    force = request.GET.get('force', '') == '1'

    if not force:
        cached = cache.get(cache_key)
        if cached:
            return JsonResponse({'insights': cached, 'cached': True})

    ollama_ok = _is_ollama_reachable()

    if not ollama_ok:
        # Insights estáticos básicos sin IA
        data = build_system_prompt_data(request.user)
        insights = _build_static_insights(data)
        if insights:
            cache.set(cache_key, insights, timeout=3600)  # 1h para estáticos
            return JsonResponse({
                'insights': insights, 'cached': False, 'fallback': True,
            })
        return JsonResponse({'insights': None, 'error': 'IA no disponible'})

    svc = get_service()
    if svc is None:
        return JsonResponse({'insights': None, 'error': 'IA no disponible'})

    system, user_prompt = build_insights_prompt(request.user)

    try:
        result = svc.generate(
            prompt=user_prompt,
            system=system,
        )
        if result:
            cache.set(cache_key, result, timeout=3600 * 6)
            return JsonResponse({'insights': result, 'cached': False})
        else:
            return JsonResponse({'insights': None, 'error': 'Sin respuesta de IA'})
    except Exception as e:
        logger.warning(f'ai_insights error: {e}')
        return JsonResponse({'insights': None, 'error': str(e)})


def _build_static_insights(data: dict) -> str | None:
    """Genera insights básicos sin IA, basados en reglas simples."""
    lines = []

    total = data.get('total_personal', 0)
    if total:
        lines.append(f'- Plantilla actual: {total} empleados activos '
                      f'({data.get("total_staff", 0)} STAFF, '
                      f'{data.get("total_rco", 0)} RCO)')

    pend = data.get('total_pendientes', 0)
    if pend:
        lines.append(f'- ⚠ Hay {pend} aprobaciones pendientes que requieren atención')

    vencer = data.get('contratos_por_vencer', 0)
    if vencer:
        lines.append(f'- ⚠ {vencer} contratos vencen en los próximos 30 días')

    faltas = data.get('asistencia_faltas', 0)
    if faltas:
        lines.append(f'- Hoy: {faltas} faltas registradas')

    vac_pend = data.get('vacaciones_pendientes', 0)
    if vac_pend:
        lines.append(f'- {vac_pend} solicitudes de vacaciones pendientes')

    cert_venc = data.get('certificaciones_vencidas', 0)
    if cert_venc:
        lines.append(f'- ⚠ {cert_venc} certificaciones vencidas')

    disc = data.get('disciplinaria_en_descargo', 0)
    if disc:
        lines.append(f'- {disc} procesos disciplinarios en descargo')

    if data.get('kpi_rotacion') is not None:
        rot = data['kpi_rotacion']
        if rot > 5:
            lines.append(f'- ⚠ Rotación {rot:.1f}% — por encima del benchmark')
        else:
            lines.append(f'- ✓ Rotación {rot:.1f}% — dentro de rangos saludables')

    if not lines:
        return None

    return '\n'.join(lines)


def _static_chart_analysis(chart_type: str, chart_data: dict) -> str:
    """Genera análisis estático básico de un gráfico sin usar IA."""
    data = chart_data if isinstance(chart_data, dict) else {}
    items = list(zip(
        data.get('labels', []),
        data.get('values', []),
    ))
    if not items:
        return '_Sin datos suficientes para análisis._'

    total = sum(v for _, v in items)
    if total == 0:
        return '_Sin datos registrados en el período._'

    top = max(items, key=lambda x: x[1])
    bottom = min(items, key=lambda x: x[1])

    analysis_map = {
        'headcount': (
            f'**Headcount**: rango de {bottom[1]} a {top[1]} empleados. '
            f'Mayor registro en {top[0]} ({top[1]}), menor en {bottom[0]} ({bottom[1]}).'
        ),
        'rotacion': (
            f'**Rotación**: mayor en {top[0]} ({top[1]}%), menor en {bottom[0]} ({bottom[1]}%). '
            f'Se recomienda analizar causas si supera el 5% mensual.'
        ),
        'asistencia': (
            f'**Asistencia**: {len(items)} indicadores. '
            f'Mayor: {top[0]} ({top[1]}), menor: {bottom[0]} ({bottom[1]}).'
        ),
        'areas': (
            f'**Distribución**: {len(items)} áreas, total {total} empleados. '
            f'Mayor concentración en {top[0]} ({top[1]}, {top[1] * 100 // total}%). '
            f'Menor en {bottom[0]} ({bottom[1]}).'
        ),
    }
    return analysis_map.get(chart_type, (
        f'**Datos**: {len(items)} categorías, total {total}. '
        f'Mayor: {top[0]} ({top[1]}), menor: {bottom[0]} ({bottom[1]}).'
    ))


# ─────────────────────────────────────────────────
# ANÁLISIS DE GRÁFICO (para dashboard ejecutivo)
# ─────────────────────────────────────────────────

@login_required
@require_POST
def ai_analyze_chart(request):
    """
    Genera análisis narrativo de un gráfico/sección de datos.
    Body JSON: {"chart": "headcount|rotacion|asistencia|areas|antiguedad|genero|edad|he_mes", "data": {...}}
    Retorna SSE streaming.
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    chart_type = body.get('chart', '')
    chart_data = body.get('data', {})

    ollama_ok = _is_ollama_reachable()

    # Fallback estático para análisis de gráficos
    if not ollama_ok:
        analysis = _static_chart_analysis(chart_type, chart_data)

        def fallback_stream():
            yield f'data: {analysis}\n\n'
            yield 'data: [DONE]\n\n'

        response = StreamingHttpResponse(
            fallback_stream(), content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response

    svc = get_service()

    # Sistema con contexto rico de la empresa (módulo personal)
    try:
        system = build_system_prompt(request.user, modules=['personal'])
        system += (
            '\n\nROL PARA ANÁLISIS DE GRÁFICO:\n'
            '- Analiza el gráfico con los datos exactos del input\n'
            '- Máximo 4-5 oraciones directas, sin bullet points\n'
            '- Usa los números exactos, calcula porcentajes cuando aplique\n'
            '- Menciona benchmarks de la industria peruana cuando sea relevante\n'
            '- Nunca digas "los datos muestran" o "se puede observar" — ve directo\n'
            '- Nunca inventes datos que no estén en el input'
        )
    except Exception:
        system = (
            'Eres Harmoni AI, analista senior de RRHH especializado en empresas Peru. '
            'Responde SIEMPRE en español. '
            'Sé directo, usa números exactos. '
            'Máximo 4-5 oraciones. Sin bullet points.'
        )

    # Calcular totales y top items para enriquecer los prompts
    labels = chart_data.get('labels', [])
    values = chart_data.get('values', [])
    total = sum(v for v in values if isinstance(v, (int, float)))
    items = list(zip(labels, values))
    top3 = sorted(items, key=lambda x: x[1], reverse=True)[:3] if items else []
    data_json = json.dumps(chart_data, ensure_ascii=False)

    prompts = {
        'headcount': (
            f'Analiza la evolución del headcount de los últimos {len(labels)} meses.\n'
            f'Datos: {data_json}\n\n'
            f'1. Identifica la tendencia: crecimiento, reducción o estabilidad\n'
            f'2. Señala el mes pico ({max(values) if values else "?"}) y el mes mínimo ({min(values) if values else "?"})\n'
            f'3. Calcula la variación entre el primer y último dato\n'
            f'4. Concluye con una interpretación estratégica para gerencia'
        ),
        'rotacion': (
            f'Analiza la tasa de rotación mensual de personal.\n'
            f'Datos: {data_json}\n\n'
            f'1. Promedio del periodo: {(sum(values)/len(values)):.2f}% mensual. '
            f'Benchmark Peru construccion/mineria: 2-4% mensual. ¿Está dentro?\n'
            f'2. ¿Qué meses tuvieron picos y cuál podría ser la causa?\n'
            f'3. Recomienda 1 acción específica basada en los datos'
        ),
        'asistencia': (
            f'Analiza la tasa de asistencia/presencia mensual.\n'
            f'Datos: {data_json}\n\n'
            f'1. Benchmark saludable: >95%. El promedio aquí es {(sum(values)/len(values)):.1f}%\n'
            f'2. ¿Hay meses con caídas? ¿Detectas estacionalidad?\n'
            f'3. Sugiere 1 acción preventiva concreta'
        ),
        'asistencia_mes': (
            f'Analiza la tasa de asistencia/presencia mensual.\n'
            f'Datos: {data_json}\n\n'
            f'1. Benchmark saludable: >95%. El promedio aquí es {(sum(values)/len(values)):.1f}% si hay datos\n'
            f'2. ¿Hay meses con caídas notables?\n'
            f'3. Sugiere 1 acción preventiva concreta'
        ),
        'areas': (
            f'Analiza la distribución de {total} empleados en {len(labels)} áreas.\n'
            f'Datos: {data_json}\n\n'
            f'Top 3 áreas: {", ".join(f"{a} ({v}, {v*100//total if total else 0}%)" for a,v in top3)}.\n'
            f'1. ¿Es razonable esta concentración para el tipo de empresa?\n'
            f'2. ¿Hay áreas subdimensionadas (<3% del total)?\n'
            f'3. Recomienda 1 ajuste organizacional'
        ),
        'antiguedad': (
            f'Analiza la distribución de antigüedad del personal ({total} empleados).\n'
            f'Datos: {data_json}\n\n'
            f'1. ¿Qué proporción tiene menos de 1 año? Eso indica rotación o crecimiento reciente\n'
            f'2. ¿Hay suficiente personal senior (>3 años) para transferencia de conocimiento?\n'
            f'3. Evalúa el riesgo operativo por concentración en ciertos rangos'
        ),
        'genero': (
            f'Analiza la distribución de género del personal ({total} empleados).\n'
            f'Datos: {data_json}\n\n'
            f'1. ¿Cuál es la proporción M/F? ¿Es típica del sector construcción/minería en Peru?\n'
            f'2. ¿Qué implica esta distribución para políticas de diversidad e inclusión?\n'
            f'3. ¿Hay oportunidades de mejora en equidad de género?'
        ),
        'edad': (
            f'Analiza la distribución etaria del personal ({total} empleados con fecha nacimiento registrada).\n'
            f'Datos: {data_json}\n\n'
            f'1. ¿Qué rango etario predomina? ¿Es una plantilla joven, madura o mixta?\n'
            f'2. Implicaciones para: capacitación, sucesión, beneficios y clima laboral\n'
            f'3. ¿Hay riesgo de pérdida de conocimiento por concentración en rangos mayores?'
        ),
        'he_mes': (
            f'Analiza las horas extra mensuales acumuladas.\n'
            f'Datos: {data_json}\n\n'
            f'1. ¿Hay tendencia creciente de HE? Eso puede indicar subdimensionamiento\n'
            f'2. ¿Hay meses con picos? ¿Coinciden con cierre de proyectos o temporadas?\n'
            f'3. Recomienda 1 acción: ¿contratar más personal o gestionar mejor la carga?'
        ),
    }

    prompt = prompts.get(
        chart_type,
        f'Analiza estos datos de RRHH, identifica 3 hallazgos clave con números exactos '
        f'y da 1 recomendación accionable para gerencia.\nDatos: {data_json}'
    )

    def event_stream():
        try:
            for chunk in svc.generate_stream(
                prompt=prompt,
                system=system,
                temperature=0.4,
                num_predict=500,
            ):
                yield f'data: {chunk}\n\n'
            yield 'data: [DONE]\n\n'
        except Exception as e:
            logger.warning(f'ai_analyze_chart error: {e}')
            analysis = _static_chart_analysis(chart_type, chart_data)
            yield f'data: {analysis}\n\n'
            yield 'data: [DONE]\n\n'

    response = StreamingHttpResponse(
        event_stream(),
        content_type='text/event-stream',
    )
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


@login_required
@require_POST
def ai_ask_data(request):
    """
    Responde pregunta libre sobre datos de RRHH.
    Body JSON: {"question": "..."}
    Retorna SSE streaming.
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    question = body.get('question', '').strip()
    if not question:
        return JsonResponse({'error': 'Pregunta vacía'}, status=400)

    # ── Archivo adjunto opcional ──
    file_context = body.get('file_context')
    ai_question = question
    if file_context:
        fc_type = file_context.get('type', '')
        fc_name = file_context.get('name', 'archivo')
        fc_content = file_context.get('content', '')
        if fc_type != 'image' and fc_content:
            trunc_note = ' [contenido truncado a 12000 chars]' if file_context.get('truncated') else ''
            ai_question = (
                f'[ARCHIVO ADJUNTO: {fc_name}{trunc_note}]\n'
                f'{fc_content}\n'
                f'[FIN ARCHIVO]\n\n'
                f'Pregunta del usuario: {question}'
            )

    svc = get_service()
    ollama_ok = _is_ollama_reachable()

    if not ollama_ok:
        fallback = responder_sin_ia(question, request.user)

        def fallback_stream():
            yield 'data: [FALLBACK]\n\n'
            if fallback:
                yield f'data: {_sse_text(fallback)}\n\n'
            else:
                yield (
                    'data: No pude responder esa consulta sin IA. '
                    'Configura Ollama para consultas avanzadas.\n\n'
                )
            yield 'data: [DONE]\n\n'

        response = StreamingHttpResponse(
            fallback_stream(), content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response

    modules = detect_module_context(question)
    system_prompt = build_system_prompt(request.user, modules=modules)

    messages = [{'role': 'user', 'content': ai_question}]

    def event_stream():
        try:
            for chunk in svc.chat_stream(
                messages=messages,
                system=system_prompt,
                temperature=0.35,
                num_predict=900,
            ):
                yield f'data: {chunk}\n\n'
            yield 'data: [DONE]\n\n'
        except Exception as e:
            logger.warning(f'ai_ask_data error: {e}')
            yield 'data: _No se pudo generar respuesta._\n\n'
            yield 'data: [DONE]\n\n'

    response = StreamingHttpResponse(
        event_stream(),
        content_type='text/event-stream',
    )
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


# ─────────────────────────────────────────────────
# EXPORT REPORTE EXCEL
# ─────────────────────────────────────────────────

@login_required
@require_POST
def ai_export_report(request):
    """
    Genera y descarga reporte ejecutivo en Excel (.xlsx).
    Body JSON: {"type": "gerencia"}
    """
    from django.http import HttpResponse

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    report_type = body.get('type', 'gerencia')

    if report_type != 'gerencia':
        return JsonResponse({'error': f'Tipo no soportado: {report_type}'}, status=400)

    try:
        from .services.ai_excel_export import ReporteGerenciaExporter
        exporter = ReporteGerenciaExporter(request.user)
        wb = exporter.generate()

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from datetime import date
        filename = f'reporte-ejecutivo-{date.today().isoformat()}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f'ai_export_report error: {e}')
        return JsonResponse({'error': str(e)}, status=500)


# ─────────────────────────────────────────────────
# UPLOAD ARCHIVO ADJUNTO
# ─────────────────────────────────────────────────

@login_required
@require_POST
def ai_upload_file(request):
    """
    Procesa archivo adjunto para el chat IA.
    Soporta: PDF (extrae texto), Excel (.xlsx), imágenes (JPG/PNG).
    Retorna JSON con contenido extraído para incluir en el contexto.
    Max tamaño: 10MB.
    """
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'ok': False, 'error': 'No se recibió archivo.'}, status=400)

    MAX_SIZE = 10 * 1024 * 1024  # 10MB
    if file.size > MAX_SIZE:
        return JsonResponse({'ok': False, 'error': 'Archivo muy grande (máx 10MB).'}, status=400)

    name = file.name or 'archivo'
    ext = name.lower().rsplit('.', 1)[-1] if '.' in name else ''
    content_bytes = file.read()

    try:
        # ── PDF ──────────────────────────────────────────────────────
        if ext == 'pdf':
            try:
                import fitz  # PyMuPDF
            except ImportError:
                return JsonResponse({'ok': False, 'error': 'PyMuPDF no instalado. pip install pymupdf'})

            doc = fitz.open(stream=content_bytes, filetype='pdf')
            pages_text = [page.get_text() for page in doc]
            doc.close()
            full_text = '\n'.join(pages_text).strip()

            if not full_text:
                return JsonResponse({
                    'ok': False,
                    'error': 'PDF escaneado sin texto extraíble. Activa OCR con Gemini en Configuración.',
                })

            # Truncate a 12000 chars (≈3000 tokens) para no saturar el contexto
            truncated = full_text[:12000]
            was_truncated = len(full_text) > 12000
            preview = full_text[:120].replace('\n', ' ').strip() + ('...' if len(full_text) > 120 else '')

            # Guardar bytes originales en cache para edición posterior (15 min TTL)
            file_id = uuid.uuid4().hex[:16]
            cache.set(f'ai_pdf_{file_id}', content_bytes, 900)

            return JsonResponse({
                'ok': True,
                'type': 'pdf',
                'name': name,
                'pages': len(pages_text),
                'content': truncated,
                'truncated': was_truncated,
                'preview': preview,
                'size_kb': round(file.size / 1024, 1),
                'file_id': file_id,  # ID para recuperar bytes y editar
            })

        # ── EXCEL ─────────────────────────────────────────────────────
        elif ext in ('xlsx', 'xls'):
            try:
                import openpyxl
                from io import BytesIO
            except ImportError:
                return JsonResponse({'ok': False, 'error': 'openpyxl no instalado.'})

            wb = openpyxl.load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
            sheets_text = []
            for sheet_name in list(wb.sheetnames)[:4]:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(max_row=80, values_only=True):
                    if any(c is not None for c in row):
                        rows.append(' | '.join(str(c or '').strip() for c in row if c is not None))
                if rows:
                    sheets_text.append(f'## Hoja: {sheet_name}\n' + '\n'.join(rows[:60]))
            wb.close()

            text = '\n\n'.join(sheets_text)
            if not text.strip():
                return JsonResponse({'ok': False, 'error': 'Excel sin datos legibles.'})

            truncated = text[:10000]
            preview = f'Excel: {name} — {len(wb.sheetnames)} hoja(s)'

            return JsonResponse({
                'ok': True,
                'type': 'excel',
                'name': name,
                'sheets': len(wb.sheetnames),
                'content': truncated,
                'truncated': len(text) > 10000,
                'preview': preview,
                'size_kb': round(file.size / 1024, 1),
            })

        # ── IMAGEN ────────────────────────────────────────────────────
        elif ext in ('jpg', 'jpeg', 'png', 'webp', 'gif'):
            import base64
            b64 = base64.b64encode(content_bytes).decode()
            mime = 'image/jpeg' if ext in ('jpg', 'jpeg') else f'image/{ext}'

            return JsonResponse({
                'ok': True,
                'type': 'image',
                'name': name,
                'content': b64,
                'mime': mime,
                'preview': f'Imagen: {name}',
                'size_kb': round(file.size / 1024, 1),
            })

        # ── TEXTO PLANO ───────────────────────────────────────────────
        elif ext in ('txt', 'csv', 'md'):
            text = content_bytes.decode('utf-8', errors='replace')
            truncated = text[:10000]
            return JsonResponse({
                'ok': True,
                'type': 'text',
                'name': name,
                'content': truncated,
                'truncated': len(text) > 10000,
                'preview': text[:100].replace('\n', ' ').strip(),
                'size_kb': round(file.size / 1024, 1),
            })

        else:
            return JsonResponse({
                'ok': False,
                'error': f'Formato .{ext} no soportado. Use: PDF, Excel, JPG, PNG o TXT.',
            }, status=400)

    except Exception as e:
        logger.warning(f'ai_upload_file error: {e}')
        return JsonResponse({'ok': False, 'error': f'Error procesando archivo: {str(e)[:120]}'}, status=500)


# ─────────────────────────────────────────────────
# DESCARGA PDF EDITADO
# ─────────────────────────────────────────────────

@login_required
@require_GET
def ai_download_edited(request):
    """
    Sirve un PDF editado desde cache temporal.
    GET /asistencia/ia/documento-editado/?id=<edit_id>
    El edit_id es generado por _apply_pdf_edit y tiene TTL de 15 min.
    """
    edit_id = request.GET.get('id', '').strip()
    if not edit_id or not edit_id.isalnum() or len(edit_id) > 32:
        return JsonResponse({'error': 'ID inválido'}, status=400)

    pdf_bytes = cache.get(f'ai_edit_{edit_id}')
    if not pdf_bytes:
        return JsonResponse(
            {'error': 'Documento expirado o no encontrado. Vuelve a subir el PDF y solicita la edición.'},
            status=404,
        )

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="documento-editado.pdf"'
    return response


# ─────────────────────────────────────────────────
# RAG — INDEXAR EMBEDDINGS
# ─────────────────────────────────────────────────

@login_required
@require_POST
def ai_index_embeddings(request):
    """
    Calcula embeddings para todos los artículos de KnowledgeArticle sin embedding.
    Requiere superusuario o permiso de configuración.

    Body JSON: {"force": false}   (force=true recalcula todos)
    Returns: {"ok": true, "indexed": N, "errors": M, "stats": {...}}
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Acceso denegado'}, status=403)

    try:
        body = json.loads(request.body) if request.body else {}
    except (json.JSONDecodeError, ValueError):
        body = {}

    force = bool(body.get('force', False))

    from core.knowledge_service import embed_text, get_embedding_stats, _get_openai_api_key
    from core.models import KnowledgeArticle
    from django.db.models import Q

    api_key = _get_openai_api_key()
    if not api_key:
        return JsonResponse({
            'ok': False,
            'error': (
                'No hay API key de OpenAI configurada. '
                'Configura ia_provider=OPENAI con su API key, o usa un provider con embeddings.'
            ),
        }, status=400)

    # Seleccionar artículos
    if force:
        qs = KnowledgeArticle.objects.filter(activo=True)
    else:
        qs = KnowledgeArticle.objects.filter(activo=True).filter(
            Q(embedding_json__isnull=True) | Q(embedding_json='')
        )

    articles = list(qs.order_by('prioridad', 'id'))

    if not articles:
        stats = get_embedding_stats()
        return JsonResponse({'ok': True, 'indexed': 0, 'errors': 0, 'already_done': True, 'stats': stats})

    ok_count = 0
    error_count = 0

    for art in articles:
        text = f'{art.titulo}\n\n{art.contenido}'
        try:
            vec = embed_text(text, api_key=api_key)
            if vec:
                import json as _json
                art.embedding_json = _json.dumps(vec)
                art.save(update_fields=['embedding_json'])
                ok_count += 1
            else:
                error_count += 1
        except Exception as exc:
            logger.warning('ai_index_embeddings: error en artículo pk=%s: %s', art.pk, exc)
            error_count += 1

    stats = get_embedding_stats()
    return JsonResponse({
        'ok': error_count == 0,
        'indexed': ok_count,
        'errors': error_count,
        'stats': stats,
    })
