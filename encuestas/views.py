"""
Vistas del módulo de Encuestas y Clima Laboral.
"""
import json
from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Avg, Count, Sum
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST

from personal.models import Personal
from .models import (
    Encuesta, PreguntaEncuesta, RespuestaEncuesta, ResultadoEncuesta,
)

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ══════════════════════════════════════════════════════════════
# ADMIN — PANEL PRINCIPAL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def encuestas_panel(request):
    """Panel principal de encuestas."""
    qs = Encuesta.objects.all()

    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '')

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo=tipo)

    # ── KPI: respuestas del mes actual ───────────────────────────────────────
    respuestas_mes = 0
    try:
        hoy = date.today()
        respuestas_mes = RespuestaEncuesta.objects.filter(
            fecha_respuesta__year=hoy.year,
            fecha_respuesta__month=hoy.month,
        ).count()
    except Exception:
        pass

    # ── KPI: eNPS score actual (ultimo resultado con enps_score) ─────────────
    enps_kpi = None
    try:
        ultimo_enps = (
            ResultadoEncuesta.objects
            .filter(enps_score__isnull=False)
            .order_by('-calculado_en')
            .values('enps_score')
            .first()
        )
        if ultimo_enps:
            enps_kpi = ultimo_enps['enps_score']
    except Exception:
        pass

    # ── KPI: tasa de participacion de la ultima encuesta cerrada ─────────────
    tasa_participacion_kpi = None
    try:
        ultimo_resultado_kpi = (
            ResultadoEncuesta.objects
            .filter(total_participantes__gt=0)
            .order_by('-calculado_en')
            .values('tasa_participacion')
            .first()
        )
        if ultimo_resultado_kpi:
            tasa_participacion_kpi = float(ultimo_resultado_kpi['tasa_participacion'])
    except Exception:
        pass

    # ── KPI: distribucion por tipo de encuesta ───────────────────────────────
    por_tipo_json = '[]'
    try:
        tipo_labels = dict(Encuesta.TIPO_CHOICES)
        tipo_qs = (
            Encuesta.objects
            .values('tipo')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        por_tipo_json = json.dumps([
            {'label': tipo_labels.get(item['tipo'], item['tipo']), 'count': item['count']}
            for item in tipo_qs
            if item['count'] > 0
        ])
    except Exception:
        pass

    # ── eNPS histórico (últimos 6 ResultadoEncuesta con enps_score) ──────────
    enps_historico_json = '[]'
    try:
        historico_qs = (
            ResultadoEncuesta.objects
            .filter(enps_score__isnull=False)
            .order_by('calculado_en')
            .select_related('encuesta')
        )[:6]
        enps_historico_json = json.dumps([
            {
                'label': r.calculado_en.strftime('%m/%y'),
                'enps': r.enps_score,
            }
            for r in historico_qs
        ])
    except Exception:
        pass

    # ── Tasa de respuesta promedio (últimas 5 encuestas con respuestas) ──────
    tasa_respuesta_promedio = None
    try:
        ultimas = (
            ResultadoEncuesta.objects
            .filter(total_participantes__gt=0)
            .order_by('-calculado_en')
        )[:5]
        tasas = [float(r.tasa_participacion) for r in ultimas if r.tasa_participacion]
        if tasas:
            tasa_respuesta_promedio = round(sum(tasas) / len(tasas), 1)
    except Exception:
        pass

    # ── Promotores / Pasivos / Detractores del resultado más reciente ────────
    promotores_neutros_detractores = None
    enps_score_actual = None
    try:
        ultimo_resultado = (
            ResultadoEncuesta.objects
            .filter(enps_score__isnull=False)
            .order_by('-calculado_en')
            .first()
        )
        if ultimo_resultado:
            enps_score_actual = ultimo_resultado.enps_score
            total_enps = (
                ultimo_resultado.enps_promotores
                + ultimo_resultado.enps_pasivos
                + ultimo_resultado.enps_detractores
            )
            promotores_neutros_detractores = {
                'promotores': ultimo_resultado.enps_promotores,
                'neutros': ultimo_resultado.enps_pasivos,
                'detractores': ultimo_resultado.enps_detractores,
                'total': total_enps,
                'encuesta': ultimo_resultado.encuesta.titulo,
            }
    except Exception:
        pass

    # ── Encuestas activas próximas a cerrar (fecha_fin <= hoy + 7 días) ──────
    encuestas_pendientes_cierre = []
    try:
        limite = date.today() + timedelta(days=7)
        encuestas_pendientes_cierre = list(
            Encuesta.objects.filter(
                estado='ACTIVA',
                fecha_fin__lte=limite,
            ).order_by('fecha_fin')
        )
    except Exception:
        pass

    context = {
        'titulo': 'Encuestas y Clima Laboral',
        'encuestas': qs[:50],
        'total': qs.count(),
        'filtro_estado': estado,
        'filtro_tipo': tipo,
        'stats': {
            'activas': Encuesta.objects.filter(estado='ACTIVA').count(),
            'cerradas': Encuesta.objects.filter(estado='CERRADA').count(),
            'borradores': Encuesta.objects.filter(estado='BORRADOR').count(),
            'total_respuestas': RespuestaEncuesta.objects.count(),
        },
        # ── KPI strip ─────────────────────────────────────────────────────
        'encuestas_activas': Encuesta.objects.filter(estado='ACTIVA').count(),
        'respuestas_mes': respuestas_mes,
        'enps_kpi': enps_kpi,
        'tasa_participacion_kpi': tasa_participacion_kpi,
        'por_tipo_json': por_tipo_json,
        # ── Analytics cards ────────────────────────────────────────────────
        'enps_historico_json': enps_historico_json,
        'tasa_respuesta_promedio': tasa_respuesta_promedio,
        'promotores_neutros_detractores': promotores_neutros_detractores,
        'enps_score_actual': enps_score_actual,
        'encuestas_pendientes_cierre': encuestas_pendientes_cierre,
    }
    return render(request, 'encuestas/panel.html', context)


@login_required
@solo_admin
def encuesta_crear(request):
    """Crear nueva encuesta."""
    if request.method == 'POST':
        try:
            enc = Encuesta.objects.create(
                titulo=request.POST['titulo'],
                descripcion=request.POST.get('descripcion', ''),
                tipo=request.POST.get('tipo', 'CLIMA'),
                anonima=request.POST.get('anonima') == 'on',
                fecha_inicio=request.POST['fecha_inicio'],
                fecha_fin=request.POST['fecha_fin'],
                aplica_grupos=request.POST.get('aplica_grupos', ''),
                estado='BORRADOR',
                creado_por=request.user,
            )
            from core.audit import log_create
            log_create(request, enc, f'Encuesta creada: {enc.titulo}')
            messages.success(request, f'Encuesta "{enc.titulo}" creada. Ahora agregue preguntas.')
            return redirect('encuesta_detalle', pk=enc.pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': 'Nueva Encuesta',
    }
    return render(request, 'encuestas/crear.html', context)


@login_required
@solo_admin
def encuesta_detalle(request, pk):
    """Detalle de encuesta con preguntas y resultados."""
    enc = get_object_or_404(Encuesta, pk=pk)
    preguntas = enc.preguntas.all()

    # Si cerrada, mostrar resultados
    resultado = None
    if enc.estado in ('CERRADA', 'ARCHIVADA'):
        resultado = ResultadoEncuesta.objects.filter(encuesta=enc).first()

    context = {
        'titulo': enc.titulo,
        'encuesta': enc,
        'preguntas': preguntas,
        'resultado': resultado,
        'total_respuestas': enc.respuestas.count(),
    }
    return render(request, 'encuestas/detalle.html', context)


@login_required
@solo_admin
@require_POST
def pregunta_agregar(request, enc_pk):
    """Agregar pregunta a encuesta."""
    enc = get_object_or_404(Encuesta, pk=enc_pk)
    try:
        opciones = request.POST.get('opciones', '')
        opciones_list = [o.strip() for o in opciones.split(',') if o.strip()] if opciones else []

        preg = PreguntaEncuesta.objects.create(
            encuesta=enc,
            texto=request.POST['texto'],
            tipo=request.POST.get('tipo', 'ESCALA_5'),
            obligatoria=request.POST.get('obligatoria', 'on') == 'on',
            opciones=opciones_list,
            categoria=request.POST.get('categoria', ''),
            orden=enc.preguntas.count() + 1,
        )
        return JsonResponse({'ok': True, 'pk': preg.pk, 'texto': preg.texto})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def pregunta_eliminar(request, pk):
    """Eliminar pregunta de encuesta."""
    preg = get_object_or_404(PreguntaEncuesta, pk=pk)
    if preg.encuesta.estado == 'BORRADOR':
        preg.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': 'Solo se pueden eliminar preguntas de encuestas en borrador.'})


@login_required
@solo_admin
@require_POST
def encuesta_activar(request, pk):
    """Activar encuesta."""
    enc = get_object_or_404(Encuesta, pk=pk)
    if enc.estado in ('BORRADOR', 'PROGRAMADA'):
        if enc.preguntas.count() == 0:
            return JsonResponse({'ok': False, 'error': 'La encuesta debe tener al menos una pregunta.'})
        enc.estado = 'ACTIVA'
        enc.save(update_fields=['estado'])
        return JsonResponse({'ok': True, 'estado': 'ACTIVA'})
    return JsonResponse({'ok': False, 'error': 'No se puede activar en este estado.'})


@login_required
@solo_admin
@require_POST
def encuesta_cerrar(request, pk):
    """Cerrar encuesta y calcular resultados."""
    enc = get_object_or_404(Encuesta, pk=pk)
    if enc.estado != 'ACTIVA':
        return JsonResponse({'ok': False, 'error': 'Solo se pueden cerrar encuestas activas.'})

    enc.estado = 'CERRADA'
    enc.save(update_fields=['estado'])

    # Calcular resultados
    resultado, _ = ResultadoEncuesta.objects.get_or_create(encuesta=enc)
    resultado.total_participantes = enc.respuestas.count()

    universo = Personal.objects.filter(estado='Activo')
    if enc.aplica_grupos:
        universo = universo.filter(grupo_tareo=enc.aplica_grupos)
    total_universo = universo.count()
    if total_universo > 0:
        resultado.tasa_participacion = round((resultado.total_participantes / total_universo) * 100, 1)

    # Puntajes por dimensión
    dimensiones = {}
    for preg in enc.preguntas.filter(tipo__in=['ESCALA_5', 'ESCALA_10']):
        cat = preg.categoria or 'General'
        if cat not in dimensiones:
            dimensiones[cat] = {'total': 0, 'count': 0}
        for resp in enc.respuestas.all():
            val = resp.respuestas.get(str(preg.pk))
            if val is not None:
                try:
                    dimensiones[cat]['total'] += float(val)
                    dimensiones[cat]['count'] += 1
                except (ValueError, TypeError):
                    pass

    puntajes = {}
    total_general = 0
    count_general = 0
    for cat, data in dimensiones.items():
        if data['count'] > 0:
            promedio = round(data['total'] / data['count'], 2)
            puntajes[cat] = promedio
            total_general += data['total']
            count_general += data['count']

    resultado.puntajes_dimension = puntajes
    if count_general > 0:
        resultado.puntaje_general = Decimal(str(round(total_general / count_general, 2)))

    # eNPS
    resultado.calcular_enps()
    resultado.save()

    return JsonResponse({
        'ok': True,
        'participantes': resultado.total_participantes,
        'tasa': float(resultado.tasa_participacion),
    })


# ══════════════════════════════════════════════════════════════
# ADMIN — RESULTADOS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def encuesta_resultados(request, pk):
    """Vista detallada de resultados de encuesta.

    ANONIMATO: el contexto NUNCA expone objetos RespuestaEncuesta ni datos
    de personal identificables — solo valores agregados y textos libres
    sin vínculo con el respondiente.
    """
    import json
    import random as _random
    from collections import Counter

    enc = get_object_or_404(Encuesta, pk=pk)
    resultado = ResultadoEncuesta.objects.filter(encuesta=enc).first()

    # Detalle por pregunta — solo valores, sin identificadores
    detalle_preguntas = []
    for preg in enc.preguntas.all():
        valores = []
        # Iterar solo sobre el JSON de respuestas, sin acceder a personal
        for resp_dict in enc.respuestas.values_list('respuestas', flat=True):
            val = resp_dict.get(str(preg.pk))
            if val is not None:
                valores.append(val)

        info = {
            'pregunta': preg,
            'total_respuestas': len(valores),
        }

        if preg.tipo in ('ESCALA_5', 'ESCALA_10'):
            nums = [float(v) for v in valores if v]
            info['promedio'] = round(sum(nums) / len(nums), 2) if nums else 0
            if preg.tipo == 'ESCALA_5':
                dist = {str(i): nums.count(float(i)) for i in range(1, 6)}
            else:
                dist = {str(i): nums.count(float(i)) for i in range(0, 11)}
            info['distribucion'] = dist
            info['distribucion_json'] = json.dumps(dist)
        elif preg.tipo in ('OPCION', 'SI_NO'):
            dist = dict(Counter(str(v) for v in valores))
            info['distribucion'] = dist
            info['distribucion_json'] = json.dumps(dist)
        elif preg.tipo == 'TEXTO':
            # Mezclar textos para eliminar cualquier orden que pueda identificar
            textos = [str(v) for v in valores if v and str(v).strip()]
            _random.shuffle(textos)
            info['textos'] = textos[:20]

        detalle_preguntas.append(info)

    # Dimensiones JSON para Chart.js
    dimensiones_json = json.dumps(resultado.puntajes_dimension) if resultado else '{}'

    context = {
        'titulo': f'Resultados: {enc.titulo}',
        'encuesta': enc,
        'resultado': resultado,
        'detalle_preguntas': detalle_preguntas,
        'dimensiones_json': dimensiones_json,
        # NUNCA incluir respuestas individuales ni objetos personal en el contexto
    }
    return render(request, 'encuestas/resultados.html', context)


# ══════════════════════════════════════════════════════════════
# PORTAL DEL TRABAJADOR
# ══════════════════════════════════════════════════════════════

@login_required
def mis_encuestas(request):
    """Portal: encuestas activas pendientes y respondidas."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    activas = Encuesta.objects.filter(estado='ACTIVA')
    respondidas_ids = []

    if empleado:
        respondidas_ids = list(RespuestaEncuesta.objects.filter(
            personal=empleado,
        ).values_list('encuesta_id', flat=True))

    pendientes = activas.exclude(pk__in=respondidas_ids)

    context = {
        'titulo': 'Mis Encuestas',
        'empleado': empleado,
        'pendientes': pendientes,
        'respondidas_count': len(respondidas_ids),
    }
    return render(request, 'encuestas/mis_encuestas.html', context)


@login_required
def responder_encuesta(request, pk):
    """Portal: responder una encuesta."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    enc = get_object_or_404(Encuesta, pk=pk, estado='ACTIVA')
    preguntas = enc.preguntas.all()

    # Verificar si ya respondió
    if empleado and enc.respuestas.filter(personal=empleado).exists():
        messages.info(request, 'Ya respondiste esta encuesta.')
        return redirect('mis_encuestas')

    if request.method == 'POST':
        try:
            respuestas_dict = {}
            for preg in preguntas:
                val = request.POST.get(f'preg_{preg.pk}')
                if val is not None:
                    respuestas_dict[str(preg.pk)] = val

            resp = RespuestaEncuesta(
                encuesta=enc,
                respuestas=respuestas_dict,
                comentarios=request.POST.get('comentarios', ''),
            )

            if enc.anonima:
                # Solo guardar metadata anónima
                if empleado:
                    resp.area_anonima = str(empleado.subarea.area) if empleado.subarea else ''
                    resp.grupo_anonimo = empleado.grupo_tareo or ''
            else:
                resp.personal = empleado

            # Capturar IP
            x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
            resp.ip = x_forwarded.split(',')[0].strip() if x_forwarded else request.META.get('REMOTE_ADDR')

            resp.save()
            messages.success(request, 'Encuesta respondida correctamente. Gracias por tu participación.')
            return redirect('mis_encuestas')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': enc.titulo,
        'encuesta': enc,
        'preguntas': preguntas,
    }
    return render(request, 'encuestas/responder.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — ACCIONES ADICIONALES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def encuesta_exportar_excel(request, pk):
    """Exportar respuestas de encuesta a Excel (.xlsx).

    ANONIMATO: encuestas anónimas solo incluyen área y grupo (sin nombre personal).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    enc = get_object_or_404(Encuesta, pk=pk)
    preguntas = list(enc.preguntas.all().order_by('orden'))
    respuestas = enc.respuestas.all().order_by('fecha_respuesta')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Respuestas'

    HEADER_COLOR = '0D2B27'
    ALT_COLOR = 'F0FDFA'
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor=HEADER_COLOR)
    alt_fill = PatternFill('solid', fgColor=ALT_COLOR)
    center = Alignment(horizontal='center', vertical='center')

    # ── Construir cabecera ─────────────────────────────────────
    headers = ['#', 'Fecha', 'Área', 'Grupo']
    if not enc.anonima:
        headers.insert(2, 'Nombre')
    for p in preguntas:
        short = p.texto[:40] + '…' if len(p.texto) > 40 else p.texto
        headers.append(f'[{p.get_tipo_display()}] {short}')
    headers.append('Comentarios')

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Filas de datos ─────────────────────────────────────────
    for idx, resp in enumerate(respuestas, 1):
        row = [idx, resp.fecha_respuesta.strftime('%d/%m/%Y %H:%M')]
        # Identidad
        if not enc.anonima and resp.personal:
            row.append(resp.personal.apellidos_nombres)
        area = resp.area_anonima or (
            str(resp.personal.subarea.area) if (resp.personal and resp.personal.subarea) else ''
        )
        grupo = resp.grupo_anonimo or (resp.personal.grupo_tareo if resp.personal else '')
        row.extend([area, grupo])

        # Respuestas por pregunta
        for p in preguntas:
            val = resp.respuestas.get(str(p.pk), '')
            row.append(val)

        row.append(resp.comentarios)
        ws.append(row)

        # Alternar colores
        if idx % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    # ── Hoja de resumen dimensiones ────────────────────────────
    ws2 = wb.create_sheet('Dimensiones')
    ws2.append(['Dimensión / Categoría', 'Promedio', 'N° Respuestas'])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    resultado = ResultadoEncuesta.objects.filter(encuesta=enc).first()
    if resultado and resultado.puntajes_dimension:
        for dim, promedio in resultado.puntajes_dimension.items():
            ws2.append([dim, promedio, resultado.total_participantes])

    # ── Auto-width ────────────────────────────────────────────
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = ''.join(c for c in enc.titulo if c.isalnum() or c in (' ', '_', '-'))[:40]
    filename = f'encuesta_{safe_title}_{date.today():%Y%m%d}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@solo_admin
@require_POST
def encuesta_duplicar(request, pk):
    """Duplicar encuesta (con todas sus preguntas) como nuevo borrador."""
    enc = get_object_or_404(Encuesta, pk=pk)

    nueva = Encuesta.objects.create(
        titulo=f'Copia de {enc.titulo}',
        descripcion=enc.descripcion,
        tipo=enc.tipo,
        anonima=enc.anonima,
        fecha_inicio=enc.fecha_inicio,
        fecha_fin=enc.fecha_fin,
        aplica_grupos=enc.aplica_grupos,
        max_respuestas=enc.max_respuestas,
        recordatorio_dias=enc.recordatorio_dias,
        estado='BORRADOR',
        creado_por=request.user,
    )

    # Copiar áreas participantes
    if enc.aplica_areas.exists():
        nueva.aplica_areas.set(enc.aplica_areas.all())

    # Copiar preguntas
    preguntas_orig = enc.preguntas.all().order_by('orden')
    for p in preguntas_orig:
        PreguntaEncuesta.objects.create(
            encuesta=nueva,
            texto=p.texto,
            tipo=p.tipo,
            obligatoria=p.obligatoria,
            opciones=p.opciones,
            categoria=p.categoria,
            orden=p.orden,
        )

    from core.audit import log_create
    log_create(request, nueva, f'Encuesta duplicada desde #{enc.pk}: {nueva.titulo}')
    messages.success(request, f'Encuesta duplicada como borrador: "{nueva.titulo}".')
    return redirect('encuesta_detalle', pk=nueva.pk)


@login_required
@solo_admin
@require_POST
def encuesta_recordatorio(request, pk):
    """Enviar recordatorio (notificación in-app) a quienes no han respondido."""
    enc = get_object_or_404(Encuesta, pk=pk)

    if enc.estado != 'ACTIVA':
        return JsonResponse({'ok': False, 'error': 'Solo se pueden enviar recordatorios a encuestas activas.'})

    # Calcular universo
    universo = Personal.objects.filter(estado='Activo')
    if enc.aplica_grupos:
        universo = universo.filter(grupo_tareo=enc.aplica_grupos)
    if enc.aplica_areas.exists():
        universo = universo.filter(subarea__area__in=enc.aplica_areas.all())

    # Quitar los que ya respondieron
    if enc.anonima:
        # En encuesta anónima no se puede cruzar personal → enviar a todo el universo
        pendientes = list(universo)
        aviso = ' (encuesta anónima: se notifica a todo el universo)'
    else:
        ya_respondieron = set(
            enc.respuestas.exclude(personal=None).values_list('personal_id', flat=True)
        )
        pendientes = [p for p in universo if p.pk not in ya_respondieron]
        aviso = ''

    if not pendientes:
        return JsonResponse({'ok': True, 'enviados': 0, 'msg': 'Todos ya respondieron. No hay pendientes.'})

    from comunicaciones.services import NotificacionService
    from django.urls import reverse

    asunto = f'Recordatorio: "{enc.titulo}"'
    url_encuesta = request.build_absolute_uri(reverse('responder_encuesta', args=[enc.pk]))
    cuerpo = (
        f'<p>Te recordamos que tienes una encuesta pendiente de responder.</p>'
        f'<p><strong>{enc.titulo}</strong></p>'
        f'<p>Fecha límite: <strong>{enc.fecha_fin:%d/%m/%Y}</strong></p>'
        f'<p><a href="{url_encuesta}" style="color:#0f766e;">Responder ahora »</a></p>'
    )

    enviados = 0
    errores = 0
    for p in pendientes:
        try:
            NotificacionService.enviar(p, asunto, cuerpo, tipo='IN_APP')
            enviados += 1
        except Exception:
            errores += 1

    return JsonResponse({
        'ok': True,
        'enviados': enviados,
        'errores': errores,
        'msg': f'{enviados} recordatorio(s) enviado(s){aviso}.',
    })
