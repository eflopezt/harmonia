"""Generate Consolidado Roster Excel - same format as example files.
Summary uses COUNTIF formulas referencing the monthly code rows below."""
import json
import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

with open(r'C:\Users\EDWIN LOPEZ\Downloads\roster_data.json') as f:
    all_data = json.load(f)

PERSONAL_INFO = {
    '21811280': {'nombre': 'MENDOZA GARCIA, JULIO CESAR', 'area': 'RESIDENCIA', 'regimen': '21x7'},
    '10105707': {'nombre': 'CASTRO LLANOS, CARLOS RAUL', 'area': 'GERENCIA', 'regimen': '21x7'},
}

DIAS_SEM = ['L', 'M', 'X', 'J', 'V', 'S', 'D']
MESES_INFO = [
    ('Enero',   1, 31),
    ('Febrero', 2, 28),
    ('Marzo',   3, 31),
    ('Abril',   4, 30),
    ('Mayo',    5, 31),
]

# Styles
FILL_HEADER = PatternFill('solid', fgColor='0F766E')
FILL_SUBHEADER = PatternFill('solid', fgColor='1A2B47')
FONT_WHITE_B = Font(bold=True, color='FFFFFF', size=9)
FONT_TITLE = Font(bold=True, size=14, color='0F766E')
FONT_SMALL = Font(size=9)
FONT_BOLD9 = Font(bold=True, size=9)
THIN = Border(
    left=Side('thin', 'D0D0D0'), right=Side('thin', 'D0D0D0'),
    top=Side('thin', 'D0D0D0'), bottom=Side('thin', 'D0D0D0'),
)
CENTER = Alignment(horizontal='center', vertical='center')

CODE_FILLS = {
    'T':   PatternFill('solid', fgColor='C6EFCE'),
    'DL':  PatternFill('solid', fgColor='BDD7EE'),
    'V':   PatternFill('solid', fgColor='D9E2F3'),
    'FA':  PatternFill('solid', fgColor='FFC7CE'),
    'DM':  PatternFill('solid', fgColor='FFE699'),
    'SS':  PatternFill('solid', fgColor='E2EFDA'),
    'DS':  PatternFill('solid', fgColor='F2F2F2'),
    'TR':  PatternFill('solid', fgColor='E8D5F5'),
    'LCG': PatternFill('solid', fgColor='DAEEF3'),
    'CHE': PatternFill('solid', fgColor='FFE0B2'),
    'F':   PatternFill('solid', fgColor='FFC7CE'),
    'I':   PatternFill('solid', fgColor='FFC7CE'),
}


def display_code(cod):
    return {'A': 'T', 'NOR': 'T', 'VAC': 'V', 'DLA': 'DL', 'FA': 'I'}.get(cod, cod)


def cell(ws, r, c, val, font=None, fill=None, alignment=CENTER, border=THIN):
    cl = ws.cell(row=r, column=c, value=val)
    if font: cl.font = font
    if fill: cl.fill = fill
    if alignment: cl.alignment = alignment
    if border: cl.border = border
    return cl


def col_letter(c):
    return get_column_letter(c)


for dni, info in PERSONAL_INFO.items():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Consolidado'

    data = all_data[dni]
    registros = data['registros']
    papeletas = data['papeletas']

    # ── Row 1: Title ──
    ws.merge_cells('A1:AF1')
    ws['A1'] = 'Consolidado Roster Ene-May 2026'
    ws['A1'].font = FONT_TITLE

    # ── Row 2: Employee info ──
    ws.merge_cells('A2:AF2')
    ws['A2'] = (f"{info['nombre']}  |  DNI: {dni}  |  "
                f"{info['area']}  |  Regimen {info['regimen']}")
    ws['A2'].font = Font(bold=True, size=11, color='555555')

    # ── Row 4: Balance header ──
    ws.merge_cells('A4:AF4')
    ws['A4'] = 'Balance de Dias Libres (1 DL por cada 3 T trabajados)'
    ws['A4'].font = Font(bold=True, size=10, color='0F766E')

    # ── Row 5: Column headers ──
    for i, h in enumerate(['Concepto', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'TOTAL']):
        cell(ws, 5, i+1, h, FONT_WHITE_B, FILL_HEADER)

    # First we need to know where each month's code row will land.
    # Structure per month: title(1) + dia_sem(1) + dia(1) + codigo(1) = 4 rows + 1 blank
    # Months start at row 31
    GRID_START = 31
    code_rows = {}  # mes_num -> row of the "Codigo" row
    r = GRID_START
    for mes_name, mes_num, num_days in MESES_INFO:
        # title row = r, dia_sem = r+1, dia = r+2, codigo = r+3
        code_rows[mes_num] = r + 3
        r += 5  # 4 rows + 1 blank

    # Helper: range string for COUNTIF for a given month
    def code_range(mes_num):
        num_days = calendar.monthrange(2026, mes_num)[1]
        cr = code_rows[mes_num]
        return f'B{cr}:{col_letter(num_days + 1)}{cr}'

    # ── Rows 6-14: Balance formulas ──
    # Row 6: Saldo DL al 31/12/2025
    cell(ws, 6, 1, 'Saldo DL al 31/12/2025', FONT_BOLD9)
    for i in range(5):
        cell(ws, 6, i+2, 0, FONT_SMALL)

    # Row 7: T Trabajados = COUNTIF from code row
    cell(ws, 7, 1, 'T Trabajados en el mes', FONT_BOLD9)
    for i, (_, mn, _) in enumerate(MESES_INFO):
        cell(ws, 7, i+2, f'=B{18 + 0}', FONT_SMALL)  # placeholder, will use row 18
    # Actually use references to Resumen row 18
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=7, column=i+2).value = f'={col}18'
    cell(ws, 7, 7, '=SUM(B7:F7)', FONT_BOLD9)

    # Row 8: DL Ganados
    cell(ws, 8, 1, 'DL Ganados (T/3)', FONT_BOLD9)
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=8, column=i+2).value = f'=ROUND({col}7/3,2)'
        ws.cell(row=8, column=i+2).font = FONT_SMALL
        ws.cell(row=8, column=i+2).alignment = CENTER
        ws.cell(row=8, column=i+2).border = THIN
    cell(ws, 8, 7, '=SUM(B8:F8)', FONT_BOLD9)

    # Row 9: DL Usados
    cell(ws, 9, 1, 'DL Usados en el mes', FONT_BOLD9)
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=9, column=i+2).value = f'={col}19'
        ws.cell(row=9, column=i+2).font = FONT_SMALL
        ws.cell(row=9, column=i+2).alignment = CENTER
        ws.cell(row=9, column=i+2).border = THIN
    cell(ws, 9, 7, '=SUM(B9:F9)', FONT_BOLD9)

    # Row 10: Vacaciones usadas
    cell(ws, 10, 1, 'Vacaciones usadas', FONT_BOLD9)
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=10, column=i+2).value = f'={col}20'
        ws.cell(row=10, column=i+2).font = FONT_SMALL
        ws.cell(row=10, column=i+2).alignment = CENTER
        ws.cell(row=10, column=i+2).border = THIN
    cell(ws, 10, 7, '=SUM(B10:F10)', FONT_BOLD9)

    # Row 11: DL Ganados acumulado
    cell(ws, 11, 1, 'DL Ganados acumulado', FONT_BOLD9)
    ws.cell(row=11, column=2).value = '=B8'
    for i in range(1, 5):
        col = col_letter(i + 2)
        prev = col_letter(i + 1)
        ws.cell(row=11, column=i+2).value = f'={prev}11+{col}8'
    for i in range(5):
        ws.cell(row=11, column=i+2).font = FONT_SMALL
        ws.cell(row=11, column=i+2).alignment = CENTER
        ws.cell(row=11, column=i+2).border = THIN

    # Row 12: DL Usados acumulado
    cell(ws, 12, 1, 'DL Usados acumulado', FONT_BOLD9)
    ws.cell(row=12, column=2).value = '=B9'
    for i in range(1, 5):
        col = col_letter(i + 2)
        prev = col_letter(i + 1)
        ws.cell(row=12, column=i+2).value = f'={prev}12+{col}9'
    for i in range(5):
        ws.cell(row=12, column=i+2).font = FONT_SMALL
        ws.cell(row=12, column=i+2).alignment = CENTER
        ws.cell(row=12, column=i+2).border = THIN

    # Row 13: DL PENDIENTES
    cell(ws, 13, 1, 'DL PENDIENTES', Font(bold=True, size=9, color='0F766E'))
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=13, column=i+2).value = f'={col}6+{col}11-{col}12'
        ws.cell(row=13, column=i+2).font = Font(bold=True, size=9, color='0F766E')
        ws.cell(row=13, column=i+2).alignment = CENTER
        ws.cell(row=13, column=i+2).border = THIN

    # Row 14: DL Pendientes redondeado
    cell(ws, 14, 1, 'DL Pendientes (redondeado)', FONT_BOLD9)
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=14, column=i+2).value = f'=INT({col}13)'
        ws.cell(row=14, column=i+2).font = FONT_BOLD9
        ws.cell(row=14, column=i+2).alignment = CENTER
        ws.cell(row=14, column=i+2).border = THIN

    # ── Row 16: Resumen por codigo ──
    ws.merge_cells('A16:G16')
    ws['A16'] = 'Resumen por codigo'
    ws['A16'].font = Font(bold=True, size=10, color='0F766E')

    # Row 17: headers
    for i, h in enumerate(['Concepto', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'TOTAL']):
        cell(ws, 17, i+1, h, FONT_WHITE_B, FILL_HEADER)

    # Rows 18-29: COUNTIF formulas
    codigos_resumen = [
        ('T - Trabajo Presencial', 'T'),
        ('DL - Dia Libre', 'DL'),
        ('V - Vacaciones', 'V'),
        ('TR - Trabajo Remoto', 'TR'),
        ('DM - Descanso Medico', 'DM'),
        ('F - Feriado No Recup.', 'F'),
        ('FC - Feriado Comp.', 'FC'),
        ('P - Permiso', 'P'),
        ('I - Inasistencia', 'I'),
        ('L - Licencia', 'L'),
        ('DOL - Comp. Horario Ext.', 'DOL'),
    ]

    for ci, (label, cod) in enumerate(codigos_resumen):
        r = 18 + ci
        cell(ws, r, 1, label, FONT_BOLD9)
        for mi, (_, mn, nd) in enumerate(MESES_INFO):
            cr = code_rows[mn]
            rng = f'B{cr}:{col_letter(nd + 1)}{cr}'
            ws.cell(row=r, column=mi+2).value = f'=COUNTIF({rng},"{cod}")'
            ws.cell(row=r, column=mi+2).font = FONT_SMALL
            ws.cell(row=r, column=mi+2).alignment = CENTER
            ws.cell(row=r, column=mi+2).border = THIN
        col_g = col_letter(7)
        ws.cell(row=r, column=7).value = f'=SUM(B{r}:F{r})'
        ws.cell(row=r, column=7).font = FONT_BOLD9
        ws.cell(row=r, column=7).alignment = CENTER
        ws.cell(row=r, column=7).border = THIN

    # Row 29: TOTAL DIAS
    r_total = 18 + len(codigos_resumen)
    cell(ws, r_total, 1, 'TOTAL DIAS', FONT_BOLD9)
    for i in range(5):
        col = col_letter(i + 2)
        ws.cell(row=r_total, column=i+2).value = f'=SUM({col}18:{col}{r_total-1})'
        ws.cell(row=r_total, column=i+2).font = FONT_BOLD9
        ws.cell(row=r_total, column=i+2).alignment = CENTER
        ws.cell(row=r_total, column=i+2).border = THIN
    ws.cell(row=r_total, column=7).value = f'=SUM(B{r_total}:F{r_total})'
    ws.cell(row=r_total, column=7).font = FONT_BOLD9
    ws.cell(row=r_total, column=7).alignment = CENTER
    ws.cell(row=r_total, column=7).border = THIN

    # ── Monthly grids starting at row 31 ──
    r = GRID_START
    for mes_name, mes_num, _ in MESES_INFO:
        num_days = calendar.monthrange(2026, mes_num)[1]

        # Title row
        ws.merge_cells(f'A{r}:{col_letter(num_days+1)}{r}')
        ws[f'A{r}'] = f'{mes_name} 2026'
        ws[f'A{r}'].font = Font(bold=True, size=11, color='0F766E')
        r += 1

        # Dia sem row
        cell(ws, r, 1, 'Dia sem.', FONT_WHITE_B, FILL_SUBHEADER)
        for d in range(1, num_days + 1):
            dt = date(2026, mes_num, d)
            cell(ws, r, d+1, DIAS_SEM[dt.weekday()],
                 Font(bold=True, color='FFFFFF', size=8), FILL_SUBHEADER)
        r += 1

        # Dia row
        cell(ws, r, 1, 'Dia', FONT_WHITE_B, FILL_HEADER)
        for d in range(1, num_days + 1):
            cell(ws, r, d+1, d, Font(bold=True, color='FFFFFF', size=9), FILL_HEADER)
        r += 1

        # Codigo row
        cell(ws, r, 1, 'Codigo', FONT_BOLD9)
        for d in range(1, num_days + 1):
            key = date(2026, mes_num, d).isoformat()
            cod = display_code(registros[key]['cod']) if key in registros else ''
            cl = cell(ws, r, d+1, cod, FONT_BOLD9)
            if cod in CODE_FILLS:
                cl.fill = CODE_FILLS[cod]
        r += 2  # blank row

    # Column widths
    ws.column_dimensions['A'].width = 28
    for col in range(2, 33):
        ws.column_dimensions[get_column_letter(col)].width = 5.2

    apellido = info['nombre'].split(',')[0].strip().replace(' ', '')
    fname = f'Consolidado_Roster_{apellido}_Ene-May2026_v2.xlsx'
    path = rf'C:\Users\EDWIN LOPEZ\Downloads\{fname}'
    wb.save(path)
    print(f'OK: {fname}')
