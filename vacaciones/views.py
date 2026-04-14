"""
Vistas del módulo de Vacaciones y Permisos.
"""
import io
import json
from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST

from personal.models import Personal
from .models import (
    SaldoVacacional, TipoPermiso, SolicitudVacacion,
    SolicitudPermiso, VentaVacaciones,
)

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')

# Paleta Harmoni
_HEADER_FILL_HEX = '0D2B27'
_ACCENT_HEX = '5EEAD4'
_ALT_ROW_HEX = 'F0FDFA'


# ══════════════════════════════════════════════════════════════
# ADMIN — VACACIONES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def vacaciones_panel(request):
    """Panel principal de vacaciones."""
    solicitudes = SolicitudVacacion.objects.select_related('personal', 'saldo').all()

    estado = request.GET.get('estado', '')
    buscar = request.GET.get('q', '')
    anio = request.GET.get('anio', str(date.today().year))

    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    if buscar:
        solicitudes = solicitudes.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar)
        )
    if anio:
        solicitudes = solicitudes.filter(fecha_inicio__year=int(anio))

    pendientes = SolicitudVacacion.objects.filter(estado='PENDIENTE').count()
    en_goce = SolicitudVacacion.objects.filter(
        estado__in=['APROBADA', 'EN_GOCE'],
        fecha_inicio__lte=date.today(),
        fecha_fin__gte=date.today(),
    ).count()

    # KPI adicionales
    con_saldo = SaldoVacacional.objects.filter(
        estado__in=['PENDIENTE', 'PARCIAL'], dias_pendientes__gt=0
    ).values('personal').distinct().count()
    dias_acumulados = SaldoVacacional.objects.filter(
        estado__in=['PENDIENTE', 'PARCIAL']
    ).aggregate(total=Sum('dias_pendientes'))['total'] or 0

    # Calendario: vacaciones aprobadas del mes actual
    hoy = date.today()
    vac_mes = SolicitudVacacion.objects.select_related('personal').filter(
        estado__in=['APROBADA', 'EN_GOCE'],
        fecha_inicio__year=hoy.year,
        fecha_inicio__month=hoy.month,
    ).order_by('fecha_inicio')

    # ── Próximas vacaciones (30 días) ─────────────────────────────
    en_30d = hoy + timedelta(days=30)
    vac_proximas = SolicitudVacacion.objects.select_related('personal').filter(
        estado__in=['APROBADA'],
        fecha_inicio__gte=hoy,
        fecha_inicio__lte=en_30d,
    ).order_by('fecha_inicio')[:8]
    proximas_30d_count = vac_proximas.count()

    # ── Top 5 áreas con más días de vacaciones (año actual) ────────
    top_areas = []
    try:
        top_areas = list(
            SolicitudVacacion.objects.filter(
                estado__in=['APROBADA', 'COMPLETADA', 'EN_GOCE'],
                fecha_inicio__year=int(anio),
                personal__subarea__area__isnull=False,
            ).values(
                'personal__subarea__area__nombre'
            ).annotate(
                total_dias=Sum('dias_calendario'),
                total_sol=Count('id'),
            ).order_by('-total_dias')[:5]
        )
    except Exception:
        pass

    # ── Distribución por estado (para chart) ──────────────────────
    try:
        dist_estado = list(
            SolicitudVacacion.objects.filter(fecha_inicio__year=int(anio))
            .values('estado')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        dist_chart = json.dumps([
            {'label': d['estado'], 'value': d['total']} for d in dist_estado
        ])
    except Exception:
        dist_chart = '[]'

    context = {
        'titulo': 'Vacaciones',
        'solicitudes': solicitudes[:100],
        'total': solicitudes.count(),
        'filtro_estado': estado,
        'buscar': buscar,
        'anio_filtro': anio,
        'stats': {
            'pendientes': pendientes,
            'en_goce': en_goce,
            'con_saldo': con_saldo,
            'dias_acumulados': dias_acumulados,
            'proximas_30d': proximas_30d_count,
        },
        'vac_mes': vac_mes,
        'mes_nombre': hoy.strftime('%B %Y'),
        'vac_proximas': vac_proximas,
        'top_areas': top_areas,
        'dist_chart': dist_chart,
    }
    return render(request, 'vacaciones/panel.html', context)


@login_required
@solo_admin
def vacacion_crear(request):
    """Registrar solicitud de vacaciones (admin)."""
    if request.method == 'POST':
        try:
            personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
            fecha_inicio = request.POST['fecha_inicio']
            fecha_fin = request.POST['fecha_fin']
            motivo = request.POST.get('motivo', '')

            # Validación: D.Leg. 713 Art. 10 — derecho tras 1 año de servicio
            if personal.fecha_alta:
                dias_servicio = (date.today() - personal.fecha_alta).days
                if dias_servicio < 365:
                    messages.warning(
                        request,
                        f'{personal.apellidos_nombres} tiene {dias_servicio} días de servicio. '
                        f'El derecho a vacaciones se genera tras 1 año completo (D.Leg. 713 Art. 10). '
                        f'Se registra la solicitud pero considere que aún no cumple el año.'
                    )

            solicitud = SolicitudVacacion(
                personal=personal,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                dias_calendario=0,  # Se calcula en save()
                motivo=motivo,
                estado='PENDIENTE',
                solicitado_por=request.user,
            )
            # Buscar saldo vacacional activo
            saldo = SaldoVacacional.objects.filter(
                personal=personal,
                estado__in=['PENDIENTE', 'PARCIAL'],
            ).order_by('periodo_inicio').first()
            if saldo:
                solicitud.saldo = saldo

            solicitud.save()

            from core.audit import log_create
            log_create(request, solicitud, f'Vacaciones registradas: {personal.apellidos_nombres} — {fecha_inicio} al {fecha_fin}')
            messages.success(request, f'Vacaciones registradas para {personal.apellidos_nombres}')
            return redirect('vacaciones_panel')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': 'Registrar Vacaciones',
        'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres'),
    }
    return render(request, 'vacaciones/crear.html', context)


@login_required
@solo_admin
@require_POST
def vacacion_aprobar(request, pk):
    """Aprobar solicitud de vacaciones."""
    solicitud = get_object_or_404(SolicitudVacacion, pk=pk)
    if solicitud.estado in ('BORRADOR', 'PENDIENTE'):
        try:
            solicitud.aprobar(request.user)
        except ValueError as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)

        from core.audit import log_update
        log_update(request, solicitud, {'estado': {'old': 'PENDIENTE', 'new': 'APROBADA'}},
                   f'Vacaciones aprobadas: {solicitud.personal.apellidos_nombres}')
        return JsonResponse({'ok': True, 'estado': 'APROBADA'})
    return JsonResponse({'ok': False, 'error': 'No se puede aprobar en este estado.'})


@login_required
@solo_admin
@require_POST
def vacacion_rechazar(request, pk):
    """Rechazar solicitud de vacaciones."""
    solicitud = get_object_or_404(SolicitudVacacion, pk=pk)
    if solicitud.estado in ('BORRADOR', 'PENDIENTE'):
        motivo = request.POST.get('motivo', '')
        solicitud.rechazar(request.user, motivo)

        from core.audit import log_update
        log_update(request, solicitud, {'estado': {'old': 'PENDIENTE', 'new': 'RECHAZADA'}},
                   f'Vacaciones rechazadas: {solicitud.personal.apellidos_nombres}')
        return JsonResponse({'ok': True, 'estado': 'RECHAZADA'})
    return JsonResponse({'ok': False, 'error': 'No se puede rechazar en este estado.'})


# ══════════════════════════════════════════════════════════════
# ADMIN — PERMISOS / LICENCIAS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def permisos_panel(request):
    """Panel de solicitudes de permisos/licencias."""
    solicitudes = SolicitudPermiso.objects.select_related('personal', 'tipo').all()

    estado = request.GET.get('estado', '')
    tipo_id = request.GET.get('tipo', '')
    buscar = request.GET.get('q', '')

    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    if tipo_id:
        solicitudes = solicitudes.filter(tipo_id=tipo_id)
    if buscar:
        solicitudes = solicitudes.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar)
        )

    hoy_perm = date.today()
    inicio_mes_perm = hoy_perm.replace(day=1)
    perm_stats = SolicitudPermiso.objects.aggregate(
        pendientes_total=Count('id', filter=Q(estado='PENDIENTE')),
        aprobadas_mes=Count('id', filter=Q(estado='APROBADA', fecha_inicio__gte=inicio_mes_perm)),
        rechazadas_mes=Count('id', filter=Q(estado='RECHAZADA', fecha_inicio__gte=inicio_mes_perm)),
        dias_aprobados_mes=Sum('dias', filter=Q(estado='APROBADA', fecha_inicio__gte=inicio_mes_perm)),
    )

    context = {
        'titulo': 'Permisos y Licencias',
        'solicitudes': solicitudes[:100],
        'total': solicitudes.count(),
        'filtro_estado': estado,
        'filtro_tipo': tipo_id,
        'buscar': buscar,
        'tipos': TipoPermiso.objects.filter(activo=True),
        'pendientes': perm_stats['pendientes_total'] or 0,
        'perm_stats': perm_stats,
    }
    return render(request, 'vacaciones/permisos_panel.html', context)


@login_required
@solo_admin
@require_POST
def permiso_aprobar(request, pk):
    """Aprobar solicitud de permiso."""
    solicitud = get_object_or_404(SolicitudPermiso, pk=pk)
    if solicitud.estado in ('BORRADOR', 'PENDIENTE'):
        solicitud.aprobar(request.user)

        from core.audit import log_update
        log_update(request, solicitud, {'estado': {'old': 'PENDIENTE', 'new': 'APROBADA'}},
                   f'Permiso aprobado: {solicitud.tipo.nombre} — {solicitud.personal.apellidos_nombres}')
        return JsonResponse({'ok': True, 'estado': 'APROBADA'})
    return JsonResponse({'ok': False, 'error': 'No se puede aprobar en este estado.'})


@login_required
@solo_admin
@require_POST
def permiso_rechazar(request, pk):
    """Rechazar solicitud de permiso."""
    solicitud = get_object_or_404(SolicitudPermiso, pk=pk)
    if solicitud.estado in ('BORRADOR', 'PENDIENTE'):
        motivo = request.POST.get('motivo', '')
        solicitud.rechazar(request.user, motivo)
        return JsonResponse({'ok': True, 'estado': 'RECHAZADA'})
    return JsonResponse({'ok': False, 'error': 'No se puede rechazar en este estado.'})


# ══════════════════════════════════════════════════════════════
# ADMIN — SALDOS VACACIONALES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def saldos_panel(request):
    """Panel de saldos vacacionales."""
    saldos = SaldoVacacional.objects.select_related('personal').all()

    buscar = request.GET.get('q', '')
    estado = request.GET.get('estado', '')

    if buscar:
        saldos = saldos.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar)
        )
    if estado:
        saldos = saldos.filter(estado=estado)

    # KPI aggregates (computed on full unfiltered queryset for summary context)
    all_saldos = SaldoVacacional.objects.all()
    kpi = all_saldos.aggregate(
        total_pendientes=Sum('dias_pendientes', filter=Q(estado__in=['PENDIENTE', 'PARCIAL'])),
        total_gozados=Sum('dias_gozados'),
        empleados_con_saldo=Count('personal', distinct=True, filter=Q(estado__in=['PENDIENTE', 'PARCIAL'])),
        empleados_vencidos=Count('id', filter=Q(estado='VENCIDO')),
    )

    context = {
        'titulo': 'Saldos Vacacionales',
        'saldos': saldos[:200],
        'total': saldos.count(),
        'buscar': buscar,
        'filtro_estado': estado,
        'kpi': kpi,
    }
    return render(request, 'vacaciones/saldos_panel.html', context)


@login_required
@solo_admin
@require_POST
def saldo_generar_masivo(request):
    """
    Genera/actualiza saldos vacacionales para todo el personal activo.

    Reglas (D.Leg. 713):
    - Art. 10: Derecho a vacaciones tras 1 año completo de servicio continuo.
    - Art. 22: Vacaciones truncas proporcionales al cese (30/12 × meses).
    - dias_derecho = 30 para año completo, proporcional para periodo incompleto.
    - dias_gozados se sincroniza con:
      a) SolicitudVacacion con estado APROBADA en el período
      b) RegistroTareo con codigo_dia='VAC' en el período
      c) RegistroPapeleta con tipo_permiso VACACIONES estado APROBADA
    """
    from asistencia.models import RegistroTareo

    personal_qs = Personal.objects.filter(estado='Activo', fecha_alta__isnull=False)
    creados = 0
    actualizados = 0
    hoy = date.today()

    from .models import SolicitudVacacion

    for emp in personal_qs:
        f_alta = emp.fecha_alta
        if isinstance(f_alta, str):
            try:
                f_alta = date.fromisoformat(f_alta[:10])
            except ValueError:
                continue

        # Generar TODOS los períodos desde la fecha de alta hasta hoy
        dias_servicio = (hoy - f_alta).days
        anios_completos = dias_servicio // 365

        # PASO 1: Calcular total de días gozados del empleado (de todas las fuentes)
        total_dias_gozados_emp = 0

        # Fuente 1: SolicitudVacacion APROBADA (todas)
        vac_sol = SolicitudVacacion.objects.filter(
            personal=emp, estado='APROBADA',
        ).aggregate(total=Sum('dias_calendario'))
        total_dias_gozados_emp += vac_sol['total'] or 0

        # Fuente 2: RegistroTareo VAC (solo si no hay solicitudes)
        if total_dias_gozados_emp == 0:
            total_dias_gozados_emp = RegistroTareo.objects.filter(
                personal=emp, codigo_dia='VAC',
            ).count()

        # PASO 2: Distribuir días gozados FIFO (período más antiguo primero)
        dias_por_distribuir = total_dias_gozados_emp
        periodos_data = []

        for anio_idx in range(anios_completos + 1):
            try:
                periodo_inicio = f_alta.replace(year=f_alta.year + anio_idx)
            except ValueError:
                periodo_inicio = f_alta.replace(year=f_alta.year + anio_idx, day=28)
            try:
                periodo_fin = periodo_inicio.replace(year=periodo_inicio.year + 1) - timedelta(days=1)
            except ValueError:
                periodo_fin = periodo_inicio.replace(year=periodo_inicio.year + 1, day=28) - timedelta(days=1)

            # Días de derecho
            if periodo_fin <= hoy:
                dias_derecho = 30
            else:
                dias_en_periodo = (hoy - periodo_inicio).days
                meses_en_periodo = dias_en_periodo / 30.0
                dias_derecho = min(30, round(30 / 12 * meses_en_periodo))

            # Días vendidos existentes
            dias_vendidos = 0
            existing = SaldoVacacional.objects.filter(
                personal=emp, periodo_inicio=periodo_inicio
            ).first()
            if existing:
                dias_vendidos = existing.dias_vendidos or 0

            periodos_data.append({
                'periodo_inicio': periodo_inicio,
                'periodo_fin': periodo_fin,
                'dias_derecho': dias_derecho,
                'dias_vendidos': dias_vendidos,
            })

        # PASO 3: Asignar gozados FIFO — primero al período más antiguo
        for pd in periodos_data:
            disponible = pd['dias_derecho'] - pd['dias_vendidos']
            if disponible <= 0:
                pd['dias_gozados'] = 0
            elif dias_por_distribuir >= disponible:
                pd['dias_gozados'] = disponible
                dias_por_distribuir -= disponible
            else:
                pd['dias_gozados'] = dias_por_distribuir
                dias_por_distribuir = 0

            pd['dias_pendientes'] = max(0, pd['dias_derecho'] - pd['dias_gozados'] - pd['dias_vendidos'])

            if pd['dias_gozados'] >= pd['dias_derecho'] and pd['dias_derecho'] > 0:
                pd['estado'] = 'GOZADO'
            elif pd['dias_gozados'] > 0:
                pd['estado'] = 'PARCIAL'
            else:
                pd['estado'] = 'PENDIENTE'

        # PASO 4: Guardar
        for pd in periodos_data:
            saldo, created = SaldoVacacional.objects.update_or_create(
                personal=emp,
                periodo_inicio=pd['periodo_inicio'],
                defaults={
                    'periodo_fin': pd['periodo_fin'],
                    'dias_derecho': pd['dias_derecho'],
                    'dias_gozados': pd['dias_gozados'],
                    'dias_vendidos': pd['dias_vendidos'],
                    'dias_pendientes': pd['dias_pendientes'],
                    'estado': pd['estado'],
                }
            )
            if created:
                creados += 1
            else:
                actualizados += 1

    return JsonResponse({
        'ok': True,
        'creados': creados,
        'actualizados': actualizados,
        'mensaje': f'{creados} nuevos, {actualizados} actualizados',
    })


@login_required
@solo_admin
def saldos_exportar_excel(request):
    """
    Exporta saldos vacacionales a Excel con estilo Harmoni.
    Columnas: DNI, Nombre, Área, Fecha Ingreso, Días Ganados, Días Tomados, Saldo Disponible.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl no está instalado.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Saldos Vacacionales'

    # Estilos
    header_fill = PatternFill('solid', fgColor=_HEADER_FILL_HEX)
    header_font = Font(color='FFFFFF', bold=True, size=10)
    alt_fill = PatternFill('solid', fgColor=_ALT_ROW_HEX)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells('A1:G1')
    titulo_cell = ws['A1']
    titulo_cell.value = f'Reporte de Saldos Vacacionales — Generado {date.today().strftime("%d/%m/%Y")}'
    titulo_cell.font = Font(color='FFFFFF', bold=True, size=12)
    titulo_cell.fill = PatternFill('solid', fgColor=_HEADER_FILL_HEX)
    titulo_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # Cabeceras
    headers = [
        'DNI', 'Apellidos y Nombres', 'Área', 'Fecha Ingreso',
        'Días Ganados', 'Días Tomados', 'Saldo Disponible',
    ]
    col_widths = [12, 42, 30, 16, 14, 14, 16]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[2].height = 18

    # Datos: un registro por empleado (sumando todos sus períodos activos)
    # Usamos el saldo más reciente activo por empleado
    saldos_qs = (
        SaldoVacacional.objects
        .select_related('personal', 'personal__subarea', 'personal__subarea__area')
        .filter(estado__in=['PENDIENTE', 'PARCIAL', 'GOZADO', 'TRUNCO'])
        .order_by('personal__apellidos_nombres', '-periodo_inicio')
    )

    # Agrupar por empleado: sumar días de todos los períodos
    from collections import defaultdict
    emp_data = defaultdict(lambda: {
        'nro_doc': '', 'nombre': '', 'area': '', 'fecha_alta': None,
        'dias_derecho': 0, 'dias_gozados': 0, 'dias_pendientes': 0,
    })
    for s in saldos_qs:
        p = s.personal
        key = p.pk
        d = emp_data[key]
        d['nro_doc'] = p.nro_doc
        d['nombre'] = p.apellidos_nombres
        d['area'] = (
            p.subarea.area.nombre if p.subarea and p.subarea.area else
            (p.subarea.nombre if p.subarea else '—')
        )
        d['fecha_alta'] = p.fecha_alta
        d['dias_derecho'] += s.dias_derecho
        d['dias_gozados'] += s.dias_gozados
        d['dias_pendientes'] += s.dias_pendientes

    row_num = 3
    for idx, (pk, d) in enumerate(emp_data.items()):
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        fecha_str = d['fecha_alta'].strftime('%d/%m/%Y') if d['fecha_alta'] else '—'
        row_data = [
            d['nro_doc'],
            d['nombre'],
            d['area'],
            fecha_str,
            d['dias_derecho'],
            d['dias_gozados'],
            d['dias_pendientes'],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = border
            cell.fill = fill
            cell.font = Font(size=9)
            # Numéricos centrados, texto a la izquierda
            if col_idx in (1, 5, 6, 7):
                cell.alignment = center
            else:
                cell.alignment = left
            # Resaltar saldo disponible bajo
            if col_idx == 7 and isinstance(value, int) and value <= 5:
                cell.font = Font(size=9, bold=True, color='CC0000')
        row_num += 1

    # Ancho de columnas
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fila de totales
    ws.cell(row=row_num, column=4, value='TOTALES').font = Font(bold=True, size=9)
    ws.cell(row=row_num, column=4).alignment = center
    for col_idx in (5, 6, 7):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row_num, column=col_idx,
                value=f'=SUM({col_letter}3:{col_letter}{row_num - 1})')
        ws.cell(row=row_num, column=col_idx).font = Font(bold=True, size=9)
        ws.cell(row=row_num, column=col_idx).alignment = center
        ws.cell(row=row_num, column=col_idx).fill = PatternFill('solid', fgColor='134E4A')
        ws.cell(row=row_num, column=col_idx).font = Font(bold=True, size=9, color='FFFFFF')

    # Freeze header
    ws.freeze_panes = 'A3'

    # Respuesta HTTP
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'saldos_vacacionales_{date.today().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@solo_admin
def saldo_detalle(request, pk):
    """
    Detalle de saldos vacacionales de un empleado.
    Muestra TODOS los períodos con desglose por período:
    solicitudes, días VAC del tareo, adelanto máximo global.
    """
    from asistencia.models import RegistroTareo
    from collections import defaultdict

    saldo = get_object_or_404(
        SaldoVacacional.objects.select_related('personal'),
        pk=pk
    )
    personal = saldo.personal
    hoy = date.today()

    # ── TODOS los períodos del empleado ──
    todos_saldos = SaldoVacacional.objects.filter(
        personal=personal
    ).order_by('periodo_inicio')

    # Para cada período: solicitudes + días VAC del tareo
    periodos_detalle = []
    total_derecho = 0
    total_gozados = 0
    total_vendidos = 0
    total_pendientes = 0

    for s in todos_saldos:
        tope = min(s.periodo_fin, hoy)

        # Solicitudes del período
        solicitudes = SolicitudVacacion.objects.filter(
            personal=personal,
            fecha_inicio__gte=s.periodo_inicio,
            fecha_inicio__lte=tope,
        ).exclude(estado='ANULADA').order_by('fecha_inicio')

        # Días VAC del tareo
        dias_vac = list(
            RegistroTareo.objects.filter(
                personal=personal,
                codigo_dia='VAC',
                fecha__gte=s.periodo_inicio,
                fecha__lte=tope,
            ).values_list('fecha', flat=True).order_by('fecha')
        )

        periodos_detalle.append({
            'saldo': s,
            'solicitudes': solicitudes,
            'dias_vac': dias_vac,
            'total_dias_vac': len(dias_vac),
            'es_actual': s.pk == saldo.pk,
        })

        total_derecho += s.dias_derecho
        total_gozados += s.dias_gozados
        total_vendidos += s.dias_vendidos or 0
        total_pendientes += s.dias_pendientes

    # ── Adelanto máximo GLOBAL (total acumulado que ha generado) ──
    dias_servicio = (hoy - personal.fecha_alta).days if personal.fecha_alta else 0
    meses_servicio = dias_servicio / 30.0
    # Total generado proporcional a toda la antigüedad
    total_generado = min(dias_servicio // 365 * 30 + min(30, round(30 / 12 * (meses_servicio % 12))), dias_servicio // 365 * 30 + 30)
    # Más simple: suma de todos los dias_derecho
    total_generado = total_derecho
    adelanto_max_global = max(0, total_generado - total_gozados - total_vendidos)

    cumple_anio = dias_servicio >= 365
    fecha_cumple_anio = personal.fecha_alta + timedelta(days=365) if personal.fecha_alta else None

    context = {
        'saldo': saldo,
        'personal': personal,
        'periodos_detalle': periodos_detalle,
        'total_periodos': len(periodos_detalle),
        'total_derecho': total_derecho,
        'total_gozados': total_gozados,
        'total_vendidos': total_vendidos,
        'total_pendientes': total_pendientes,
        'adelanto_max_global': adelanto_max_global,
        'cumple_anio': cumple_anio,
        'fecha_cumple_anio': fecha_cumple_anio,
        'dias_servicio': dias_servicio,
        'meses_servicio': round(meses_servicio, 1),
    }
    return render(request, 'vacaciones/saldo_detalle.html', context)


@login_required
@solo_admin
def vacaciones_calendario(request):
    """Vista de calendario mensual de vacaciones aprobadas."""
    hoy = date.today()
    try:
        anio = int(request.GET.get('anio', hoy.year))
        mes = int(request.GET.get('mes', hoy.month))
    except (ValueError, TypeError):
        anio, mes = hoy.year, hoy.month

    import calendar as cal_mod
    # Primero y último día del mes
    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, cal_mod.monthrange(anio, mes)[1])

    vac_aprobadas = SolicitudVacacion.objects.select_related(
        'personal', 'personal__subarea', 'personal__subarea__area'
    ).filter(
        estado__in=['APROBADA', 'EN_GOCE', 'COMPLETADA'],
        fecha_inicio__lte=ultimo_dia,
        fecha_fin__gte=primer_dia,
    ).order_by('fecha_inicio', 'personal__apellidos_nombres')

    # Mes anterior / siguiente para navegación
    if mes == 1:
        mes_ant, anio_ant = 12, anio - 1
    else:
        mes_ant, anio_ant = mes - 1, anio
    if mes == 12:
        mes_sig, anio_sig = 1, anio + 1
    else:
        mes_sig, anio_sig = mes + 1, anio

    nombre_mes = primer_dia.strftime('%B %Y').capitalize()

    meses_lista = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre'),
    ]

    context = {
        'titulo': f'Calendario de Vacaciones — {nombre_mes}',
        'vac_aprobadas': vac_aprobadas,
        'anio': anio,
        'mes': mes,
        'nombre_mes': nombre_mes,
        'mes_ant': mes_ant,
        'anio_ant': anio_ant,
        'mes_sig': mes_sig,
        'anio_sig': anio_sig,
        'total': vac_aprobadas.count(),
        'meses_lista': meses_lista,
    }
    return render(request, 'vacaciones/calendario.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — TIPOS DE PERMISO (CONFIG)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def tipos_permiso(request):
    """Configuración de tipos de permiso."""
    tipos = TipoPermiso.objects.all()
    context = {
        'titulo': 'Tipos de Permiso',
        'tipos': tipos,
    }
    return render(request, 'vacaciones/tipos_permiso.html', context)


@login_required
@solo_admin
@require_POST
def tipo_permiso_crear(request):
    """Crear tipo de permiso."""
    try:
        t = TipoPermiso.objects.create(
            nombre=request.POST['nombre'],
            codigo=request.POST['codigo'],
            descripcion=request.POST.get('descripcion', ''),
            base_legal=request.POST.get('base_legal', ''),
            dias_max=int(request.POST.get('dias_max', 0) or 0),
            pagado=request.POST.get('pagado') == 'on',
            requiere_sustento=request.POST.get('requiere_sustento') == 'on',
            descuenta_vacaciones=request.POST.get('descuenta_vacaciones') == 'on',
        )
        return JsonResponse({'ok': True, 'pk': t.pk, 'nombre': str(t)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def tipo_permiso_editar(request, pk):
    """Editar tipo de permiso."""
    t = get_object_or_404(TipoPermiso, pk=pk)
    try:
        t.nombre = request.POST.get('nombre', t.nombre)
        t.codigo = request.POST.get('codigo', t.codigo)
        t.descripcion = request.POST.get('descripcion', t.descripcion)
        t.base_legal = request.POST.get('base_legal', t.base_legal)
        t.dias_max = int(request.POST.get('dias_max', t.dias_max) or 0)
        t.pagado = request.POST.get('pagado') == 'on'
        t.requiere_sustento = request.POST.get('requiere_sustento') == 'on'
        t.descuenta_vacaciones = request.POST.get('descuenta_vacaciones') == 'on'
        t.activo = request.POST.get('activo', 'on') == 'on'
        t.save()
        return JsonResponse({'ok': True, 'pk': t.pk, 'nombre': str(t)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


# ══════════════════════════════════════════════════════════════
# PORTAL DEL TRABAJADOR
# ══════════════════════════════════════════════════════════════

@login_required
def mis_vacaciones(request):
    """Portal: mis vacaciones y saldos."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    saldos = []
    solicitudes = []
    if empleado:
        saldos = SaldoVacacional.objects.filter(personal=empleado).order_by('-periodo_inicio')
        solicitudes = SolicitudVacacion.objects.filter(personal=empleado).exclude(estado='ANULADA')

    context = {
        'titulo': 'Mis Vacaciones',
        'empleado': empleado,
        'saldos': saldos,
        'solicitudes': solicitudes,
    }
    return render(request, 'vacaciones/mis_vacaciones.html', context)


@login_required
def vacacion_solicitar(request):
    """Portal: solicitar vacaciones."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    if not empleado:
        messages.error(request, 'No tienes un perfil de empleado vinculado.')
        return redirect('mis_vacaciones')

    if request.method == 'POST':
        try:
            fecha_inicio = request.POST['fecha_inicio']
            fecha_fin = request.POST['fecha_fin']
            motivo = request.POST.get('motivo', '')

            saldo = SaldoVacacional.objects.filter(
                personal=empleado,
                estado__in=['PENDIENTE', 'PARCIAL'],
            ).order_by('periodo_inicio').first()

            solicitud = SolicitudVacacion(
                personal=empleado,
                saldo=saldo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                dias_calendario=0,
                motivo=motivo,
                estado='PENDIENTE',
                solicitado_por=request.user,
            )
            solicitud.save()

            messages.success(request, f'Solicitud de vacaciones registrada ({solicitud.dias_calendario} días)')
            return redirect('mis_vacaciones')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    saldo_actual = SaldoVacacional.objects.filter(
        personal=empleado,
        estado__in=['PENDIENTE', 'PARCIAL'],
    ).order_by('periodo_inicio').first()

    context = {
        'titulo': 'Solicitar Vacaciones',
        'empleado': empleado,
        'saldo_actual': saldo_actual,
    }
    return render(request, 'vacaciones/solicitar.html', context)


@login_required
def mis_permisos(request):
    """Portal: mis permisos/licencias."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    solicitudes = []
    if empleado:
        solicitudes = SolicitudPermiso.objects.filter(
            personal=empleado
        ).select_related('tipo').exclude(estado='ANULADA')

    context = {
        'titulo': 'Mis Permisos',
        'empleado': empleado,
        'solicitudes': solicitudes,
        'tipos': TipoPermiso.objects.filter(activo=True),
    }
    return render(request, 'vacaciones/mis_permisos.html', context)


@login_required
def permiso_solicitar(request):
    """Portal: solicitar permiso."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    if not empleado:
        messages.error(request, 'No tienes un perfil de empleado vinculado.')
        return redirect('mis_permisos')

    if request.method == 'POST':
        try:
            tipo = get_object_or_404(TipoPermiso, pk=request.POST['tipo_id'])
            solicitud = SolicitudPermiso(
                personal=empleado,
                tipo=tipo,
                fecha_inicio=request.POST['fecha_inicio'],
                fecha_fin=request.POST['fecha_fin'],
                dias=0,  # Se calcula en save()
                motivo=request.POST.get('motivo', ''),
                estado='PENDIENTE',
                solicitado_por=request.user,
            )
            if request.FILES.get('sustento'):
                solicitud.sustento = request.FILES['sustento']
            solicitud.save()

            messages.success(request, f'Solicitud de {tipo.nombre} registrada ({solicitud.dias} días)')
            return redirect('mis_permisos')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': 'Solicitar Permiso',
        'empleado': empleado,
        'tipos': TipoPermiso.objects.filter(activo=True),
    }
    return render(request, 'vacaciones/solicitar_permiso.html', context)


@login_required
@require_POST
def solicitud_anular(request, tipo, pk):
    """Portal: anular solicitud de vacación o permiso."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    if tipo == 'vacacion':
        obj = get_object_or_404(SolicitudVacacion, pk=pk, personal=empleado)
    else:
        obj = get_object_or_404(SolicitudPermiso, pk=pk, personal=empleado)

    if obj.estado in ('BORRADOR', 'PENDIENTE'):
        obj.estado = 'ANULADA'
        obj.save(update_fields=['estado'])
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': 'No se puede anular en este estado.'})


# ─────────────────────────────────────────────────
# VENTA DE VACACIONES (DL 713 Art. 19)
# ─────────────────────────────────────────────────

solo_admin = user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')


@login_required
@solo_admin
def venta_vacaciones_lista(request):
    """Lista de ventas de vacaciones."""
    ventas = VentaVacaciones.objects.select_related(
        'personal', 'saldo', 'aprobado_por'
    ).order_by('-fecha')

    buscar = request.GET.get('q', '').strip()
    anio = request.GET.get('anio', '')
    if buscar:
        ventas = ventas.filter(
            Q(personal__apellidos_nombres__icontains=buscar)
            | Q(personal__nro_doc__icontains=buscar)
        )
    if anio:
        ventas = ventas.filter(fecha__year=int(anio))

    totales = ventas.aggregate(
        total_dias=Sum('dias_vendidos'),
        total_monto=Sum('monto'),
    )

    return render(request, 'vacaciones/venta_lista.html', {
        'ventas': ventas,
        'total': ventas.count(),
        'totales': totales,
        'buscar': buscar,
        'anio': anio,
    })


@login_required
@solo_admin
def venta_vacaciones_crear(request):
    """Crear nueva venta de vacaciones."""
    if request.method == 'POST':
        personal_id = request.POST.get('personal')
        dias = int(request.POST.get('dias_vendidos', 0))

        if dias < 1 or dias > 15:
            messages.error(request, 'Los días a vender deben ser entre 1 y 15.')
            return redirect('venta_vacaciones_crear')

        personal = get_object_or_404(Personal, pk=personal_id)
        saldo = personal.saldos_vacacionales.filter(
            estado__in=['PENDIENTE', 'PARCIAL']
        ).first()

        if not saldo:
            messages.error(request, f'{personal} no tiene saldo vacacional disponible.')
            return redirect('venta_vacaciones_crear')

        if dias > saldo.dias_pendientes:
            messages.error(request, f'Solo tiene {saldo.dias_pendientes} días disponibles.')
            return redirect('venta_vacaciones_crear')

        # Calcular monto: remuneración diaria × días
        rem_diaria = (personal.sueldo_base or Decimal('0')) / Decimal('30')
        monto = rem_diaria * Decimal(str(dias))

        venta = VentaVacaciones.objects.create(
            personal=personal,
            saldo=saldo,
            dias_vendidos=dias,
            monto=monto,
            aprobado_por=request.user,
        )

        # Actualizar saldo
        saldo.dias_vendidos = (saldo.dias_vendidos or 0) + dias
        saldo.recalcular()

        messages.success(
            request,
            f'Venta registrada: {personal.apellidos_nombres} — '
            f'{dias} días por S/ {monto:,.2f}'
        )
        return redirect('venta_vacaciones_lista')

    # GET: mostrar formulario
    empleados = Personal.objects.filter(
        estado='Activo'
    ).order_by('apellidos_nombres')

    return render(request, 'vacaciones/venta_crear.html', {
        'empleados': empleados,
    })
