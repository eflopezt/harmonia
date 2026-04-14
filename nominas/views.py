"""
Nóminas — vistas principales.
Cubre: panel períodos, crear período, detalle, generar, aprobar,
exportar CSV, detalle registro, editar registro, conceptos, mis recibos (portal),
resumen estadístico AJAX, descarga masiva boletas ZIP.
"""
import csv
import io
import json
import logging
import zipfile
from datetime import date
from decimal import Decimal

logger = logging.getLogger('nominas.views')

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET

from .models import (
    ConceptoRemunerativo, LineaNomina, PeriodoNomina, RegistroNomina,
    PresupuestoPlanilla, PlanPlantilla, LineaPlan,
)
from . import engine
from .flujo_caja_engine import proyectar_flujo_caja, proyectar_desde_plan

solo_admin = user_passes_test(lambda u: u.is_superuser)


# ─── Calendario procesos especiales ────────────────────────────────────────────

# Definición del calendario legal peruano de procesos especiales
_SCHEDULE = [
    # tipo,            mes, label,          base legal
    ('GRATIFICACION',  7,  'Gratif. Julio',  'Ley 27735'),
    ('GRATIFICACION', 12,  'Gratif. Diciembre', 'Ley 27735'),
    ('CTS',            5,  'CTS Mayo',       'D.Leg. 650'),
    ('CTS',           11,  'CTS Noviembre',  'D.Leg. 650'),
    ('UTILIDADES',     3,  'Utilidades',     'D.Leg. 892'),
]

_TIPO_META = {
    'GRATIFICACION': {'icon': 'fas fa-gift',              'color_base': '#ccfbf1', 'color_icon': '#0f766e'},
    'CTS':           {'icon': 'fas fa-piggy-bank',        'color_base': '#dbeafe', 'color_icon': '#1d4ed8'},
    'UTILIDADES':    {'icon': 'fas fa-chart-line',        'color_base': '#fef9c3', 'color_icon': '#a16207'},
    'LIQUIDACION':   {'icon': 'fas fa-file-invoice-dollar','color_base': '#fee2e2', 'color_icon': '#b91c1c'},
}

_ESTADO_META = {
    'PENDIENTE':  {'label': 'Por iniciar',  'badge': 'secondary', 'icon': 'far fa-circle'},
    'BORRADOR':   {'label': 'Borrador',     'badge': 'info',      'icon': 'fas fa-pen'},
    'CALCULADO':  {'label': 'Calculado',    'badge': 'primary',   'icon': 'fas fa-calculator'},
    'APROBADO':   {'label': 'Aprobado',     'badge': 'success',   'icon': 'fas fa-check-circle'},
    'CERRADO':    {'label': 'Cerrado',      'badge': 'dark',      'icon': 'fas fa-lock'},
    'ANULADO':    {'label': 'Anulado',      'badge': 'danger',    'icon': 'fas fa-times-circle'},
}


def _urgency(dias):
    """Clasifica urgencia por días restantes."""
    if dias < 0:
        return 'pasado'
    if dias <= 15:
        return 'critico'
    if dias <= 30:
        return 'urgente'
    if dias <= 60:
        return 'proximo'
    return 'planificado'


def _build_procesos_calendar(hoy, ultimo_regular):
    """
    Construye el calendario de procesos especiales para los últimos 3 años + siguiente.
    Devuelve dict {anio: [slot_dict, ...]} ordenado cronológicamente por año DESC.

    Cada slot tiene:
      tipo, label, ley, mes, anio, fecha_pago, dias, urgencia,
      estado, monto, monto_estimado, trabajadores, pk, creado,
      icon, color_base, color_icon, estado_label, estado_badge, estado_icon
    """
    # Estimaciones basadas en último período regular aprobado
    masa = float(ultimo_regular.total_neto or 0) if ultimo_regular else 0
    trab = int(ultimo_regular.total_trabajadores or 0) if ultimo_regular else 0
    estimados = {
        'GRATIFICACION': masa,           # equivalente a 1 sueldo mensual
        'CTS':           masa * 0.5,     # medio sueldo semestral
        'UTILIDADES':    masa * 0.18,    # ~18% de la masa (estimación conservadora)
    }

    # Índice de períodos existentes: (tipo, anio, mes) → PeriodoNomina
    existentes = {
        (p.tipo, p.anio, p.mes): p
        for p in PeriodoNomina.objects.filter(
            tipo__in=['GRATIFICACION', 'CTS', 'UTILIDADES', 'LIQUIDACION']
        ).only('tipo', 'anio', 'mes', 'estado', 'total_neto', 'total_trabajadores', 'pk')
    }

    anios = [hoy.year + 1, hoy.year, hoy.year - 1, hoy.year - 2]
    calendar = {}

    for anio in anios:
        slots = []
        for tipo, mes, label, ley in _SCHEDULE:
            try:
                fecha_pago = date(anio, mes, 15)
            except ValueError:
                fecha_pago = date(anio, mes, 28)

            dias = (fecha_pago - hoy).days
            periodo = existentes.get((tipo, anio, mes))
            meta_tipo   = _TIPO_META.get(tipo, _TIPO_META['GRATIFICACION'])
            monto_est   = estimados.get(tipo, 0)

            if periodo:
                estado = periodo.estado
                monto  = float(periodo.total_neto or 0)
                n_trab = int(periodo.total_trabajadores or trab)
                pk     = periodo.pk
                creado = True
            else:
                # Sólo mostrar años futuros o actuales como "por iniciar"
                # Años pasados sin período: mostrar igual pero como histórico perdido
                estado = 'PENDIENTE'
                monto  = 0.0
                n_trab = trab
                pk     = None
                creado = False

            meta_estado = _ESTADO_META.get(estado, _ESTADO_META['PENDIENTE'])

            slots.append({
                # Identificadores
                'tipo':          tipo,
                'mes':           mes,
                'anio':          anio,
                'label':         label,
                'ley':           ley,
                # Fechas
                'fecha_pago':    fecha_pago,
                'dias':          dias,
                'urgencia':      _urgency(dias),
                # Estado
                'estado':        estado,
                'estado_label':  meta_estado['label'],
                'estado_badge':  meta_estado['badge'],
                'estado_icon':   meta_estado['icon'],
                # Montos
                'monto':         monto,
                'monto_estimado': monto_est,
                'trabajadores':  n_trab,
                # Navegación
                'pk':            pk,
                'creado':        creado,
                # Visual
                'icon':          meta_tipo['icon'],
                'color_base':    meta_tipo['color_base'],
                'color_icon':    meta_tipo['color_icon'],
            })

        # Ordenar: primero los próximos (dias >= 0), luego los pasados
        slots.sort(key=lambda s: s['fecha_pago'])
        calendar[anio] = slots

    return calendar


# ─── Panel principal ───────────────────────────────────────────────────────────

@login_required
@solo_admin
def nominas_panel(request):
    """Lista de períodos + estadísticas + KPIs comparativos."""
    periodos = PeriodoNomina.objects.select_related('generado_por').order_by('-anio', '-mes')

    # Filtros
    anio = request.GET.get('anio', '')
    tipo = request.GET.get('tipo', '')
    if anio:
        periodos = periodos.filter(anio=anio)
    if tipo:
        periodos = periodos.filter(tipo=tipo)

    # Stats globales del último período APROBADO/CERRADO
    ultimo = PeriodoNomina.objects.filter(
        tipo='REGULAR', estado__in=['APROBADO', 'CERRADO']
    ).order_by('-anio', '-mes').first()

    anios_disp = PeriodoNomina.objects.values_list('anio', flat=True).distinct().order_by('-anio')

    # ── KPI 1: Comparativa últimos 3 períodos regulares ───────────────────────
    comparativa_meses = []
    try:
        ultimos_3 = list(
            PeriodoNomina.objects.filter(
                tipo='REGULAR',
                estado__in=['CALCULADO', 'APROBADO', 'CERRADO'],
            ).order_by('-anio', '-mes')[:3]
        )
        for i, p in enumerate(ultimos_3):
            neto = float(p.total_neto or 0)
            variacion_pct = None
            if i < len(ultimos_3) - 1:
                neto_ant = float(ultimos_3[i + 1].total_neto or 0)
                if neto_ant > 0:
                    variacion_pct = round(((neto - neto_ant) / neto_ant) * 100, 1)
            comparativa_meses.append({
                'periodo': str(p),
                'neto': neto,
                'trabajadores': p.total_trabajadores,
                'variacion_pct': variacion_pct,
                'estado': p.estado,
                'color_estado': p.color_estado,
            })
    except Exception:
        comparativa_meses = []

    # ── KPI 2: Distribución de conceptos del último período ───────────────────
    distribucion_conceptos_json = json.dumps({
        'remuneraciones': 0, 'descuentos': 0, 'aportes_empresa': 0,
    })
    try:
        if ultimo:
            agg = ultimo.registros.aggregate(
                remuneraciones=Sum('total_ingresos'),
                descuentos=Sum('total_descuentos'),
                aportes_empresa=Sum('aporte_essalud'),
            )
            distribucion_conceptos_json = json.dumps({
                'remuneraciones': float(agg['remuneraciones'] or 0),
                'descuentos':     float(agg['descuentos'] or 0),
                'aportes_empresa': float(agg['aportes_empresa'] or 0),
            })
    except Exception:
        pass

    # ── KPI 3: Top 5 mayores netos del último período ─────────────────────────
    top_sueldos = []
    try:
        if ultimo:
            top_sueldos = list(
                RegistroNomina.objects.filter(periodo=ultimo)
                .select_related('personal')
                .order_by('-neto_a_pagar')[:5]
            )
    except Exception:
        top_sueldos = []

    # ── KPI 4: Tendencia masa salarial — últimos 6 períodos ──────────────────
    masa_salarial_json = json.dumps([])
    try:
        ultimos_6 = list(
            PeriodoNomina.objects.filter(
                tipo='REGULAR',
                estado__in=['CALCULADO', 'APROBADO', 'CERRADO'],
            ).order_by('-anio', '-mes')[:6]
        )
        # Reverse to display chronologically (oldest → newest)
        ultimos_6.reverse()
        masa_data = [
            {
                'label': f'{p.mes:02d}/{str(p.anio)[2:]}',
                'neto': float(p.total_neto or 0),
            }
            for p in ultimos_6
        ]
        masa_salarial_json = json.dumps(masa_data)
    except Exception:
        pass

    # ── Calendario de Procesos Especiales ────────────────────────────────────
    hoy_d = date.today()
    _cal_dict = _build_procesos_calendar(hoy_d, ultimo)
    # Convertir a lista de dicts para que Django templates pueda iterar sin filtros custom
    cal_anios_list = [hoy_d.year + 1, hoy_d.year, hoy_d.year - 1, hoy_d.year - 2]
    cal_data = [
        {'anio': anio_k, 'slots': _cal_dict[anio_k]}
        for anio_k in cal_anios_list
    ]

    return render(request, 'nominas/panel.html', {
        'titulo': 'Nóminas',
        'periodos': periodos,
        'ultimo': ultimo,
        'tipo_choices': PeriodoNomina.TIPO_CHOICES,
        'estado_choices': PeriodoNomina.ESTADO_CHOICES,
        'anios_disp': anios_disp,
        'anio_filtro': anio,
        'tipo_filtro': tipo,
        # KPIs comparativos
        'comparativa_meses': comparativa_meses,
        'distribucion_conceptos_json': distribucion_conceptos_json,
        'top_sueldos': top_sueldos,
        'masa_salarial_json': masa_salarial_json,
        # Calendario procesos especiales
        'cal_data': cal_data,
        'cal_anio_actual': hoy_d.year,
        'cal_anios': cal_anios_list,
    })


# ─── Período ──────────────────────────────────────────────────────────────────

@login_required
@solo_admin
def periodo_crear(request):
    """Formulario para crear un nuevo período."""
    if request.method == 'POST':
        tipo  = request.POST.get('tipo', 'REGULAR')
        try:
            anio  = int(request.POST.get('anio', timezone.now().year))
            mes   = int(request.POST.get('mes', timezone.now().month))
        except (ValueError, TypeError):
            messages.error(request, 'Año o mes inválido.')
            return redirect('nominas_panel')
        desc  = request.POST.get('descripcion', '')
        fi    = date.fromisoformat(request.POST['fecha_inicio']) if request.POST.get('fecha_inicio') else None
        ff    = date.fromisoformat(request.POST['fecha_fin']) if request.POST.get('fecha_fin') else None
        fp    = date.fromisoformat(request.POST['fecha_pago']) if request.POST.get('fecha_pago') else None

        if PeriodoNomina.objects.filter(tipo=tipo, anio=anio, mes=mes).exists():
            messages.error(request, 'Ya existe un período con ese tipo/año/mes.')
            return redirect('nominas_panel')

        p = PeriodoNomina.objects.create(
            tipo=tipo, anio=anio, mes=mes,
            descripcion=desc,
            fecha_inicio=fi, fecha_fin=ff, fecha_pago=fp,
        )
        messages.success(request, f'Período {p.descripcion or p} creado.')
        return redirect('nominas_periodo_detalle', pk=p.pk)

    # GET
    hoy = timezone.now()
    return render(request, 'nominas/periodo_form.html', {
        'titulo': 'Nuevo período de nómina',
        'tipo_choices': PeriodoNomina.TIPO_CHOICES,
        'anio_actual': hoy.year,
        'mes_actual':  hoy.month,
        'meses': range(1, 13),
        'anios': range(hoy.year - 1, hoy.year + 2),
    })


@login_required
@solo_admin
def periodo_detalle(request, pk):
    """Detalle del período: lista de registros de nómina."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    registros = (
        periodo.registros
        .select_related('personal')
        .order_by('personal__apellidos_nombres')
    )

    # Filtros
    grupo = request.GET.get('grupo', '')
    q = request.GET.get('q', '')
    if grupo:
        registros = registros.filter(grupo=grupo)
    if q:
        registros = registros.filter(personal__apellidos_nombres__icontains=q)

    grupos = periodo.registros.values_list('grupo', flat=True).distinct().order_by('grupo')

    # Calcular totales de las filas visibles (tras filtros server-side)
    agg = registros.aggregate(
        sueldo_base=Sum('sueldo_base'),
        total_ingresos=Sum('total_ingresos'),
        total_descuentos=Sum('total_descuentos'),
        neto_a_pagar=Sum('neto_a_pagar'),
        costo_total_empresa=Sum('costo_total_empresa'),
    )
    totales = {
        'sueldo_base':        agg['sueldo_base'] or Decimal('0'),
        'total_ingresos':     agg['total_ingresos'] or Decimal('0'),
        'total_descuentos':   agg['total_descuentos'] or Decimal('0'),
        'neto_a_pagar':       agg['neto_a_pagar'] or Decimal('0'),
        'costo_total_empresa':agg['costo_total_empresa'] or Decimal('0'),
    }

    return render(request, 'nominas/periodo_detalle.html', {
        'titulo': f'Nómina — {periodo}',
        'periodo': periodo,
        'registros': registros,
        'grupos': grupos,
        'grupo_filtro': grupo,
        'q': q,
        'totales': totales,
        'puede_generar': periodo.estado in ('BORRADOR', 'CALCULADO'),
        'puede_aprobar': periodo.estado == 'CALCULADO',
        'puede_exportar': periodo.estado in ('CALCULADO', 'APROBADO', 'CERRADO'),
    })


@login_required
@solo_admin
@require_POST
def periodo_generar(request, pk):
    """Genera/recalcula todos los registros del período."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    if periodo.estado in ('APROBADO', 'CERRADO', 'ANULADO'):
        messages.error(request, 'No se puede recalcular un período aprobado o cerrado.')
        return redirect('nominas_periodo_detalle', pk=pk)

    grupo = request.POST.get('grupo') or None
    stats = engine.generar_periodo(periodo, usuario=request.user, grupo=grupo)

    if stats.get('errores'):
        messages.warning(
            request,
            f"Generado con {stats['errores']} error(es). "
            f"Generados: {stats['generados']} | Actualizados: {stats['actualizados']}. "
            f"Errores: {'; '.join(stats.get('detalle_errores', [])[:3])}"
        )
    else:
        messages.success(
            request,
            f"Período generado: {stats['generados']} nuevos + "
            f"{stats['actualizados']} actualizados. "
            f"Neto total: S/ {stats['total_neto']:,.2f}"
        )
    return redirect('nominas_periodo_detalle', pk=pk)


@login_required
@solo_admin
@require_POST
def periodo_aprobar(request, pk):
    """Aprueba el período (lo bloquea para edición)."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    if periodo.estado != 'CALCULADO':
        messages.error(request, 'Solo se puede aprobar un período en estado CALCULADO.')
        return redirect('nominas_periodo_detalle', pk=pk)

    periodo.estado = 'APROBADO'
    periodo.aprobado_por = request.user
    periodo.aprobado_en  = timezone.now()
    periodo.save()
    messages.success(request, f'Período {periodo} aprobado correctamente.')
    return redirect('nominas_periodo_detalle', pk=pk)


@login_required
@solo_admin
@require_POST
def periodo_cerrar(request, pk):
    """Cierra el período definitivamente (no permite más cambios)."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    if periodo.estado != 'APROBADO':
        messages.error(request, 'Solo se puede cerrar un período en estado APROBADO.')
        return redirect('nominas_periodo_detalle', pk=pk)

    periodo.estado = 'CERRADO'
    periodo.save(update_fields=['estado'])
    messages.success(request, f'Período {periodo} cerrado definitivamente.')
    return redirect('nominas_periodo_detalle', pk=pk)


@login_required
@solo_admin
def periodo_exportar(request, pk):
    """Exporta el período a CSV compatible con Excel (BOM UTF-8)."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    registros = (
        periodo.registros
        .select_related('personal')
        .prefetch_related('lineas__concepto')
        .order_by('personal__apellidos_nombres')
    )

    output = io.StringIO()
    output.write('\ufeff')  # BOM para Excel
    writer = csv.writer(output)

    # Encabezado
    writer.writerow([
        'DNI', 'Apellidos y Nombres', 'Grupo', 'AFP/ONP', 'AFP',
        'Días', 'Sueldo Base',
        'Sueldo Prop.', 'Asig. Familiar', 'HE 25%', 'HE 35%', 'HE 100%',
        'Otros Ingresos', 'Total Ingresos',
        'AFP Aporte', 'AFP Comisión', 'AFP Seguro', 'ONP',
        'IR 5ta', 'Desc. Préstamo', 'Otros Descuentos', 'Total Descuentos',
        'Neto a Pagar',
        'EsSalud', 'Prov. Gratif.', 'Prov. CTS', 'Costo Empresa',
    ])

    for r in registros:
        lineas_map = {l.concepto.codigo: l for l in r.lineas.all()}

        def _m(codigo):
            l = lineas_map.get(codigo)
            return float(l.monto) if l else 0.0

        writer.writerow([
            r.personal.nro_doc or '',
            r.personal.apellidos_nombres,
            r.grupo,
            r.regimen_pension,
            r.afp or '',
            r.dias_trabajados,
            float(r.sueldo_base),
            _m('sueldo-basico'),
            _m('asig-familiar'),
            _m('he-25'),
            _m('he-35'),
            _m('he-100'),
            _m('otros-ingresos'),
            float(r.total_ingresos),
            _m('afp-aporte'),
            _m('afp-comision'),
            _m('afp-seguro'),
            _m('onp'),
            _m('ir-5ta'),
            _m('descto-prestamo'),
            _m('otros-descuentos'),
            float(r.total_descuentos),
            float(r.neto_a_pagar),
            _m('essalud'),
            _m('prov-gratificacion'),
            _m('prov-cts'),
            float(r.costo_total_empresa),
        ])

    content = output.getvalue()
    nombre = f'nomina_{periodo.anio}_{str(periodo.mes).zfill(2)}_{periodo.tipo}.csv'
    response = HttpResponse(content.encode('utf-8-sig'), content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


# ─── Exportaciones PLAME / T-Registro SUNAT ───────────────────────────────────

@login_required
@solo_admin
def periodo_exportar_plame(request, pk):
    """
    Descarga archivo ZIP con los archivos PLAME del periodo.
    Incluye: remuneraciones, retenciones 5ta, jornada laboral.
    """
    from .exports_plame import generar_plame_completo

    periodo = get_object_or_404(PeriodoNomina, pk=pk)

    if periodo.estado not in ('CALCULADO', 'APROBADO', 'CERRADO'):
        messages.warning(request, 'El período debe estar calculado o aprobado para exportar PLAME.')
        return redirect('nominas_periodo_detalle', pk=pk)

    resultado = generar_plame_completo(periodo)
    periodo_str = resultado['periodo_str']

    # Generar ZIP con todos los archivos
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        rem_content, rem_count = resultado['remuneraciones']
        if rem_count > 0:
            zf.writestr(
                f'0601_remuneraciones_{periodo_str}.txt',
                rem_content.encode('latin-1', errors='replace')
            )

        ret_content, ret_count = resultado['retenciones_5ta']
        if ret_count > 0:
            zf.writestr(
                f'0605_retenciones_5ta_{periodo_str}.txt',
                ret_content.encode('latin-1', errors='replace')
            )

        jor_content, jor_count = resultado['jornada']
        if jor_count > 0:
            zf.writestr(
                f'0701_jornada_{periodo_str}.txt',
                jor_content.encode('latin-1', errors='replace')
            )

        # Archivo resumen informativo
        resumen = (
            f'PLAME - Periodo {periodo_str}\r\n'
            f'Generado: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}\r\n'
            f'Periodo: {periodo}\r\n'
            f'Estado: {periodo.get_estado_display()}\r\n'
            f'---\r\n'
            f'Archivo 0601 (Remuneraciones): {rem_count} registros\r\n'
            f'Archivo 0605 (Retenciones 5ta): {ret_count} registros\r\n'
            f'Archivo 0701 (Jornada): {jor_count} registros\r\n'
        )
        zf.writestr(f'_RESUMEN_{periodo_str}.txt', resumen.encode('utf-8'))

    zip_buffer.seek(0)
    nombre_zip = f'PLAME_{periodo_str}_{periodo.get_tipo_display().replace(" ", "_")}.zip'

    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'

    logger.info(
        f'PLAME exportado: periodo={periodo_str}, '
        f'registros={resultado["total_registros"]}, '
        f'usuario={request.user}'
    )
    messages.success(
        request,
        f'PLAME exportado: {resultado["total_registros"]} trabajadores. '
        f'Archivos: remuneraciones ({rem_count}), retenciones 5ta ({ret_count}), '
        f'jornada ({jor_count}).'
    )

    return response


@login_required
@solo_admin
def periodo_exportar_tregistro(request, pk):
    """
    Descarga archivo T-Registro (altas) de los trabajadores del periodo.
    """
    from .exports_tregistro import generar_tregistro_desde_periodo

    periodo = get_object_or_404(PeriodoNomina, pk=pk)

    if periodo.estado not in ('CALCULADO', 'APROBADO', 'CERRADO'):
        messages.warning(request, 'El período debe estar calculado o aprobado para exportar T-Registro.')
        return redirect('nominas_periodo_detalle', pk=pk)

    content, count = generar_tregistro_desde_periodo(periodo)
    periodo_str = f'{periodo.anio}{periodo.mes:02d}'

    if count == 0:
        messages.warning(request, 'No hay trabajadores activos para exportar en T-Registro.')
        return redirect('nominas_periodo_detalle', pk=pk)

    nombre = f'T-Registro_Altas_{periodo_str}.txt'
    response = HttpResponse(
        content.encode('latin-1', errors='replace'),
        content_type='text/plain; charset=iso-8859-1'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'

    logger.info(
        f'T-Registro exportado: periodo={periodo_str}, '
        f'registros={count}, usuario={request.user}'
    )

    return response


# ─── Registro individual ───────────────────────────────────────────────────────

@login_required
@solo_admin
def registro_detalle(request, pk):
    """Detalle de un registro individual de nómina."""
    reg = get_object_or_404(
        RegistroNomina.objects.select_related('personal', 'periodo')
        .prefetch_related('lineas__concepto'),
        pk=pk
    )
    lineas_ing  = reg.lineas.filter(concepto__tipo='INGRESO').order_by('concepto__orden')
    lineas_desc = reg.lineas.filter(concepto__tipo='DESCUENTO').order_by('concepto__orden')
    lineas_apc  = reg.lineas.filter(concepto__tipo='APORTE_EMPLEADOR').order_by('concepto__orden')
    lineas_prov = reg.lineas.filter(concepto__subtipo='PROVISION').order_by('concepto__orden')

    return render(request, 'nominas/registro_detalle.html', {
        'titulo': f'Boleta — {reg.personal}',
        'reg': reg,
        'lineas_ing':  lineas_ing,
        'lineas_desc': lineas_desc,
        'lineas_apc':  lineas_apc,
        'lineas_prov': lineas_prov,
        'puede_editar': reg.periodo.estado in ('BORRADOR', 'CALCULADO'),
    })


@login_required
def explicar_boleta_ia(request, pk):
    """
    Explica una boleta de pago en lenguaje simple usando IA.
    Similar a BUK AI — interpretación automática de boletas.
    Accesible tanto para admin como para el trabajador (portal).
    """
    reg = get_object_or_404(
        RegistroNomina.objects.select_related('personal', 'periodo'),
        pk=pk
    )

    # Seguridad: solo admin o el propio trabajador
    if not request.user.is_staff:
        if not hasattr(request.user, 'personal') or request.user.personal != reg.personal:
            return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)

    from asistencia.services.ai_service import get_service
    svc = get_service()
    if not svc:
        return JsonResponse({'ok': False, 'error': 'IA no configurada'})

    # Construir datos de la boleta
    lineas_txt = ''
    for l in reg.lineas.select_related('concepto').all():
        lineas_txt += f'  {l.concepto.nombre}: S/ {l.monto}\n'

    prompt = (
        f'Explica esta boleta de pago de forma simple y clara para el trabajador.\n\n'
        f'Trabajador: {reg.personal.apellidos_nombres}\n'
        f'Periodo: {reg.periodo}\n'
        f'Sueldo Base: S/ {reg.sueldo_base}\n'
        f'Dias Trabajados: {reg.dias_trabajados}\n'
        f'Total Ingresos: S/ {reg.total_ingresos}\n'
        f'Total Descuentos: S/ {reg.total_descuentos}\n'
        f'Neto a Pagar: S/ {reg.neto_a_pagar}\n'
        f'EsSalud (empleador): S/ {reg.aporte_essalud}\n'
        f'Regimen Pension: {reg.regimen_pension}\n\n'
        f'Detalle:\n{lineas_txt}\n\n'
        f'Explica cada concepto de forma que un trabajador sin conocimiento contable '
        f'entienda: que es cada descuento, por que se cobra, y cuanto le queda. '
        f'Responde en espanol, tono amigable, maximo 200 palabras.'
    )

    system = (
        'Eres un asistente de RRHH que explica boletas de pago a trabajadores peruanos. '
        'Usa lenguaje simple, amigable y directo. Evita tecnicismos.'
    )

    try:
        resultado = svc.generate(prompt, system=system)
        return JsonResponse({'ok': True, 'explicacion': resultado or 'Sin explicacion disponible'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)[:200]})


@login_required
@solo_admin
def registro_editar(request, pk):
    """Editar ajustes manuales de un registro y recalcular."""
    reg = get_object_or_404(
        RegistroNomina.objects.select_related('personal', 'periodo'),
        pk=pk,
    )
    if reg.periodo.estado in ('APROBADO', 'CERRADO', 'ANULADO'):
        messages.error(request, 'El período está aprobado o cerrado, no se puede editar.')
        return redirect('nominas_registro_detalle', pk=pk)

    if request.method == 'POST':
        try:
            reg.dias_trabajados   = int(request.POST.get('dias_trabajados', 30))
            reg.horas_extra_25    = Decimal(request.POST.get('horas_extra_25', '0'))
            reg.horas_extra_35    = Decimal(request.POST.get('horas_extra_35', '0'))
            reg.horas_extra_100   = Decimal(request.POST.get('horas_extra_100', '0'))
            reg.asignacion_familiar = 'asignacion_familiar' in request.POST
            reg.otros_ingresos    = Decimal(request.POST.get('otros_ingresos', '0'))
            reg.descuento_prestamo= Decimal(request.POST.get('descuento_prestamo', '0'))
            reg.otros_descuentos  = Decimal(request.POST.get('otros_descuentos', '0'))
        except (ValueError, ArithmeticError):
            messages.error(request, 'Valores numéricos inválidos.')
            return redirect('nominas_registro_editar', pk=pk)

        with transaction.atomic():
            reg.save()

            # Marcar la cuota de préstamo como pagada si hay descuento registrado
            if reg.descuento_prestamo and reg.descuento_prestamo > 0 and reg.personal_id:
                try:
                    from prestamos.models import CuotaPrestamo
                    cuota = CuotaPrestamo.objects.filter(
                        prestamo__personal_id=reg.personal_id,
                        prestamo__estado='EN_CURSO',
                        estado='PENDIENTE',
                        periodo__year=reg.periodo.anio,
                        periodo__month=reg.periodo.mes,
                    ).first()
                    if cuota:
                        cuota.registrar_pago(
                            monto=reg.descuento_prestamo,
                            referencia=f'Nómina {reg.periodo}',
                        )
                except Exception:
                    pass  # No interrumpir el guardado de nómina por un error de préstamo

            # Recalcular
            conceptos = ConceptoRemunerativo.objects.filter(activo=True).order_by('tipo', 'orden')
            resultado = engine.calcular_registro(reg, conceptos)

            reg.lineas.all().delete()
            for l in resultado['lineas']:
                LineaNomina.objects.create(
                    registro=reg,
                    concepto=l['concepto'],
                    base_calculo=l['base_calculo'],
                    porcentaje_aplicado=l['porcentaje_aplicado'],
                    monto=l['monto'],
                    observacion=l['observacion'],
                )
            reg.total_ingresos      = resultado['total_ingresos']
            reg.total_descuentos    = resultado['total_descuentos']
            reg.neto_a_pagar        = resultado['neto_a_pagar']
            reg.aporte_essalud      = resultado['aporte_essalud']
            reg.costo_total_empresa = resultado['costo_total_empresa']
            reg.estado = 'CALCULADO'
            reg.save()

            # Resync period aggregates after individual edit
            from django.db.models import Sum
            agg = reg.periodo.registros.aggregate(
                t_bruto=Sum('total_ingresos'),
                t_desc=Sum('total_descuentos'),
                t_neto=Sum('neto_a_pagar'),
                t_costo=Sum('costo_total_empresa'),
            )
            reg.periodo.total_bruto = agg['t_bruto'] or Decimal('0')
            reg.periodo.total_descuentos = agg['t_desc'] or Decimal('0')
            reg.periodo.total_neto = agg['t_neto'] or Decimal('0')
            reg.periodo.total_costo_empresa = agg['t_costo'] or Decimal('0')
            reg.periodo.save(update_fields=[
                'total_bruto', 'total_descuentos', 'total_neto', 'total_costo_empresa'
            ])

        messages.success(request, 'Registro recalculado correctamente.')
        return redirect('nominas_registro_detalle', pk=pk)

    return render(request, 'nominas/registro_editar.html', {
        'titulo': f'Editar — {reg.personal}',
        'reg': reg,
    })


# ─── Conceptos remunerativos ───────────────────────────────────────────────────

@login_required
@solo_admin
def conceptos_panel(request):
    """Panel de conceptos remunerativos."""
    conceptos = ConceptoRemunerativo.objects.all().order_by('tipo', 'orden', 'nombre')
    return render(request, 'nominas/conceptos.html', {
        'titulo': 'Conceptos Remunerativos',
        'conceptos': conceptos,
        'tipo_choices': ConceptoRemunerativo.TIPO_CHOICES,
    })


@login_required
@solo_admin
@require_POST
def concepto_crear(request):
    """Crea un concepto remunerativo (AJAX)."""
    try:
        c = ConceptoRemunerativo.objects.create(
            codigo   = request.POST['codigo'].strip().lower(),
            nombre   = request.POST['nombre'].strip(),
            tipo     = request.POST.get('tipo', 'INGRESO'),
            subtipo  = request.POST.get('subtipo', 'REMUNERATIVO'),
            formula  = request.POST.get('formula', 'MANUAL'),
            porcentaje=Decimal(request.POST.get('porcentaje', '0')),
            orden    = int(request.POST.get('orden', 99)),
            activo   = True,
            es_sistema=False,
        )
        return JsonResponse({'ok': True, 'id': c.pk, 'nombre': c.nombre})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def concepto_eliminar(request, pk):
    """Elimina un concepto no-sistema (AJAX)."""
    c = get_object_or_404(ConceptoRemunerativo, pk=pk)
    if c.es_sistema:
        return JsonResponse({'ok': False, 'error': 'No se puede eliminar un concepto del sistema.'}, status=400)
    nombre = c.nombre
    c.delete()
    return JsonResponse({'ok': True, 'nombre': nombre})


# ─── Portal: Mis Recibos ───────────────────────────────────────────────────────

@login_required
def mis_recibos(request):
    """Portal del trabajador: historial de recibos de nómina."""
    from personal.models import Personal
    try:
        empleado = Personal.objects.get(usuario=request.user)
    except Personal.DoesNotExist:
        empleado = None

    registros = []
    if empleado:
        registros = (
            RegistroNomina.objects
            .filter(personal=empleado, periodo__estado__in=['APROBADO', 'CERRADO'])
            .select_related('periodo')
            .order_by('-periodo__anio', '-periodo__mes')
        )

    return render(request, 'nominas/mis_recibos.html', {
        'titulo': 'Mis Recibos de Nómina',
        'empleado': empleado,
        'registros': registros,
    })


# ─── Resumen estadístico AJAX ──────────────────────────────────────────────────

@login_required
@solo_admin
def periodo_resumen_ajax(request, pk):
    """
    Devuelve JSON con estadísticas resumidas del período de nómina.
    Útil para mostrar el panel de resumen en la vista de detalle sin recargar.
    """
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    registros = periodo.registros.select_related('personal')

    # Totales globales del período
    agg = registros.aggregate(
        total_bruto=Sum('total_ingresos'),
        total_neto=Sum('neto_a_pagar'),
        total_essalud=Sum('aporte_essalud'),
    )

    # AFP y ONP: sumar montos de líneas por fórmula del concepto
    total_afp = Decimal('0')
    total_onp = Decimal('0')
    for reg in registros.prefetch_related('lineas__concepto'):
        for linea in reg.lineas.all():
            formula = linea.concepto.formula
            if formula in ('AFP_APORTE', 'AFP_COMISION', 'AFP_SEGURO'):
                total_afp += linea.monto or Decimal('0')
            elif formula == 'ONP':
                total_onp += linea.monto or Decimal('0')

    # Distribución por régimen pensionario
    from django.db.models import Count
    por_regimen_qs = (
        registros.values('regimen_pension')
        .annotate(total=Count('id'))
        .order_by('regimen_pension')
    )
    por_regimen = {item['regimen_pension']: item['total'] for item in por_regimen_qs}

    # Distribución por grupo
    por_grupo_qs = (
        registros.values('grupo')
        .annotate(total=Count('id'))
        .order_by('grupo')
    )
    por_grupo = {(item['grupo'] or '—'): item['total'] for item in por_grupo_qs}

    return JsonResponse({
        'trabajadores':  registros.count(),
        'total_bruto':   float(agg['total_bruto'] or 0),
        'total_neto':    float(agg['total_neto'] or 0),
        'total_essalud': float(agg['total_essalud'] or 0),
        'total_afp':     float(total_afp),
        'total_onp':     float(total_onp),
        'por_regimen':   por_regimen,
        'por_grupo':     por_grupo,
    })


# ─── Boletas ZIP masivo ─────────────────────────────────────────────────────────

@login_required
@solo_admin
def periodo_boletas_zip(request, pk):
    """
    Genera un archivo ZIP con todas las boletas PDF del período.
    Solo disponible para períodos en estado CALCULADO, APROBADO o CERRADO.
    """
    periodo = get_object_or_404(PeriodoNomina, pk=pk)

    if periodo.estado not in ('CALCULADO', 'APROBADO', 'CERRADO'):
        messages.error(request, 'Solo se pueden descargar boletas de períodos calculados o aprobados.')
        return redirect('nominas_periodo_detalle', pk=pk)

    registros = (
        periodo.registros
        .select_related('personal')
        .prefetch_related('lineas__concepto')
        .order_by('personal__apellidos_nombres')
    )

    try:
        from .pdf import generar_boleta_pdf
    except ImportError:
        messages.error(request, 'El módulo de PDF no está disponible.')
        return redirect('nominas_periodo_detalle', pk=pk)

    zip_buffer = io.BytesIO()
    errores = []

    with zipfile.ZipFile(zip_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for reg in registros:
            try:
                pdf_bytes = generar_boleta_pdf(reg)
                nro_doc = reg.personal.nro_doc or f'reg{reg.pk}'
                nombre_archivo = (
                    f'Boleta_{nro_doc}_'
                    f'{periodo.anio}{str(periodo.mes).zfill(2)}.pdf'
                )
                zf.writestr(nombre_archivo, pdf_bytes)
            except Exception as e:
                logger.warning('Error adding boleta reg %s to ZIP: %s', reg.pk, e)
                errores.append(f'{reg.personal}: {e}')

    if errores:
        messages.warning(
            request,
            f'Se generaron las boletas con {len(errores)} error(es): '
            + '; '.join(errores[:5])
        )

    zip_buffer.seek(0)
    nombre_zip = f'Boletas_{periodo.anio}{str(periodo.mes).zfill(2)}_{periodo.tipo}.zip'
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
    return response


# ─── Boleta PDF ────────────────────────────────────────────────────────────────

@login_required
def boleta_pdf(request, pk):
    """
    Descarga la boleta de pago en PDF para un RegistroNomina.
    - Superusers/staff pueden ver cualquier boleta.
    - Trabajadores solo pueden ver sus propias boletas (via portal).
    """
    registro = get_object_or_404(
        RegistroNomina.objects.select_related('personal', 'periodo'),
        pk=pk,
    )

    # Control acceso: solo admin o el propio trabajador
    if not (request.user.is_superuser or request.user.is_staff):
        if not hasattr(request.user, 'personal') or request.user.personal != registro.personal:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('No tienes permiso para ver esta boleta.')

    try:
        from .pdf import generar_boleta_pdf
        pdf_bytes = generar_boleta_pdf(registro)
    except Exception as e:
        messages.error(request, f'Error generando PDF: {e}')
        return redirect('nominas_registro_detalle', pk=pk)

    nombre = (
        f'Boleta_{registro.personal.nro_doc}_'
        f'{registro.periodo.anio}{registro.periodo.mes:02d}.pdf'
    )
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nombre}"'
    return response


# ─── Gratificaciones Panel ─────────────────────────────────────────────────────

@login_required
@solo_admin
def gratificacion_panel(request):
    """Redirige al panel principal con la sección de procesos especiales."""
    from django.urls import reverse
    return redirect(reverse('nominas_panel') + '#gratificaciones')


# ─── Crear Período Especial (Gratificación / CTS) ─────────────────────────────

@login_required
@solo_admin
@require_POST
def crear_periodo_especial(request):
    """
    Crea un período especial de tipo GRATIFICACION o CTS con las fechas
    legalmente correctas según el mes/año indicado.

    Para GRATIFICACION:
      - mes=7  (Julio):    período June 1 – June 30, pago July 15
      - mes=12 (Dic):      período Dec 1 – Dec 15,   pago Dec 15

    Para CTS:
      - mes=5  (Mayo):     período Nov 1 año ant – Apr 30 año act, pago May 15
      - mes=11 (Nov):      período May 1 – Oct 31, pago Nov 15
    """
    tipo = request.POST.get('tipo', '').upper()
    try:
        anio = int(request.POST.get('anio', timezone.now().year))
        mes  = int(request.POST.get('mes', 7))
    except (ValueError, TypeError):
        messages.error(request, 'Año o mes inválido.')
        return redirect('nominas_gratificaciones')

    if tipo not in ('GRATIFICACION', 'CTS'):
        messages.error(request, 'Tipo de período inválido.')
        return redirect('nominas_gratificaciones')

    # Calcular fechas según tipo y mes
    if tipo == 'GRATIFICACION':
        if mes == 7:
            fecha_inicio = date(anio, 6, 1)
            fecha_fin    = date(anio, 6, 30)
            fecha_pago   = date(anio, 7, 15)
            descripcion  = f'Gratificación Julio {anio}'
        else:
            # Diciembre: período Dec 1 - Dec 15
            fecha_inicio = date(anio, 12, 1)
            fecha_fin    = date(anio, 12, 15)
            fecha_pago   = date(anio, 12, 15)
            descripcion  = f'Gratificación Diciembre {anio}'
    else:
        # CTS
        if mes == 5:
            # Mayo: período Nov 1 (año ant) – Apr 30 (año act)
            fecha_inicio = date(anio - 1, 11, 1)
            fecha_fin    = date(anio, 4, 30)
            fecha_pago   = date(anio, 5, 15)
            descripcion  = f'CTS Mayo {anio} (Nov {anio-1}–Abr {anio})'
        else:
            # Noviembre: período May 1 – Oct 31
            fecha_inicio = date(anio, 5, 1)
            fecha_fin    = date(anio, 10, 31)
            fecha_pago   = date(anio, 11, 15)
            descripcion  = f'CTS Noviembre {anio} (May–Oct {anio})'

    # Verificar que no exista ya
    if PeriodoNomina.objects.filter(tipo=tipo, anio=anio, mes=mes).exists():
        messages.warning(
            request,
            f'Ya existe un período de {tipo} para {anio}/{mes:02d}. '
            f'Revisa la lista de períodos.'
        )
        redirect_url = 'nominas_gratificaciones' if tipo == 'GRATIFICACION' else 'nominas_panel'
        return redirect(redirect_url)

    periodo = PeriodoNomina.objects.create(
        tipo=tipo,
        anio=anio,
        mes=mes,
        descripcion=descripcion,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        fecha_pago=fecha_pago,
    )

    messages.success(
        request,
        f'Período "{descripcion}" creado correctamente. '
        f'Período: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}. '
        f'Fecha de pago: {fecha_pago.strftime("%d/%m/%Y")}. '
        f'Haz clic en "Generar planilla" para calcular los montos.'
    )
    return redirect('nominas_periodo_detalle', pk=periodo.pk)


# ─── IR 5ta Panel ──────────────────────────────────────────────────────────────

@login_required
@solo_admin
def ir5ta_panel(request):
    """
    Panel de IR 5ta Categoría — Perú 2026.
    Muestra la escala de retención y proyección anual por empleado activo.
    Base legal: Art. 53° + 75° TUO LIR — escala progresiva, 7 UIT deducción.
    """
    from personal.models import Personal

    uit = engine._get_uit()  # S/ 5,500 (DS 233-2025-EF, vigente 2026)
    deduccion = engine.IR_5TA_DEDUCCION_UITS * uit  # 7 × 5,500 = S/ 38,500

    # Escala IR 5ta con límites en soles para la tabla visual
    # Tramos acumulativos: (limite_acum_uits, tasa%)
    # 0-5 UIT → 8%  |  5-20 UIT → 14%  |  20-35 UIT → 17%
    # 35-45 UIT → 20%  |  >45 UIT → 30%
    escala_visual = [
        {'tramo': 'Hasta 5 UIT',   'uits': '0 — 5',   'limite_s': uit * 5,  'tasa': Decimal('8'),  'color': 'success'},
        {'tramo': '5 a 20 UIT',    'uits': '5 — 20',  'limite_s': uit * 20, 'tasa': Decimal('14'), 'color': 'info'},
        {'tramo': '20 a 35 UIT',   'uits': '20 — 35', 'limite_s': uit * 35, 'tasa': Decimal('17'), 'color': 'warning'},
        {'tramo': '35 a 45 UIT',   'uits': '35 — 45', 'limite_s': uit * 45, 'tasa': Decimal('20'), 'color': 'orange'},
        {'tramo': 'Más de 45 UIT', 'uits': '> 45',    'limite_s': None,     'tasa': Decimal('30'), 'color': 'danger'},
    ]

    # Empleados activos con sueldo_base > 0
    empleados_qs = (
        Personal.objects
        .filter(estado='Activo', sueldo_base__isnull=False, sueldo_base__gt=0)
        .order_by('-sueldo_base', 'apellidos_nombres')
        .values(
            'pk', 'apellidos_nombres', 'nro_doc', 'sueldo_base',
            'grupo_tareo', 'asignacion_familiar',
            # nuevos campos nómina
            'tiene_eps', 'eps_descuento_mensual', 'viaticos_mensual',
        )
    )

    asig_fam_monto = engine.ASIG_FAM  # S/ 102.50

    empleados = []
    total_ir_mensual = Decimal('0')

    for emp in empleados_qs:
        sueldo = Decimal(str(emp['sueldo_base']))
        asig   = asig_fam_monto if emp.get('asignacion_familiar') else Decimal('0')
        rem_computable = sueldo + asig

        # Proyección anual = sueldo computable × 14
        # (12 meses + gratificación julio + gratificación diciembre)
        anual = rem_computable * Decimal('14')

        # EPS: co-pago anual del trabajador (deducible del base IR 5ta)
        tiene_eps         = bool(emp.get('tiene_eps', False))
        eps_mensual       = Decimal(str(emp.get('eps_descuento_mensual') or '0'))
        eps_anual         = eps_mensual * Decimal('12')

        # Viáticos: monto mensual fijo (NO remunerativo, excluido de base IR 5ta)
        viaticos_mensual  = Decimal(str(emp.get('viaticos_mensual') or '0'))

        # Renta neta imponible = anual - EPS anual - 7 UIT
        rni = max(anual - eps_anual - deduccion, Decimal('0'))

        # IR anual y mensual (pasando deducción EPS al engine)
        ir_mensual = engine.calcular_ir_5ta_mensual(anual, deduccion_eps_anual=eps_anual)
        ir_anual   = ir_mensual * Decimal('12')

        # Clasificar para color de fila
        if ir_mensual == 0:
            fila_clase = 'no-ir'
        elif rni <= uit * 5:
            fila_clase = 'ir-8'
        elif rni <= uit * 20:
            fila_clase = 'ir-14'
        else:
            fila_clase = 'ir-alto'

        empleados.append({
            'pk':               emp['pk'],
            'apellidos_nombres':emp['apellidos_nombres'],
            'numero_documento': emp['nro_doc'],
            'sueldo_base':      sueldo,
            'asig_fam':         asig,
            'asignacion_familiar': bool(emp.get('asignacion_familiar')),
            'rem_computable':   rem_computable,
            'grupo':            emp.get('grupo_tareo', ''),
            'anual_proyectado': anual,
            'rni':              rni,
            'ir_anual':         ir_anual,
            'ir_mensual':       ir_mensual,
            'fila_clase':       fila_clase,
            # EPS / viáticos
            'tiene_eps':        tiene_eps,
            'eps_mensual':      eps_mensual,
            'eps_anual':        eps_anual,
            'viaticos_mensual': viaticos_mensual,
        })
        total_ir_mensual += ir_mensual

    # ── Resumen por tramo (para cards de distribución) ──────────────────────
    resumen = {
        'sin_ir':     sum(1 for e in empleados if e['ir_mensual'] == 0),
        'tramo_8':    sum(1 for e in empleados if e['fila_clase'] == 'ir-8'),
        'tramo_14':   sum(1 for e in empleados if e['fila_clase'] == 'ir-14'),
        'tramo_17_mas': sum(1 for e in empleados if e['fila_clase'] == 'ir-alto'),
        'total_con_ir': sum(1 for e in empleados if e['ir_mensual'] > 0),
        'con_eps':    sum(1 for e in empleados if e['tiene_eps']),
        'con_viaticos': sum(1 for e in empleados if e['viaticos_mensual'] > 0),
    }

    return render(request, 'nominas/ir5ta_panel.html', {
        'titulo': 'IR 5ta Categoría — 2026',
        'uit': uit,
        'deduccion': deduccion,
        'escala_visual': escala_visual,
        'empleados': empleados,
        'total_ir_mensual': total_ir_mensual,
        'total_empleados': len(empleados),
        'pagan_ir': sum(1 for e in empleados if e['ir_mensual'] > 0),
        'resumen': resumen,
    })


# ─── IR 5ta AJAX por RegistroNomina ───────────────────────────────────────────

@login_required
@solo_admin
def registro_ir5ta_ajax(request, pk):
    """
    Devuelve JSON con el desglose de IR 5ta para un RegistroNomina específico.
    Útil para mostrar detalles en modal desde la vista de período.
    """
    reg = get_object_or_404(
        RegistroNomina.objects.select_related('personal', 'periodo'),
        pk=pk
    )

    uit = engine.UIT_2026
    deduccion_uits = engine.IR_5TA_DEDUCCION_UITS

    # Remuneración computable mensual y proyección anual
    sueldo = reg.sueldo_base
    asig_fam = engine.ASIG_FAM if reg.asignacion_familiar else Decimal('0')
    rem_computable = sueldo + asig_fam

    # Proyección anual (mismo criterio que el engine)
    rem_anual = rem_computable * Decimal('12')

    deduccion = deduccion_uits * uit
    rni = max(rem_anual - deduccion, Decimal('0'))
    ir_anual = engine.calcular_ir_5ta_mensual(rem_anual) * Decimal('12')
    ir_mensual = engine.calcular_ir_5ta_mensual(rem_anual)

    # Desglose por tramos para transparencia
    tramos = []
    anterior = Decimal('0')
    for limite_uits, tasa in engine.IR_5TA_ESCALA:
        if rni <= 0:
            break
        if limite_uits is None:
            limite_s = None
            exceso = max(rni - anterior, Decimal('0'))
            label = f'> {int(anterior / uit)} UIT'
        else:
            limite_s = limite_uits * uit
            exceso = max(min(rni, limite_s) - anterior, Decimal('0'))
            label = f'{int(anterior / uit) if anterior > 0 else 0}–{int(limite_uits)} UIT'

        if exceso > 0:
            impuesto_tramo = (exceso * tasa / Decimal('100')).quantize(Decimal('0.01'))
            tramos.append({
                'tramo': label,
                'base': float(exceso),
                'tasa': float(tasa),
                'impuesto': float(impuesto_tramo),
            })

        if limite_uits is None:
            break
        anterior = limite_s
        if rni <= anterior:
            break

    return JsonResponse({
        'trabajador':        str(reg.personal),
        'sueldo_base':       float(sueldo),
        'asig_familiar':     float(asig_fam),
        'rem_computable':    float(rem_computable),
        'rem_anual':         float(rem_anual),
        'uit':               float(uit),
        'deduccion_7uits':   float(deduccion),
        'rni':               float(rni),
        'ir_anual':          float(ir_anual),
        'ir_mensual':        float(ir_mensual),
        'tramos':            tramos,
    })


# ─── Flujo de Caja de Planilla ─────────────────────────────────────────────────

@login_required
@solo_admin
def flujo_caja_panel(request):
    """
    Panel de proyección de flujo de caja de planilla.

    Parámetros GET:
        hasta=YYYY-MM  → proyecta hasta ese mes inclusive
        meses=N        → proyecta N meses desde hoy (legacy)
        plan=PK        → usa un PlanPlantilla como fuente (en lugar de empleados reales)
    """
    from dateutil.relativedelta import relativedelta as _rdelta

    hoy_inicio = date.today().replace(day=1)

    # ── Resolver horizonte ────────────────────────────────────────────
    hasta_param = request.GET.get('hasta', '').strip()
    if hasta_param:
        try:
            yr, mo = hasta_param.split('-')
            hasta_date = date(int(yr), int(mo), 1)
            diff = _rdelta(hasta_date, hoy_inicio)
            n_meses = diff.years * 12 + diff.months + 1
            n_meses = max(1, n_meses)
        except (ValueError, AttributeError):
            hasta_param = ''
            try:
                n_meses = int(request.GET.get('meses', 18))
            except (ValueError, TypeError):
                n_meses = 18
    else:
        try:
            n_meses = int(request.GET.get('meses', 18))
        except (ValueError, TypeError):
            n_meses = 18
    n_meses = max(1, min(n_meses, 60))

    # Fecha fin proyección
    hasta_date  = hoy_inicio + _rdelta(months=n_meses - 1)
    hasta_str   = hasta_date.strftime('%Y-%m')
    mes_min_str = hoy_inicio.strftime('%Y-%m')
    _MESES_ES_L = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                   'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    hasta_label = f"{_MESES_ES_L[hasta_date.month]}-{str(hasta_date.year)[2:]}"
    atajos = {
        '6':  (hoy_inicio + _rdelta(months=5)).strftime('%Y-%m'),
        '12': (hoy_inicio + _rdelta(months=11)).strftime('%Y-%m'),
        '18': (hoy_inicio + _rdelta(months=17)).strftime('%Y-%m'),
        '24': (hoy_inicio + _rdelta(months=23)).strftime('%Y-%m'),
        '36': (hoy_inicio + _rdelta(months=35)).strftime('%Y-%m'),
    }

    # ── Fuente: REAL o PLAN ───────────────────────────────────────────
    plan_id   = request.GET.get('plan', '').strip()
    modo      = 'REAL'
    plan_activo = None
    planes_disponibles = PlanPlantilla.objects.order_by('nombre')

    if plan_id:
        try:
            plan_activo = PlanPlantilla.objects.get(pk=int(plan_id))
            modo = 'PLAN'
            meses, empleados = proyectar_desde_plan(plan_activo, n_meses=n_meses)
        except (ValueError, PlanPlantilla.DoesNotExist):
            plan_activo = None

    if plan_activo is None:
        meses, empleados = proyectar_flujo_caja(n_meses=n_meses)

    # ── Presupuesto (solo en modo REAL) ───────────────────────────────
    tiene_presupuesto = False
    if modo == 'REAL':
        presup_map = {(p.anio, p.mes): p for p in PresupuestoPlanilla.objects.all()}
        for mes in meses:
            key = (mes['fecha'].year, mes['fecha'].month)
            presup = presup_map.get(key)
            if presup:
                tiene_presupuesto = True
                mes['presup_total'] = presup.presup_total
                mes['variacion']    = mes['total_desembolso'] - presup.presup_total
                mes['variacion_pct'] = (
                    mes['variacion'] / presup.presup_total * 100
                    if presup.presup_total else None
                )

    # ── Totales ───────────────────────────────────────────────────────
    total_18m        = sum(m['total_desembolso'] for m in meses)
    promedio_mensual = total_18m / len(meses) if meses else Decimal('0')
    headcount_actual = meses[0]['headcount'] if meses else 0
    meses_con_vencimientos = sum(1 for m in meses if m['liquidaciones'] > 0)

    _hoy = date.today()
    _horizon = _hoy.replace(day=1) + _rdelta(months=n_meses)
    if modo == 'REAL':
        empleados_por_vencer = sum(
            1 for e in empleados
            if e.get('fecha_fin_contrato') and _hoy.replace(day=1) <= e['fecha_fin_contrato'] < _horizon
        )
    else:
        # En modo PLAN: puestos que terminan dentro del horizonte
        empleados_por_vencer = sum(
            1 for e in empleados
            if e.get('fecha_fin_puesto') and _hoy.replace(day=1) <= e['fecha_fin_puesto'] < _horizon
        )

    # ── Charts ────────────────────────────────────────────────────────
    chart_labels    = json.dumps([m['mes_label'] for m in meses])
    chart_neto      = json.dumps([float(m['neto'])             for m in meses])
    chart_cond      = json.dumps([float(m['cond_trabajo'])     for m in meses])
    chart_alim      = json.dumps([float(m['alimentacion'])     for m in meses])
    chart_viaticos  = json.dumps([float(m.get('viaticos', 0)) for m in meses])
    chart_essalud   = json.dumps([float(m['essalud'])          for m in meses])
    chart_gratif    = json.dumps([float(m['gratificaciones'])  for m in meses])
    chart_cts       = json.dumps([float(m['cts'])              for m in meses])
    chart_liq       = json.dumps([float(m['liquidaciones'])    for m in meses])
    chart_headcount = json.dumps([m['headcount']               for m in meses])
    chart_total     = json.dumps([float(m['total_desembolso']) for m in meses])
    chart_presup    = json.dumps([
        float(m['presup_total']) if m.get('presup_total') is not None else None
        for m in meses
    ])

    return render(request, 'nominas/flujo_caja.html', {
        'meses':            meses,
        'empleados':        empleados,
        'n_meses':          n_meses,
        'modo':             modo,
        'plan_activo':      plan_activo,
        'planes_disponibles': planes_disponibles,
        'tiene_presupuesto':      tiene_presupuesto,
        'total_18m':              total_18m,
        'promedio_mensual':       promedio_mensual,
        'headcount_actual':       headcount_actual,
        'meses_con_vencimientos': meses_con_vencimientos,
        'empleados_por_vencer':   empleados_por_vencer,
        # Selector de período
        'hasta_str':    hasta_str,
        'hasta_label':  hasta_label,
        'mes_min_str':  mes_min_str,
        'atajos':       atajos,
        # Charts
        'chart_labels':    chart_labels,
        'chart_neto':      chart_neto,
        'chart_cond':      chart_cond,
        'chart_alim':      chart_alim,
        'chart_viaticos':  chart_viaticos,
        'chart_essalud':   chart_essalud,
        'chart_gratif':    chart_gratif,
        'chart_cts':       chart_cts,
        'chart_liq':       chart_liq,
        'chart_headcount': chart_headcount,
        'chart_total':     chart_total,
        'chart_presup':    chart_presup,
    })


@login_required
@solo_admin
def presupuesto_guardar(request):
    """
    API — guarda o actualiza el presupuesto de un mes.
    POST JSON: {anio, mes, presup_total, presup_rem_bruta?,
                presup_cond_trabajo?, presup_alimentacion?,
                presup_essalud?, presup_gratif?, presup_cts?,
                presup_liquidaciones?, observaciones?}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        anio = int(data['anio'])
        mes  = int(data['mes'])

        def _d(key, default='0'):
            return Decimal(str(data.get(key, default) or '0'))

        presup, created = PresupuestoPlanilla.objects.update_or_create(
            anio=anio, mes=mes, empresa=None,
            defaults={
                'presup_rem_bruta':     _d('presup_rem_bruta'),
                'presup_cond_trabajo':  _d('presup_cond_trabajo'),
                'presup_alimentacion':  _d('presup_alimentacion'),
                'presup_essalud':       _d('presup_essalud'),
                'presup_gratif':        _d('presup_gratif'),
                'presup_cts':           _d('presup_cts'),
                'presup_liquidaciones': _d('presup_liquidaciones'),
                'presup_total':         _d('presup_total'),
                'observaciones':        data.get('observaciones', ''),
                'creado_por':           request.user,
            }
        )
        return JsonResponse({'ok': True, 'created': created, 'id': presup.pk})

    except (KeyError, ValueError, Exception) as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@solo_admin
def presupuesto_eliminar(request, anio, mes):
    """Elimina el presupuesto de un mes específico."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    deleted, _ = PresupuestoPlanilla.objects.filter(anio=anio, mes=mes, empresa=None).delete()
    return JsonResponse({'ok': True, 'deleted': deleted})


# ═══════════════════════════════════════════════════════════════════════════════
#  PLANES DE PLANTILLA
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@solo_admin
def planes_panel(request):
    """Panel con lista de planes y botón de crear."""
    planes = PlanPlantilla.objects.prefetch_related('lineas').order_by('-creado_en')
    return render(request, 'nominas/planes_panel.html', {'planes': planes})


@login_required
@solo_admin
def plan_crear(request):
    """Crear un nuevo PlanPlantilla vía POST JSON."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        plan = PlanPlantilla.objects.create(
            nombre      = data['nombre'].strip(),
            tipo        = data.get('tipo', 'OBRA'),
            descripcion = data.get('descripcion', ''),
            fecha_inicio= data['fecha_inicio'],
            fecha_fin   = data.get('fecha_fin') or None,
            estado      = 'BORRADOR',
            creado_por  = request.user,
        )
        return JsonResponse({'ok': True, 'id': plan.pk, 'url': f'/nominas/planes/{plan.pk}/'})
    except (KeyError, Exception) as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@solo_admin
def plan_detalle(request, pk):
    """Vista Gantt + tabla de líneas + totales mensuales."""
    plan = get_object_or_404(PlanPlantilla.objects.prefetch_related('lineas__area', 'lineas__personal'), pk=pk)

    # Proyección del plan
    meses, posiciones = proyectar_desde_plan(plan)

    # ── Preparar datos para el template ──────────────────────────────
    chart_labels = [m['mes_label'] for m in meses]
    chart_total  = [float(m['total_desembolso']) for m in meses]
    chart_headcount = [m['headcount'] for m in meses]

    # Datos chart stacked (igual que flujo_caja_panel)
    chart_neto   = [float(m['neto'])           for m in meses]
    chart_cond   = [float(m['cond_trabajo'])   for m in meses]
    chart_alim   = [float(m['alimentacion'])   for m in meses]
    chart_essalud= [float(m['essalud'])        for m in meses]
    chart_gratif = [float(m['gratificaciones'])for m in meses]
    chart_cts    = [float(m['cts'])            for m in meses]
    chart_liq    = [float(m['liquidaciones'])  for m in meses]

    # Total del horizonte
    total_horizonte = sum(m['total_desembolso'] for m in meses)
    headcount_pico  = max((m['headcount'] for m in meses), default=0)

    # Áreas disponibles para el modal de línea
    from personal.models import Area
    areas = Area.objects.filter(activa=True).order_by('nombre')

    # ── Agregados para la pestaña Presupuesto ─────────────────────────
    from collections import defaultdict
    _by_cargo: dict = defaultdict(lambda: {'cantidad': 0, 'costo_mes': Decimal('0'), 'cargos': 0})
    _by_area:  dict = defaultdict(lambda: {'cantidad': 0, 'costo_mes': Decimal('0')})
    for pos in posiciones:
        cargo_key = pos['cargo']
        area_key  = pos.get('area__nombre') or 'Sin área'
        n         = pos['cantidad']
        costo_mes = (Decimal(str(pos['sueldo_base'])) * Decimal('1.09')) * n
        _by_cargo[cargo_key]['cantidad']  += n
        _by_cargo[cargo_key]['costo_mes'] += costo_mes
        _by_cargo[cargo_key]['cargos']    += 1
        _by_area[area_key]['cantidad']    += n
        _by_area[area_key]['costo_mes']   += costo_mes

    by_cargo_list = sorted(_by_cargo.items(), key=lambda x: -float(x[1]['costo_mes']))
    by_area_list  = sorted(_by_area.items(),  key=lambda x: -float(x[1]['costo_mes']))

    return render(request, 'nominas/plan_detalle.html', {
        'plan':             plan,
        'meses':            meses,
        'posiciones':       posiciones,
        'chart_labels':     json.dumps(chart_labels),
        'chart_neto':       json.dumps(chart_neto),
        'chart_cond':       json.dumps(chart_cond),
        'chart_alim':       json.dumps(chart_alim),
        'chart_essalud':    json.dumps(chart_essalud),
        'chart_gratif':     json.dumps(chart_gratif),
        'chart_cts':        json.dumps(chart_cts),
        'chart_liq':        json.dumps(chart_liq),
        'chart_total':      json.dumps(chart_total),
        'chart_headcount':  json.dumps(chart_headcount),
        'total_horizonte':  total_horizonte,
        'headcount_pico':   headcount_pico,
        'areas':            areas,
        'by_cargo_list':    by_cargo_list,
        'by_area_list':     by_area_list,
    })


@login_required
@solo_admin
def plan_linea_upsert(request, plan_pk):
    """Crear o actualizar una LineaPlan. POST JSON."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    plan = get_object_or_404(PlanPlantilla, pk=plan_pk)
    try:
        data    = json.loads(request.body)
        linea_id = data.get('id')

        def _d(key, default='0'):
            v = data.get(key, default)
            return Decimal(str(v or 0))

        from personal.models import Area
        area_id = data.get('area_id')
        area    = Area.objects.filter(pk=area_id).first() if area_id else None

        defaults = {
            'cargo':                 data.get('cargo', '').strip(),
            'area':                  area,
            'cantidad':              int(data.get('cantidad', 1)),
            'sueldo_base':           _d('sueldo_base'),
            'asignacion_familiar':   bool(data.get('asignacion_familiar', False)),
            'regimen_pension':       data.get('regimen_pension', 'AFP'),
            'afp':                   data.get('afp', ''),
            'cond_trabajo_mensual':  _d('cond_trabajo_mensual'),
            'alimentacion_mensual':  _d('alimentacion_mensual'),
            'fecha_inicio_puesto':   data['fecha_inicio_puesto'],
            'fecha_fin_puesto':      data.get('fecha_fin_puesto') or None,
            'notas':                 data.get('notas', ''),
            'orden':                 int(data.get('orden', 0)),
        }

        if linea_id:
            linea = get_object_or_404(LineaPlan, pk=linea_id, plan=plan)
            for k, v in defaults.items():
                setattr(linea, k, v)
            linea.save()
            created = False
        else:
            linea   = LineaPlan.objects.create(plan=plan, **defaults)
            created = True

        return JsonResponse({'ok': True, 'id': linea.pk, 'created': created})
    except (KeyError, Exception) as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@solo_admin
def plan_linea_eliminar(request, plan_pk, linea_pk):
    """Eliminar una LineaPlan. POST."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    plan  = get_object_or_404(PlanPlantilla, pk=plan_pk)
    linea = get_object_or_404(LineaPlan, pk=linea_pk, plan=plan)
    linea.delete()
    return JsonResponse({'ok': True})


@login_required
@solo_admin
def plan_actualizar_estado(request, pk):
    """Cambia el estado de un plan. POST JSON: {estado: 'APROBADO'}."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    plan = get_object_or_404(PlanPlantilla, pk=pk)
    try:
        data   = json.loads(request.body)
        estado = data['estado']
        if estado not in dict(PlanPlantilla.ESTADO_CHOICES):
            return JsonResponse({'error': 'Estado inválido'}, status=400)
        plan.estado = estado
        plan.save(update_fields=['estado', 'actualizado_en'])
        return JsonResponse({'ok': True, 'estado': plan.estado, 'badge': plan.badge_estado})
    except (KeyError, Exception) as e:
        return JsonResponse({'error': str(e)}, status=400)


# ═══════════════════════════════════════════════════════════════════════════════
# PLAN PLANTILLA — Excel Import / Export
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@solo_admin
def plan_plantilla_excel(request):
    """Descarga una plantilla XLSX en blanco para importar líneas al plan."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Líneas Plan'

    headers = [
        'Cargo / Puesto *', 'Área (nombre)', 'N° Posiciones *',
        'Sueldo Base *', 'Asig. Familiar (Si/No)', 'Cond. Trabajo Mensual',
        'Alimentación Mensual', 'Régimen (AFP/ONP/SIN_PENSION)',
        'AFP (Habitat/Integra/Prima/Profuturo)', 'Fecha Inicio (YYYY-MM-DD) *',
        'Fecha Fin (YYYY-MM-DD)', 'Notas',
    ]
    fill_h  = PatternFill('solid', fgColor='0D2B27')
    fnt_h   = Font(color='5EEAD4', bold=True, size=10)
    brd_h   = Border(bottom=Side(style='thin', color='5EEAD4'))
    aln_c   = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill_h; cell.font = fnt_h
        cell.alignment = aln_c; cell.border = brd_h
    ws.row_dimensions[1].height = 32

    # Fila de ejemplo
    example = [
        'Ing. de Campo', 'Operaciones', 2, 4500.00, 'Si', 300.00,
        200.00, 'AFP', 'Integra', '2026-04-01', '2026-12-31', 'Personal foráneo',
    ]
    fnt_ex = Font(size=9, color='64748B', italic=True)
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = fnt_ex
        cell.alignment = Alignment(horizontal='center')

    # Anchos de columna
    widths = [22, 18, 12, 12, 16, 16, 16, 22, 24, 18, 18, 22]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_plan.xlsx"'
    return response


@login_required
@solo_admin
def plan_export_excel(request, pk):
    """Exporta el plan completo como XLSX: hoja Líneas + hoja Proyección mensual."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    plan   = get_object_or_404(PlanPlantilla.objects.prefetch_related('lineas__area'), pk=pk)
    meses, _ = proyectar_desde_plan(plan)

    wb     = openpyxl.Workbook()
    fill_h = PatternFill('solid', fgColor='0D2B27')
    fnt_h  = Font(color='5EEAD4', bold=True, size=10)
    fnt_d  = Font(size=9)
    aln_c  = Alignment(horizontal='center', vertical='center')
    aln_l  = Alignment(horizontal='left', vertical='center')

    # ── Hoja 1: Líneas ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Líneas'

    hdrs1 = ['ID', 'Cargo', 'Área', 'Cantidad', 'Sueldo Base',
             'Asig. Familiar', 'Cond. Trabajo', 'Alimentación',
             'Régimen', 'AFP', 'Inicio Puesto', 'Fin Puesto', 'Notas']
    for c, h in enumerate(hdrs1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = fill_h; cell.font = fnt_h; cell.alignment = aln_c
    ws1.row_dimensions[1].height = 22

    lineas = plan.lineas.select_related('area').order_by('orden', 'pk')
    for row, ln in enumerate(lineas, 2):
        vals = [
            ln.pk, ln.cargo,
            ln.area.nombre if ln.area else '',
            ln.cantidad, float(ln.sueldo_base),
            'Sí' if ln.asignacion_familiar else 'No',
            float(ln.cond_trabajo_mensual),
            float(ln.alimentacion_mensual),
            ln.regimen_pension, ln.afp or '',
            ln.fecha_inicio_puesto.strftime('%Y-%m-%d'),
            ln.fecha_fin_puesto.strftime('%Y-%m-%d') if ln.fecha_fin_puesto else '',
            ln.notas or '',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            cell.font = fnt_d
            cell.alignment = aln_c if c != 2 else aln_l
            if c == 5:
                cell.number_format = '"S/ "#,##0.00'
            elif c in (7, 8):
                cell.number_format = '"S/ "#,##0.00'

    for c, w in enumerate([6,22,18,8,13,10,13,13,10,12,12,12,22], 1):
        ws1.column_dimensions[get_column_letter(c)].width = w

    # ── Hoja 2: Proyección mensual ───────────────────────────────────
    ws2 = wb.create_sheet('Proyección')

    hdrs2 = ['Mes', 'Headcount', 'Rem. Bruta', 'Neto Empleados',
             'Cond. Trabajo', 'Alimentación', 'EsSalud (9%)',
             'Gratificaciones', 'CTS', 'Liquidaciones',
             'Total Desembolso', 'Acumulado']
    for c, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = fill_h; cell.font = fnt_h; cell.alignment = aln_c
    ws2.row_dimensions[1].height = 22

    for row, m in enumerate(meses, 2):
        vals = [
            m['mes_label'], m['headcount'],
            float(m.get('rem_bruta', 0)), float(m['neto']),
            float(m['cond_trabajo']), float(m['alimentacion']),
            float(m['essalud']), float(m['gratificaciones']),
            float(m['cts']), float(m['liquidaciones']),
            float(m['total_desembolso']), float(m.get('acumulado', 0)),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.font = fnt_d; cell.alignment = aln_c
            if c >= 3:
                cell.number_format = '"S/ "#,##0.00'

    # Fila TOTAL con fórmulas
    tot = len(meses) + 2
    fnt_tot  = Font(bold=True, size=9)
    fill_tot = PatternFill('solid', fgColor='D1FAE5')
    ws2.cell(row=tot, column=1, value='TOTAL').font = fnt_tot
    for c in range(2, 13):
        col_l = get_column_letter(c)
        v = f'=SUM({col_l}2:{col_l}{tot-1})' if c >= 3 else ''
        cell = ws2.cell(row=tot, column=c, value=v)
        cell.fill = fill_tot; cell.font = fnt_tot; cell.alignment = aln_c
        if c >= 3:
            cell.number_format = '"S/ "#,##0.00'

    for c, w in enumerate([12,10,14,14,14,14,14,14,14,14,16,16], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = plan.nombre.replace(' ', '_').replace('/', '-')[:40]
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Plan_{safe}.xlsx"'
    return response


@login_required
@solo_admin
def plan_import_excel(request, pk):
    """Importa líneas desde un archivo XLSX. POST multipart con campo 'archivo'."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    plan    = get_object_or_404(PlanPlantilla, pk=pk)
    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'error': 'No se recibió archivo'}, status=400)
    if not archivo.name.lower().endswith(('.xlsx', '.xlsm')):
        return JsonResponse({'error': 'Solo se aceptan archivos .xlsx'}, status=400)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'error': f'No se pudo leer el archivo: {e}'}, status=400)

    from personal.models import Area
    from datetime import date as _date
    import re as _re
    import django.db.models as _models

    area_map  = {a.nombre.strip().lower(): a.pk for a in Area.objects.all()}
    max_orden = LineaPlan.objects.filter(plan=plan).aggregate(
        m=_models.Max('orden'))['m'] or 0

    to_create = []
    errors    = []

    def parse_date(val):
        if val is None:
            return None
        if isinstance(val, _date):
            return val
        if hasattr(val, 'date') and callable(val.date):
            return val.date()
        s = str(val).strip()
        m = _re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            return _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        try:
            cargo     = str(row[0] or '').strip()
            area_name = str(row[1] or '').strip()
            cantidad  = int(row[2] or 1)
            sueldo    = Decimal(str(row[3] or 0))
            asig_fam  = str(row[4] or 'No').strip().lower() in ('si', 'sí', 'yes', '1', 'true')
            cond_trab = Decimal(str(row[5] or 0))
            alim      = Decimal(str(row[6] or 0))
            regimen   = str(row[7] or 'AFP').strip().upper()
            afp_val   = str(row[8] or '').strip()
            fecha_ini = parse_date(row[9])
            fecha_fin = parse_date(row[10])
            notas     = str(row[11] or '').strip()[:300]

            if not cargo:
                errors.append({'fila': row_idx, 'msg': 'Cargo vacío — fila ignorada'})
                continue
            if sueldo <= 0:
                errors.append({'fila': row_idx, 'msg': f'"{cargo}": sueldo inválido'})
                continue
            if regimen not in ('AFP', 'ONP', 'SIN_PENSION'):
                regimen = 'AFP'
            if not fecha_ini:
                fecha_ini = plan.fecha_inicio

            max_orden += 1
            to_create.append(LineaPlan(
                plan=plan,
                cargo=cargo[:150],
                area_id=area_map.get(area_name.lower()),
                cantidad=max(1, min(cantidad, 999)),
                sueldo_base=sueldo,
                asignacion_familiar=asig_fam,
                cond_trabajo_mensual=cond_trab,
                alimentacion_mensual=alim,
                regimen_pension=regimen,
                afp=afp_val[:20] if regimen == 'AFP' else '',
                fecha_inicio_puesto=fecha_ini,
                fecha_fin_puesto=fecha_fin,
                notas=notas,
                orden=max_orden,
            ))
        except Exception as exc:
            errors.append({'fila': row_idx, 'msg': str(exc)})

    if to_create:
        LineaPlan.objects.bulk_create(to_create)

    return JsonResponse({'ok': True, 'created': len(to_create), 'errors': errors})


# ─────────────────────────────────────────────────
# RECARGAS DE ALIMENTACIÓN (Edenred / Sodexo)
# ─────────────────────────────────────────────────

@login_required
@solo_admin
def alimentacion_panel(request):
    """Panel de recargas de tarjetas de alimentación."""
    from .models import RecargaAlimentacion

    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))

    recargas = RecargaAlimentacion.objects.filter(anio=anio, mes=mes).select_related('personal')
    totales = recargas.aggregate(
        total_monto=Sum('monto'), total_comision=Sum('comision'), total_total=Sum('total'),
        pendientes=Count('id', filter=Q(estado='PENDIENTE')),
        procesadas=Count('id', filter=Q(estado='PROCESADA')),
    )

    MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    return render(request, 'nominas/alimentacion_panel.html', {
        'recargas': recargas,
        'totales': totales,
        'anio': anio, 'mes': mes,
        'mes_nombre': MESES[mes-1],
        'total_recargas': recargas.count(),
    })


@login_required
@solo_admin
@require_POST
def alimentacion_generar(request):
    """Genera recargas para todos los empleados con alimentación > 0."""
    from .models import RecargaAlimentacion
    from personal.models import Personal

    anio = int(request.POST.get('anio', date.today().year))
    mes = int(request.POST.get('mes', date.today().month))
    comision_pct = Decimal(request.POST.get('comision_pct', '0.55')) / 100

    activos = Personal.objects.filter(estado='Activo')
    creados = 0
    for p in activos:
        # alimentacion_mensual field or similar
        alim = getattr(p, 'alimentacion_mensual', None)
        if not alim:
            # Try to get from a field pattern
            for field_name in ['alimentacion_mensual', 'alimentacion']:
                alim = getattr(p, field_name, None)
                if alim and alim > 0:
                    break
        if not alim or alim <= 0:
            continue

        monto = Decimal(str(alim))
        comision = (monto * comision_pct).quantize(Decimal('0.01'))

        _, created = RecargaAlimentacion.objects.update_or_create(
            personal=p, anio=anio, mes=mes,
            defaults={
                'monto': monto,
                'comision': comision,
                'proveedor': 'EDENRED',
            }
        )
        if created:
            creados += 1

    messages.success(request, f'Recargas generadas: {creados} empleados para {mes:02d}/{anio}')
    return redirect(f'/nominas/alimentacion/?anio={anio}&mes={mes}')


@login_required
@solo_admin
def alimentacion_exportar(request):
    """Exportar recargas a Excel para enviar al proveedor."""
    from .models import RecargaAlimentacion
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))

    recargas = RecargaAlimentacion.objects.filter(
        anio=anio, mes=mes
    ).select_related('personal').order_by('personal__apellidos_nombres')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Recarga {mes:02d}-{anio}'

    headers = ['N', 'DNI', 'NOMBRES', 'TARJETA', 'MONTO', 'COMISION', 'TOTAL', 'ESTADO']
    hdr_fill = PatternFill('solid', fgColor='0F766E')
    hdr_font = Font(bold=True, color='FFFFFF')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font

    for i, r in enumerate(recargas, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=r.personal.nro_doc)
        ws.cell(row=i+1, column=3, value=r.personal.apellidos_nombres)
        ws.cell(row=i+1, column=4, value=r.numero_tarjeta)
        ws.cell(row=i+1, column=5, value=float(r.monto)).number_format = '#,##0.00'
        ws.cell(row=i+1, column=6, value=float(r.comision)).number_format = '#,##0.00'
        ws.cell(row=i+1, column=7, value=float(r.total)).number_format = '#,##0.00'
        ws.cell(row=i+1, column=8, value=r.get_estado_display())

    for c in range(1, 9): ws.column_dimensions[chr(64+c)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Recarga_Alimentacion_{mes:02d}_{anio}.xlsx'
    return response


@login_required
@solo_admin
@require_POST
def alimentacion_procesar(request):
    """Marcar recargas pendientes como procesadas."""
    from .models import RecargaAlimentacion

    anio = int(request.POST.get('anio', date.today().year))
    mes = int(request.POST.get('mes', date.today().month))

    updated = RecargaAlimentacion.objects.filter(
        anio=anio, mes=mes, estado='PENDIENTE'
    ).update(estado='PROCESADA', procesado_en=timezone.now())

    messages.success(request, f'{updated} recargas marcadas como procesadas.')
    return redirect(f'/nominas/alimentacion/?anio={anio}&mes={mes}')
