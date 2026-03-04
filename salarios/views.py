"""
Vistas del módulo de Estructura Salarial.
Gestión de bandas salariales, historial de remuneraciones y simulaciones de incremento.
"""
from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
import statistics

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Q, Sum, Avg, Count, StdDev
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST

from core.audit import log_create, log_update
from personal.models import Personal
from .models import BandaSalarial, HistorialSalarial, SimulacionIncremento, DetalleSimulacion

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ══════════════════════════════════════════════════════════════
# ADMIN — BANDAS SALARIALES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def bandas_panel(request):
    """Panel de bandas salariales con filtros."""
    qs = BandaSalarial.objects.all()

    cargo = request.GET.get('cargo', '')
    nivel = request.GET.get('nivel', '')
    activa = request.GET.get('activa', '')

    if cargo:
        qs = qs.filter(cargo__icontains=cargo)
    if nivel:
        qs = qs.filter(nivel=nivel)
    if activa:
        qs = qs.filter(activa=(activa == '1'))

    # Listado de cargos únicos para el filtro
    cargos_unicos = BandaSalarial.objects.values_list(
        'cargo', flat=True
    ).distinct().order_by('cargo')

    # Enriquecer cada banda con conteo de empleados por rango y compa-ratio
    bandas_enriquecidas = []
    empleados_activos = Personal.objects.filter(
        estado='Activo', sueldo_base__isnull=False
    ).values('cargo', 'sueldo_base')

    # Construir lookup de sueldos por cargo
    sueldos_por_cargo = defaultdict(list)
    for emp in empleados_activos:
        if emp['cargo']:
            sueldos_por_cargo[emp['cargo'].strip()].append(emp['sueldo_base'])

    for banda in qs:
        sueldos = sueldos_por_cargo.get(banda.cargo, [])
        bajo = sum(1 for s in sueldos if s < banda.minimo)
        en_banda = sum(1 for s in sueldos if banda.minimo <= s <= banda.maximo)
        sobre = sum(1 for s in sueldos if s > banda.maximo)
        total_emp = len(sueldos)

        # Compa-ratio promedio de empleados en esta banda
        compa_ratios = []
        for s in sueldos:
            if banda.medio and banda.medio > 0:
                compa_ratios.append(round(s / banda.medio, 3))

        compa_ratio_prom = (
            round(sum(compa_ratios) / len(compa_ratios), 3)
            if compa_ratios else None
        )

        bandas_enriquecidas.append({
            'banda': banda,
            'total_empleados': total_emp,
            'bajo_banda': bajo,
            'en_banda': en_banda,
            'sobre_banda': sobre,
            'compa_ratio_prom': compa_ratio_prom,
        })

    # KPI summary for header strip
    try:
        import json as _json
        total_empleados_cubiertos = sum(item['total_empleados'] for item in bandas_enriquecidas)
        total_en_banda = sum(item['en_banda'] for item in bandas_enriquecidas)
        pct_en_banda = round(total_en_banda / total_empleados_cubiertos * 100, 1) if total_empleados_cubiertos > 0 else 0
        # Avg compa-ratio across all bands that have data
        ratios_validos = [item['compa_ratio_prom'] for item in bandas_enriquecidas if item['compa_ratio_prom'] is not None]
        compa_ratio_global = round(sum(ratios_validos) / len(ratios_validos), 3) if ratios_validos else None
        # Masa salarial activos
        from django.db.models import Sum as _Sum
        masa_result = Personal.objects.filter(estado='Activo', sueldo_base__isnull=False).aggregate(total=_Sum('sueldo_base'))
        masa_salarial_mensual = float(masa_result['total'] or 0)
        # Chart data: por nivel -- avg min/mid/max
        from .models import BandaSalarial as _BS
        niveles_data = []
        for val, label in BandaSalarial.NIVEL_CHOICES:
            bandas_nivel = BandaSalarial.objects.filter(nivel=val, activa=True)
            if bandas_nivel.exists():
                avg = bandas_nivel.aggregate(
                    avg_min=Avg('minimo'), avg_mid=Avg('medio'), avg_max=Avg('maximo')
                )
                niveles_data.append({
                    'label': label,
                    'min': float(avg['avg_min'] or 0),
                    'mid': float(avg['avg_mid'] or 0),
                    'max': float(avg['avg_max'] or 0),
                })
        bandas_niveles_json = _json.dumps(niveles_data)
    except Exception:
        total_empleados_cubiertos = 0
        pct_en_banda = 0
        compa_ratio_global = None
        masa_salarial_mensual = 0
        bandas_niveles_json = '[]'

    context = {
        'titulo': 'Bandas Salariales',
        'bandas': qs,
        'bandas_enriquecidas': bandas_enriquecidas,
        'total': qs.count(),
        'filtro_cargo': cargo,
        'filtro_nivel': nivel,
        'filtro_activa': activa,
        'cargos_unicos': cargos_unicos,
        'niveles': BandaSalarial.NIVEL_CHOICES,
        'total_empleados_cubiertos': total_empleados_cubiertos,
        'pct_en_banda': pct_en_banda,
        'compa_ratio_global': compa_ratio_global,
        'masa_salarial_mensual': masa_salarial_mensual,
        'bandas_niveles_json': bandas_niveles_json,
    }
    return render(request, 'salarios/bandas_panel.html', context)


@login_required
@solo_admin
@require_POST
def banda_crear(request):
    """Crear banda salarial (AJAX)."""
    try:
        minimo = Decimal(request.POST['minimo'])
        medio = Decimal(request.POST['medio'])
        maximo = Decimal(request.POST['maximo'])

        # Validación: mínimo <= medio <= máximo
        if not (minimo <= medio <= maximo):
            return JsonResponse(
                {'ok': False, 'error': 'Debe cumplirse: Mínimo <= Medio <= Máximo'},
                status=400,
            )

        banda = BandaSalarial.objects.create(
            cargo=request.POST['cargo'].strip(),
            nivel=request.POST['nivel'],
            minimo=minimo,
            medio=medio,
            maximo=maximo,
            moneda=request.POST.get('moneda', 'PEN'),
            activa=True,
        )
        log_create(request, banda, f'Banda salarial creada: {banda}')
        return JsonResponse({
            'ok': True, 'pk': banda.pk, 'nombre': str(banda),
        })
    except (InvalidOperation, KeyError) as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def banda_editar(request, pk):
    """Editar banda salarial (AJAX)."""
    banda = get_object_or_404(BandaSalarial, pk=pk)
    try:
        cambios = {}
        campos = {
            'cargo': str, 'nivel': str, 'moneda': str,
            'minimo': Decimal, 'medio': Decimal, 'maximo': Decimal,
        }
        for campo, tipo in campos.items():
            nuevo_val = request.POST.get(campo)
            if nuevo_val is not None:
                nuevo_val = tipo(nuevo_val.strip()) if tipo == str else tipo(nuevo_val)
                old_val = getattr(banda, campo)
                if str(old_val) != str(nuevo_val):
                    cambios[campo] = {'old': old_val, 'new': nuevo_val}
                    setattr(banda, campo, nuevo_val)

        activa_val = request.POST.get('activa')
        if activa_val is not None:
            new_activa = activa_val == '1'
            if banda.activa != new_activa:
                cambios['activa'] = {'old': banda.activa, 'new': new_activa}
                banda.activa = new_activa

        # Validación: mínimo <= medio <= máximo
        if not (banda.minimo <= banda.medio <= banda.maximo):
            return JsonResponse(
                {'ok': False, 'error': 'Debe cumplirse: Mínimo <= Medio <= Máximo'},
                status=400,
            )

        banda.save()
        if cambios:
            log_update(request, banda, cambios, f'Banda salarial editada: {banda}')
        return JsonResponse({'ok': True, 'pk': banda.pk})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


# ══════════════════════════════════════════════════════════════
# ADMIN — HISTORIAL SALARIAL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def historial_panel(request):
    """Panel de historial salarial con todos los cambios."""
    qs = HistorialSalarial.objects.select_related(
        'personal', 'aprobado_por'
    ).all()

    buscar = request.GET.get('q', '')
    motivo = request.GET.get('motivo', '')
    anio = request.GET.get('anio', '')

    if buscar:
        qs = qs.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar)
        )
    if motivo:
        qs = qs.filter(motivo=motivo)
    if anio:
        qs = qs.filter(fecha_efectiva__year=int(anio))

    # Stats
    total_incrementos = qs.count()
    incremento_promedio = qs.aggregate(
        prom=Avg('remuneracion_nueva') - Avg('remuneracion_anterior')
    )

    # Analytics
    import json as _json
    try:
        from django.db.models.functions import TruncMonth as _TM
        from datetime import date as _date
        _hoy = _date.today()
        # Movimientos por mes -- ultimos 6 meses
        movs_mes = (
            HistorialSalarial.objects
            .filter(fecha_efectiva__gte=_date(_hoy.year - 1 if _hoy.month <= 6 else _hoy.year,
                                              (_hoy.month - 6) % 12 or 12, 1))
            .annotate(mes=_TM('fecha_efectiva'))
            .values('mes')
            .annotate(total=Count('id'), promedio_inc=Avg('porcentaje_incremento'))
            .order_by('mes')
        )
        movs_mes_json = _json.dumps([
            {'label': m['mes'].strftime('%b %Y'), 'total': m['total'],
             'avg_inc': round(float(m['promedio_inc'] or 0), 1)}
            for m in movs_mes if m['mes']
        ])
        # Por motivo
        motivos_data = list(
            HistorialSalarial.objects.values('motivo')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        MOTIVO_COLORS = {'INGRESO': '#16a34a', 'INCREMENTO': '#0f766e', 'PROMOCION': '#0ea5e9', 'AJUSTE': '#d97706', 'OTRO': '#94a3b8'}
        MOTIVO_LABELS = {'INGRESO': 'Ingreso', 'INCREMENTO': 'Incremento', 'PROMOCION': 'Promoción', 'AJUSTE': 'Ajuste', 'OTRO': 'Otro'}
        motivos_json = _json.dumps([
            {'label': MOTIVO_LABELS.get(m['motivo'], m['motivo']), 'value': m['total'],
             'color': MOTIVO_COLORS.get(m['motivo'], '#94a3b8')}
            for m in motivos_data
        ])
        # KPIs anio actual
        anio_actual = _hoy.year
        movs_anio = HistorialSalarial.objects.filter(fecha_efectiva__year=anio_actual)
        total_anio = movs_anio.count()
        avg_inc_anio = movs_anio.aggregate(avg=Avg('porcentaje_incremento'))['avg']
        avg_inc_anio = round(float(avg_inc_anio), 1) if avg_inc_anio else None
        _anio_filtro = anio or str(anio_actual)
    except Exception:
        movs_mes_json = '[]'
        motivos_json = '[]'
        total_anio = 0
        avg_inc_anio = None
        _anio_filtro = anio

    context = {
        'titulo': 'Historial Salarial',
        'registros': qs[:200],
        'total': total_incrementos,
        'buscar': buscar,
        'filtro_motivo': motivo,
        'filtro_anio': anio,
        'motivos': HistorialSalarial.MOTIVO_CHOICES,
        'stats': {
            'total': total_incrementos,
        },
        'movs_mes_json': movs_mes_json,
        'motivos_json': motivos_json,
        'total_anio': total_anio,
        'avg_inc_anio': avg_inc_anio,
        'anio_filtro': _anio_filtro,
    }
    return render(request, 'salarios/historial_panel.html', context)


@login_required
@solo_admin
@require_POST
def historial_crear(request):
    """Registrar un cambio salarial individual (AJAX)."""
    try:
        personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
        remuneracion_anterior = Decimal(request.POST['remuneracion_anterior'])
        remuneracion_nueva = Decimal(request.POST['remuneracion_nueva'])

        registro = HistorialSalarial.objects.create(
            personal=personal,
            fecha_efectiva=request.POST['fecha_efectiva'],
            remuneracion_anterior=remuneracion_anterior,
            remuneracion_nueva=remuneracion_nueva,
            motivo=request.POST['motivo'],
            observaciones=request.POST.get('observaciones', ''),
            aprobado_por=request.user,
        )

        # Actualizar sueldo base del personal
        personal.sueldo_base = remuneracion_nueva
        personal.save(update_fields=['sueldo_base'])

        log_create(
            request, registro,
            f'Cambio salarial registrado: {personal.apellidos_nombres} '
            f'{remuneracion_anterior} -> {remuneracion_nueva}'
        )
        return JsonResponse({
            'ok': True, 'pk': registro.pk,
            'porcentaje': str(registro.porcentaje_incremento),
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


# ══════════════════════════════════════════════════════════════
# ADMIN — SIMULACIONES DE INCREMENTO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def simulacion_panel(request):
    """Panel de simulaciones de incremento."""
    qs = SimulacionIncremento.objects.select_related('creado_por').all()

    estado = request.GET.get('estado', '')
    if estado:
        qs = qs.filter(estado=estado)

    context = {
        'titulo': 'Simulaciones de Incremento',
        'simulaciones': qs,
        'total': qs.count(),
        'filtro_estado': estado,
        'stats': {
            'borradores': qs.filter(estado='BORRADOR').count(),
            'aprobadas': qs.filter(estado='APROBADA').count(),
            'aplicadas': qs.filter(estado='APLICADA').count(),
        },
    }
    return render(request, 'salarios/simulacion_panel.html', context)


@login_required
@solo_admin
def simulacion_crear(request):
    """Crear simulación de incremento."""
    if request.method == 'POST':
        try:
            presupuesto = request.POST.get('presupuesto_total')
            simulacion = SimulacionIncremento.objects.create(
                nombre=request.POST['nombre'].strip(),
                fecha=request.POST['fecha'],
                descripcion=request.POST.get('descripcion', ''),
                tipo=request.POST.get('tipo', 'PORCENTAJE'),
                presupuesto_total=Decimal(presupuesto) if presupuesto else None,
                creado_por=request.user,
            )

            # Si se proporcionó un valor de incremento global, crear detalles
            incremento_global = request.POST.get('incremento_global')
            area_id = request.POST.get('area_id')

            if incremento_global:
                incremento_global = Decimal(incremento_global)
                personal_qs = Personal.objects.filter(
                    estado='Activo', sueldo_base__isnull=False,
                )
                if area_id:
                    personal_qs = personal_qs.filter(subarea__area_id=int(area_id))

                detalles = []
                for emp in personal_qs:
                    if simulacion.tipo == 'PORCENTAJE':
                        monto_incremento = round(emp.sueldo_base * incremento_global / 100, 2)
                    else:
                        monto_incremento = incremento_global

                    detalles.append(DetalleSimulacion(
                        simulacion=simulacion,
                        personal=emp,
                        remuneracion_actual=emp.sueldo_base,
                        incremento_propuesto=monto_incremento,
                        aprobado=True,
                    ))
                DetalleSimulacion.objects.bulk_create(detalles)

            log_create(request, simulacion,
                       f'Simulación creada: {simulacion.nombre} ({simulacion.total_empleados} empleados)')
            messages.success(request, f'Simulación "{simulacion.nombre}" creada con {simulacion.total_empleados} empleados.')
            return redirect('simulacion_detalle', pk=simulacion.pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    from personal.models import Area
    context = {
        'titulo': 'Nueva Simulación de Incremento',
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
    }
    return render(request, 'salarios/simulacion_crear.html', context)


@login_required
@solo_admin
def simulacion_detalle(request, pk):
    """Detalle de una simulación con sus líneas."""
    simulacion = get_object_or_404(
        SimulacionIncremento.objects.select_related('creado_por'),
        pk=pk,
    )
    detalles = simulacion.detalles.select_related('personal').all()

    # Resumen estadístico
    aprobados = detalles.filter(aprobado=True)
    costo_mensual = aprobados.aggregate(total=Sum('incremento_propuesto'))['total'] or Decimal('0')

    context = {
        'titulo': f'Simulación: {simulacion.nombre}',
        'simulacion': simulacion,
        'detalles': detalles,
        'stats': {
            'total_empleados': detalles.count(),
            'aprobados': aprobados.count(),
            'rechazados': detalles.filter(aprobado=False).count(),
            'costo_mensual': costo_mensual,
            'costo_anual': costo_mensual * 12,
            'dentro_presupuesto': simulacion.dentro_presupuesto,
        },
    }
    return render(request, 'salarios/simulacion_detalle.html', context)


@login_required
@solo_admin
@require_POST
def simulacion_agregar_detalle(request, pk):
    """Agregar un trabajador a la simulación (AJAX)."""
    simulacion = get_object_or_404(SimulacionIncremento, pk=pk)
    if simulacion.estado != 'BORRADOR':
        return JsonResponse(
            {'ok': False, 'error': 'Solo se pueden modificar simulaciones en borrador.'},
            status=400,
        )

    try:
        personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
        incremento = Decimal(request.POST['incremento_propuesto'])

        detalle, created = DetalleSimulacion.objects.update_or_create(
            simulacion=simulacion,
            personal=personal,
            defaults={
                'remuneracion_actual': personal.sueldo_base or Decimal('0'),
                'incremento_propuesto': incremento,
                'aprobado': True,
            },
        )
        return JsonResponse({
            'ok': True, 'pk': detalle.pk, 'created': created,
            'remuneracion_nueva': str(detalle.remuneracion_nueva),
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def simulacion_toggle_detalle(request, pk, detalle_pk):
    """Aprobar/rechazar un detalle individual de simulación (AJAX)."""
    simulacion = get_object_or_404(SimulacionIncremento, pk=pk)
    if simulacion.estado != 'BORRADOR':
        return JsonResponse(
            {'ok': False, 'error': 'Solo se pueden modificar simulaciones en borrador.'},
            status=400,
        )

    detalle = get_object_or_404(DetalleSimulacion, pk=detalle_pk, simulacion=simulacion)
    detalle.aprobado = not detalle.aprobado
    detalle.save(update_fields=['aprobado'])
    return JsonResponse({'ok': True, 'aprobado': detalle.aprobado})


@login_required
@solo_admin
@require_POST
def simulacion_aprobar(request, pk):
    """Aprobar una simulación (cambia estado a APROBADA)."""
    simulacion = get_object_or_404(SimulacionIncremento, pk=pk)
    if simulacion.estado != 'BORRADOR':
        return JsonResponse(
            {'ok': False, 'error': 'Solo se pueden aprobar simulaciones en borrador.'},
            status=400,
        )

    simulacion.estado = 'APROBADA'
    simulacion.save(update_fields=['estado'])

    log_update(request, simulacion,
               {'estado': {'old': 'BORRADOR', 'new': 'APROBADA'}},
               f'Simulación aprobada: {simulacion.nombre}')
    return JsonResponse({'ok': True, 'estado': 'APROBADA'})


@login_required
@solo_admin
@require_POST
def simulacion_aplicar(request, pk):
    """Aplicar simulación: crea registros de HistorialSalarial y actualiza sueldos."""
    simulacion = get_object_or_404(SimulacionIncremento, pk=pk)
    if simulacion.estado not in ('BORRADOR', 'APROBADA'):
        messages.error(request, 'Esta simulación ya fue aplicada.')
        return redirect('simulacion_detalle', pk=pk)

    detalles_aprobados = simulacion.detalles.filter(aprobado=True).select_related('personal')
    if not detalles_aprobados.exists():
        messages.error(request, 'No hay detalles aprobados para aplicar.')
        return redirect('simulacion_detalle', pk=pk)

    aplicados = 0
    errores = []

    with transaction.atomic():
        for detalle in detalles_aprobados:
            try:
                # Crear registro de historial salarial
                historial = HistorialSalarial.objects.create(
                    personal=detalle.personal,
                    fecha_efectiva=simulacion.fecha,
                    remuneracion_anterior=detalle.remuneracion_actual,
                    remuneracion_nueva=detalle.remuneracion_nueva,
                    motivo='INCREMENTO',
                    observaciones=f'Aplicado desde simulación: {simulacion.nombre}',
                    aprobado_por=request.user,
                )
                log_create(
                    request, historial,
                    f'Incremento aplicado: {detalle.personal.apellidos_nombres} '
                    f'{detalle.remuneracion_actual} -> {detalle.remuneracion_nueva}'
                )

                # Actualizar sueldo base del personal
                detalle.personal.sueldo_base = detalle.remuneracion_nueva
                detalle.personal.save(update_fields=['sueldo_base'])
                aplicados += 1
            except Exception as e:
                errores.append(f'{detalle.personal.apellidos_nombres}: {e}')

        # Marcar simulación como aplicada
        simulacion.estado = 'APLICADA'
        simulacion.save(update_fields=['estado'])
        log_update(request, simulacion,
                   {'estado': {'old': 'APROBADA', 'new': 'APLICADA'}},
                   f'Simulación aplicada: {simulacion.nombre} ({aplicados} incrementos)')

    if errores:
        messages.warning(request, f'{aplicados} incrementos aplicados. Errores: {"; ".join(errores)}')
    else:
        messages.success(request, f'{aplicados} incrementos salariales aplicados exitosamente.')

    return redirect('simulacion_detalle', pk=pk)


# ══════════════════════════════════════════════════════════════
# PORTAL — MI HISTORIAL SALARIAL
# ══════════════════════════════════════════════════════════════

@login_required
def mi_historial_salarial(request):
    """Vista portal: el trabajador ve su propio historial salarial."""
    empleado = getattr(request.user, 'personal_data', None)

    registros = []
    if empleado:
        registros = HistorialSalarial.objects.filter(
            personal=empleado,
        ).order_by('-fecha_efectiva')

    context = {
        'titulo': 'Mi Historial Salarial',
        'empleado': empleado,
        'registros': registros,
    }
    return render(request, 'salarios/mi_historial.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — EXPORTAR ANÁLISIS SALARIAL (EXCEL)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def exportar_analisis_salarial(request):
    """
    Exporta en Excel el análisis salarial de todos los empleados activos:
    nombre, cargo, banda asignada, sueldo actual, compa-ratio y posición en banda.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Análisis Salarial'

    # ── Estilos ──────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='0F766E')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    fill_bajo    = PatternFill('solid', fgColor='FFF3CD')   # Amarillo — bajo banda
    fill_en      = PatternFill('solid', fgColor='D1FAE5')   # Verde — en banda
    fill_sobre   = PatternFill('solid', fgColor='FEE2E2')   # Rojo — sobre banda
    fill_sin     = PatternFill('solid', fgColor='F3F4F6')   # Gris — sin banda

    # ── Encabezado principal ─────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = 'ANÁLISIS SALARIAL — HARMONI ERP'
    ws['A1'].font = Font(bold=True, size=13, color='0F766E')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Generado: {date.today().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(italic=True, size=9, color='6B7280')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 14

    # ── Cabeceras de columna ─────────────────────────────────
    headers = [
        'N°', 'Trabajador', 'DNI', 'Cargo', 'Banda asignada',
        'Sueldo actual (S/)', 'Mínimo banda', 'Medio banda', 'Máximo banda',
        'Compa-ratio', 'Posición en banda', 'Estado',
    ]
    # Ajustamos el merge al nuevo número de columnas
    ws.merge_cells('A1:L1')
    ws.merge_cells('A2:L2')

    col_widths = [5, 36, 12, 28, 36, 16, 16, 16, 16, 13, 18, 14]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 30

    # ── Construir diccionario de bandas por cargo ────────────
    bandas_dict = {}
    for b in BandaSalarial.objects.filter(activa=True):
        bandas_dict[b.cargo.strip()] = b

    # ── Datos ────────────────────────────────────────────────
    empleados = Personal.objects.filter(
        estado='Activo'
    ).select_related('subarea__area').order_by('apellidos_nombres')

    row_num = 4
    for idx, emp in enumerate(empleados, start=1):
        sueldo = emp.sueldo_base or Decimal('0')
        banda = bandas_dict.get(emp.cargo.strip() if emp.cargo else '')

        if banda and banda.medio > 0:
            compa_ratio = round(sueldo / banda.medio, 3)
            posicion = round((sueldo - banda.minimo) / (banda.maximo - banda.minimo) * 100, 1) \
                       if (banda.maximo - banda.minimo) > 0 else Decimal('0')
            nombre_banda = str(banda)
            minimo = float(banda.minimo)
            medio  = float(banda.medio)
            maximo = float(banda.maximo)

            if sueldo < banda.minimo:
                estado_banda = 'Bajo banda'
                row_fill = fill_bajo
            elif sueldo > banda.maximo:
                estado_banda = 'Sobre banda'
                row_fill = fill_sobre
            else:
                estado_banda = 'En banda'
                row_fill = fill_en
        else:
            compa_ratio  = None
            posicion     = None
            nombre_banda = '— Sin banda asignada —'
            minimo = medio = maximo = None
            estado_banda = 'Sin banda'
            row_fill = fill_sin

        row_data = [
            idx,
            emp.apellidos_nombres,
            emp.nro_doc or '',
            emp.cargo or '',
            nombre_banda,
            float(sueldo),
            minimo,
            medio,
            maximo,
            float(compa_ratio) if compa_ratio is not None else '',
            float(posicion) if posicion is not None else '',
            estado_banda,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            # Formato numérico
            if col_idx in (6, 7, 8, 9):          # montos
                cell.number_format = '#,##0.00'
                cell.alignment = right
            elif col_idx == 10:                    # compa-ratio
                cell.number_format = '0.000'
                cell.alignment = center
            elif col_idx == 11:                    # posición %
                cell.number_format = '0.0'
                cell.alignment = center
            elif col_idx == 1:
                cell.alignment = center
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        row_num += 1

    # ── Fila de totales ───────────────────────────────────────
    ws.cell(row=row_num, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row_num, column=1).alignment = center
    ws.cell(row=row_num, column=2, value=f'{empleados.count()} empleados').font = Font(bold=True)

    # ── Freeze y filtros ──────────────────────────────────────
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:L{row_num - 1}'

    # ── Segunda hoja: Resumen por banda ──────────────────────
    ws2 = wb.create_sheet(title='Resumen por Banda')
    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 14

    headers2 = ['Banda', 'N° Empleados', 'Bajo banda', 'En banda', 'Sobre banda', 'Compa-ratio prom.']
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    sueldos_por_cargo_local = defaultdict(list)
    for emp in empleados:
        if emp.cargo and emp.sueldo_base:
            sueldos_por_cargo_local[emp.cargo.strip()].append(emp.sueldo_base)

    r2 = 2
    for banda in BandaSalarial.objects.filter(activa=True).order_by('cargo', 'nivel'):
        sueldos = sueldos_por_cargo_local.get(banda.cargo, [])
        bajo    = sum(1 for s in sueldos if s < banda.minimo)
        en_b    = sum(1 for s in sueldos if banda.minimo <= s <= banda.maximo)
        sobre   = sum(1 for s in sueldos if s > banda.maximo)
        cr_list = [round(s / banda.medio, 3) for s in sueldos if banda.medio > 0]
        cr_prom = round(sum(cr_list) / len(cr_list), 3) if cr_list else ''

        fila = [str(banda), len(sueldos), bajo, en_b, sobre, float(cr_prom) if cr_prom != '' else '']
        for ci, val in enumerate(fila, start=1):
            cell = ws2.cell(row=r2, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = center if ci != 1 else Alignment(vertical='center')
        r2 += 1

    # ── Respuesta HTTP ────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f'analisis_salarial_{date.today().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ══════════════════════════════════════════════════════════════
# ADMIN — SIMULADOR COMPARATIVO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def simulacion_comparativa(request):
    """
    Simulador comparativo: selecciona múltiples empleados y una banda objetivo,
    muestra tabla con sueldo actual vs. propuesto, incremento S/ y %, y costo total.
    Permite exportar a Excel.
    """
    try:
        from personal.models import Area
        areas = Area.objects.filter(activa=True).order_by('nombre')
    except Exception:
        areas = []

    bandas = BandaSalarial.objects.filter(activa=True).order_by('cargo', 'nivel')

    # Base queryset de empleados activos con sueldo
    empleados_qs = Personal.objects.filter(
        estado='Activo', sueldo_base__isnull=False
    ).select_related('subarea__area').order_by('apellidos_nombres')

    # Construir lookup banda actual por cargo
    bandas_dict = {b.cargo.strip(): b for b in BandaSalarial.objects.filter(activa=True)}

    resultados = []
    banda_destino = None
    empleados_sel_pks = []
    masa_actual = Decimal('0')
    masa_propuesta = Decimal('0')
    costo_ajuste = Decimal('0')

    if request.method == 'POST':
        empleados_sel_pks = [
            int(pk) for pk in request.POST.getlist('empleados') if pk.isdigit()
        ]
        banda_id = request.POST.get('banda_destino')

        if empleados_sel_pks and banda_id:
            try:
                banda_destino = BandaSalarial.objects.get(pk=int(banda_id))
            except (BandaSalarial.DoesNotExist, ValueError):
                banda_destino = None

        if empleados_sel_pks and banda_destino:
            sel_empleados = empleados_qs.filter(pk__in=empleados_sel_pks)

            for emp in sel_empleados:
                sueldo_actual = emp.sueldo_base or Decimal('0')
                banda_actual = bandas_dict.get(emp.cargo.strip() if emp.cargo else '')

                # Compa-ratio actual (respecto a banda actual)
                cr_actual = None
                if banda_actual and banda_actual.medio > 0:
                    cr_actual = round(sueldo_actual / banda_actual.medio, 3)

                # Sueldo propuesto = medio de la banda destino (ajuste al midpoint)
                sueldo_propuesto = banda_destino.medio
                incremento_monto = sueldo_propuesto - sueldo_actual
                incremento_pct = (
                    round(incremento_monto / sueldo_actual * 100, 2)
                    if sueldo_actual > 0 else Decimal('0')
                )

                # Compa-ratio propuesto
                cr_propuesto = round(
                    sueldo_propuesto / banda_destino.medio, 3
                ) if banda_destino.medio > 0 else None

                resultados.append({
                    'emp': emp,
                    'sueldo_actual': sueldo_actual,
                    'banda_actual_nombre': str(banda_actual) if banda_actual else '—',
                    'cr_actual': cr_actual,
                    'sueldo_propuesto': sueldo_propuesto,
                    'incremento_monto': incremento_monto,
                    'incremento_pct': incremento_pct,
                    'cr_propuesto': cr_propuesto,
                })

                masa_actual += sueldo_actual
                masa_propuesta += sueldo_propuesto
                costo_ajuste += incremento_monto

        # Exportar a Excel
        if request.POST.get('exportar') == '1' and resultados:
            return _exportar_simulacion_comparativa_excel(
                resultados, banda_destino, masa_actual, masa_propuesta, costo_ajuste
            )

    # Masa salarial total de todos los activos (para comparación)
    masa_total_empresa = empleados_qs.aggregate(
        total=Sum('sueldo_base')
    )['total'] or Decimal('0')

    context = {
        'titulo': 'Simulador Comparativo',
        'empleados': empleados_qs,
        'bandas': bandas,
        'areas': areas,
        'resultados': resultados,
        'banda_destino': banda_destino,
        'empleados_sel_pks': empleados_sel_pks,
        'masa_actual': masa_actual,
        'masa_propuesta': masa_propuesta,
        'costo_ajuste': costo_ajuste,
        'masa_total_empresa': masa_total_empresa,
        'pct_sobre_masa': (
            round(costo_ajuste / masa_total_empresa * 100, 2)
            if masa_total_empresa > 0 and costo_ajuste else Decimal('0')
        ),
    }
    return render(request, 'salarios/simulacion_comparativa.html', context)


def _exportar_simulacion_comparativa_excel(
    resultados, banda_destino, masa_actual, masa_propuesta, costo_ajuste
):
    """Genera el Excel del simulador comparativo."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Simulación Comparativa'

    header_fill = PatternFill('solid', fgColor='0D2B27')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    subheader_fill = PatternFill('solid', fgColor='0F766E')
    subheader_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    fill_pos = PatternFill('solid', fgColor='D1FAE5')
    fill_neg = PatternFill('solid', fgColor='FEE2E2')
    fill_neu = PatternFill('solid', fgColor='F3F4F6')

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'SIMULACIÓN COMPARATIVA SALARIAL — HARMONI ERP'
    ws['A1'].font = Font(bold=True, size=13, color='0F766E')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:H2')
    ws['A2'] = (
        f'Banda destino: {banda_destino}  |  '
        f'Generado: {date.today().strftime("%d/%m/%Y")}'
    )
    ws['A2'].font = Font(italic=True, size=9, color='6B7280')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 14

    headers = [
        'N°', 'Trabajador', 'Cargo actual',
        'Sueldo actual (S/)', 'Compa-ratio actual',
        'Sueldo propuesto (S/)', 'Incremento S/', 'Incremento %',
    ]
    col_widths = [5, 36, 28, 18, 16, 20, 16, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30

    for idx, r in enumerate(resultados, 1):
        row_fill = (
            fill_pos if r['incremento_monto'] > 0
            else fill_neg if r['incremento_monto'] < 0
            else fill_neu
        )
        row_data = [
            idx,
            r['emp'].apellidos_nombres,
            r['emp'].cargo or '',
            float(r['sueldo_actual']),
            float(r['cr_actual']) if r['cr_actual'] is not None else '',
            float(r['sueldo_propuesto']),
            float(r['incremento_monto']),
            float(r['incremento_pct']),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=3 + idx, column=ci, value=val)
            cell.fill = row_fill
            cell.border = thin
            if ci in (4, 6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = right
            elif ci in (5, 8):
                cell.number_format = '0.000' if ci == 5 else '0.00'
                cell.alignment = center
            elif ci == 1:
                cell.alignment = center
            else:
                cell.alignment = left

    # Resumen
    total_row = 3 + len(resultados) + 2
    ws.cell(row=total_row, column=1, value='RESUMEN').font = Font(bold=True, color='0D2B27')
    summaries = [
        ('Masa salarial actual (S/)', float(masa_actual)),
        ('Masa salarial propuesta (S/)', float(masa_propuesta)),
        ('Costo total del ajuste (S/)', float(costo_ajuste)),
    ]
    for i, (label, val) in enumerate(summaries):
        ws.cell(row=total_row + 1 + i, column=1, value=label).font = Font(bold=True)
        cell_val = ws.cell(row=total_row + 1 + i, column=2, value=val)
        cell_val.number_format = '#,##0.00'
        cell_val.alignment = right

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:H{3 + len(resultados)}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f'simulacion_comparativa_{date.today().strftime("%Y%m%d")}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


# ══════════════════════════════════════════════════════════════
# ADMIN — GRÁFICO VISUAL DE BANDAS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def bandas_grafico(request):
    """
    Vista con Chart.js que muestra todas las bandas activas como rangos horizontales.
    Superpone los puntos salariales de empleados actuales dentro/fuera de cada banda.
    """
    bandas = BandaSalarial.objects.filter(activa=True).order_by('cargo', 'nivel')

    # Construir lookup de sueldos por cargo
    empleados_activos = Personal.objects.filter(
        estado='Activo', sueldo_base__isnull=False
    ).values('apellidos_nombres', 'cargo', 'sueldo_base')

    sueldos_por_cargo = defaultdict(list)
    for emp in empleados_activos:
        if emp['cargo']:
            sueldos_por_cargo[emp['cargo'].strip()].append({
                'nombre': emp['apellidos_nombres'],
                'sueldo': float(emp['sueldo_base']),
            })

    # Construir datos para Chart.js
    chart_labels = []
    chart_min = []
    chart_range = []    # maximo - minimo (para barra de rango)
    chart_ref = []      # punto medio (referencia)
    empleados_puntos = []  # lista de {x, y_index, nombre, estado}

    for idx, banda in enumerate(bandas):
        label = f"{banda.cargo} / {banda.get_nivel_display()}"
        chart_labels.append(label)
        chart_min.append(float(banda.minimo))
        chart_range.append(float(banda.maximo - banda.minimo))
        chart_ref.append(float(banda.medio))

        for emp_info in sueldos_por_cargo.get(banda.cargo.strip(), []):
            s = emp_info['sueldo']
            if s < float(banda.minimo):
                estado = 'below'
            elif s > float(banda.maximo):
                estado = 'above'
            else:
                estado = 'within'
            empleados_puntos.append({
                'x': s,
                'y': idx,
                'nombre': emp_info['nombre'],
                'estado': estado,
            })

    import json
    context = {
        'titulo': 'Gráfico Visual de Bandas Salariales',
        'bandas': bandas,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_min_json': json.dumps(chart_min),
        'chart_range_json': json.dumps(chart_range),
        'chart_ref_json': json.dumps(chart_ref),
        'empleados_puntos_json': json.dumps(empleados_puntos),
        'total_bandas': bandas.count(),
        'total_empleados_graficados': len(empleados_puntos),
    }
    return render(request, 'salarios/bandas_grafico.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — ANÁLISIS DE EQUIDAD SALARIAL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def equidad_salarial(request):
    """
    Análisis de equidad salarial por género, tipo de trabajador y área.
    Calcula brecha salarial de género y muestra gráfico de barras agrupadas Chart.js.
    """
    try:
        from personal.models import Area
        areas_qs = Area.objects.filter(activa=True).order_by('nombre')
    except Exception:
        areas_qs = []

    # Filtros opcionales
    filtro_area = request.GET.get('area', '')
    filtro_grupo = request.GET.get('grupo', '')

    empleados = Personal.objects.filter(
        estado='Activo', sueldo_base__isnull=False, sueldo_base__gt=0
    ).select_related('subarea__area')

    if filtro_area:
        try:
            empleados = empleados.filter(subarea__area_id=int(filtro_area))
        except (ValueError, TypeError):
            pass
    if filtro_grupo:
        empleados = empleados.filter(grupo_tareo=filtro_grupo)

    # ── Análisis por género ────────────────────────────────────
    def calcular_stats(sueldos):
        if not sueldos:
            return None
        n = len(sueldos)
        prom = round(sum(sueldos) / n, 2)
        med = round(statistics.median(sueldos), 2)
        mn = round(min(sueldos), 2)
        mx = round(max(sueldos), 2)
        desv = round(statistics.stdev(sueldos), 2) if n > 1 else Decimal('0')
        return {'n': n, 'promedio': prom, 'mediana': med, 'minimo': mn, 'maximo': mx, 'desviacion': desv}

    sueldos_m = [float(e.sueldo_base) for e in empleados if e.sexo == 'M']
    sueldos_f = [float(e.sueldo_base) for e in empleados if e.sexo == 'F']

    stats_genero = {
        'M': calcular_stats(sueldos_m),
        'F': calcular_stats(sueldos_f),
    }

    # Brecha salarial de género (%)
    brecha_genero = None
    if stats_genero['M'] and stats_genero['F']:
        prom_m = stats_genero['M']['promedio']
        prom_f = stats_genero['F']['promedio']
        if prom_m > 0:
            brecha_genero = round((prom_m - prom_f) / prom_m * 100, 2)

    # ── Análisis por grupo_tareo ───────────────────────────────
    grupos = ['STAFF', 'RCO', 'OTRO']
    stats_grupo = {}
    for g in grupos:
        sueldos_g = [float(e.sueldo_base) for e in empleados if e.grupo_tareo == g]
        if sueldos_g:
            stats_grupo[g] = calcular_stats(sueldos_g)

    # ── Análisis por área (para Chart.js) ─────────────────────
    area_data = defaultdict(lambda: {'M': [], 'F': [], 'todos': []})
    for emp in empleados:
        area_nombre = (
            emp.subarea.area.nombre
            if emp.subarea and emp.subarea.area
            else 'Sin área'
        )
        sueldo_f = float(emp.sueldo_base)
        area_data[area_nombre]['todos'].append(sueldo_f)
        if emp.sexo == 'M':
            area_data[area_nombre]['M'].append(sueldo_f)
        elif emp.sexo == 'F':
            area_data[area_nombre]['F'].append(sueldo_f)

    # Ordenar por nombre de área y construir datos para Chart.js
    areas_ordenadas = sorted(area_data.keys())
    chart_labels = []
    chart_prom_m = []
    chart_prom_f = []
    chart_prom_total = []
    tabla_areas = []

    for area_n in areas_ordenadas:
        d = area_data[area_n]
        prom_m_a = round(sum(d['M']) / len(d['M']), 2) if d['M'] else None
        prom_f_a = round(sum(d['F']) / len(d['F']), 2) if d['F'] else None
        prom_tot = round(sum(d['todos']) / len(d['todos']), 2) if d['todos'] else None

        chart_labels.append(area_n)
        chart_prom_m.append(prom_m_a)
        chart_prom_f.append(prom_f_a)
        chart_prom_total.append(prom_tot)

        brecha_a = None
        if prom_m_a and prom_f_a and prom_m_a > 0:
            brecha_a = round((prom_m_a - prom_f_a) / prom_m_a * 100, 2)

        tabla_areas.append({
            'area': area_n,
            'n_total': len(d['todos']),
            'n_m': len(d['M']),
            'n_f': len(d['F']),
            'prom_m': prom_m_a,
            'prom_f': prom_f_a,
            'prom_total': prom_tot,
            'brecha': brecha_a,
        })

    import json
    context = {
        'titulo': 'Análisis de Equidad Salarial',
        'areas': areas_qs,
        'filtro_area': filtro_area,
        'filtro_grupo': filtro_grupo,
        'stats_genero': stats_genero,
        'brecha_genero': brecha_genero,
        'stats_grupo': stats_grupo,
        'tabla_areas': tabla_areas,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_prom_m_json': json.dumps(chart_prom_m),
        'chart_prom_f_json': json.dumps(chart_prom_f),
        'chart_prom_total_json': json.dumps(chart_prom_total),
        'total_analizados': empleados.count(),
        'grupos_choices': Personal.GRUPO_TAREO_CHOICES,
    }
    return render(request, 'salarios/equidad_salarial.html', context)
