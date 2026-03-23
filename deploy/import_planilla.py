"""Importar planilla (normal + RIA) desde Excel."""
import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.production')
django.setup()

import openpyxl
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from django.contrib.auth.models import User
from nominas.models import PeriodoNomina, RegistroNomina
from personal.models import Personal
from empresas.models import Empresa

MES = int(sys.argv[1])
ANIO = int(sys.argv[2])
ARCHIVO_NORMAL = sys.argv[3]
ARCHIVO_RIA = sys.argv[4] if len(sys.argv) > 4 else None

empresa = Empresa.objects.first()
admin = User.objects.get(username='admin')
personal_cache = {}
for p in Personal.objects.all():
    personal_cache[p.nro_doc] = p
    personal_cache[p.nro_doc.lstrip('0')] = p  # sin ceros
    if len(p.nro_doc) < 8:
        personal_cache[p.nro_doc.zfill(8)] = p  # con ceros
        personal_cache[p.nro_doc.zfill(9)] = p

def D(val):
    """Safe Decimal."""
    if not val or val == '-' or val == 'None':
        return Decimal('0')
    try:
        return Decimal(str(val)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return Decimal('0')

def safe_int(val):
    if not val: return 0
    try: return int(float(val))
    except: return 0

# Crear periodo
ultimo_dia = 28 if MES == 2 else 30 if MES in (4, 6, 9, 11) else 31
periodo, created = PeriodoNomina.objects.get_or_create(
    empresa=empresa,
    anio=ANIO,
    mes=MES,
    tipo='mensual',
    defaults={
        'descripcion': f'Planilla {MES:02d}/{ANIO}',
        'fecha_inicio': date(ANIO, MES, 1),
        'fecha_fin': date(ANIO, MES, ultimo_dia),
        'fecha_pago': date(ANIO, MES, ultimo_dia),
        'estado': 'cerrado',
        'generado_por': admin,
        'total_trabajadores': 0,
        'total_bruto': Decimal('0'),
        'total_descuentos': Decimal('0'),
        'total_neto': Decimal('0'),
        'total_costo_empresa': Decimal('0'),
    }
)
print(f'Periodo {MES:02d}/{ANIO}: {"creado" if created else "ya existía"} (ID: {periodo.id})')

total_creados = 0
total_errores = 0
errores = []
totales = {'bruto': Decimal('0'), 'descuentos': Decimal('0'), 'neto': Decimal('0'), 'costo': Decimal('0')}

def parse_afp(regimen_str):
    """Extraer nombre de AFP."""
    if not regimen_str:
        return '', 'AFP'
    r = str(regimen_str).upper()
    if 'ONP' in r:
        return '', 'ONP'
    if 'HABITAT' in r:
        return 'Habitat', 'AFP'
    if 'INTEGRA' in r:
        return 'Integra', 'AFP'
    if 'PRIMA' in r:
        return 'Prima', 'AFP'
    if 'PROFUTURO' in r:
        return 'Profuturo', 'AFP'
    return '', 'AFP'

def importar_planilla(archivo, grupo_default='STAFF'):
    global total_creados, total_errores

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)

    # Buscar hoja VistaNomina
    vista_sheet = None
    for sn in wb.sheetnames:
        if sn.startswith('VistaNomina'):
            vista_sheet = sn
            break

    if not vista_sheet:
        print(f'  [WARN] No se encontró hoja VistaNomina en {archivo}')
        wb.close()
        return

    ws = wb[vista_sheet]
    print(f'  Procesando {vista_sheet} ({ws.max_row} filas)...')

    # Leer headers
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for i, col in enumerate(row):
            if col:
                headers[str(col).strip()] = i

    # Mapeo de columnas
    col_map = {
        'dni': next((headers[k] for k in headers if 'Documento Identidad' in k), None),
        'nombre': next((headers[k] for k in headers if k == 'Nombre'), None),
        'area': next((headers[k] for k in headers if k == 'Area'), None),
        'regimen_pension': next((headers[k] for k in headers if 'Regimen' in k and 'Pensi' in k), None),
        'cuspp': next((headers[k] for k in headers if 'Cuspp' in k), None),
        'horas_oficina': next((headers[k] for k in headers if 'HORAS OFICINA' in k or 'HRS. NORMALES' in k), None),
        'he_25': next((headers[k] for k in headers if 'EXTRAS 25%' in k and 'Hora' in k), None),
        'he_35': next((headers[k] for k in headers if 'EXTRAS 35%' in k and 'Hora' in k), None),
        'he_100': next((headers[k] for k in headers if 'EXTRAS 100%' in k and 'Hora' in k), None),
        'dias_basico': next((headers[k] for k in headers if 'DIAS MES BASICO' in k or 'DIA TRABAJO' in k), None),
        'dias_dm': next((headers[k] for k in headers if 'DESCANSO MEDICO' in k and 'a' in k.lower()), None),
        'dias_lsg': next((headers[k] for k in headers if 'LICENCIA S/GOCE' in k), None),
        'dias_falta': next((headers[k] for k in headers if 'FALTA' in k and 'a' in k.lower()), None),
        'dias_vac': next((headers[k] for k in headers if 'DESCANSO VACACIONAL' in k), None),
        'sueldo_basico': next((headers[k] for k in headers if 'REMUNERACION O JORNAL' in k), None),
        'rem_mensual': next((headers[k] for k in headers if 'REMUNERACION MENSUAL' in k), None),
        'condicion_trabajo': next((headers[k] for k in headers if 'CONDICION DE TRABAJO' in k and 'ADELANTO' not in k), None),
        'asig_familiar': next((headers[k] for k in headers if 'ASIGNACION FAMILIAR' in k), None),
        'he_25_soles': next((headers[k] for k in headers if 'HE 25%' in k and 'S/' in k), None),
        'he_35_soles': next((headers[k] for k in headers if 'HE 35%' in k and 'S/' in k), None),
        'he_100_soles': next((headers[k] for k in headers if 'HE 100%' in k and 'S/' in k and 'REINT' not in k), None),
        'otros_ingresos': next((headers[k] for k in headers if 'OTROS INGRESOS AFECTOS' in k), None),
        'total_rem': next((headers[k] for k in headers if 'TOTAL REMUNERACION' in k), None),
        'onp': next((headers[k] for k in headers if k.startswith('ONP')), None),
        'afp_pension': next((headers[k] for k in headers if 'AFP APORTE PENSION' in k), None),
        'afp_seguros': next((headers[k] for k in headers if 'AFP PRIMA SEGUROS' in k), None),
        'afp_comision': next((headers[k] for k in headers if 'AFP COMISION' in k), None),
        'dscto_prestamo': next((headers[k] for k in headers if 'DSCTO. PRESTAMO' in k or 'DSCTO PRESTAMO' in k), None),
        'ir_5ta': next((headers[k] for k in headers if '5TA' in k), None),
        'total_descuento': next((headers[k] for k in headers if 'TOTAL DESCUENTO' in k), None),
        'total_neto': next((headers[k] for k in headers if 'TOTAL NETO' in k), None),
        'essalud': next((headers[k] for k in headers if 'ESSALUD' in k and 'BASE' not in k and 'EPS' not in k), None),
        'costo_empresa': next((headers[k] for k in headers if 'Costo empresa' in k or 'costo empresa' in k.lower()), None),
    }

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col_map['dni']]:
            continue

        dni = str(row[col_map['dni']]).strip()
        personal = personal_cache.get(dni)
        if not personal:
            # Intentar sin ceros iniciales
            dni_clean = dni.lstrip('0')
            personal = personal_cache.get(dni_clean)
            if not personal:
                total_errores += 1
                if len(errores) < 10:
                    errores.append(f'{dni}: No encontrado en Personal')
                continue

        afp_nombre, regimen = parse_afp(row[col_map['regimen_pension']] if col_map['regimen_pension'] is not None else '')

        # Calcular totales
        he_25_s = D(row[col_map['he_25_soles']]) if col_map['he_25_soles'] is not None else Decimal('0')
        he_35_s = D(row[col_map['he_35_soles']]) if col_map['he_35_soles'] is not None else Decimal('0')
        he_100_s = D(row[col_map['he_100_soles']]) if col_map['he_100_soles'] is not None else Decimal('0')
        otros_ing = D(row[col_map['otros_ingresos']]) if col_map['otros_ingresos'] is not None else Decimal('0')
        otros_ing = otros_ing + he_25_s + he_35_s + he_100_s

        condicion_t = D(row[col_map['condicion_trabajo']]) if col_map['condicion_trabajo'] is not None else Decimal('0')
        asig_fam = D(row[col_map['asig_familiar']]) if col_map['asig_familiar'] is not None else Decimal('0')

        total_ingresos = D(row[col_map['total_rem']]) if col_map['total_rem'] is not None else Decimal('0')
        total_desc = D(row[col_map['total_descuento']]) if col_map['total_descuento'] is not None else Decimal('0')
        neto = D(row[col_map['total_neto']]) if col_map['total_neto'] is not None else Decimal('0')
        essalud = D(row[col_map['essalud']]) if col_map['essalud'] is not None else Decimal('0')
        costo = D(row[col_map['costo_empresa']]) if col_map['costo_empresa'] is not None else Decimal('0')
        dscto_prest = D(row[col_map['dscto_prestamo']]) if col_map['dscto_prestamo'] is not None else Decimal('0')

        # ONP o AFP
        onp_val = D(row[col_map['onp']]) if col_map['onp'] is not None else Decimal('0')
        afp_val = D(row[col_map['afp_pension']]) if col_map['afp_pension'] is not None else Decimal('0')
        ir_5ta = D(row[col_map['ir_5ta']]) if col_map['ir_5ta'] is not None else Decimal('0')

        otros_desc = total_desc - dscto_prest - onp_val - afp_val - ir_5ta
        if col_map['afp_seguros'] is not None:
            afp_seg = D(row[col_map['afp_seguros']])
            otros_desc -= afp_seg
        if col_map['afp_comision'] is not None:
            afp_com = D(row[col_map['afp_comision']])
            otros_desc -= afp_com
        if otros_desc < 0:
            otros_desc = Decimal('0')

        try:
            reg, was_created = RegistroNomina.objects.update_or_create(
                periodo=periodo,
                personal=personal,
                defaults={
                    'sueldo_base': D(row[col_map['sueldo_basico']]) if col_map['sueldo_basico'] is not None else personal.sueldo_base,
                    'regimen_pension': regimen,
                    'afp': afp_nombre,
                    'grupo': grupo_default,
                    'dias_trabajados': safe_int(row[col_map['dias_basico']]) if col_map['dias_basico'] is not None else 30,
                    'dias_descanso': safe_int(row[col_map['dias_vac']]) if col_map['dias_vac'] is not None else 0,
                    'dias_falta': safe_int(row[col_map['dias_falta']]) if col_map['dias_falta'] is not None else 0,
                    'horas_extra_25': D(row[col_map['he_25']]) if col_map['he_25'] is not None else Decimal('0'),
                    'horas_extra_35': D(row[col_map['he_35']]) if col_map['he_35'] is not None else Decimal('0'),
                    'horas_extra_100': D(row[col_map['he_100']]) if col_map['he_100'] is not None else Decimal('0'),
                    'asignacion_familiar': asig_fam > 0,
                    'descuento_prestamo': dscto_prest,
                    'otros_ingresos': otros_ing + condicion_t + asig_fam,
                    'otros_descuentos': otros_desc,
                    'total_ingresos': total_ingresos,
                    'total_descuentos': total_desc,
                    'neto_a_pagar': neto,
                    'aporte_essalud': essalud,
                    'costo_total_empresa': costo,
                    'estado': 'cerrado',
                }
            )
            if was_created:
                created += 1

            totales['bruto'] += total_ingresos
            totales['descuentos'] += total_desc
            totales['neto'] += neto
            totales['costo'] += costo

        except Exception as e:
            total_errores += 1
            if len(errores) < 10:
                errores.append(f'{dni}: {e}')

    total_creados += created
    print(f'  {created} registros creados')
    wb.close()


# --- Ejecutar ---
print(f'\nImportando planilla {MES:02d}/{ANIO}')
print(f'Personal en BD: {len(personal_cache)}')

print(f'\n1. Planilla Normal: {ARCHIVO_NORMAL}')
importar_planilla(ARCHIVO_NORMAL, 'STAFF')

if ARCHIVO_RIA:
    print(f'\n2. Planilla RIA: {ARCHIVO_RIA}')
    importar_planilla(ARCHIVO_RIA, 'RIA')

# Actualizar totales del periodo
periodo.total_trabajadores = RegistroNomina.objects.filter(periodo=periodo).count()
periodo.total_bruto = totales['bruto']
periodo.total_descuentos = totales['descuentos']
periodo.total_neto = totales['neto']
periodo.total_costo_empresa = totales['costo']
periodo.save()

print(f'\n=== RESULTADO {MES:02d}/{ANIO} ===')
print(f'Registros creados: {total_creados}')
print(f'Errores: {total_errores}')
print(f'Total trabajadores: {periodo.total_trabajadores}')
print(f'Total bruto: S/ {periodo.total_bruto:,.2f}')
print(f'Total descuentos: S/ {periodo.total_descuentos:,.2f}')
print(f'Total neto: S/ {periodo.total_neto:,.2f}')
print(f'Costo empresa: S/ {periodo.total_costo_empresa:,.2f}')
if errores:
    print('\nErrores:')
    for e in errores:
        print(f'  {e}')
