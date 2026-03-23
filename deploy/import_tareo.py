"""Importar tareo desde Excel de incidencias (STAFF + RCO)."""
import json, os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.production')
django.setup()

import openpyxl
from datetime import datetime, date, time
from decimal import Decimal
from django.contrib.auth.models import User
from asistencia.models import RegistroTareo, TareoImportacion
from personal.models import Personal

# Args: archivo mes año
ARCHIVO = sys.argv[1]
MES = int(sys.argv[2])
ANIO = int(sys.argv[3])

wb = openpyxl.load_workbook(ARCHIVO, read_only=True, data_only=True)

# Crear registro de importación
admin = User.objects.get(username='admin')
importacion = TareoImportacion.objects.create(
    tipo='EXCEL',
    archivo_nombre=os.path.basename(ARCHIVO),
    periodo_inicio=date(ANIO, MES, 1),
    periodo_fin=date(ANIO, MES, 28 if MES == 2 else 30 if MES in (4,6,9,11) else 31),
    estado='procesando',
    usuario=admin,
    total_registros=0,
    registros_ok=0,
    registros_error=0,
    registros_sin_match=0,
    errores=[],
    advertencias=[],
    metadata={'fuente': 'import_tareo.py'},
)

# Cache de personal por DNI
personal_cache = {p.nro_doc: p for p in Personal.objects.filter(estado='Activo')}

total_creados = 0
total_errores = 0
errores = []

def safe_decimal(val):
    if not val or val == '-' or val == 'None': return Decimal('0')
    try: return Decimal(str(val))
    except: return Decimal('0')

def safe_time(val):
    if not val: return None
    if isinstance(val, time): return val
    if isinstance(val, datetime): return val.time()
    try:
        s = str(val).strip()
        if ':' in s:
            parts = s.split(':')
            return time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
    except: pass
    return None

def parse_staff_sheet(ws, sheet_name):
    """Parse CSRT STAFF - códigos diarios por empleado."""
    global total_creados, total_errores

    # Encontrar fila de fechas y fila de datos
    fecha_row = None
    data_start = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), start=1):
        # Buscar fila con fechas (datetime en col 8+)
        for cell in row[8:50]:
            if isinstance(cell.value, datetime):
                fecha_row = i
                break
        if fecha_row: break

    if not fecha_row:
        print(f'  [WARN] No se encontró fila de fechas en {sheet_name}')
        return

    # Leer fechas de las columnas
    fechas = {}
    for row in ws.iter_rows(min_row=fecha_row, max_row=fecha_row, values_only=False):
        for cell in row[8:]:
            if isinstance(cell.value, datetime):
                d = cell.value.date()
                if d.month == MES and d.year == ANIO:
                    fechas[cell.column - 1] = d  # 0-indexed

    if not fechas:
        print(f'  [WARN] No hay fechas de {MES:02d}/{ANIO} en {sheet_name}')
        return

    print(f'  Fechas encontradas: {len(fechas)} días del mes {MES:02d}/{ANIO}')

    # Leer empleados (filas después de fecha_row + 1)
    created = 0
    for row in ws.iter_rows(min_row=fecha_row + 2, values_only=False):
        dni = str(row[1].value).strip() if row[1].value else None
        if not dni or not dni.replace(' ', '').isdigit():
            continue
        dni = dni.strip()

        personal = personal_cache.get(dni)
        if not personal:
            continue

        condicion = str(row[3].value).strip() if row[3].value else ''

        for col_idx, fecha in fechas.items():
            codigo = str(row[col_idx].value).strip().upper() if row[col_idx].value else ''
            if not codigo:
                continue

            try:
                reg, was_created = RegistroTareo.objects.update_or_create(
                    personal=personal,
                    fecha=fecha,
                    defaults={
                        'dni': dni,
                        'codigo_dia': codigo,
                        'condicion': condicion,
                        'grupo': 'STAFF',
                        'dia_semana': fecha.weekday(),
                        'horas_normales': Decimal('8.5') if codigo == 'NOR' else Decimal('0'),
                        'horas_efectivas': Decimal('8.5') if codigo == 'NOR' else Decimal('0'),
                        'fuente_codigo': 'EXCEL',
                        'nombre_archivo': os.path.basename(ARCHIVO),
                        'importacion': importacion,
                    }
                )
                if was_created:
                    created += 1
            except Exception as e:
                total_errores += 1
                if len(errores) < 20:
                    errores.append(f'{dni} {fecha}: {e}')

    total_creados += created
    print(f'  {sheet_name}: {created} registros creados')


def parse_rco_sheet(ws, sheet_name):
    """Parse CSRT RCO - con horas extras detalladas."""
    global total_creados, total_errores

    # Encontrar fila de fechas (fila 6 tiene fechas en cols cada 9)
    fecha_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        for cell in row[8:]:
            if isinstance(cell.value, datetime):
                fecha_row = i
                break
        if fecha_row: break

    if not fecha_row:
        print(f'  [WARN] No se encontró fila de fechas en {sheet_name}')
        return

    # Las fechas están cada 9 columnas: Entrada|Salida|Día|TotalH|HorasEfec|Normal|0.25|0.35|1
    fechas = {}  # col_base -> fecha
    for row in ws.iter_rows(min_row=fecha_row, max_row=fecha_row, values_only=False):
        for cell in row[8:]:
            if isinstance(cell.value, datetime):
                d = cell.value.date()
                if d.month == MES and d.year == ANIO:
                    fechas[cell.column - 1] = d  # col de "Entrada"

    if not fechas:
        print(f'  [WARN] No hay fechas de {MES:02d}/{ANIO} en {sheet_name}')
        return

    print(f'  Fechas encontradas: {len(fechas)} días')

    # Fila de headers (fecha_row + 1)
    header_row = fecha_row + 1

    # Leer empleados
    created = 0
    for row in ws.iter_rows(min_row=header_row + 2, values_only=False):
        dni = str(row[1].value).strip() if row[1].value else None
        if not dni or not dni.replace(' ', '').isdigit():
            continue

        personal = personal_cache.get(dni)
        if not personal:
            continue

        condicion = str(row[3].value).strip() if row[3].value else ''

        for col_base, fecha in fechas.items():
            # Estructura: col_base=Entrada, +1=Salida, +2=Día, +3=TotalH, +4=HorasEfec, +5=Normal, +6=0.25, +7=0.35, +8=1
            entrada = safe_time(row[col_base].value) if col_base < len(row) else None
            salida = safe_time(row[col_base + 1].value) if col_base + 1 < len(row) else None
            codigo = str(row[col_base + 2].value).strip().upper() if col_base + 2 < len(row) and row[col_base + 2].value else ''
            total_h = safe_decimal(row[col_base + 3].value) if col_base + 3 < len(row) else Decimal('0')
            horas_efec = safe_decimal(row[col_base + 4].value) if col_base + 4 < len(row) else Decimal('0')
            horas_norm = safe_decimal(row[col_base + 5].value) if col_base + 5 < len(row) else Decimal('0')
            he_25 = safe_decimal(row[col_base + 6].value) if col_base + 6 < len(row) else Decimal('0')
            he_35 = safe_decimal(row[col_base + 7].value) if col_base + 7 < len(row) else Decimal('0')
            he_100 = safe_decimal(row[col_base + 8].value) if col_base + 8 < len(row) else Decimal('0')

            if not codigo:
                continue

            try:
                reg, was_created = RegistroTareo.objects.update_or_create(
                    personal=personal,
                    fecha=fecha,
                    defaults={
                        'dni': dni,
                        'codigo_dia': codigo,
                        'condicion': condicion,
                        'grupo': 'RCO',
                        'dia_semana': fecha.weekday(),
                        'hora_entrada_real': entrada,
                        'hora_salida_real': salida,
                        'horas_marcadas': total_h,
                        'horas_efectivas': horas_efec if horas_efec else Decimal('0'),
                        'horas_normales': horas_norm,
                        'he_25': he_25,
                        'he_35': he_35,
                        'he_100': he_100,
                        'fuente_codigo': 'EXCEL',
                        'nombre_archivo': os.path.basename(ARCHIVO),
                        'importacion': importacion,
                    }
                )
                if was_created:
                    created += 1
            except Exception as e:
                total_errores += 1
                if len(errores) < 20:
                    errores.append(f'{dni} {fecha}: {e}')

    total_creados += created
    print(f'  {sheet_name}: {created} registros creados')


# --- Procesar hojas ---
print(f'\nImportando tareo de {ARCHIVO}')
print(f'Periodo: {MES:02d}/{ANIO}')
print(f'Personal en BD: {len(personal_cache)}')

if 'CSRT STAFF' in wb.sheetnames:
    print('\nProcesando CSRT STAFF...')
    parse_staff_sheet(wb['CSRT STAFF'], 'CSRT STAFF')

if 'CSRT RCO' in wb.sheetnames:
    print('\nProcesando CSRT RCO...')
    parse_rco_sheet(wb['CSRT RCO'], 'CSRT RCO')

wb.close()

# Actualizar importación
from django.utils import timezone
importacion.estado = 'completado' if total_errores == 0 else 'completado_con_errores'
importacion.total_registros = total_creados + total_errores
importacion.registros_ok = total_creados
importacion.registros_error = total_errores
importacion.errores = errores[:50]
importacion.procesado_en = timezone.now()
importacion.save()

print(f'\n=== RESULTADO ===')
print(f'Total registros creados: {total_creados}')
print(f'Total errores: {total_errores}')
print(f'Total tareos en BD: {RegistroTareo.objects.count()}')
if errores:
    print('Errores:')
    for e in errores:
        print(f'  {e}')
