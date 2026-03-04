"""
Vistas del módulo Documentos - Legajo Digital del Trabajador.
"""
import json
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Q, Exists, OuterRef
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.views.decorators.http import require_POST

from documentos.models import (
    CategoriaDocumento, TipoDocumento, DocumentoTrabajador, PlantillaConstancia,
    ConstanciaGenerada, BoletaPago, DocumentoLaboral, EntregaDocumento,
    PlantillaDossier, PlantillaDossierItem, Dossier, DossierPersonal, DossierItem,
)

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ── Panel Principal ──────────────────────────────────────────────

@login_required
@solo_admin
def panel_documentos(request):
    """Panel principal de documentos: estadísticas y accesos rápidos."""
    from personal.models import Personal

    _recalcular_vencimientos()

    try:
        total_docs = DocumentoTrabajador.objects.exclude(estado="ANULADO").count()
        vencidos = DocumentoTrabajador.objects.filter(estado="VENCIDO").count()
        por_vencer = DocumentoTrabajador.objects.filter(estado="POR_VENCER").count()
        total_documentos = total_docs
    except Exception:
        total_docs = vencidos = por_vencer = total_documentos = 0

    try:
        from datetime import date as _date
        hoy = _date.today()
        docs_este_mes = DocumentoTrabajador.objects.filter(
            creado_en__year=hoy.year,
            creado_en__month=hoy.month,
        ).exclude(estado="ANULADO").count()
    except Exception:
        docs_este_mes = 0

    try:
        personal_qs = Personal.objects.filter(estado="Activo").order_by("apellidos_nombres")
        total_personal = personal_qs.count()
    except Exception:
        personal_qs = Personal.objects.none()
        total_personal = 0

    try:
        empleados_con_legajo = personal_qs.filter(
            Exists(DocumentoTrabajador.objects.filter(
                personal=OuterRef("pk"),
                estado__in=["VIGENTE", "POR_VENCER"],
            ))
        ).count()
        empleados_sin_legajo = personal_qs.filter(
            ~Exists(DocumentoTrabajador.objects.filter(
                personal=OuterRef("pk"),
                estado__in=["VIGENTE", "POR_VENCER"],
            ))
        ).count()
    except Exception:
        empleados_con_legajo = empleados_sin_legajo = 0

    PALETTE = [
        "#0f766e", "#14b8a6", "#5eead4", "#0d9488", "#134e4a",
        "#2dd4bf", "#99f6e4", "#6ee7b7", "#a7f3d0", "#d1fae5",
    ]
    try:
        docs_por_tipo = list(
            DocumentoTrabajador.objects
            .exclude(estado="ANULADO")
            .values("tipo__nombre")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        docs_por_tipo_json = json.dumps([
            {"label": (item["tipo__nombre"] or "Sin tipo"), "value": item["total"], "color": PALETTE[i % len(PALETTE)]}
            for i, item in enumerate(docs_por_tipo)
        ])
    except Exception:
        docs_por_tipo = []
        docs_por_tipo_json = json.dumps([])

    AREA_PALETTE = ["#0f766e", "#14b8a6", "#5eead4", "#0d9488", "#134e4a"]
    try:
        docs_por_area_raw = list(
            DocumentoTrabajador.objects
            .exclude(estado="ANULADO")
            .filter(personal__subarea__area__isnull=False)
            .values("personal__subarea__area__nombre")
            .annotate(total=Count("id"))
            .order_by("-total")[:5]
        )
        docs_por_area_json = json.dumps([
            {"label": (item["personal__subarea__area__nombre"] or "Sin area"), "value": item["total"], "color": AREA_PALETTE[i % len(AREA_PALETTE)]}
            for i, item in enumerate(docs_por_area_raw)
        ])
    except Exception:
        docs_por_area_json = json.dumps([])

    try:
        docs_recientes = DocumentoTrabajador.objects.select_related(
            "personal", "tipo", "subido_por",
        ).exclude(estado="ANULADO").order_by("-creado_en")[:5]
    except Exception:
        docs_recientes = []

    try:
        tipos_oblig_total = TipoDocumento.objects.filter(obligatorio=True, activo=True).count()
        personal_con_docs = personal_qs.annotate(
            num_docs=Count("documentos", filter=Q(documentos__estado__in=["VIGENTE", "POR_VENCER"])),
            tiene_vencido=Exists(DocumentoTrabajador.objects.filter(personal=OuterRef("pk"), estado="VENCIDO")),
            tiene_por_vencer=Exists(DocumentoTrabajador.objects.filter(personal=OuterRef("pk"), estado="POR_VENCER")),
        )
        buscar = request.GET.get("buscar", "")
        if buscar:
            personal_con_docs = personal_con_docs.filter(
                Q(apellidos_nombres__icontains=buscar) | Q(nro_doc__icontains=buscar)
            )
        personal_list_raw = list(personal_con_docs[:200])
        if tipos_oblig_total > 0:
            for emp in personal_list_raw:
                if emp.grupo_tareo == "STAFF":
                    tipos_aplicables_count = TipoDocumento.objects.filter(obligatorio=True, activo=True, aplica_staff=True).count()
                else:
                    tipos_aplicables_count = TipoDocumento.objects.filter(obligatorio=True, activo=True, aplica_rco=True).count()
                docs_oblig_count = DocumentoTrabajador.objects.filter(
                    personal=emp,
                    tipo__obligatorio=True,
                    estado__in=["VIGENTE", "POR_VENCER"],
                ).values("tipo_id").distinct().count()
                emp.pct_completo = min(100, round(docs_oblig_count / tipos_aplicables_count * 100)) if tipos_aplicables_count > 0 else 100
        else:
            for emp in personal_list_raw:
                emp.pct_completo = 100
    except Exception:
        personal_list_raw = []
        buscar = ""

    try:
        recientes = DocumentoTrabajador.objects.select_related(
            "personal", "tipo", "subido_por",
        ).exclude(estado="ANULADO").order_by("-creado_en")[:10]
    except Exception:
        recientes = []

    try:
        alertas = DocumentoTrabajador.objects.filter(
            estado__in=["VENCIDO", "POR_VENCER"],
        ).select_related("personal", "tipo").order_by("fecha_vencimiento")[:20]
    except Exception:
        alertas = []

    try:
        lista_sin_legajo = list(
            personal_qs.filter(
                ~Exists(DocumentoTrabajador.objects.filter(
                    personal=OuterRef("pk"),
                    estado__in=["VIGENTE", "POR_VENCER"],
                ))
            ).values("apellidos_nombres", "nro_doc", "grupo_tareo", "pk")[:10]
        )
    except Exception:
        lista_sin_legajo = []

    try:
        tipos = TipoDocumento.objects.filter(activo=True)
        categorias = CategoriaDocumento.objects.filter(activa=True)
    except Exception:
        tipos = []
        categorias = []

    context = {
        "titulo": "Documentos",
        "total_docs": total_docs,
        "total_documentos": total_documentos,
        "vencidos": vencidos,
        "por_vencer": por_vencer,
        "docs_este_mes": docs_este_mes,
        "empleados_con_legajo": empleados_con_legajo,
        "empleados_sin_legajo": empleados_sin_legajo,
        "docs_por_tipo_json": docs_por_tipo_json,
        "docs_por_area_json": docs_por_area_json,
        "docs_por_tipo": docs_por_tipo,
        "docs_recientes": docs_recientes,
        "personal_list": personal_list_raw,
        "recientes": recientes,
        "alertas": alertas,
        "lista_sin_legajo": lista_sin_legajo,
        "tipos": tipos,
        "categorias": categorias,
        "buscar": buscar,
        "total_personal": total_personal,
    }
    return render(request, "documentos/panel.html", context)

# ── Legajo de un Trabajador ──────────────────────────────────────

@login_required
@solo_admin
def legajo_trabajador(request, personal_id):
    """Legajo digital completo de un trabajador."""
    from personal.models import Personal

    empleado = get_object_or_404(Personal, pk=personal_id)
    docs = DocumentoTrabajador.objects.filter(
        personal=empleado,
    ).exclude(estado='ANULADO').select_related(
        'tipo', 'tipo__categoria', 'subido_por',
    ).order_by('tipo__categoria__orden', 'tipo__orden', '-version')

    # Agrupar por categoría
    categorias_dict = {}
    for doc in docs:
        cat_nombre = doc.tipo.categoria.nombre if doc.tipo.categoria else 'Sin Categoría'
        cat_icono = doc.tipo.categoria.icono if doc.tipo.categoria else 'fa-folder'
        key = cat_nombre
        if key not in categorias_dict:
            categorias_dict[key] = {'icono': cat_icono, 'docs': []}
        categorias_dict[key]['docs'].append(doc)

    # Documentos faltantes obligatorios
    tipos_obligatorios = TipoDocumento.objects.filter(obligatorio=True, activo=True)
    if empleado.grupo_tareo == 'STAFF':
        tipos_obligatorios = tipos_obligatorios.filter(aplica_staff=True)
    else:
        tipos_obligatorios = tipos_obligatorios.filter(aplica_rco=True)

    tipos_existentes = set(docs.values_list('tipo_id', flat=True))
    faltantes = [t for t in tipos_obligatorios if t.pk not in tipos_existentes]

    tipos = TipoDocumento.objects.filter(activo=True)

    context = {
        'titulo': f'Legajo - {empleado.apellidos_nombres}',
        'empleado': empleado,
        'categorias_dict': categorias_dict,
        'faltantes': faltantes,
        'total_docs': docs.count(),
        'tipos': tipos,
    }
    return render(request, 'documentos/legajo.html', context)


# ── CRUD Documentos ──────────────────────────────────────────────

@login_required
@solo_admin
@require_POST
def documento_subir(request):
    """Subir un documento al legajo de un trabajador."""
    from personal.models import Personal

    try:
        personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
        tipo = get_object_or_404(TipoDocumento, pk=request.POST['tipo_id'])
        archivo = request.FILES.get('archivo')

        if not archivo:
            return JsonResponse({'ok': False, 'error': 'Debe seleccionar un archivo.'}, status=400)

        # Verificar tamaño (max 10MB)
        if archivo.size > 10 * 1024 * 1024:
            return JsonResponse({'ok': False, 'error': 'El archivo no puede superar 10 MB.'}, status=400)

        # Calcular versión
        ultima_version = DocumentoTrabajador.objects.filter(
            personal=personal, tipo=tipo,
        ).order_by('-version').values_list('version', flat=True).first() or 0

        doc = DocumentoTrabajador.objects.create(
            personal=personal,
            tipo=tipo,
            archivo=archivo,
            nombre_archivo=request.POST.get('nombre', '') or archivo.name,
            fecha_emision=request.POST.get('fecha_emision') or None,
            fecha_vencimiento=request.POST.get('fecha_vencimiento') or None,
            notas=request.POST.get('notas', '').strip(),
            version=ultima_version + 1,
            subido_por=request.user,
        )

        from core.audit import log_create
        log_create(request, doc, f'Documento subido: {doc.tipo.nombre} para {personal.apellidos_nombres}')

        return JsonResponse({
            'ok': True,
            'pk': doc.pk,
            'nombre': doc.nombre_archivo,
            'tipo': doc.tipo.nombre,
            'estado': doc.estado,
            'version': doc.version,
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def documento_eliminar(request, pk):
    """Eliminar (anular) un documento."""
    doc = get_object_or_404(DocumentoTrabajador, pk=pk)
    doc.estado = 'ANULADO'
    doc.save(update_fields=['estado', 'actualizado_en'])

    from core.audit import log_delete
    log_delete(request, doc, f'Documento anulado: {doc.tipo.nombre} de {doc.personal.apellidos_nombres}')
    return JsonResponse({'ok': True})


# ── Tipos de Documento (Configuración) ───────────────────────────

@login_required
@solo_admin
def tipos_documento(request):
    """Configuración de tipos y categorías de documento."""
    tipos = TipoDocumento.objects.select_related('categoria').all()
    categorias = CategoriaDocumento.objects.all()

    context = {
        'titulo': 'Tipos de Documento',
        'tipos': tipos,
        'categorias': categorias,
    }
    return render(request, 'documentos/tipos.html', context)


@login_required
@solo_admin
@require_POST
def tipo_crear(request):
    """Crear un tipo de documento."""
    try:
        cat_id = request.POST.get('categoria_id')
        t = TipoDocumento.objects.create(
            nombre=request.POST['nombre'],
            categoria_id=cat_id if cat_id else None,
            obligatorio=request.POST.get('obligatorio') == 'on',
            vence=request.POST.get('vence') == 'on',
            dias_alerta_vencimiento=int(request.POST.get('dias_alerta', 30) or 30),
            aplica_staff=request.POST.get('aplica_staff', 'on') == 'on',
            aplica_rco=request.POST.get('aplica_rco', 'on') == 'on',
        )
        return JsonResponse({'ok': True, 'pk': t.pk, 'nombre': t.nombre})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def tipo_editar(request, pk):
    """Editar un tipo de documento."""
    t = get_object_or_404(TipoDocumento, pk=pk)
    try:
        t.nombre = request.POST.get('nombre', t.nombre)
        cat_id = request.POST.get('categoria_id')
        t.categoria_id = cat_id if cat_id else None
        t.obligatorio = request.POST.get('obligatorio') == 'on'
        t.vence = request.POST.get('vence') == 'on'
        t.dias_alerta_vencimiento = int(request.POST.get('dias_alerta', t.dias_alerta_vencimiento) or 30)
        t.aplica_staff = request.POST.get('aplica_staff') == 'on'
        t.aplica_rco = request.POST.get('aplica_rco') == 'on'
        t.activo = request.POST.get('activo', 'on') == 'on'
        t.save()
        return JsonResponse({'ok': True, 'pk': t.pk, 'nombre': t.nombre})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def tipo_eliminar(request, pk):
    """Desactivar un tipo de documento."""
    t = get_object_or_404(TipoDocumento, pk=pk)
    t.activo = False
    t.save(update_fields=['activo'])
    return JsonResponse({'ok': True})


# ── Categorías CRUD ──────────────────────────────────────────────

@login_required
@solo_admin
@require_POST
def categoria_crear(request):
    """Crear una categoría de documento."""
    try:
        c = CategoriaDocumento.objects.create(
            nombre=request.POST['nombre'],
            icono=request.POST.get('icono', 'fa-folder'),
            orden=int(request.POST.get('orden', 0) or 0),
        )
        return JsonResponse({'ok': True, 'pk': c.pk, 'nombre': c.nombre})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def categoria_editar(request, pk):
    """Editar una categoría."""
    c = get_object_or_404(CategoriaDocumento, pk=pk)
    try:
        c.nombre = request.POST.get('nombre', c.nombre)
        c.icono = request.POST.get('icono', c.icono)
        c.orden = int(request.POST.get('orden', c.orden) or 0)
        c.save()
        return JsonResponse({'ok': True, 'pk': c.pk, 'nombre': c.nombre})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def categoria_eliminar(request, pk):
    """Desactivar una categoría."""
    c = get_object_or_404(CategoriaDocumento, pk=pk)
    c.activa = False
    c.save(update_fields=['activa'])
    return JsonResponse({'ok': True})


# ── Reporte: Documentos Faltantes ────────────────────────────────

@login_required
@solo_admin
def documentos_faltantes(request):
    """Matriz de personal x documentos obligatorios faltantes."""
    from personal.models import Personal

    tipos_oblig = TipoDocumento.objects.filter(obligatorio=True, activo=True).order_by('orden', 'nombre')
    personal_qs = Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')

    grupo = request.GET.get('grupo', '')
    if grupo:
        personal_qs = personal_qs.filter(grupo_tareo=grupo)

    # Construir matriz
    matriz = []
    for emp in personal_qs[:300]:
        if emp.grupo_tareo == 'STAFF':
            tipos_aplicables = tipos_oblig.filter(aplica_staff=True)
        else:
            tipos_aplicables = tipos_oblig.filter(aplica_rco=True)

        docs_existentes = set(
            DocumentoTrabajador.objects.filter(
                personal=emp, estado__in=['VIGENTE', 'POR_VENCER'],
            ).values_list('tipo_id', flat=True)
        )

        faltantes_emp = [t for t in tipos_aplicables if t.pk not in docs_existentes]
        if faltantes_emp:
            matriz.append({
                'personal': emp,
                'faltantes': faltantes_emp,
                'total_faltantes': len(faltantes_emp),
            })

    context = {
        'titulo': 'Documentos Faltantes',
        'tipos_oblig': tipos_oblig,
        'matriz': sorted(matriz, key=lambda x: -x['total_faltantes']),
        'filtro_grupo': grupo,
    }
    return render(request, 'documentos/faltantes.html', context)


# ── AJAX ─────────────────────────────────────────────────────────

@login_required
@solo_admin
def ajax_stats(request):
    """Estadísticas rápidas para widgets."""
    total = DocumentoTrabajador.objects.exclude(estado='ANULADO').count()
    vencidos = DocumentoTrabajador.objects.filter(estado='VENCIDO').count()
    por_vencer = DocumentoTrabajador.objects.filter(estado='POR_VENCER').count()
    return JsonResponse({
        'total': total,
        'vencidos': vencidos,
        'por_vencer': por_vencer,
    })


# ── Exportar Inventario Excel ─────────────────────────────────────

@login_required
@solo_admin
def exportar_inventario_excel(request):
    """
    Genera un archivo .xlsx con el inventario completo de documentos del legajo.
    Solo accesible para superusuarios.

    Sheet 1 - Inventario Documentos: fila por cada DocumentoTrabajador.
    Sheet 2 - Resumen por Empleado: una fila por empleado con conteo por tipo.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'openpyxl no está instalado. Ejecute: pip install openpyxl',
            status=500,
            content_type='text/plain',
        )

    from personal.models import Personal

    # ── Paleta Harmoni ────────────────────────────────────────────
    COLOR_HEADER = '0D2B27'   # sidebar teal oscuro
    COLOR_ACCENT = '5EEAD4'   # teal claro para subheader
    COLOR_WHITE  = 'FFFFFF'
    COLOR_ROW_ALT = 'F0FDF9'  # fondo alterno muy suave

    def _header_font():
        return Font(bold=True, color=COLOR_WHITE, size=10)

    def _header_fill():
        return PatternFill('solid', fgColor=COLOR_HEADER)

    def _subheader_font():
        return Font(bold=True, color=COLOR_HEADER, size=9)

    def _subheader_fill():
        return PatternFill('solid', fgColor=COLOR_ACCENT)

    def _alt_fill():
        return PatternFill('solid', fgColor=COLOR_ROW_ALT)

    def _thin_border():
        side = Side(style='thin', color='D1D5DB')
        return Border(left=side, right=side, top=side, bottom=side)

    def _center():
        return Alignment(horizontal='center', vertical='center')

    def _left():
        return Alignment(horizontal='left', vertical='center', wrap_text=False)

    def _autofit(ws, min_width=10, max_width=50):
        for col_cells in ws.columns:
            length = max(
                len(str(cell.value or ''))
                for cell in col_cells
            )
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = max(min_width, min(length + 4, max_width))

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════
    # Sheet 1: Inventario Documentos
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Inventario Documentos'

    # Título fusionado
    ws1.merge_cells('A1:H1')
    title_cell = ws1['A1']
    title_cell.value = 'INVENTARIO DE DOCUMENTOS - LEGAJO DIGITAL'
    title_cell.font = Font(bold=True, color=COLOR_WHITE, size=12)
    title_cell.fill = PatternFill('solid', fgColor=COLOR_HEADER)
    title_cell.alignment = _center()
    ws1.row_dimensions[1].height = 22

    # Subtítulo con fecha
    ws1.merge_cells('A2:H2')
    sub_cell = ws1['A2']
    from django.utils import timezone
    sub_cell.value = f'Generado: {timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")}  |  Usuario: {request.user.get_full_name() or request.user.username}'
    sub_cell.font = Font(italic=True, color=COLOR_HEADER, size=9)
    sub_cell.fill = PatternFill('solid', fgColor=COLOR_ACCENT)
    sub_cell.alignment = _center()
    ws1.row_dimensions[2].height = 14

    # Cabeceras
    headers1 = ['DNI', 'Apellidos y Nombres', 'Área', 'Tipo de Documento',
                 'Nombre Archivo', 'Fecha Subida', 'Estado', 'Tamaño (KB)']
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()
    ws1.row_dimensions[3].height = 16
    ws1.freeze_panes = 'A4'

    # Datos
    docs_qs = (
        DocumentoTrabajador.objects
        .exclude(estado='ANULADO')
        .select_related('personal', 'personal__subarea__area', 'tipo')
        .order_by('personal__apellidos_nombres', 'tipo__nombre')
    )

    for row_idx, doc in enumerate(docs_qs, start=4):
        area_nombre = ''
        try:
            area_nombre = doc.personal.subarea.area.nombre
        except Exception:
            pass

        tamano_kb = ''
        try:
            tamano_kb = round(doc.archivo.size / 1024, 1)
        except Exception:
            pass

        row_data = [
            doc.personal.nro_doc,
            doc.personal.apellidos_nombres,
            area_nombre,
            doc.tipo.nombre,
            doc.nombre_archivo or '',
            doc.creado_en.strftime('%d/%m/%Y %H:%M') if doc.creado_en else '',
            doc.get_estado_display(),
            tamano_kb,
        ]
        fill = _alt_fill() if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.alignment = _left()
            cell.border = _thin_border()
            if fill:
                cell.fill = fill

    _autofit(ws1)

    # ════════════════════════════════════════════════════════════
    # Sheet 2: Resumen por Empleado
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Resumen por Empleado')

    # Obtener lista de tipos activos para columnas dinámicas
    tipos_activos = list(
        TipoDocumento.objects.filter(activo=True).order_by('categoria__orden', 'orden', 'nombre')
    )
    n_tipos = len(tipos_activos)

    last_col = get_column_letter(4 + n_tipos)

    # Título fusionado
    ws2.merge_cells(f'A1:{last_col}1')
    t2 = ws2['A1']
    t2.value = 'RESUMEN DE DOCUMENTOS POR EMPLEADO'
    t2.font = Font(bold=True, color=COLOR_WHITE, size=12)
    t2.fill = PatternFill('solid', fgColor=COLOR_HEADER)
    t2.alignment = _center()
    ws2.row_dimensions[1].height = 22

    ws2.merge_cells(f'A2:{last_col}2')
    s2 = ws2['A2']
    s2.value = f'Total de tipos de documentos activos: {n_tipos}'
    s2.font = Font(italic=True, color=COLOR_HEADER, size=9)
    s2.fill = PatternFill('solid', fgColor=COLOR_ACCENT)
    s2.alignment = _center()
    ws2.row_dimensions[2].height = 14

    # Cabeceras fijas
    fixed_headers = ['DNI', 'Apellidos y Nombres', 'Área', 'Total Docs']
    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()

    # Cabeceras de tipos de documentos
    for tipo_idx, tipo in enumerate(tipos_activos):
        col_idx = 5 + tipo_idx
        cell = ws2.cell(row=3, column=col_idx, value=tipo.nombre)
        cell.font = _subheader_font()
        cell.fill = _subheader_fill()
        cell.alignment = _center()
        cell.border = _thin_border()
    ws2.row_dimensions[3].height = 16
    ws2.freeze_panes = 'E4'

    # Precalcular documentos por empleado y tipo de forma eficiente
    personal_activo = Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')
    tipos_ids = [t.pk for t in tipos_activos]

    # Dict: {personal_id: {tipo_id: count}}
    from collections import defaultdict
    docs_por_emp_tipo = defaultdict(lambda: defaultdict(int))
    for doc in (DocumentoTrabajador.objects
                .exclude(estado='ANULADO')
                .filter(tipo_id__in=tipos_ids)
                .values('personal_id', 'tipo_id')):
        docs_por_emp_tipo[doc['personal_id']][doc['tipo_id']] += 1

    for row_idx, emp in enumerate(personal_activo, start=4):
        area_nombre = ''
        try:
            area_nombre = emp.subarea.area.nombre
        except Exception:
            pass

        tipo_counts = docs_por_emp_tipo.get(emp.pk, {})
        total_emp = sum(tipo_counts.values())

        row_data = [emp.nro_doc, emp.apellidos_nombres, area_nombre, total_emp]
        fill = _alt_fill() if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.alignment = _left() if col_idx == 2 else _center()
            cell.border = _thin_border()
            if fill:
                cell.fill = fill

        for tipo_idx, tipo in enumerate(tipos_activos):
            col_idx = 5 + tipo_idx
            count = tipo_counts.get(tipo.pk, 0)
            cell = ws2.cell(row=row_idx, column=col_idx, value=count if count > 0 else '')
            cell.font = Font(size=9, color='0F766E' if count > 0 else '9CA3AF')
            cell.alignment = _center()
            cell.border = _thin_border()
            if fill:
                cell.fill = fill

    _autofit(ws2, min_width=8, max_width=30)

    # ── Respuesta HTTP ────────────────────────────────────────────
    from django.utils import timezone as tz
    fecha_str = tz.localtime(tz.now()).strftime('%Y%m%d_%H%M')
    filename = f'inventario_legajo_{fecha_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Helpers ──────────────────────────────────────────────────────

def _recalcular_vencimientos():
    """Recalcula el estado de documentos con fecha de vencimiento."""
    hoy = date.today()
    # Marcar vencidos
    DocumentoTrabajador.objects.filter(
        tipo__vence=True,
        fecha_vencimiento__lt=hoy,
        estado__in=['VIGENTE', 'POR_VENCER'],
    ).update(estado='VENCIDO')

    # Marcar por vencer (próximos 30 días por defecto)
    DocumentoTrabajador.objects.filter(
        tipo__vence=True,
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timedelta(days=30),
        estado='VIGENTE',
    ).update(estado='POR_VENCER')


# ══════════════════════════════════════════════════════════════
# CONSTANCIAS / GENERADOR DE DOCUMENTOS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def constancias_panel(request):
    """Panel de plantillas de constancias disponibles + historial reciente."""
    plantillas = PlantillaConstancia.objects.filter(activa=True)

    # Historial reciente de generaciones (últimas 30)
    historial = (
        ConstanciaGenerada.objects
        .select_related('plantilla', 'personal', 'generado_por')
        .order_by('-fecha_generacion')[:30]
    )

    # Estadísticas rápidas por plantilla (para badges)
    from django.db.models import Count
    stats = (
        ConstanciaGenerada.objects
        .values('plantilla_id')
        .annotate(total=Count('id'))
    )
    stats_map = {s['plantilla_id']: s['total'] for s in stats}

    context = {
        'titulo': 'Generador de Constancias',
        'plantillas': plantillas,
        'historial': historial,
        'stats_map': stats_map,
    }
    return render(request, 'documentos/constancias_panel.html', context)


@login_required
@solo_admin
def constancia_generar(request, plantilla_id):
    """Seleccionar empleado y generar constancia en PDF."""
    from personal.models import Personal

    plantilla = get_object_or_404(PlantillaConstancia, pk=plantilla_id, activa=True)

    personal_id = request.GET.get('personal_id')
    if personal_id:
        # Generar PDF directamente
        from documentos.services import generar_constancia_pdf
        from django.http import HttpResponse

        personal = get_object_or_404(Personal, pk=personal_id)
        try:
            pdf_bytes = generar_constancia_pdf(plantilla, personal)

            # Registrar generación en historial
            ip = request.META.get('REMOTE_ADDR')
            ConstanciaGenerada.objects.create(
                plantilla=plantilla,
                personal=personal,
                generado_por=request.user,
                origen='ADMIN',
                ip_solicitud=ip,
            )

            filename = f'{plantilla.codigo}_{personal.nro_doc}.pdf'
            response = HttpResponse(pdf_bytes, content_type='application/pdf')

            # Si piden descarga directa
            if request.GET.get('download'):
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            else:
                response['Content-Disposition'] = f'inline; filename="{filename}"'
            return response
        except Exception as e:
            return render(request, 'documentos/constancia_generar.html', {
                'titulo': f'Generar: {plantilla.nombre}',
                'plantilla': plantilla,
                'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')[:300],
                'error': str(e),
            })

    # Mostrar selector de empleado
    buscar = request.GET.get('buscar', '')
    personal_qs = Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')
    if buscar:
        personal_qs = personal_qs.filter(
            Q(apellidos_nombres__icontains=buscar) | Q(nro_doc__icontains=buscar)
        )

    context = {
        'titulo': f'Generar: {plantilla.nombre}',
        'plantilla': plantilla,
        'personal_list': personal_qs[:300],
        'buscar': buscar,
    }
    return render(request, 'documentos/constancia_generar.html', context)


@login_required
@solo_admin
def constancia_preview(request, plantilla_id):
    """Preview HTML de la constancia (sin generar PDF)."""
    from personal.models import Personal
    from documentos.services import generar_constancia_pdf, _fecha_texto, _calcular_antiguedad

    plantilla = get_object_or_404(PlantillaConstancia, pk=plantilla_id, activa=True)
    personal_id = request.GET.get('personal_id')

    if not personal_id:
        return JsonResponse({'ok': False, 'error': 'Falta personal_id'}, status=400)

    personal = get_object_or_404(Personal, pk=personal_id)
    hoy = date.today()

    # Renderizar solo el HTML (sin PDF)
    from django.template import Template, Context

    ctx = {
        'personal': personal,
        'hoy': hoy,
        'hoy_texto': _fecha_texto(hoy),
        'antiguedad': _calcular_antiguedad(personal),
        'fecha_alta_texto': _fecha_texto(personal.fecha_alta) if personal.fecha_alta else 'N/A',
        'fecha_cese_texto': _fecha_texto(personal.fecha_cese) if personal.fecha_cese else None,
        'empresa': {'nombre': 'EMPRESA S.A.C.', 'ruc': '', 'direccion': ''},
    }

    try:
        from asistencia.models import ConfiguracionSistema
        config = ConfiguracionSistema.objects.first()
        if config:
            ctx['empresa'] = {
                'nombre': getattr(config, 'empresa_nombre', 'EMPRESA S.A.C.'),
                'ruc': getattr(config, 'empresa_ruc', ''),
                'direccion': getattr(config, 'empresa_direccion', ''),
            }
    except Exception:
        pass

    tpl = Template(plantilla.contenido_html)
    html = tpl.render(Context(ctx))

    return JsonResponse({'ok': True, 'html': html})


# ── Portal del Trabajador: Constancias ──────────────────────────

def _get_empleado_doc(request):
    """Obtiene el Personal vinculado al usuario autenticado (portal)."""
    from personal.models import Personal
    try:
        return Personal.objects.get(usuario=request.user)
    except Exception:
        return None


@login_required
def mis_constancias(request):
    """
    Portal del trabajador: ver plantillas disponibles y mi historial
    de constancias generadas.
    """
    personal = _get_empleado_doc(request)
    plantillas = PlantillaConstancia.objects.filter(activa=True).order_by('orden', 'nombre')

    historial = []
    if personal:
        historial = (
            ConstanciaGenerada.objects
            .filter(personal=personal)
            .select_related('plantilla', 'generado_por')
            .order_by('-fecha_generacion')[:20]
        )

    context = {
        'titulo': 'Mis Constancias',
        'plantillas': plantillas,
        'historial': historial,
        'personal': personal,
    }
    return render(request, 'documentos/mis_constancias.html', context)


@login_required
def portal_generar_constancia(request, plantilla_id):
    """
    Portal del trabajador: genera una constancia para sí mismo directamente,
    sin necesitar intervención del administrador.
    Las constancias son documentos factuales - autoservicio está permitido.
    """
    from documentos.services import generar_constancia_pdf
    from django.http import HttpResponse

    personal = _get_empleado_doc(request)
    if not personal:
        from django.contrib import messages
        messages.error(request, 'Tu usuario no está vinculado a un empleado.')
        return redirect('mis_constancias')

    plantilla = get_object_or_404(PlantillaConstancia, pk=plantilla_id, activa=True)

    try:
        pdf_bytes = generar_constancia_pdf(plantilla, personal)

        # Registrar generación en historial
        ip = request.META.get('REMOTE_ADDR')
        ConstanciaGenerada.objects.create(
            plantilla=plantilla,
            personal=personal,
            generado_por=request.user,
            origen='PORTAL',
            ip_solicitud=ip,
        )

        filename = f'{plantilla.codigo}_{personal.nro_doc}.pdf'
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        if request.GET.get('download'):
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        else:
            response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Error al generar la constancia: {e}')
        return redirect('mis_constancias')


# ══════════════════════════════════════════════════════════════
# BOLETAS DE PAGO DIGITAL (DS 009-2011-TR + DS 003-2013-TR)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def boletas_panel(request):
    """Panel de gestión de boletas de pago."""
    from personal.models import Personal
    from documentos.models import BoletaPago

    qs = BoletaPago.objects.select_related('personal').exclude(estado='ANULADA')

    # Filtros
    periodo = request.GET.get('periodo', '')
    estado = request.GET.get('estado', '')
    buscar = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')

    if periodo:
        try:
            anio, mes = periodo.split('-')
            qs = qs.filter(periodo__year=int(anio), periodo__month=int(mes))
        except (ValueError, IndexError):
            pass
    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo=tipo)
    if buscar:
        qs = qs.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar)
        )

    # Stats
    total_boletas = qs.count()
    publicadas = qs.filter(estado='PUBLICADA').count()
    leidas = qs.filter(estado='LEIDA').count()
    borradores = qs.filter(estado='BORRADOR').count()

    context = {
        'titulo': 'Boletas de Pago',
        'boletas': qs[:200],
        'total': total_boletas,
        'stats': {
            'publicadas': publicadas,
            'leidas': leidas,
            'borradores': borradores,
            'tasa_lectura': round((leidas / publicadas * 100) if publicadas else 0),
        },
        'filtro_periodo': periodo,
        'filtro_estado': estado,
        'filtro_tipo': tipo,
        'buscar': buscar,
        'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres'),
    }
    return render(request, 'documentos/boletas_panel.html', context)


@login_required
@solo_admin
@require_POST
def boleta_subir(request):
    """Subir boleta de pago (individual o masiva)."""
    from documentos.models import BoletaPago
    from personal.models import Personal

    try:
        personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({'ok': False, 'error': 'Debe seleccionar un archivo.'}, status=400)

        periodo_str = request.POST.get('periodo', '')
        if not periodo_str:
            return JsonResponse({'ok': False, 'error': 'Debe indicar el período.'}, status=400)

        # Parsear período (YYYY-MM)
        from datetime import datetime
        periodo = datetime.strptime(periodo_str + '-01', '%Y-%m-%d').date()

        tipo = request.POST.get('tipo', 'MENSUAL')

        boleta, created = BoletaPago.objects.update_or_create(
            personal=personal, periodo=periodo, tipo=tipo,
            defaults={
                'archivo': archivo,
                'nombre_archivo': archivo.name,
                'remuneracion_bruta': request.POST.get('rem_bruta') or None,
                'descuentos': request.POST.get('descuentos') or None,
                'neto_pagar': request.POST.get('neto') or None,
                'observaciones': request.POST.get('observaciones', ''),
                'subido_por': request.user,
            }
        )

        from core.audit import log_create
        log_create(request, boleta, f'Boleta subida: {personal.apellidos_nombres} - {periodo.strftime("%m/%Y")}')

        return JsonResponse({
            'ok': True, 'pk': boleta.pk, 'created': created,
            'nombre': boleta.nombre_archivo,
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
@require_POST
def boleta_publicar(request, pk):
    """Publicar una boleta (hacerla visible al trabajador)."""
    from documentos.models import BoletaPago

    boleta = get_object_or_404(BoletaPago, pk=pk)
    if boleta.estado == 'BORRADOR':
        boleta.publicar()

        from core.audit import log_update
        log_update(request, boleta, {'estado': {'old': 'BORRADOR', 'new': 'PUBLICADA'}},
                   f'Boleta publicada: {boleta.personal.apellidos_nombres} - {boleta.periodo.strftime("%m/%Y")}')

        return JsonResponse({'ok': True, 'estado': 'PUBLICADA'})
    return JsonResponse({'ok': False, 'error': 'La boleta no está en estado borrador.'})


@login_required
@solo_admin
@require_POST
def boletas_publicar_masivo(request):
    """Publicar todas las boletas borrador de un período."""
    from documentos.models import BoletaPago
    from django.utils import timezone

    periodo = request.POST.get('periodo', '')
    if not periodo:
        return JsonResponse({'ok': False, 'error': 'Falta período.'}, status=400)

    try:
        from datetime import datetime
        fecha_periodo = datetime.strptime(periodo + '-01', '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Período inválido.'}, status=400)

    count = BoletaPago.objects.filter(
        periodo=fecha_periodo, estado='BORRADOR',
    ).update(estado='PUBLICADA', fecha_publicacion=timezone.now())

    return JsonResponse({'ok': True, 'publicadas': count})


@login_required
@solo_admin
@require_POST
def boleta_anular(request, pk):
    """Anular una boleta."""
    from documentos.models import BoletaPago

    boleta = get_object_or_404(BoletaPago, pk=pk)
    boleta.estado = 'ANULADA'
    boleta.save(update_fields=['estado'])

    from core.audit import log_delete
    log_delete(request, boleta, f'Boleta anulada: {boleta.personal.apellidos_nombres} - {boleta.periodo.strftime("%m/%Y")}')
    return JsonResponse({'ok': True})


# ── Portal: Mis Boletas ──
@login_required
def mis_boletas(request):
    """Portal del trabajador: ver boletas de pago publicadas."""
    from documentos.models import BoletaPago
    from portal.views import _get_empleado

    empleado = _get_empleado(request.user)
    boletas = []
    stats = {}

    if empleado:
        qs = BoletaPago.objects.filter(
            personal=empleado,
            estado__in=['PUBLICADA', 'LEIDA'],
        ).order_by('-periodo')

        anio = request.GET.get('anio', str(date.today().year))
        if anio:
            qs = qs.filter(periodo__year=int(anio))

        boletas = qs
        stats = {
            'total': qs.count(),
            'leidas': qs.filter(estado='LEIDA').count(),
            'pendientes': qs.filter(estado='PUBLICADA').count(),
        }

    context = {
        'titulo': 'Mis Boletas de Pago',
        'empleado': empleado,
        'boletas': boletas,
        'stats': stats,
        'anio_actual': date.today().year,
        'anio_filtro': request.GET.get('anio', str(date.today().year)),
    }
    return render(request, 'documentos/mis_boletas.html', context)


@login_required
@require_POST
def boleta_confirmar_lectura(request, pk):
    """Trabajador confirma recepción de boleta (constancia DS 009-2011-TR)."""
    from documentos.models import BoletaPago
    from portal.views import _get_empleado

    empleado = _get_empleado(request.user)
    boleta = get_object_or_404(BoletaPago, pk=pk, personal=empleado)

    ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
    if ',' in ip:
        ip = ip.split(',')[0].strip()

    boleta.registrar_lectura(ip=ip)
    boleta.confirmar_recepcion()

    return JsonResponse({
        'ok': True,
        'estado': 'LEIDA',
        'fecha': boleta.fecha_lectura.strftime('%d/%m/%Y %H:%M') if boleta.fecha_lectura else '',
    })


@login_required
@solo_admin
def plantilla_list(request):
    """CRUD de plantillas de constancias."""
    plantillas = PlantillaConstancia.objects.all()
    context = {
        'titulo': 'Plantillas de Constancias',
        'plantillas': plantillas,
    }
    return render(request, 'documentos/plantilla_list.html', context)


@login_required
@solo_admin
def plantilla_editar(request, pk=None):
    """Crear o editar una plantilla de constancia."""
    if pk:
        plantilla = get_object_or_404(PlantillaConstancia, pk=pk)
    else:
        plantilla = None

    if request.method == 'POST':
        if plantilla is None:
            plantilla = PlantillaConstancia()
        plantilla.nombre = request.POST.get('nombre', '')
        plantilla.codigo = request.POST.get('codigo', '')
        plantilla.categoria = request.POST.get('categoria', 'CONSTANCIA')
        plantilla.descripcion = request.POST.get('descripcion', '')
        plantilla.contenido_html = request.POST.get('contenido_html', '')
        plantilla.activa = request.POST.get('activa') == 'on'
        plantilla.orden = int(request.POST.get('orden', 0) or 0)
        try:
            plantilla.save()
            return redirect('plantilla_list')
        except Exception as e:
            return render(request, 'documentos/plantilla_form.html', {
                'titulo': 'Editar Plantilla' if pk else 'Nueva Plantilla',
                'plantilla': plantilla,
                'error': str(e),
            })

    context = {
        'titulo': f'Editar: {plantilla.nombre}' if plantilla else 'Nueva Plantilla',
        'plantilla': plantilla,
    }
    return render(request, 'documentos/plantilla_form.html', context)


# ══════════════════════════════════════════════════════════════
# DOCUMENTOS LABORALES (Fase 6.3)
# Distribución de políticas, reglamentos, memos con constancia
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def docs_laborales_panel(request):
    """Panel principal de Documentos Laborales."""
    docs = DocumentoLaboral.objects.select_related('creado_por').annotate(
        total_entregas=Count('entregas', distinct=True),
        total_confirmados=Count('entregas', filter=Q(entregas__confirmado=True), distinct=True),
    ).order_by('-creado_en')

    # Filtros
    tipo_f = request.GET.get('tipo', '')
    estado_f = request.GET.get('estado', 'PUBLICADO')
    buscar = request.GET.get('buscar', '')

    if tipo_f:
        docs = docs.filter(tipo=tipo_f)
    if estado_f and estado_f != 'TODOS':
        docs = docs.filter(estado=estado_f)
    if buscar:
        docs = docs.filter(titulo__icontains=buscar)

    # Stats globales
    stats = {
        'total': DocumentoLaboral.objects.count(),
        'publicados': DocumentoLaboral.objects.filter(estado='PUBLICADO').count(),
        'borradores': DocumentoLaboral.objects.filter(estado='BORRADOR').count(),
        'pendientes_conf': EntregaDocumento.objects.filter(confirmado=False, documento__estado='PUBLICADO').count(),
    }

    from personal.models import Area
    context = {
        'docs': docs,
        'tipo_f': tipo_f,
        'estado_f': estado_f,
        'buscar': buscar,
        'stats': stats,
        'tipos': DocumentoLaboral.TIPO_CHOICES,
        'estados': DocumentoLaboral.ESTADO_CHOICES,
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
    }
    return render(request, 'documentos/docs_laborales_panel.html', context)


@login_required
@solo_admin
def doc_laboral_crear(request):
    """Crear nuevo documento laboral."""
    from personal.models import Area, Personal

    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        tipo = request.POST.get('tipo', 'COMUNICADO')
        descripcion = request.POST.get('descripcion', '').strip()
        destinatarios_tipo = request.POST.get('destinatarios_tipo', 'TODOS')
        requiere_confirmacion = request.POST.get('requiere_confirmacion') == '1'
        vigente_hasta = request.POST.get('vigente_hasta') or None
        contenido_html = request.POST.get('contenido_html', '').strip()

        if not titulo:
            from django.contrib import messages
            messages.error(request, 'El título es obligatorio.')
        else:
            doc = DocumentoLaboral(
                titulo=titulo, tipo=tipo, descripcion=descripcion,
                destinatarios_tipo=destinatarios_tipo,
                requiere_confirmacion=requiere_confirmacion,
                vigente_hasta=vigente_hasta or None,
                contenido_html=contenido_html,
                creado_por=request.user,
            )
            if request.FILES.get('archivo'):
                doc.archivo = request.FILES['archivo']
            doc.save()

            # Áreas
            area_ids = request.POST.getlist('areas')
            if area_ids:
                doc.areas.set(area_ids)
            # Personal específico
            personal_ids = request.POST.getlist('personal_especifico')
            if personal_ids:
                doc.personal_especifico.set(personal_ids)

            from django.contrib import messages
            messages.success(request, f'Documento "{titulo}" creado como borrador.')
            return redirect('doc_laboral_detalle', pk=doc.pk)

    context = {
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres'),
        'tipos': DocumentoLaboral.TIPO_CHOICES,
        'dest_tipos': DocumentoLaboral.DESTINATARIOS_CHOICES,
    }
    return render(request, 'documentos/doc_laboral_form.html', context)


@login_required
@solo_admin
def doc_laboral_detalle(request, pk):
    """Detalle de documento laboral con estadísticas de entrega."""
    doc = get_object_or_404(DocumentoLaboral, pk=pk)
    entregas = doc.entregas.select_related('personal__subarea__area').order_by(
        'confirmado', 'personal__apellidos_nombres'
    )

    total = entregas.count()
    confirmados = entregas.filter(confirmado=True).count()
    vistos = entregas.filter(visto=True).count()
    pendientes = total - confirmados

    context = {
        'doc': doc,
        'entregas': entregas[:200],
        'total': total,
        'confirmados': confirmados,
        'vistos': vistos,
        'pendientes': pendientes,
        'tasa': round(confirmados / total * 100) if total else 0,
    }
    return render(request, 'documentos/doc_laboral_detalle.html', context)


@login_required
@solo_admin
@require_POST
def doc_laboral_publicar(request, pk):
    """Publica el documento y genera entregas para todos los destinatarios."""
    from django.contrib import messages
    doc = get_object_or_404(DocumentoLaboral, pk=pk, estado='BORRADOR')
    nuevas = doc.publicar(usuario=request.user)
    messages.success(
        request,
        f'Documento publicado. Se generaron {nuevas} entrega(s) para los destinatarios.'
    )
    return redirect('doc_laboral_detalle', pk=doc.pk)


@login_required
@solo_admin
@require_POST
def doc_laboral_archivar(request, pk):
    """Archiva el documento."""
    from django.contrib import messages
    doc = get_object_or_404(DocumentoLaboral, pk=pk)
    doc.estado = 'ARCHIVADO'
    doc.save(update_fields=['estado'])
    messages.info(request, f'Documento "{doc.titulo}" archivado.')
    return redirect('docs_laborales_panel')


# ── Portal trabajador ─────────────────────────────────────────

@login_required
def mis_documentos_laborales(request):
    """Portal: documentos laborales pendientes y confirmados del trabajador."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)
    if not empleado:
        from django.contrib import messages
        messages.warning(request, 'Tu usuario no está vinculado a un empleado.')
        return redirect('portal_home')

    entregas = EntregaDocumento.objects.filter(
        personal=empleado,
        documento__estado='PUBLICADO',
    ).select_related('documento').order_by('-documento__fecha_publicacion')

    pendientes = entregas.filter(confirmado=False).count()

    context = {
        'entregas': entregas,
        'pendientes': pendientes,
        'empleado': empleado,
    }
    return render(request, 'documentos/mis_documentos_laborales.html', context)


@login_required
def doc_laboral_ver(request, pk):
    """Portal: ver documento laboral (marca como visto)."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)
    if not empleado:
        return redirect('portal_home')

    entrega = get_object_or_404(
        EntregaDocumento, documento_id=pk, personal=empleado
    )
    entrega.marcar_visto()
    doc = entrega.documento

    context = {
        'doc': doc,
        'entrega': entrega,
    }
    return render(request, 'documentos/doc_laboral_ver.html', context)


@login_required
@require_POST
def doc_laboral_confirmar(request, pk):
    """Portal: confirmar recepción del documento (constancia legal)."""
    from portal.views import _get_empleado
    from django.contrib import messages
    empleado = _get_empleado(request.user)
    if not empleado:
        return redirect('portal_home')

    entrega = get_object_or_404(
        EntregaDocumento, documento_id=pk, personal=empleado
    )
    ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
    if ',' in ip:
        ip = ip.split(',')[0].strip()
    entrega.confirmar(ip=ip)
    messages.success(
        request,
        f'Has confirmado la recepción de "{entrega.documento.titulo}". '
        'Se ha registrado tu confirmación con fecha y hora.'
    )
    return redirect('mis_documentos_laborales')


# ═══════════════════════════════════════════════════════════════
# DOSSIER DOCUMENTARIO
# ═══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def dossier_list(request):
    """Lista de dossiers con filtros por estado y proyecto."""
    from django.db.models import Prefetch

    qs = Dossier.objects.select_related('plantilla', 'responsable').annotate(
        total_p=Count('personal_dossier', distinct=True),
    )

    estado = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')
    if estado:
        qs = qs.filter(estado=estado)
    if buscar:
        qs = qs.filter(
            Q(nombre__icontains=buscar) | Q(proyecto__icontains=buscar) |
            Q(cliente__icontains=buscar)
        )

    context = {
        'titulo': 'Dossiers Documentarios',
        'dossiers': qs.order_by('-creado_en')[:100],
        'estado_sel': estado,
        'buscar': buscar,
        'ESTADO_CHOICES': Dossier.ESTADO_CHOICES,
        'total': qs.count(),
        # KPI
        'total_borrador': Dossier.objects.filter(estado='BORRADOR').count(),
        'total_revision': Dossier.objects.filter(estado='EN_REVISION').count(),
        'total_aprobado': Dossier.objects.filter(estado='APROBADO').count(),
        'total_entregado': Dossier.objects.filter(estado='ENTREGADO').count(),
    }
    return render(request, 'documentos/dossier_list.html', context)


@login_required
@solo_admin
def dossier_crear(request):
    """Crear un nuevo dossier."""
    from django.contrib import messages
    from personal.models import Personal

    plantillas = PlantillaDossier.objects.filter(activa=True).order_by('nombre')

    if request.method == 'POST':
        nombre    = request.POST.get('nombre', '').strip()
        proyecto  = request.POST.get('proyecto', '').strip()
        cliente   = request.POST.get('cliente', '').strip()
        plantilla_id = request.POST.get('plantilla', '')
        fecha_inicio = request.POST.get('fecha_inicio') or None
        fecha_prevista = request.POST.get('fecha_entrega_prevista') or None
        observaciones = request.POST.get('observaciones', '').strip()

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        else:
            dossier = Dossier.objects.create(
                nombre=nombre,
                proyecto=proyecto,
                cliente=cliente,
                plantilla_id=plantilla_id if plantilla_id else None,
                fecha_inicio=fecha_inicio,
                fecha_entrega_prevista=fecha_prevista,
                observaciones=observaciones,
                responsable=request.user,
                creado_por=request.user,
            )
            messages.success(request, f'Dossier "{dossier.nombre}" creado.')
            return redirect('dossier_detalle', pk=dossier.pk)

    context = {
        'titulo': 'Nuevo Dossier',
        'plantillas': plantillas,
    }
    return render(request, 'documentos/dossier_form.html', context)


@login_required
@solo_admin
def dossier_detalle(request, pk):
    """Detalle del dossier: progreso por trabajador, ítems, acciones."""
    from personal.models import Personal

    dossier = get_object_or_404(Dossier, pk=pk)
    progreso_personal = dossier.progreso_por_personal()

    # Personal disponible para agregar (no en el dossier aún)
    ids_ya = dossier.personal_dossier.values_list('personal_id', flat=True)
    personal_disponible = Personal.objects.filter(
        estado='Activo'
    ).exclude(id__in=ids_ya).order_by('apellidos_nombres')

    # Secciones del dossier (para vista de items por trabajador seleccionado)
    personal_sel_id = request.GET.get('personal')
    items_personal = None
    personal_sel = None
    if personal_sel_id:
        personal_sel = get_object_or_404(Personal, pk=personal_sel_id)
        items_personal = dossier.items.filter(
            personal=personal_sel
        ).select_related('tipo_documento', 'documento', 'tipo_documento__categoria').order_by(
            'orden', 'tipo_documento__nombre'
        )

    context = {
        'titulo': dossier.nombre,
        'dossier': dossier,
        'progreso_personal': progreso_personal,
        'personal_disponible': personal_disponible,
        'personal_sel': personal_sel,
        'items_personal': items_personal,
    }
    return render(request, 'documentos/dossier_detalle.html', context)


@login_required
@solo_admin
@require_POST
def dossier_agregar_personal(request, pk):
    """Agrega trabajadores al dossier y genera sus ítems."""
    from django.contrib import messages
    from personal.models import Personal

    dossier = get_object_or_404(Dossier, pk=pk)
    ids = request.POST.getlist('personal_ids')

    agregados = 0
    for pid in ids:
        _, created = DossierPersonal.objects.get_or_create(
            dossier=dossier,
            personal_id=pid,
        )
        if created:
            agregados += 1

    # Generar ítems y vincular documentos existentes
    if agregados and dossier.plantilla:
        creados, _ = dossier.generar_items()
        vinculados = dossier.vincular_documentos()
        messages.success(
            request,
            f'{agregados} trabajador(es) agregado(s). '
            f'{creados} ítems generados, {vinculados} documentos vinculados automáticamente.'
        )
    elif agregados:
        messages.success(request, f'{agregados} trabajador(es) agregado(s).')
    else:
        messages.info(request, 'No se agregaron trabajadores nuevos.')

    return redirect('dossier_detalle', pk=pk)


@login_required
@solo_admin
@require_POST
def dossier_generar_items(request, pk):
    """Genera/regenera ítems del dossier desde la plantilla (AJAX)."""
    dossier = get_object_or_404(Dossier, pk=pk)
    if not dossier.plantilla:
        return JsonResponse({'ok': False, 'msg': 'Este dossier no tiene plantilla asignada.'})
    creados, existentes = dossier.generar_items()
    vinculados = dossier.vincular_documentos()
    return JsonResponse({
        'ok': True,
        'creados': creados,
        'existentes': existentes,
        'vinculados': vinculados,
        'progreso': dossier.progreso,
        'msg': f'{creados} ítems generados, {vinculados} documentos vinculados.',
    })


@login_required
@solo_admin
@require_POST
def dossier_vincular(request, pk):
    """Re-vincula automáticamente documentos del legajo a ítems pendientes (AJAX)."""
    dossier = get_object_or_404(Dossier, pk=pk)
    vinculados = dossier.vincular_documentos()
    return JsonResponse({
        'ok': True,
        'vinculados': vinculados,
        'progreso': dossier.progreso,
        'msg': f'{vinculados} documento(s) vinculado(s) automáticamente.',
    })


@login_required
@solo_admin
@require_POST
def dossier_item_estado(request, item_pk):
    """Actualiza el estado de un DossierItem (AJAX)."""
    item = get_object_or_404(DossierItem, pk=item_pk)
    nuevo_estado = request.POST.get('estado')
    observacion  = request.POST.get('observacion', '').strip()

    estados_validos = [c[0] for c in DossierItem.ESTADO_CHOICES]
    if nuevo_estado not in estados_validos:
        return JsonResponse({'ok': False, 'msg': 'Estado inválido.'})

    item.estado = nuevo_estado
    item.observacion = observacion
    item.save(update_fields=['estado', 'observacion', 'actualizado_en'])

    return JsonResponse({
        'ok': True,
        'estado': item.estado,
        'estado_display': item.get_estado_display(),
        'color': item.color_estado,
        'progreso_dossier': item.dossier.progreso,
    })


@login_required
@solo_admin
@require_POST
def dossier_cambiar_estado(request, pk):
    """Cambia el estado del dossier."""
    from django.contrib import messages
    dossier = get_object_or_404(Dossier, pk=pk)
    nuevo = request.POST.get('estado')
    estados = [c[0] for c in Dossier.ESTADO_CHOICES]
    if nuevo in estados:
        dossier.estado = nuevo
        if nuevo == 'ENTREGADO':
            dossier.fecha_entrega_real = date.today()
        dossier.save(update_fields=['estado', 'fecha_entrega_real'])
        messages.success(request, f'Estado actualizado a: {dossier.get_estado_display()}')
    return redirect('dossier_detalle', pk=pk)


# ── Plantillas Dossier ────────────────────────────────────────────────────────

@login_required
@solo_admin
def plantilla_dossier_list(request):
    """Lista de plantillas de dossier."""
    plantillas = PlantillaDossier.objects.annotate(
        n_items=Count('items'),
        n_dossiers=Count('dossiers'),
    ).order_by('nombre')
    context = {
        'titulo': 'Plantillas de Dossier',
        'plantillas': plantillas,
    }
    return render(request, 'documentos/plantilla_dossier_list.html', context)


@login_required
@solo_admin
def plantilla_dossier_form(request, pk=None):
    """Crear o editar una PlantillaDossier."""
    from django.contrib import messages

    plantilla = get_object_or_404(PlantillaDossier, pk=pk) if pk else None
    tipos = TipoDocumento.objects.filter(activo=True).select_related('categoria').order_by(
        'categoria__orden', 'nombre'
    )
    items_existentes = list(plantilla.items.select_related('tipo_documento').order_by('orden')) if plantilla else []

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        tipo   = request.POST.get('tipo', 'PROYECTO')
        desc   = request.POST.get('descripcion', '').strip()
        activa = request.POST.get('activa') == 'on'

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        else:
            if plantilla:
                plantilla.nombre = nombre
                plantilla.tipo = tipo
                plantilla.descripcion = desc
                plantilla.activa = activa
                plantilla.save()
            else:
                plantilla = PlantillaDossier.objects.create(
                    nombre=nombre, tipo=tipo, descripcion=desc,
                    activa=activa, creado_por=request.user,
                )

            # Actualizar ítems: se envía lista de tipo_id + seccion + orden
            tipo_ids   = request.POST.getlist('item_tipo[]')
            secciones  = request.POST.getlist('item_seccion[]')
            ordenes    = request.POST.getlist('item_orden[]')
            obligat    = request.POST.getlist('item_obligatorio[]')
            instruc    = request.POST.getlist('item_instruccion[]')

            # Borrar los que no están en la nueva lista
            nuevos_ids = [int(x) for x in tipo_ids if x]
            plantilla.items.exclude(tipo_documento_id__in=nuevos_ids).delete()

            for i, tid in enumerate(tipo_ids):
                if not tid:
                    continue
                orden_val = int(ordenes[i]) if i < len(ordenes) and ordenes[i].isdigit() else i + 1
                sec_val   = secciones[i] if i < len(secciones) else ''
                obl_val   = str(tid) in obligat
                ins_val   = instruc[i] if i < len(instruc) else ''

                PlantillaDossierItem.objects.update_or_create(
                    plantilla=plantilla,
                    tipo_documento_id=tid,
                    defaults={
                        'orden': orden_val,
                        'seccion': sec_val,
                        'obligatorio': obl_val,
                        'instruccion': ins_val,
                    },
                )

            messages.success(request, f'Plantilla "{plantilla.nombre}" guardada.')
            return redirect('plantilla_dossier_list')

    context = {
        'titulo': 'Nueva Plantilla' if not plantilla else f'Editar: {plantilla.nombre}',
        'plantilla': plantilla,
        'tipos': tipos,
        'items_existentes': items_existentes,
        'TIPO_CHOICES': PlantillaDossier.TIPO_CHOICES,
    }
    return render(request, 'documentos/plantilla_dossier_form.html', context)
