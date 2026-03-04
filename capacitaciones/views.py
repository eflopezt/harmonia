"""
Vistas del módulo de Capacitaciones / LMS Ligero.
"""
import io
import json
from calendar import monthcalendar
from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Sum, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST

from personal.models import Personal, Area
from .models import (
    CategoriaCapacitacion, Capacitacion, AsistenciaCapacitacion,
    RequerimientoCapacitacion, CertificacionTrabajador,
)

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ══════════════════════════════════════════════════════════════
# ADMIN — PANEL PRINCIPAL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def capacitaciones_panel(request):
    """Panel principal de capacitaciones."""
    qs = Capacitacion.objects.select_related('categoria').all()

    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '')
    buscar = request.GET.get('q', '')
    anio = request.GET.get('anio', str(date.today().year))

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo=tipo)
    if buscar:
        qs = qs.filter(Q(titulo__icontains=buscar) | Q(instructor__icontains=buscar))
    if anio:
        qs = qs.filter(fecha_inicio__year=int(anio))

    # Stats del filtro actual
    total_horas = qs.filter(estado='COMPLETADA').aggregate(t=Sum('horas'))['t'] or Decimal('0.0')
    total_costo = qs.filter(estado='COMPLETADA').aggregate(t=Sum('costo'))['t'] or Decimal('0.00')
    anio_int = int(anio) if anio else date.today().year
    inscritos_total = AsistenciaCapacitacion.objects.filter(
        capacitacion__fecha_inicio__year=anio_int,
    ).count()

    # KPIs globales del año
    total_personal_activo = Personal.objects.filter(estado='Activo').count()
    personal_capacitado_ids = AsistenciaCapacitacion.objects.filter(
        capacitacion__fecha_inicio__year=anio_int,
        estado__in=['ASISTIO', 'PARCIAL'],
    ).values_list('personal_id', flat=True).distinct()
    num_personal_capacitado = personal_capacitado_ids.count()
    pct_capacitados = (
        round((num_personal_capacitado / total_personal_activo) * 100)
        if total_personal_activo else 0
    )
    caps_pendientes = Capacitacion.objects.filter(
        fecha_inicio__gte=date.today(),
        estado='PROGRAMADA',
    ).count()
    total_horas_anio = Capacitacion.objects.filter(
        fecha_inicio__year=anio_int,
        estado='COMPLETADA',
    ).aggregate(t=Sum('horas'))['t'] or Decimal('0.0')

    hoy = date.today()
    completadas_anio = Capacitacion.objects.filter(
        fecha_inicio__year=anio_int, estado='COMPLETADA'
    ).count()
    este_mes = Capacitacion.objects.filter(
        fecha_inicio__year=hoy.year, fecha_inicio__month=hoy.month
    ).count()

    # ── Analytics extras ──────────────────────────────────────────────────────
    # Completadas este mes
    try:
        completadas_mes = Capacitacion.objects.filter(
            estado='COMPLETADA',
            fecha_inicio__year=hoy.year,
            fecha_inicio__month=hoy.month,
        ).count()
    except Exception:
        completadas_mes = 0

    # Horas impartidas este mes (capacitaciones completadas con fecha_inicio en el mes)
    try:
        horas_mes_raw = Capacitacion.objects.filter(
            estado='COMPLETADA',
            fecha_inicio__year=hoy.year,
            fecha_inicio__month=hoy.month,
        ).aggregate(t=Sum('horas'))['t']
        horas_capacitacion_mes = horas_mes_raw or Decimal('0.0')
    except Exception:
        horas_capacitacion_mes = Decimal('0.0')

    # Certificados por vencer (estado=POR_VENCER)
    try:
        certificados_por_vencer = CertificacionTrabajador.objects.filter(
            estado='POR_VENCER'
        ).count()
    except Exception:
        certificados_por_vencer = 0

    # En curso este momento
    try:
        en_curso_count = Capacitacion.objects.filter(estado='EN_CURSO').count()
    except Exception:
        en_curso_count = 0

    # Chart 1: Capacitaciones por categoría (top 6, incluye sin categoría)
    try:
        cats_data = (
            Capacitacion.objects
            .values('categoria__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')[:6]
        )
        tasa_completitud_json = json.dumps([
            {
                'label': item['categoria__nombre'] or 'Sin categoría',
                'value': item['total'],
            }
            for item in cats_data
        ])
    except Exception:
        tasa_completitud_json = json.dumps([])

    # Chart 2: Participantes/asistencias por tipo de capacitación
    try:
        tipos_data = (
            AsistenciaCapacitacion.objects
            .values('capacitacion__tipo')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        TIPO_LABELS = {
            'INTERNA': 'Interna',
            'EXTERNA': 'Externa',
            'ELEARNING': 'E-Learning',
            'INDUCCION': 'Inducción',
            'SSOMA': 'SSOMA',
        }
        participantes_por_tipo_json = json.dumps([
            {
                'label': TIPO_LABELS.get(item['capacitacion__tipo'], item['capacitacion__tipo'] or 'Sin tipo'),
                'value': item['total'],
            }
            for item in tipos_data
        ])
    except Exception:
        participantes_por_tipo_json = json.dumps([])

    # Top 3 capacitaciones EN_CURSO para el ranking
    try:
        caps_en_curso = list(
            Capacitacion.objects
            .filter(estado='EN_CURSO')
            .select_related('categoria')
            .annotate(inscritos_count=Count('asistencias'))
            .order_by('-inscritos_count')[:3]
        )
    except Exception:
        caps_en_curso = []

    # ── Nuevos KPIs enriquecidos ───────────────────────────────────────────

    # Cursos activos: programados + en curso
    try:
        cursos_activos = Capacitacion.objects.filter(
            estado__in=['PROGRAMADA', 'EN_CURSO']
        ).count()
    except Exception:
        cursos_activos = 0

    # Asistentes reales este mes (estado ASISTIO o PARCIAL)
    try:
        asistentes_mes = AsistenciaCapacitacion.objects.filter(
            capacitacion__fecha_inicio__year=hoy.year,
            capacitacion__fecha_inicio__month=hoy.month,
            estado__in=['ASISTIO', 'PARCIAL'],
        ).count()
    except Exception:
        asistentes_mes = 0

    # Tasa de asistencia del mes (asistentes / total inscritos este mes)
    try:
        total_inscritos_mes = AsistenciaCapacitacion.objects.filter(
            capacitacion__fecha_inicio__year=hoy.year,
            capacitacion__fecha_inicio__month=hoy.month,
        ).count()
        tasa_asistencia = (
            round((asistentes_mes / total_inscritos_mes) * 100)
            if total_inscritos_mes > 0 else 0
        )
    except Exception:
        tasa_asistencia = 0

    # Certificados emitidos vigentes
    try:
        certificados_emitidos = CertificacionTrabajador.objects.filter(
            estado='VIGENTE'
        ).count()
    except Exception:
        certificados_emitidos = 0

    # Tendencia 6 meses: capacitaciones por mes (completadas)
    try:
        MESES_CORTOS = [
            '', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
            'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'
        ]
        tendencia_labels = []
        tendencia_values = []
        for delta in range(5, -1, -1):
            # Calculate month offset from today
            ref_month = hoy.month - delta
            ref_year = hoy.year
            while ref_month < 1:
                ref_month += 12
                ref_year -= 1
            count = Capacitacion.objects.filter(
                fecha_inicio__year=ref_year,
                fecha_inicio__month=ref_month,
            ).count()
            tendencia_labels.append(MESES_CORTOS[ref_month])
            tendencia_values.append(count)
        tendencia_6m_json = json.dumps({
            'labels': tendencia_labels,
            'data': tendencia_values,
        })
    except Exception:
        tendencia_6m_json = json.dumps({'labels': [], 'data': []})

    context = {
        'titulo': 'Capacitaciones',
        'capacitaciones': qs[:100],
        'total': qs.count(),
        'filtro_estado': estado,
        'filtro_tipo': tipo,
        'buscar': buscar,
        'anio_filtro': anio,
        'categorias': CategoriaCapacitacion.objects.filter(activa=True),
        'stats': {
            'total_horas': total_horas,
            'total_costo': total_costo,
            'inscritos': inscritos_total,
            'programadas': Capacitacion.objects.filter(estado='PROGRAMADA').count(),
        },
        'kpis': {
            'total_anio': Capacitacion.objects.filter(fecha_inicio__year=anio_int).count(),
            'total_horas_anio': total_horas_anio,
            'pct_capacitados': pct_capacitados,
            'num_personal_capacitado': num_personal_capacitado,
            'total_personal_activo': total_personal_activo,
            'caps_pendientes': caps_pendientes,
            'completadas_anio': completadas_anio,
            'este_mes': este_mes,
        },
        'anio_actual': anio_int,
        # Analytics extras
        'completadas_mes': completadas_mes,
        'horas_capacitacion_mes': horas_capacitacion_mes,
        'certificados_por_vencer': certificados_por_vencer,
        'en_curso_count': en_curso_count,
        'tasa_completitud_json': tasa_completitud_json,
        'participantes_por_tipo_json': participantes_por_tipo_json,
        'caps_en_curso': caps_en_curso,
        # Nuevos KPIs enriquecidos
        'cursos_activos': cursos_activos,
        'asistentes_mes': asistentes_mes,
        'tasa_asistencia': tasa_asistencia,
        'certificados_emitidos': certificados_emitidos,
        'tendencia_6m_json': tendencia_6m_json,
    }
    return render(request, 'capacitaciones/panel.html', context)


@login_required
@solo_admin
def capacitacion_crear(request):
    """Crear nueva capacitación."""
    if request.method == 'POST':
        try:
            cap = Capacitacion.objects.create(
                titulo=request.POST['titulo'],
                descripcion=request.POST.get('descripcion', ''),
                categoria_id=request.POST.get('categoria_id') or None,
                tipo=request.POST.get('tipo', 'INTERNA'),
                instructor=request.POST.get('instructor', ''),
                lugar=request.POST.get('lugar', ''),
                fecha_inicio=request.POST['fecha_inicio'],
                fecha_fin=request.POST.get('fecha_fin') or None,
                horas=Decimal(request.POST.get('horas', '1')),
                costo=Decimal(request.POST.get('costo', '0')),
                max_participantes=int(request.POST.get('max_participantes', 0) or 0),
                obligatoria=request.POST.get('obligatoria') == 'on',
                creado_por=request.user,
            )
            if request.FILES.get('material'):
                cap.material_archivo = request.FILES['material']
                cap.save(update_fields=['material_archivo'])

            from core.audit import log_create
            log_create(request, cap, f'Capacitación creada: {cap.titulo}')
            messages.success(request, f'Capacitación "{cap.titulo}" creada.')
            return redirect('capacitacion_detalle', pk=cap.pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': 'Nueva Capacitación',
        'categorias': CategoriaCapacitacion.objects.filter(activa=True),
    }
    return render(request, 'capacitaciones/crear.html', context)


@login_required
@solo_admin
def capacitacion_detalle(request, pk):
    """Detalle de una capacitación con lista de participantes."""
    cap = get_object_or_404(
        Capacitacion.objects.select_related('categoria', 'creado_por'),
        pk=pk
    )
    asistencias = cap.asistencias.select_related('personal').all()

    context = {
        'titulo': cap.titulo,
        'cap': cap,
        'asistencias': asistencias,
        'personal_disponible': Personal.objects.filter(
            estado='Activo',
        ).exclude(
            pk__in=asistencias.values_list('personal_id', flat=True)
        ).order_by('apellidos_nombres')[:200],
    }
    return render(request, 'capacitaciones/detalle.html', context)


@login_required
@solo_admin
@require_POST
def participante_agregar(request, cap_pk):
    """Agregar participante a capacitación."""
    cap = get_object_or_404(Capacitacion, pk=cap_pk)
    personal = get_object_or_404(Personal, pk=request.POST['personal_id'])

    obj, created = AsistenciaCapacitacion.objects.get_or_create(
        capacitacion=cap, personal=personal,
    )
    if created:
        return JsonResponse({'ok': True, 'nombre': personal.apellidos_nombres})
    return JsonResponse({'ok': False, 'error': 'Ya está inscrito.'})


@login_required
@solo_admin
@require_POST
def participante_asistencia(request, pk):
    """Registrar asistencia/nota de un participante."""
    asist = get_object_or_404(AsistenciaCapacitacion, pk=pk)
    asist.estado = request.POST.get('estado', asist.estado)
    nota = request.POST.get('nota')
    if nota:
        asist.nota = Decimal(nota)
    asist.aprobado = request.POST.get('aprobado') == 'on'
    asist.observaciones = request.POST.get('observaciones', asist.observaciones)
    if request.FILES.get('certificado'):
        asist.certificado = request.FILES['certificado']
    asist.save()
    return JsonResponse({'ok': True, 'estado': asist.get_estado_display(), 'aprobado': asist.aprobado})


@login_required
@solo_admin
@require_POST
def capacitacion_completar(request, pk):
    """Marcar capacitación como completada."""
    cap = get_object_or_404(Capacitacion, pk=pk)
    cap.estado = 'COMPLETADA'
    cap.save(update_fields=['estado'])

    from core.audit import log_update
    log_update(request, cap, {'estado': {'old': 'EN_CURSO', 'new': 'COMPLETADA'}})
    return JsonResponse({'ok': True})


# ══════════════════════════════════════════════════════════════
# ADMIN — REQUERIMIENTOS Y CUMPLIMIENTO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def requerimientos_panel(request):
    """Panel de requerimientos de capacitación."""
    reqs = RequerimientoCapacitacion.objects.select_related('categoria').filter(activo=True)

    context = {
        'titulo': 'Requerimientos de Capacitación',
        'requerimientos': reqs,
        'categorias': CategoriaCapacitacion.objects.filter(activa=True),
    }
    return render(request, 'capacitaciones/requerimientos.html', context)


@login_required
@solo_admin
def incumplimientos_panel(request):
    """Reporte de incumplimientos: quién falta capacitarse."""
    reqs = RequerimientoCapacitacion.objects.filter(activo=True, obligatorio=True)
    personal_qs = Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')

    grupo = request.GET.get('grupo', '')
    if grupo:
        personal_qs = personal_qs.filter(grupo_tareo=grupo)

    matriz = []
    for emp in personal_qs[:300]:
        faltantes = []
        for req in reqs:
            # Verificar si aplica
            if emp.grupo_tareo == 'STAFF' and not req.aplica_staff:
                continue
            if emp.grupo_tareo == 'RCO' and not req.aplica_rco:
                continue

            # Verificar certificación vigente
            cert = CertificacionTrabajador.objects.filter(
                personal=emp, requerimiento=req, estado='VIGENTE',
            ).first()
            if not cert:
                faltantes.append(req)

        if faltantes:
            matriz.append({
                'personal': emp,
                'faltantes': faltantes,
                'total': len(faltantes),
            })

    context = {
        'titulo': 'Incumplimientos de Capacitación',
        'matriz': sorted(matriz, key=lambda x: -x['total']),
        'requerimientos': reqs,
        'filtro_grupo': grupo,
    }
    return render(request, 'capacitaciones/incumplimientos.html', context)


# ══════════════════════════════════════════════════════════════
# EXPORTACIONES Y DOCUMENTOS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def exportar_asistentes_excel(request, capacitacion_pk):
    """Exporta la lista de asistentes de una capacitación a Excel con estilos Harmoni."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cap = get_object_or_404(Capacitacion, pk=capacitacion_pk)
    asistencias = cap.asistencias.select_related(
        'personal', 'personal__subarea', 'personal__subarea__area'
    ).order_by('personal__apellidos_nombres')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistentes"

    # Estilos Harmoni
    COLOR_HEADER = "0D2B27"
    COLOR_SUBHEADER = "134E4A"
    COLOR_ACCENT = "5EEAD4"
    COLOR_FILA_PAR = "F0FFFE"

    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    subheader_font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    subheader_fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")

    fila_par_fill = PatternFill(start_color=COLOR_FILA_PAR, end_color=COLOR_FILA_PAR, fill_type="solid")
    centro = Alignment(horizontal='center', vertical='center')
    borde_fino = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Fila 1: Título principal
    ws.merge_cells('A1:F1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"LISTA DE ASISTENTES — {cap.titulo.upper()}"
    titulo_cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=13)
    titulo_cell.fill = header_fill
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2: Subtítulo con info de capacitación
    ws.merge_cells('A2:F2')
    fecha_str = cap.fecha_inicio.strftime('%d/%m/%Y')
    if cap.fecha_fin and cap.fecha_fin != cap.fecha_inicio:
        fecha_str += f" al {cap.fecha_fin.strftime('%d/%m/%Y')}"
    info_str = (
        f"Tipo: {cap.get_tipo_display()}  |  "
        f"Instructor: {cap.instructor or '—'}  |  "
        f"Lugar: {cap.lugar or '—'}  |  "
        f"Fecha: {fecha_str}  |  "
        f"Horas: {cap.horas}h"
    )
    sub_cell = ws['A2']
    sub_cell.value = info_str
    sub_cell.font = Font(name='Calibri', color="FFFFFF", size=9)
    sub_cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 20

    # Fila 3: encabezados de columna
    columnas = ["DNI", "Apellidos y Nombres", "Área", "Asistió", "Nota", "Certificado"]
    anchos =   [14,     38,                    28,     12,         10,     14]
    for col_idx, (col_name, ancho) in enumerate(zip(columnas, anchos), start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = header_align
        cell.border = borde_fino
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[3].height = 22

    # Datos
    for fila_idx, asist in enumerate(asistencias, start=4):
        p = asist.personal
        area_nombre = ""
        if hasattr(p, 'subarea') and p.subarea:
            area_nombre = p.subarea.area.nombre if p.subarea.area else p.subarea.nombre
        elif hasattr(p, 'area') and p.area:
            area_nombre = str(p.area)

        asistio = "Sí" if asist.estado == 'ASISTIO' else ("Parcial" if asist.estado == 'PARCIAL' else "No")
        nota_str = str(asist.nota) if asist.nota is not None else "—"
        cert_str = "Sí" if asist.certificado else "No"

        fila_data = [
            p.nro_doc,
            p.apellidos_nombres,
            area_nombre,
            asistio,
            nota_str,
            cert_str,
        ]

        usa_fila_par = (fila_idx % 2 == 0)
        for col_idx, valor in enumerate(fila_data, start=1):
            cell = ws.cell(row=fila_idx, column=col_idx, value=valor)
            cell.font = Font(name='Calibri', size=10)
            cell.border = borde_fino
            if col_idx in (4, 5, 6):
                cell.alignment = centro
            else:
                cell.alignment = Alignment(vertical='center')
            if usa_fila_par:
                cell.fill = fila_par_fill
            # Color condicional para Asistió
            if col_idx == 4:
                if asistio == "Sí":
                    cell.font = Font(name='Calibri', size=10, color="1A6B41", bold=True)
                elif asistio == "No":
                    cell.font = Font(name='Calibri', size=10, color="B91C1C")

        ws.row_dimensions[fila_idx].height = 18

    # Fila resumen al final
    fila_total = len(list(asistencias)) + 4
    ws.cell(row=fila_total, column=1, value="TOTALES").font = Font(bold=True, size=10)
    ws.cell(row=fila_total, column=1).fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    ws.cell(row=fila_total, column=4, value=f"{cap.num_inscritos} inscritos / {cap.num_aprobados} aprobados")
    ws.cell(row=fila_total, column=4).font = Font(bold=True, size=10)
    ws.cell(row=fila_total, column=4).fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    # Congelar encabezado
    ws.freeze_panes = 'A4'

    # Respuesta HTTP
    nombre_archivo = f"asistentes_{cap.pk}_{cap.fecha_inicio.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
def generar_certificado_pdf(request, asistencia_pk):
    """
    Genera un certificado de participación en PDF para una asistencia aprobada.
    Usa ReportLab. El empleado solo puede ver su propio certificado; admin ve cualquiera.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    asist = get_object_or_404(
        AsistenciaCapacitacion.objects.select_related('personal', 'capacitacion'),
        pk=asistencia_pk
    )

    # Seguridad: el trabajador solo puede ver su propio certificado
    if not request.user.is_superuser:
        try:
            from portal.views import _get_empleado
            empleado = _get_empleado(request.user)
            if not empleado or empleado.pk != asist.personal.pk:
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied
        except ImportError:
            if not request.user.is_superuser:
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied

    if not asist.aprobado:
        messages.warning(request, "El participante no ha sido marcado como aprobado.")
        return redirect('capacitacion_detalle', pk=asist.capacitacion.pk)

    # Datos de empresa
    try:
        from asistencia.models import ConfiguracionSistema
        config = ConfiguracionSistema.get_config()
        empresa_nombre = config.empresa_nombre or "Harmoni ERP"
        empresa_direccion = config.empresa_direccion or ""
        empresa_ruc = config.ruc or ""
    except Exception:
        empresa_nombre = "Harmoni ERP"
        empresa_direccion = ""
        empresa_ruc = ""

    cap = asist.capacitacion
    personal = asist.personal
    fecha_cert = asist.fecha_certificado or cap.fecha_fin or cap.fecha_inicio
    TEAL_DARK = colors.HexColor('#0D2B27')
    TEAL_MID = colors.HexColor('#0F766E')
    TEAL_LIGHT = colors.HexColor('#5EEAD4')
    GRIS = colors.HexColor('#64748B')
    BLANCO = colors.white

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()

    estilo_empresa = ParagraphStyle(
        'empresa', fontSize=10, textColor=GRIS,
        alignment=TA_CENTER, spaceAfter=2,
    )
    estilo_titulo_cert = ParagraphStyle(
        'titulo_cert', fontSize=28, textColor=TEAL_DARK,
        alignment=TA_CENTER, spaceAfter=4, fontName='Helvetica-Bold',
        leading=34,
    )
    estilo_subtitulo = ParagraphStyle(
        'subtitulo', fontSize=13, textColor=TEAL_MID,
        alignment=TA_CENTER, spaceAfter=8, fontName='Helvetica',
    )
    estilo_otorga = ParagraphStyle(
        'otorga', fontSize=11, textColor=GRIS,
        alignment=TA_CENTER, spaceAfter=4,
    )
    estilo_nombre = ParagraphStyle(
        'nombre', fontSize=24, textColor=TEAL_DARK,
        alignment=TA_CENTER, spaceAfter=6, fontName='Helvetica-Bold',
        leading=30,
    )
    estilo_por = ParagraphStyle(
        'por', fontSize=11, textColor=GRIS,
        alignment=TA_CENTER, spaceAfter=4,
    )
    estilo_capacitacion = ParagraphStyle(
        'capacitacion', fontSize=16, textColor=TEAL_MID,
        alignment=TA_CENTER, spaceAfter=6, fontName='Helvetica-Bold',
        leading=22,
    )
    estilo_detalles = ParagraphStyle(
        'detalles', fontSize=10, textColor=GRIS,
        alignment=TA_CENTER, spaceAfter=4,
    )
    estilo_firma_label = ParagraphStyle(
        'firma_label', fontSize=9, textColor=GRIS,
        alignment=TA_CENTER,
    )

    fecha_formateada = fecha_cert.strftime('%d de %B de %Y') if fecha_cert else date.today().strftime('%d de %B de %Y')
    # Traducción de meses al español
    meses = {
        'January': 'enero', 'February': 'febrero', 'March': 'marzo',
        'April': 'abril', 'May': 'mayo', 'June': 'junio',
        'July': 'julio', 'August': 'agosto', 'September': 'septiembre',
        'October': 'octubre', 'November': 'noviembre', 'December': 'diciembre',
    }
    for en, es in meses.items():
        fecha_formateada = fecha_formateada.replace(en, es)

    detalles_partes = [f"Duración: {cap.horas} horas"]
    if cap.instructor:
        detalles_partes.append(f"Instructor: {cap.instructor}")
    if cap.lugar:
        detalles_partes.append(f"Lugar: {cap.lugar}")
    detalles_str = "  |  ".join(detalles_partes)

    elementos = []

    # Encabezado empresa
    elementos.append(Paragraph(empresa_nombre.upper(), ParagraphStyle(
        'emp_nombre', fontSize=14, textColor=TEAL_DARK,
        alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=2,
    )))
    if empresa_ruc:
        elementos.append(Paragraph(f"RUC: {empresa_ruc}", estilo_empresa))
    if empresa_direccion:
        elementos.append(Paragraph(empresa_direccion, estilo_empresa))

    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(HRFlowable(width="100%", thickness=2, color=TEAL_DARK, spaceAfter=0.3*cm))

    # Título
    elementos.append(Paragraph("CERTIFICADO DE PARTICIPACIÓN", estilo_titulo_cert))
    elementos.append(HRFlowable(width="60%", thickness=1, color=TEAL_LIGHT, spaceAfter=0.4*cm))

    # Texto central
    elementos.append(Paragraph("Se certifica que:", estilo_otorga))
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(personal.apellidos_nombres, estilo_nombre))
    if personal.nro_doc:
        elementos.append(Paragraph(f"DNI: {personal.nro_doc}", estilo_detalles))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph("participó satisfactoriamente en la capacitación:", estilo_por))
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(f'"{cap.titulo}"', estilo_capacitacion))
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(detalles_str, estilo_detalles))
    if asist.nota is not None:
        nota_str = f"Calificación obtenida: {asist.nota}"
        elementos.append(Paragraph(nota_str, ParagraphStyle(
            'nota', fontSize=10, textColor=TEAL_MID,
            alignment=TA_CENTER, fontName='Helvetica-Bold',
        )))

    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(Paragraph(
        f"Lima, {fecha_formateada}",
        ParagraphStyle('fecha', fontSize=11, textColor=GRIS, alignment=TA_CENTER)
    ))

    elementos.append(Spacer(1, 0.8*cm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E2E8F0'), spaceAfter=0.3*cm))

    # Firmas
    firma_style = ParagraphStyle('firma_linea', fontSize=10, textColor=TEAL_DARK,
                                  alignment=TA_CENTER, fontName='Helvetica-Bold')
    tabla_firmas = Table(
        [
            [Paragraph("_______________________________", firma_style),
             Paragraph("_______________________________", firma_style)],
            [Paragraph("Jefe de Recursos Humanos", estilo_firma_label),
             Paragraph("Gerente General", estilo_firma_label)],
            [Paragraph(empresa_nombre, estilo_firma_label),
             Paragraph(empresa_nombre, estilo_firma_label)],
        ],
        colWidths=[10*cm, 10*cm],
    )
    tabla_firmas.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla_firmas)

    doc.build(elementos)
    buffer.seek(0)

    nombre_archivo = (
        f"certificado_{personal.nro_doc}_{cap.pk}.pdf"
    )
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'
    return response


# ══════════════════════════════════════════════════════════════
# PORTAL DEL TRABAJADOR
# ══════════════════════════════════════════════════════════════

@login_required
def mis_capacitaciones(request):
    """Portal: mis capacitaciones y certificaciones."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    asistencias = []
    certificaciones = []
    stats = {}

    if empleado:
        asistencias = AsistenciaCapacitacion.objects.filter(
            personal=empleado,
        ).select_related('capacitacion', 'capacitacion__categoria').order_by('-capacitacion__fecha_inicio')

        certificaciones = CertificacionTrabajador.objects.filter(
            personal=empleado,
        ).select_related('requerimiento').order_by('-fecha_obtencion')

        total_horas = asistencias.filter(
            estado='ASISTIO',
        ).aggregate(t=Sum('capacitacion__horas'))['t'] or Decimal('0.0')

        stats = {
            'total_capacitaciones': asistencias.count(),
            'total_horas': total_horas,
            'aprobadas': asistencias.filter(aprobado=True).count(),
            'certificaciones_vigentes': certificaciones.filter(estado='VIGENTE').count(),
        }

    context = {
        'titulo': 'Mis Capacitaciones',
        'empleado': empleado,
        'asistencias': asistencias,
        'certificaciones': certificaciones,
        'stats': stats,
    }
    return render(request, 'capacitaciones/mis_capacitaciones.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — ASIGNACIÓN MASIVA
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def asignacion_masiva(request):
    """Asignar masivamente empleados a una capacitación."""
    capacitaciones = Capacitacion.objects.exclude(estado='CANCELADA').order_by('-fecha_inicio')
    areas = Area.objects.filter(activa=True).order_by('nombre')

    if request.method == 'POST':
        cap_id = request.POST.get('capacitacion')
        filtro_tipo = request.POST.get('filtro_tipo', 'todos')
        filtro_valor = request.POST.get('filtro_valor', '')

        cap = get_object_or_404(Capacitacion, pk=cap_id)
        qs = Personal.objects.filter(estado='Activo')

        if filtro_tipo == 'area':
            qs = qs.filter(subarea__area_id=filtro_valor)
        elif filtro_tipo == 'grupo':
            qs = qs.filter(grupo_tareo=filtro_valor)
        elif filtro_tipo == 'cargo':
            qs = qs.filter(cargo__icontains=filtro_valor)
        # 'todos' — no additional filter

        creados = 0
        ya_inscritos = 0
        asignados = []
        for emp in qs:
            obj, created = AsistenciaCapacitacion.objects.get_or_create(
                capacitacion=cap,
                personal=emp,
            )
            if created:
                creados += 1
                asignados.append(emp.apellidos_nombres)
            else:
                ya_inscritos += 1

        messages.success(
            request,
            f'{creados} empleado(s) asignado(s) a "{cap.titulo}". '
            f'{ya_inscritos} ya estaban inscritos.'
        )
        return redirect('capacitacion_detalle', pk=cap.pk)

    # AJAX: preview count
    if request.GET.get('preview') == '1':
        filtro_tipo = request.GET.get('filtro_tipo', 'todos')
        filtro_valor = request.GET.get('filtro_valor', '')
        cap_id = request.GET.get('capacitacion', '')

        qs = Personal.objects.filter(estado='Activo')
        if filtro_tipo == 'area' and filtro_valor:
            qs = qs.filter(subarea__area_id=filtro_valor)
        elif filtro_tipo == 'grupo' and filtro_valor:
            qs = qs.filter(grupo_tareo=filtro_valor)
        elif filtro_tipo == 'cargo' and filtro_valor:
            qs = qs.filter(cargo__icontains=filtro_valor)

        ya_inscritos = 0
        nuevos = 0
        if cap_id:
            inscritos_ids = AsistenciaCapacitacion.objects.filter(
                capacitacion_id=cap_id
            ).values_list('personal_id', flat=True)
            ya_inscritos = qs.filter(pk__in=inscritos_ids).count()
            nuevos = qs.exclude(pk__in=inscritos_ids).count()
        else:
            nuevos = qs.count()

        return JsonResponse({'total': qs.count(), 'nuevos': nuevos, 'ya_inscritos': ya_inscritos})

    context = {
        'titulo': 'Asignación Masiva de Capacitación',
        'capacitaciones': capacitaciones,
        'areas': areas,
    }
    return render(request, 'capacitaciones/asignacion_masiva.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — ESTADÍSTICAS DE CAPACITACIÓN
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def capacitacion_estadisticas(request, pk):
    """Estadísticas detalladas de una capacitación."""
    cap = get_object_or_404(
        Capacitacion.objects.select_related('categoria'),
        pk=pk
    )
    asistencias = cap.asistencias.select_related(
        'personal', 'personal__subarea', 'personal__subarea__area'
    ).all()

    total_inscritos = asistencias.count()
    asistieron = asistencias.filter(estado__in=['ASISTIO', 'PARCIAL']).count()
    ausentes = asistencias.filter(estado='NO_ASISTIO').count()
    aprobados = asistencias.filter(aprobado=True).count()
    reprobados = total_inscritos - aprobados
    certificados = asistencias.exclude(certificado='').exclude(certificado__isnull=True).count()

    promedio_nota = asistencias.filter(nota__isnull=False).aggregate(
        avg=Avg('nota')
    )['avg']
    promedio_nota = round(float(promedio_nota), 2) if promedio_nota else None

    # Por área
    areas_data = {}
    for asist in asistencias:
        p = asist.personal
        area_nombre = 'Sin Área'
        if p.subarea and p.subarea.area:
            area_nombre = p.subarea.area.nombre
        elif p.subarea:
            area_nombre = p.subarea.nombre
        if area_nombre not in areas_data:
            areas_data[area_nombre] = {'inscritos': 0, 'asistieron': 0, 'aprobados': 0}
        areas_data[area_nombre]['inscritos'] += 1
        if asist.estado in ('ASISTIO', 'PARCIAL'):
            areas_data[area_nombre]['asistieron'] += 1
        if asist.aprobado:
            areas_data[area_nombre]['aprobados'] += 1

    # Por grupo_tareo
    grupos_data = {}
    for asist in asistencias:
        grupo = asist.personal.grupo_tareo or 'SIN GRUPO'
        if grupo not in grupos_data:
            grupos_data[grupo] = {'inscritos': 0, 'asistieron': 0, 'aprobados': 0}
        grupos_data[grupo]['inscritos'] += 1
        if asist.estado in ('ASISTIO', 'PARCIAL'):
            grupos_data[grupo]['asistieron'] += 1
        if asist.aprobado:
            grupos_data[grupo]['aprobados'] += 1

    # Ranking por nota (solo los que tienen nota)
    ranking = asistencias.filter(nota__isnull=False).order_by('-nota')

    context = {
        'titulo': f'Estadísticas — {cap.titulo}',
        'cap': cap,
        'stats': {
            'total_inscritos': total_inscritos,
            'asistieron': asistieron,
            'ausentes': ausentes,
            'aprobados': aprobados,
            'reprobados': reprobados,
            'certificados': certificados,
            'promedio_nota': promedio_nota,
            'pct_asistencia': round((asistieron / total_inscritos * 100)) if total_inscritos else 0,
            'pct_aprobacion': round((aprobados / total_inscritos * 100)) if total_inscritos else 0,
        },
        'areas_data': sorted(areas_data.items()),
        'grupos_data': grupos_data,
        'ranking': ranking,
    }
    return render(request, 'capacitaciones/estadisticas.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — MIS REQUERIMIENTOS (vista alternativa)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def mis_requerimientos(request):
    """Panel de requerimientos con filtro pendientes/completados."""
    estado_filtro = request.GET.get('estado', 'pendientes')
    area_filtro = request.GET.get('area', '')

    reqs = RequerimientoCapacitacion.objects.select_related('categoria').filter(activo=True)

    if area_filtro:
        reqs = reqs.filter(aplica_areas__id=area_filtro)

    areas = Area.objects.filter(activa=True).order_by('nombre')

    reqs_con_estado = []
    for req in reqs:
        # Contar certificaciones vigentes para este requerimiento
        certs_vigentes = CertificacionTrabajador.objects.filter(
            requerimiento=req, estado='VIGENTE'
        ).count()
        # Empleados a quienes aplica
        total_aplica = Personal.objects.filter(estado='Activo')
        if not req.aplica_todos:
            filtros = Q()
            if req.aplica_staff:
                filtros |= Q(grupo_tareo='STAFF')
            if req.aplica_rco:
                filtros |= Q(grupo_tareo='RCO')
            total_aplica = total_aplica.filter(filtros)
        total_aplica_count = total_aplica.count()
        pendientes_count = total_aplica_count - certs_vigentes

        req_data = {
            'req': req,
            'certs_vigentes': certs_vigentes,
            'total_aplica': total_aplica_count,
            'pendientes': max(pendientes_count, 0),
            'completado': pendientes_count <= 0,
        }
        if estado_filtro == 'pendientes' and pendientes_count > 0:
            reqs_con_estado.append(req_data)
        elif estado_filtro == 'completados' and pendientes_count <= 0:
            reqs_con_estado.append(req_data)
        elif estado_filtro == 'todos':
            reqs_con_estado.append(req_data)

    context = {
        'titulo': 'Requerimientos de Capacitación',
        'reqs_con_estado': reqs_con_estado,
        'estado_filtro': estado_filtro,
        'area_filtro': area_filtro,
        'areas': areas,
    }
    return render(request, 'capacitaciones/mis_requerimientos.html', context)


# ══════════════════════════════════════════════════════════════
# CALENDARIO DE CAPACITACIONES
# ══════════════════════════════════════════════════════════════

@login_required
def calendario_capacitaciones(request):
    """Vista de calendario mensual de capacitaciones."""
    hoy = date.today()
    mes = int(request.GET.get('mes', hoy.month))
    anio = int(request.GET.get('anio', hoy.year))

    # Navegación prev/next
    if mes == 1:
        mes_prev, anio_prev = 12, anio - 1
    else:
        mes_prev, anio_prev = mes - 1, anio

    if mes == 12:
        mes_next, anio_next = 1, anio + 1
    else:
        mes_next, anio_next = mes + 1, anio

    # Capacitaciones del mes
    caps_mes = Capacitacion.objects.filter(
        fecha_inicio__year=anio,
        fecha_inicio__month=mes,
    ).exclude(estado='CANCELADA').select_related('categoria').order_by('fecha_inicio')

    # Agrupar por semana
    semanas_data = []
    cal = monthcalendar(anio, mes)
    for semana in cal:
        dias_no_cero = [d for d in semana if d != 0]
        if not dias_no_cero:
            continue
        inicio_sem = date(anio, mes, min(dias_no_cero))
        fin_sem = date(anio, mes, max(dias_no_cero))
        caps_semana = [c for c in caps_mes if inicio_sem <= c.fecha_inicio <= fin_sem]
        if caps_semana:
            semanas_data.append({
                'inicio': inicio_sem,
                'fin': fin_sem,
                'caps': caps_semana,
            })

    # Próximas capacitaciones (para sidebar)
    proximas = Capacitacion.objects.filter(
        fecha_inicio__gte=hoy,
        estado='PROGRAMADA',
    ).select_related('categoria').order_by('fecha_inicio')[:5]

    MESES_ES = [
        '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    context = {
        'titulo': 'Calendario de Capacitaciones',
        'mes': mes,
        'anio': anio,
        'mes_nombre': MESES_ES[mes],
        'mes_prev': mes_prev,
        'anio_prev': anio_prev,
        'mes_next': mes_next,
        'anio_next': anio_next,
        'semanas_data': semanas_data,
        'caps_mes': caps_mes,
        'proximas': proximas,
        'hoy': hoy,
    }
    return render(request, 'capacitaciones/calendario.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — DUPLICAR CAPACITACIÓN
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
@require_POST
def capacitacion_duplicar(request, pk):
    """Duplica una capacitación existente, poniendo estado PROGRAMADA."""
    cap = get_object_or_404(Capacitacion, pk=pk)

    nueva = Capacitacion.objects.create(
        titulo=f'[Copia] {cap.titulo}',
        descripcion=cap.descripcion,
        categoria=cap.categoria,
        tipo=cap.tipo,
        instructor=cap.instructor,
        lugar=cap.lugar,
        fecha_inicio=cap.fecha_inicio,
        fecha_fin=cap.fecha_fin,
        horas=cap.horas,
        costo=cap.costo,
        max_participantes=cap.max_participantes,
        obligatoria=cap.obligatoria,
        material_url=cap.material_url,
        estado='PROGRAMADA',
        creado_por=request.user,
    )

    from core.audit import log_create
    log_create(request, nueva, f'Capacitación duplicada desde #{cap.pk}: {nueva.titulo}')

    messages.success(request, f'Capacitación duplicada como "{nueva.titulo}".')
    return redirect('capacitacion_detalle', pk=nueva.pk)
