"""
Vistas del módulo de Evaluaciones de Desempeño.
"""
import json
from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Avg, Count, Max
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST

try:
    from personal.models import Personal, Area
except ImportError:
    Personal = None
    Area = None

from .models import (
    Competencia, PlantillaEvaluacion, PlantillaCompetencia,
    CicloEvaluacion, Evaluacion, RespuestaEvaluacion,
    ResultadoConsolidado, PlanDesarrollo, AccionDesarrollo,
    ObjetivoClave, ResultadoClave, CheckInOKR,
)

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ══════════════════════════════════════════════════════════════
# ADMIN — CICLOS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def ciclos_panel(request):
    """Panel principal de ciclos de evaluación."""
    qs = CicloEvaluacion.objects.all()

    estado = request.GET.get('estado', '')
    anio = request.GET.get('anio', str(date.today().year))

    if estado:
        qs = qs.filter(estado=estado)
    if anio:
        qs = qs.filter(fecha_inicio__year=int(anio))

    # ── KPI globales ──────────────────────────────────────────
    total_evaluaciones = Evaluacion.objects.count()
    evaluaciones_activas = Evaluacion.objects.filter(
        ciclo__estado__in=['ABIERTO', 'EN_EVALUACION'],
        estado__in=['PENDIENTE', 'EN_PROGRESO'],
    ).count()

    # Promedio general: puntaje_total relativo a escala_max de la plantilla
    completadas_qs = Evaluacion.objects.filter(
        estado='COMPLETADA',
        puntaje_total__isnull=False,
        ciclo__plantilla__isnull=False,
    ).select_related('ciclo__plantilla')

    promedio_general = None
    if completadas_qs.exists():
        total_ratio = Decimal('0')
        count_ratio = 0
        for ev in completadas_qs:
            escala = Decimal(str(ev.ciclo.plantilla.escala_max)) if ev.ciclo.plantilla else Decimal('5')
            if escala > 0:
                total_ratio += ev.puntaje_total / escala * 100
                count_ratio += 1
        if count_ratio:
            promedio_general = round(total_ratio / count_ratio, 1)

    ciclos_cerrados = CicloEvaluacion.objects.filter(estado='CERRADO').count()

    # ── Distribución 9-Box (últimos resultados activos) ───────
    # Keys are strings so Django template dot-notation (nine_box_counts.9) resolves correctly
    nine_box_counts = {str(i): 0 for i in range(1, 10)}
    nine_box_results = ResultadoConsolidado.objects.filter(
        nine_box_position__isnull=False,
    ).values('nine_box_position').annotate(total=Count('id'))
    for row in nine_box_results:
        pos = row['nine_box_position']
        if 1 <= pos <= 9:
            nine_box_counts[str(pos)] = row['total']
    hay_nine_box = any(v > 0 for v in nine_box_counts.values())

    # ── Evaluaciones recientes ────────────────────────────────
    recientes = Evaluacion.objects.select_related(
        'evaluado', 'ciclo',
    ).order_by('-creado_en')[:5]

    context = {
        'titulo': 'Evaluaciones de Desempeño',
        'ciclos': qs[:50],
        'total': qs.count(),
        'filtro_estado': estado,
        'anio_filtro': anio,
        # KPI cards
        'total_evaluaciones': total_evaluaciones,
        'evaluaciones_activas': evaluaciones_activas,
        'promedio_general': promedio_general,
        'ciclos_cerrados': ciclos_cerrados,
        # 9-Box miniatura
        'nine_box_counts': nine_box_counts,
        'hay_nine_box': hay_nine_box,
        # Recientes
        'recientes': recientes,
        'stats': {
            'activos': CicloEvaluacion.objects.filter(estado__in=['ABIERTO', 'EN_EVALUACION']).count(),
            'cerrados_anio': CicloEvaluacion.objects.filter(estado='CERRADO', fecha_fin__year=date.today().year).count(),
        },
    }
    return render(request, 'evaluaciones/ciclos_panel.html', context)


@login_required
@solo_admin
def ciclo_crear(request):
    """Crear nuevo ciclo de evaluación."""
    if request.method == 'POST':
        try:
            ciclo = CicloEvaluacion.objects.create(
                nombre=request.POST['nombre'],
                tipo=request.POST.get('tipo', '180'),
                plantilla_id=request.POST.get('plantilla_id') or None,
                fecha_inicio=request.POST['fecha_inicio'],
                fecha_fin=request.POST['fecha_fin'],
                descripcion=request.POST.get('descripcion', ''),
                estado='BORRADOR',
                creado_por=request.user,
            )
            from core.audit import log_create
            log_create(request, ciclo, f'Ciclo creado: {ciclo.nombre}')
            messages.success(request, f'Ciclo "{ciclo.nombre}" creado.')
            return redirect('ciclo_detalle', pk=ciclo.pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': 'Nuevo Ciclo de Evaluación',
        'plantillas': PlantillaEvaluacion.objects.filter(activa=True),
    }
    return render(request, 'evaluaciones/ciclo_crear.html', context)


@login_required
@solo_admin
def ciclo_detalle(request, pk):
    """Detalle de ciclo con evaluaciones y avance."""
    ciclo = get_object_or_404(CicloEvaluacion, pk=pk)
    evaluaciones = ciclo.evaluaciones.select_related('evaluado', 'evaluador').all()

    # Stats por relación
    stats_relacion = evaluaciones.values('relacion').annotate(
        total=Count('id'),
        completadas=Count('id', filter=Q(estado='COMPLETADA')),
    )

    context = {
        'titulo': ciclo.nombre,
        'ciclo': ciclo,
        'evaluaciones': evaluaciones[:200],
        'stats_relacion': {s['relacion']: s for s in stats_relacion},
        'personal_disponible': Personal.objects.filter(
            estado='Activo',
        ).exclude(
            pk__in=evaluaciones.filter(relacion='AUTO').values_list('evaluado_id', flat=True)
        ).order_by('apellidos_nombres')[:200],
    }
    return render(request, 'evaluaciones/ciclo_detalle.html', context)


@login_required
@solo_admin
@require_POST
def ciclo_generar_evaluaciones(request, pk):
    """Genera evaluaciones automáticas para todo el personal del ciclo."""
    ciclo = get_object_or_404(CicloEvaluacion, pk=pk)
    if ciclo.estado != 'BORRADOR':
        return JsonResponse({'ok': False, 'error': 'Solo se pueden generar en estado Borrador.'})

    personal_qs = Personal.objects.filter(estado='Activo')
    areas = ciclo.aplica_areas.all()
    if areas.exists():
        personal_qs = personal_qs.filter(subarea__area__in=areas)

    creadas = 0
    for emp in personal_qs:
        # Autoevaluación
        if ciclo.plantilla and ciclo.plantilla.aplica_autoevaluacion:
            _, created = Evaluacion.objects.get_or_create(
                ciclo=ciclo, evaluado=emp, relacion='AUTO',
                defaults={'estado': 'PENDIENTE'},
            )
            if created:
                creadas += 1

        # Evaluación jefe (placeholder — se asigna evaluador después)
        if ciclo.plantilla and ciclo.plantilla.aplica_jefe:
            _, created = Evaluacion.objects.get_or_create(
                ciclo=ciclo, evaluado=emp, relacion='JEFE',
                defaults={'estado': 'PENDIENTE'},
            )
            if created:
                creadas += 1

    return JsonResponse({'ok': True, 'creadas': creadas})


@login_required
@solo_admin
@require_POST
def ciclo_abrir(request, pk):
    """Abrir ciclo para evaluación."""
    ciclo = get_object_or_404(CicloEvaluacion, pk=pk)
    if ciclo.estado == 'BORRADOR':
        ciclo.estado = 'ABIERTO'
        ciclo.save(update_fields=['estado'])
        return JsonResponse({'ok': True, 'estado': 'ABIERTO'})
    return JsonResponse({'ok': False, 'error': 'Solo se puede abrir desde Borrador.'})


@login_required
@solo_admin
@require_POST
def ciclo_cerrar(request, pk):
    """Cerrar ciclo y consolidar resultados."""
    ciclo = get_object_or_404(CicloEvaluacion, pk=pk)
    if ciclo.estado not in ('ABIERTO', 'EN_EVALUACION', 'CALIBRACION'):
        return JsonResponse({'ok': False, 'error': 'No se puede cerrar en este estado.'})

    ciclo.estado = 'CERRADO'
    ciclo.save(update_fields=['estado'])

    # Consolidar resultados
    evaluados = ciclo.evaluaciones.values_list('evaluado', flat=True).distinct()
    consolidados = 0
    for emp_id in evaluados:
        evals = ciclo.evaluaciones.filter(evaluado_id=emp_id, estado='COMPLETADA')
        if not evals.exists():
            continue

        puntaje_jefe = evals.filter(relacion='JEFE').aggregate(p=Avg('puntaje_total'))['p']
        puntaje_auto = evals.filter(relacion='AUTO').aggregate(p=Avg('puntaje_total'))['p']
        puntaje_pares = evals.filter(relacion='PAR').aggregate(p=Avg('puntaje_total'))['p']
        promedio = evals.aggregate(p=Avg('puntaje_total'))['p']

        ResultadoConsolidado.objects.update_or_create(
            ciclo=ciclo, personal_id=emp_id,
            defaults={
                'puntaje_promedio': promedio,
                'puntaje_jefe': puntaje_jefe,
                'puntaje_auto': puntaje_auto,
                'puntaje_pares': puntaje_pares,
                'consolidado_por': request.user,
                'fecha_consolidacion': timezone.now(),
            },
        )
        consolidados += 1

    return JsonResponse({'ok': True, 'consolidados': consolidados})


# ══════════════════════════════════════════════════════════════
# ADMIN — 9-BOX GRID
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def nine_box(request, ciclo_pk):
    """Vista 9-Box Grid para un ciclo."""
    ciclo = get_object_or_404(CicloEvaluacion, pk=ciclo_pk)
    resultados = ciclo.resultados.select_related('personal').all()

    # Organizar en grid 3x3
    grid = {i: [] for i in range(1, 10)}
    for r in resultados:
        if r.nine_box_position:
            grid[r.nine_box_position].append(r)

    context = {
        'titulo': f'9-Box — {ciclo.nombre}',
        'ciclo': ciclo,
        'grid': grid,
        'resultados': resultados,
    }
    return render(request, 'evaluaciones/nine_box.html', context)


@login_required
@solo_admin
@require_POST
def resultado_clasificar(request, pk):
    """Clasificar resultado en 9-Box (desempeño + potencial)."""
    resultado = get_object_or_404(ResultadoConsolidado, pk=pk)
    resultado.clasificacion_desempeno = request.POST.get('desempeno', '')
    resultado.clasificacion_potencial = request.POST.get('potencial', '')
    resultado.calcular_nine_box()
    return JsonResponse({
        'ok': True,
        'nine_box': resultado.nine_box_position,
        'desempeno': resultado.clasificacion_desempeno,
        'potencial': resultado.clasificacion_potencial,
    })


# ══════════════════════════════════════════════════════════════
# ADMIN — COMPETENCIAS Y PLANTILLAS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def competencias_panel(request):
    """Panel de competencias."""
    competencias = Competencia.objects.all()
    context = {
        'titulo': 'Competencias',
        'competencias': competencias,
    }
    return render(request, 'evaluaciones/competencias.html', context)


@login_required
@solo_admin
@require_POST
def competencia_crear(request):
    """Crear competencia."""
    try:
        c = Competencia.objects.create(
            nombre=request.POST['nombre'],
            codigo=request.POST['codigo'],
            descripcion=request.POST.get('descripcion', ''),
            categoria=request.POST.get('categoria', 'CORE'),
        )
        return JsonResponse({'ok': True, 'pk': c.pk, 'nombre': str(c)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@solo_admin
def plantillas_panel(request):
    """Panel de plantillas de evaluación."""
    plantillas = PlantillaEvaluacion.objects.prefetch_related('items__competencia').filter(activa=True)
    context = {
        'titulo': 'Plantillas de Evaluación',
        'plantillas': plantillas,
    }
    return render(request, 'evaluaciones/plantillas.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — PLANES DE DESARROLLO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def planes_panel(request):
    """Panel de planes de desarrollo individual."""
    planes = PlanDesarrollo.objects.select_related('personal', 'ciclo').all()

    estado = request.GET.get('estado', '')
    buscar = request.GET.get('q', '')

    if estado:
        planes = planes.filter(estado=estado)
    if buscar:
        planes = planes.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(titulo__icontains=buscar)
        )

    context = {
        'titulo': 'Planes de Desarrollo',
        'planes': planes[:100],
        'total': planes.count(),
        'filtro_estado': estado,
        'buscar': buscar,
    }
    return render(request, 'evaluaciones/planes_panel.html', context)


@login_required
@solo_admin
def plan_detalle(request, pk):
    """Detalle de un PDI con acciones."""
    plan = get_object_or_404(
        PlanDesarrollo.objects.select_related('personal', 'ciclo'),
        pk=pk,
    )
    acciones = plan.acciones.all()

    context = {
        'titulo': f'PDI: {plan.personal.apellidos_nombres}',
        'plan': plan,
        'acciones': acciones,
    }
    return render(request, 'evaluaciones/plan_detalle.html', context)


@login_required
@solo_admin
@require_POST
def accion_completar(request, pk):
    """Marcar acción de desarrollo como completada."""
    accion = get_object_or_404(AccionDesarrollo, pk=pk)
    accion.completada = not accion.completada
    accion.fecha_completada = date.today() if accion.completada else None
    accion.save(update_fields=['completada', 'fecha_completada'])
    return JsonResponse({'ok': True, 'completada': accion.completada})


# ══════════════════════════════════════════════════════════════
# EVALUADOR — COMPLETAR EVALUACIÓN
# ══════════════════════════════════════════════════════════════

@login_required
def evaluacion_completar(request, pk):
    """Vista para que el evaluador complete una evaluación."""
    evaluacion = get_object_or_404(
        Evaluacion.objects.select_related('ciclo', 'evaluado', 'ciclo__plantilla'),
        pk=pk,
    )

    # Verificar acceso: evaluador o superuser
    if not request.user.is_superuser:
        if evaluacion.evaluador_usuario != request.user:
            # Autoevaluación: el evaluado puede completar
            from portal.views import _get_empleado
            emp = _get_empleado(request.user)
            if not (evaluacion.relacion == 'AUTO' and emp == evaluacion.evaluado):
                messages.error(request, 'No tiene permiso para completar esta evaluación.')
                return redirect('mis_evaluaciones')

    plantilla = evaluacion.ciclo.plantilla
    items = plantilla.items.select_related('competencia').all() if plantilla else []

    if request.method == 'POST':
        try:
            for item in items:
                puntaje = request.POST.get(f'puntaje_{item.pk}')
                comentario = request.POST.get(f'comentario_{item.pk}', '')
                if puntaje:
                    RespuestaEvaluacion.objects.update_or_create(
                        evaluacion=evaluacion,
                        competencia_plantilla=item,
                        defaults={
                            'puntaje': Decimal(puntaje),
                            'comentario': comentario,
                        },
                    )

            evaluacion.comentario_general = request.POST.get('comentario_general', '')
            evaluacion.fortalezas = request.POST.get('fortalezas', '')
            evaluacion.areas_mejora = request.POST.get('areas_mejora', '')
            evaluacion.estado = 'COMPLETADA'
            evaluacion.fecha_completada = timezone.now()
            evaluacion.save()
            evaluacion.calcular_puntaje()

            messages.success(request, 'Evaluación completada correctamente.')
            if request.user.is_superuser:
                return redirect('ciclo_detalle', pk=evaluacion.ciclo.pk)
            return redirect('mis_evaluaciones')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    context = {
        'titulo': f'Evaluar: {evaluacion.evaluado.apellidos_nombres}',
        'evaluacion': evaluacion,
        'items': items,
        'escala_max': plantilla.escala_max if plantilla else 5,
    }
    return render(request, 'evaluaciones/evaluacion_completar.html', context)


# ══════════════════════════════════════════════════════════════
# PORTAL DEL TRABAJADOR
# ══════════════════════════════════════════════════════════════

@login_required
def mis_evaluaciones(request):
    """Portal: mis evaluaciones pendientes y completadas."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    pendientes = []
    completadas = []
    planes = []

    if empleado:
        # Evaluaciones donde soy evaluado (autoevaluación)
        mis_evals = Evaluacion.objects.filter(
            evaluado=empleado, relacion='AUTO',
        ).select_related('ciclo').order_by('-ciclo__fecha_inicio')

        pendientes = mis_evals.filter(estado__in=['PENDIENTE', 'EN_PROGRESO'])
        completadas = mis_evals.filter(estado__in=['COMPLETADA', 'CALIBRADA'])

        # Evaluaciones que debo realizar a otros
        a_evaluar = Evaluacion.objects.filter(
            evaluador=empleado,
            estado__in=['PENDIENTE', 'EN_PROGRESO'],
        ).select_related('ciclo', 'evaluado')

        # Mis PDI
        planes = PlanDesarrollo.objects.filter(
            personal=empleado,
            estado__in=['BORRADOR', 'ACTIVO'],
        ).order_by('-fecha_inicio')

    context = {
        'titulo': 'Mis Evaluaciones',
        'empleado': empleado,
        'pendientes': pendientes,
        'completadas': completadas,
        'a_evaluar': a_evaluar if empleado else [],
        'planes': planes,
    }
    return render(request, 'evaluaciones/mis_evaluaciones.html', context)


# ══════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def exportar_evaluacion_excel(request, pk):
    """
    Exporta los resultados de un ciclo de evaluación a Excel (.xlsx).
    Hoja "Resultados": evaluado, evaluadores, puntaje por competencia, total, fecha.
    Hoja "9-Box": clasificación de potencial/desempeño de resultados consolidados.
    Estilo Harmoni: headers en #0D2B27 con texto blanco.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'openpyxl no está instalado. Ejecuta: pip install openpyxl',
            status=500,
        )

    ciclo = get_object_or_404(CicloEvaluacion, pk=pk)

    wb = Workbook()

    # ── Estilos Harmoni ────────────────────────────────────────
    COLOR_HEADER = '0D2B27'
    COLOR_ACCENT = '5EEAD4'
    COLOR_SUB    = '134E4A'
    COLOR_LIGHT  = 'F0FDF4'

    fill_header = PatternFill('solid', fgColor=COLOR_HEADER)
    fill_sub    = PatternFill('solid', fgColor=COLOR_SUB)
    fill_alt    = PatternFill('solid', fgColor=COLOR_LIGHT)
    font_header = Font(color='FFFFFF', bold=True, size=10)
    font_sub    = Font(color=COLOR_ACCENT, bold=True, size=9)
    font_title  = Font(color='FFFFFF', bold=True, size=12)
    font_normal = Font(size=9)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    thin = Side(style='thin', color='DDDDDD')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _apply_header_row(ws, row_num, values, fill=None):
        """Aplica estilo de header a una fila completa."""
        f = fill or fill_header
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = f
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_thin

    # ══════════════════════════════════════════════════════════
    # HOJA 1 — RESULTADOS
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Resultados'
    ws.sheet_view.showGridLines = False

    # Fila título
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f'Resultados — {ciclo.nombre}'
    title_cell.fill = fill_header
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 28

    # Fila subtítulo
    ws.merge_cells('A2:J2')
    sub_cell = ws['A2']
    sub_cell.value = (
        f'{ciclo.get_tipo_display()} | '
        f'{ciclo.fecha_inicio.strftime("%d/%m/%Y")} — {ciclo.fecha_fin.strftime("%d/%m/%Y")} | '
        f'Estado: {ciclo.get_estado_display()}'
    )
    sub_cell.fill = fill_sub
    sub_cell.font = font_sub
    sub_cell.alignment = align_center
    ws.row_dimensions[2].height = 18

    # Obtener competencias de la plantilla
    competencias = []
    if ciclo.plantilla:
        competencias = list(
            ciclo.plantilla.items.select_related('competencia').order_by('orden')
        )

    # Cabeceras
    headers = [
        'Evaluado', 'DNI', 'Cargo', 'Área',
        'Tipo Evaluación', 'Evaluador',
    ]
    for item in competencias:
        headers.append(item.competencia.nombre)
    headers += ['Puntaje Total', 'Puntaje Calibrado', 'Estado', 'Fecha Completada']

    _apply_header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 36

    # Datos
    evaluaciones = ciclo.evaluaciones.select_related(
        'evaluado', 'evaluado__subarea', 'evaluado__subarea__area',
        'evaluador', 'ciclo',
    ).prefetch_related('respuestas__competencia_plantilla__competencia').order_by(
        'evaluado__apellidos_nombres', 'relacion',
    )

    for row_idx, ev in enumerate(evaluaciones, start=4):
        row_fill = fill_alt if row_idx % 2 == 0 else None

        area_nombre = '—'
        try:
            area_nombre = ev.evaluado.subarea.area.nombre if ev.evaluado.subarea else '—'
        except Exception:
            pass

        evaluador_nombre = '—'
        if ev.evaluador:
            evaluador_nombre = ev.evaluado.apellidos_nombres if ev.relacion == 'AUTO' else ev.evaluador.apellidos_nombres
        elif ev.relacion == 'AUTO':
            evaluador_nombre = ev.evaluado.apellidos_nombres

        row_data = [
            ev.evaluado.apellidos_nombres,
            getattr(ev.evaluado, 'dni', '—') or '—',
            getattr(ev.evaluado, 'cargo', '—') or '—',
            area_nombre,
            ev.get_relacion_display(),
            evaluador_nombre,
        ]

        # Puntaje por competencia
        respuestas_map = {
            r.competencia_plantilla_id: r.puntaje
            for r in ev.respuestas.all()
        }
        for item in competencias:
            puntaje = respuestas_map.get(item.pk)
            row_data.append(float(puntaje) if puntaje is not None else '—')

        # Totales y estado
        row_data.append(float(ev.puntaje_total) if ev.puntaje_total else '—')
        row_data.append(float(ev.puntaje_calibrado) if ev.puntaje_calibrado else '—')
        row_data.append(ev.get_estado_display())
        row_data.append(
            ev.fecha_completada.strftime('%d/%m/%Y %H:%M') if ev.fecha_completada else '—'
        )

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_fill:
                cell.fill = row_fill
            cell.font = font_normal
            cell.alignment = align_left if col_idx in (1, 3, 4, 6) else align_center
            cell.border = border_thin

    # Anchos de columna
    col_widths = [30, 12, 22, 20, 16, 26]
    col_widths += [14] * len(competencias)
    col_widths += [14, 16, 14, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Inmovilizar filas de encabezado
    ws.freeze_panes = 'A4'

    # ══════════════════════════════════════════════════════════
    # HOJA 2 — 9-BOX
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title='9-Box')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:F1')
    t2 = ws2['A1']
    t2.value = f'9-Box Grid — {ciclo.nombre}'
    t2.fill = fill_header
    t2.font = font_title
    t2.alignment = align_center
    ws2.row_dimensions[1].height = 28

    headers_9box = [
        'Colaborador', 'DNI', 'Cargo',
        'Desempeño', 'Potencial', 'Posición 9-Box',
        'Puntaje Promedio', 'Puntaje Jefe', 'Puntaje Auto',
        'Observaciones',
    ]
    _apply_header_row(ws2, 2, headers_9box)
    ws2.row_dimensions[2].height = 32

    resultados = ciclo.resultados.select_related('personal').order_by(
        '-nine_box_position', '-puntaje_promedio',
    )

    # Colores semafóricos por posición 9-Box
    BOX_COLORS = {
        9: 'FFD700',  # estrella — amarillo
        8: '5EEAD4',  # alto potencial — teal
        6: '5EEAD4',
        7: '86EFAC',  # alto desempeño — verde claro
        5: 'E2E8F0',  # promedio — gris claro
        4: 'E2E8F0',
        2: 'E2E8F0',
        3: 'FCA5A5',  # bajo potencial — rojo claro
        1: 'FCA5A5',  # bajo rendimiento — rojo
    }

    for row_idx, res in enumerate(resultados, start=3):
        pos = res.nine_box_position
        row_fill_9box = PatternFill('solid', fgColor=BOX_COLORS.get(pos, 'FFFFFF')) if pos else None

        row_data = [
            res.personal.apellidos_nombres,
            getattr(res.personal, 'dni', '—') or '—',
            getattr(res.personal, 'cargo', '—') or '—',
            res.get_clasificacion_desempeno_display() if res.clasificacion_desempeno else '—',
            res.get_clasificacion_potencial_display() if res.clasificacion_potencial else '—',
            pos or '—',
            float(res.puntaje_promedio) if res.puntaje_promedio else '—',
            float(res.puntaje_jefe) if res.puntaje_jefe else '—',
            float(res.puntaje_auto) if res.puntaje_auto else '—',
            res.observaciones or '—',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_fill_9box:
                cell.fill = row_fill_9box
            cell.font = font_normal
            cell.alignment = align_left if col_idx in (1, 3, 10) else align_center
            cell.border = border_thin

    col_widths_9box = [30, 12, 22, 14, 14, 14, 16, 14, 14, 30]
    for col_idx, width in enumerate(col_widths_9box, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.freeze_panes = 'A3'

    # ── Respuesta HTTP ─────────────────────────────────────────
    nombre_archivo = f'evaluacion_{ciclo.pk}_{ciclo.nombre[:30].replace(" ", "_")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# ══════════════════════════════════════════════════════════════
# DASHBOARD EJECUTIVO DE EVALUACIONES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def evaluaciones_dashboard(request):
    """
    Dashboard ejecutivo del módulo de evaluaciones.
    Muestra KPIs globales, distribución por tipo, 9-Box, top performers y actividad reciente.
    """
    hoy = date.today()
    hace_30_dias = hoy - timedelta(days=30)

    # ── KPIs globales ──────────────────────────────────────────
    total_evaluaciones = Evaluacion.objects.count()
    completadas = Evaluacion.objects.filter(estado='COMPLETADA').count()
    en_progreso = Evaluacion.objects.filter(estado='EN_PROGRESO').count()
    pendientes = Evaluacion.objects.filter(estado='PENDIENTE').count()
    calibradas = Evaluacion.objects.filter(estado='CALIBRADA').count()
    tasa_completado = round((completadas / total_evaluaciones * 100), 1) if total_evaluaciones > 0 else 0

    total_ciclos = CicloEvaluacion.objects.count()
    ciclos_activos = CicloEvaluacion.objects.filter(estado__in=['ABIERTO', 'EN_EVALUACION']).count()
    ciclos_cerrados = CicloEvaluacion.objects.filter(estado='CERRADO').count()

    # Promedio ponderado global
    promedio_global = None
    completadas_qs = Evaluacion.objects.filter(
        estado='COMPLETADA',
        puntaje_total__isnull=False,
        ciclo__plantilla__isnull=False,
    ).select_related('ciclo__plantilla')
    if completadas_qs.exists():
        total_ratio = Decimal('0')
        count_ratio = 0
        for ev in completadas_qs:
            escala = Decimal(str(ev.ciclo.plantilla.escala_max)) if ev.ciclo.plantilla else Decimal('5')
            if escala > 0:
                total_ratio += ev.puntaje_total / escala * 100
                count_ratio += 1
        if count_ratio:
            promedio_global = round(float(total_ratio / count_ratio), 1)

    stats = {
        'total_evaluaciones': total_evaluaciones,
        'completadas': completadas,
        'en_progreso': en_progreso,
        'pendientes': pendientes,
        'calibradas': calibradas,
        'tasa_completado': tasa_completado,
        'total_ciclos': total_ciclos,
        'ciclos_activos': ciclos_activos,
        'ciclos_cerrados': ciclos_cerrados,
        'promedio_global': promedio_global,
        'total_resultados': ResultadoConsolidado.objects.count(),
        'total_planes': PlanDesarrollo.objects.filter(estado__in=['BORRADOR', 'ACTIVO']).count(),
        'total_competencias': Competencia.objects.filter(activa=True).count(),
    }

    # ── Distribución por tipo de ciclo ────────────────────────
    por_tipo_qs = CicloEvaluacion.objects.values('tipo').annotate(
        total=Count('id'),
    ).order_by('-total')
    tipo_labels = []
    tipo_data = []
    tipo_display = dict(CicloEvaluacion.TIPO_CHOICES)
    for row in por_tipo_qs:
        tipo_labels.append(tipo_display.get(row['tipo'], row['tipo']))
        tipo_data.append(row['total'])
    por_tipo = {
        'labels': json.dumps(tipo_labels),
        'data': json.dumps(tipo_data),
    }

    # ── Distribución de estados de evaluaciones ───────────────
    estados_data = [pendientes, en_progreso, completadas, calibradas]
    estados_labels = ['Pendiente', 'En Progreso', 'Completada', 'Calibrada']
    estados_colors = ['#f59e0b', '#6366f1', '#10b981', '#0f766e']

    # ── 9-Box: distribución global de ResultadoConsolidado ────
    nine_box_counts = {str(i): 0 for i in range(1, 10)}
    nine_box_qs = ResultadoConsolidado.objects.filter(
        nine_box_position__isnull=False,
    ).values('nine_box_position').annotate(total=Count('id'))
    for row in nine_box_qs:
        pos = row['nine_box_position']
        if 1 <= pos <= 9:
            nine_box_counts[str(pos)] = row['total']
    hay_nine_box = any(v > 0 for v in nine_box_counts.values())
    total_clasificados = sum(nine_box_counts.values())
    ninebox_dist = {
        'counts': nine_box_counts,
        'hay_datos': hay_nine_box,
        'total': total_clasificados,
    }

    # ── Top performers (puntaje promedio más alto) ────────────
    top_performers = ResultadoConsolidado.objects.filter(
        puntaje_promedio__isnull=False,
    ).select_related('personal', 'ciclo').order_by('-puntaje_promedio')[:10]

    # ── Colaboradores con mayor oportunidad de mejora ─────────
    mejora = ResultadoConsolidado.objects.filter(
        puntaje_promedio__isnull=False,
        clasificacion_desempeno='BAJO',
    ).select_related('personal', 'ciclo').order_by('puntaje_promedio')[:5]

    # ── Actividad reciente (últimos 30 días) ──────────────────
    recientes = Evaluacion.objects.filter(
        creado_en__date__gte=hace_30_dias,
    ).select_related('evaluado', 'ciclo', 'evaluador').order_by('-creado_en')[:20]

    # Completadas en los últimos 30 días (para gráfico de tendencia por semana)
    semanas_labels = []
    semanas_data = []
    for i in range(4, -1, -1):
        fin_semana = hoy - timedelta(weeks=i)
        inicio_semana = fin_semana - timedelta(days=6)
        count = Evaluacion.objects.filter(
            estado='COMPLETADA',
            fecha_completada__date__range=(inicio_semana, fin_semana),
        ).count()
        semanas_labels.append(inicio_semana.strftime('%d/%m'))
        semanas_data.append(count)

    # ── Ciclos activos con avance ─────────────────────────────
    ciclos_activos_qs = CicloEvaluacion.objects.filter(
        estado__in=['ABIERTO', 'EN_EVALUACION'],
    ).order_by('-fecha_inicio')[:6]

    # ── Promedio de puntaje por ciclo (últimos 4 ciclos cerrados) ─
    promedio_scores_labels = []
    promedio_scores_data = []
    try:
        ultimos_ciclos = CicloEvaluacion.objects.filter(
            estado='CERRADO',
        ).order_by('-fecha_fin')[:4]
        for ciclo in reversed(list(ultimos_ciclos)):
            avg = Evaluacion.objects.filter(
                ciclo=ciclo,
                estado='COMPLETADA',
                puntaje_total__isnull=False,
            ).aggregate(avg=Avg('puntaje_total'))['avg']
            promedio_scores_labels.append(ciclo.nombre[:22])
            promedio_scores_data.append(round(float(avg), 2) if avg else 0)
    except Exception:
        pass

    # ── Distribución por tipo para gráfico ───────────────────
    por_tipo_chart_labels = tipo_labels
    por_tipo_chart_data = tipo_data
    por_tipo_chart_colors = [
        '#0f766e', '#5eead4', '#6366f1', '#f59e0b', '#f43f5e',
    ][:len(tipo_labels)]

    context = {
        'titulo': 'Dashboard de Evaluaciones',
        'stats': stats,
        'por_tipo': por_tipo,
        'estados_data': json.dumps(estados_data),
        'estados_labels': json.dumps(estados_labels),
        'estados_colors': json.dumps(estados_colors),
        'ninebox_dist': ninebox_dist,
        'top_performers': top_performers,
        'mejora': mejora,
        'recientes': recientes,
        'semanas_labels': json.dumps(semanas_labels),
        'semanas_data': json.dumps(semanas_data),
        'ciclos_activos_qs': ciclos_activos_qs,
        'promedio_scores_labels': json.dumps(promedio_scores_labels),
        'promedio_scores_data': json.dumps(promedio_scores_data),
        'por_tipo_chart_labels': json.dumps(por_tipo_chart_labels),
        'por_tipo_chart_data': json.dumps(por_tipo_chart_data),
        'por_tipo_chart_colors': json.dumps(por_tipo_chart_colors),
    }
    return render(request, 'evaluaciones/dashboard.html', context)


# ══════════════════════════════════════════════════════════════
# 9-BOX GRID GLOBAL (todos los ciclos / por ciclo)
# ══════════════════════════════════════════════════════════════

# Nombres canónicos de cada celda del 9-Box
NINEBOX_NOMBRES = {
    1: ('Bajo Rendimiento', 'Bajo desempeño, bajo potencial'),
    2: ('Dilema', 'Bajo desempeño, potencial medio'),
    3: ('Enigma', 'Bajo desempeño, alto potencial'),
    4: ('Trabajador Confiable', 'Desempeño medio, bajo potencial'),
    5: ('Core', 'Desempeño medio, potencial medio'),
    6: ('Estrella Prometedora', 'Desempeño medio, alto potencial'),
    7: ('Alto Impacto', 'Alto desempeño, bajo potencial'),
    8: ('Alto Impacto +', 'Alto desempeño, potencial medio'),
    9: ('Top Talent', 'Alto desempeño, alto potencial'),
}

# Paleta de colores semafórica por posición
NINEBOX_COLORES = {
    1: '#fee2e2',  # rojo claro
    2: '#fef3c7',  # amarillo claro
    3: '#dbeafe',  # azul claro
    4: '#fef3c7',  # amarillo claro
    5: '#e0e7ff',  # indigo claro
    6: '#d1fae5',  # verde claro
    7: '#dbeafe',  # azul claro
    8: '#d1fae5',  # verde claro
    9: '#fef9c3',  # dorado claro
}

NINEBOX_TEXTO = {
    1: '#991b1b', 2: '#92400e', 3: '#1e40af',
    4: '#92400e', 5: '#3730a3', 6: '#065f46',
    7: '#1e40af', 8: '#065f46', 9: '#854d0e',
}


@login_required
@solo_admin
def ninebox_grid(request):
    """
    Vista del 9-Box Grid global con filtros por ciclo y área.
    Muestra el grid 3x3 con empleados clasificados, conteos y leyenda.
    """
    # Filtros
    ciclo_pk = request.GET.get('ciclo', '')
    area_pk = request.GET.get('area', '')

    ciclos_qs = CicloEvaluacion.objects.order_by('-fecha_inicio')
    areas_qs = Area.objects.order_by('nombre') if Area else []

    resultados_qs = ResultadoConsolidado.objects.select_related(
        'personal', 'personal__subarea', 'personal__subarea__area', 'ciclo',
    ).filter(nine_box_position__isnull=False)

    ciclo_sel = None
    if ciclo_pk:
        try:
            ciclo_sel = CicloEvaluacion.objects.get(pk=int(ciclo_pk))
            resultados_qs = resultados_qs.filter(ciclo=ciclo_sel)
        except (CicloEvaluacion.DoesNotExist, ValueError):
            pass

    area_sel = None
    if area_pk and Area:
        try:
            area_sel = Area.objects.get(pk=int(area_pk))
            resultados_qs = resultados_qs.filter(personal__subarea__area=area_sel)
        except Exception:
            pass

    # Construir grid con max 5 empleados visibles por celda
    grid = {}
    for pos in range(1, 10):
        empleados_pos = [r for r in resultados_qs if r.nine_box_position == pos]
        grid[pos] = {
            'nombre': NINEBOX_NOMBRES[pos][0],
            'descripcion': NINEBOX_NOMBRES[pos][1],
            'color': NINEBOX_COLORES[pos],
            'texto': NINEBOX_TEXTO[pos],
            'total': len(empleados_pos),
            'empleados': empleados_pos[:5],
            'hay_mas': len(empleados_pos) > 5,
        }

    total_clasificados = sum(grid[p]['total'] for p in range(1, 10))
    sin_clasificar = ResultadoConsolidado.objects.filter(nine_box_position__isnull=True)
    if ciclo_sel:
        sin_clasificar = sin_clasificar.filter(ciclo=ciclo_sel)

    # Datos para gráfico de barras (distribución)
    bar_labels = [NINEBOX_NOMBRES[p][0] for p in range(1, 10)]
    bar_data = [grid[p]['total'] for p in range(1, 10)]
    bar_colors = [NINEBOX_COLORES[p] for p in range(1, 10)]
    bar_borders = [NINEBOX_TEXTO[p] for p in range(1, 10)]

    context = {
        'titulo': 'Grid 9-Box — Talento y Desempeño',
        'grid': grid,
        'ciclos': ciclos_qs,
        'ciclo_sel': ciclo_sel,
        'areas': areas_qs,
        'area_sel': area_sel,
        'total_clasificados': total_clasificados,
        'sin_clasificar': sin_clasificar[:20],
        'sin_clasificar_total': sin_clasificar.count(),
        'bar_labels': json.dumps(bar_labels),
        'bar_data': json.dumps(bar_data),
        'bar_colors': json.dumps(bar_colors),
        'bar_borders': json.dumps(bar_borders),
    }
    return render(request, 'evaluaciones/ninebox_grid.html', context)


# ══════════════════════════════════════════════════════════════
# COMPARATIVA DE COMPETENCIAS (Radar Chart)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def comparativa_competencias(request):
    """
    Comparativa de puntajes por competencia.
    Radar chart con Chart.js: comparación entre áreas o entre ciclos.
    Filtros: ciclo, área (hasta 5 áreas comparables).
    """
    ciclos_qs = CicloEvaluacion.objects.filter(estado='CERRADO').order_by('-fecha_inicio')
    areas_qs = Area.objects.order_by('nombre') if Area else []

    ciclo_pk = request.GET.get('ciclo', '')
    modo = request.GET.get('modo', 'areas')  # 'areas' o 'ciclos'

    ciclo_sel = None
    if ciclo_pk:
        try:
            ciclo_sel = CicloEvaluacion.objects.get(pk=int(ciclo_pk))
        except (CicloEvaluacion.DoesNotExist, ValueError):
            pass

    # Sin ciclo seleccionado: usar el último cerrado
    if not ciclo_sel and ciclos_qs.exists():
        ciclo_sel = ciclos_qs.first()

    radar_data = None
    competencias_nombres = []
    datasets = []

    if ciclo_sel and ciclo_sel.plantilla:
        plantilla = ciclo_sel.plantilla
        items_plantilla = list(
            plantilla.items.select_related('competencia').order_by('orden')
        )
        competencias_nombres = [item.competencia.nombre for item in items_plantilla]

        if modo == 'areas' and Area:
            # Comparar por áreas: promedio de puntaje por competencia por área
            areas_con_datos = []
            COLORES_RADAR = [
                ('rgba(15,118,110,0.25)', 'rgba(15,118,110,1)'),    # teal
                ('rgba(99,102,241,0.25)', 'rgba(99,102,241,1)'),    # indigo
                ('rgba(245,158,11,0.25)', 'rgba(245,158,11,1)'),    # amber
                ('rgba(239,68,68,0.25)',  'rgba(239,68,68,1)'),     # red
                ('rgba(16,185,129,0.25)', 'rgba(16,185,129,1)'),    # green
            ]

            # Áreas que tienen respuestas en este ciclo
            areas_en_ciclo = Area.objects.filter(
                subareas__personal__evaluaciones_recibidas__ciclo=ciclo_sel,
            ).distinct().order_by('nombre')[:5]

            for idx, area in enumerate(areas_en_ciclo):
                puntos = []
                for item in items_plantilla:
                    avg = RespuestaEvaluacion.objects.filter(
                        evaluacion__ciclo=ciclo_sel,
                        evaluacion__evaluado__subarea__area=area,
                        competencia_plantilla=item,
                        evaluacion__estado='COMPLETADA',
                    ).aggregate(p=Avg('puntaje'))['p']
                    puntos.append(round(float(avg), 2) if avg else 0)

                color_bg, color_border = COLORES_RADAR[idx % len(COLORES_RADAR)]
                datasets.append({
                    'label': area.nombre,
                    'data': puntos,
                    'backgroundColor': color_bg,
                    'borderColor': color_border,
                    'borderWidth': 2,
                    'pointBackgroundColor': color_border,
                    'pointRadius': 4,
                })
                areas_con_datos.append(area)

        elif modo == 'ciclos':
            # Comparar ciclos cerrados (últimos 3) para las mismas competencias
            COLORES_RADAR = [
                ('rgba(15,118,110,0.25)', 'rgba(15,118,110,1)'),
                ('rgba(99,102,241,0.25)', 'rgba(99,102,241,1)'),
                ('rgba(245,158,11,0.25)', 'rgba(245,158,11,1)'),
            ]
            ciclos_comparar = CicloEvaluacion.objects.filter(
                estado='CERRADO',
                plantilla=plantilla,
            ).order_by('-fecha_inicio')[:3]

            for idx, ciclo_c in enumerate(ciclos_comparar):
                puntos = []
                for item in items_plantilla:
                    avg = RespuestaEvaluacion.objects.filter(
                        evaluacion__ciclo=ciclo_c,
                        competencia_plantilla=item,
                        evaluacion__estado='COMPLETADA',
                    ).aggregate(p=Avg('puntaje'))['p']
                    puntos.append(round(float(avg), 2) if avg else 0)

                color_bg, color_border = COLORES_RADAR[idx % len(COLORES_RADAR)]
                datasets.append({
                    'label': ciclo_c.nombre,
                    'data': puntos,
                    'backgroundColor': color_bg,
                    'borderColor': color_border,
                    'borderWidth': 2,
                    'pointBackgroundColor': color_border,
                    'pointRadius': 4,
                })

        escala_max = float(plantilla.escala_max)
        radar_data = {
            'labels': competencias_nombres,
            'datasets': datasets,
            'escala_max': escala_max,
        }

    # Ranking de competencias: promedio global por competencia (ciclo seleccionado)
    ranking_competencias = []
    if ciclo_sel and ciclo_sel.plantilla:
        for item in items_plantilla:
            avg = RespuestaEvaluacion.objects.filter(
                evaluacion__ciclo=ciclo_sel,
                competencia_plantilla=item,
                evaluacion__estado='COMPLETADA',
            ).aggregate(p=Avg('puntaje'))['p']
            if avg is not None:
                pct = round(float(avg) / float(ciclo_sel.plantilla.escala_max) * 100, 1)
                ranking_competencias.append({
                    'nombre': item.competencia.nombre,
                    'categoria': item.competencia.get_categoria_display(),
                    'promedio': round(float(avg), 2),
                    'porcentaje': pct,
                    'escala_max': ciclo_sel.plantilla.escala_max,
                })
        ranking_competencias.sort(key=lambda x: x['promedio'], reverse=True)

    context = {
        'titulo': 'Comparativa de Competencias',
        'ciclos': ciclos_qs,
        'ciclo_sel': ciclo_sel,
        'areas': areas_qs,
        'modo': modo,
        'radar_labels': json.dumps(competencias_nombres),
        'radar_datasets': json.dumps(datasets),
        'radar_escala_max': float(ciclo_sel.plantilla.escala_max) if ciclo_sel and ciclo_sel.plantilla else 5.0,
        'hay_datos': bool(datasets),
        'ranking_competencias': ranking_competencias,
    }
    return render(request, 'evaluaciones/comparativa_competencias.html', context)
