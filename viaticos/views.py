"""
Vistas del módulo de Viáticos y CDT.
"""
import json
from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect

from personal.models import Personal
from .models import ConceptoViatico, AsignacionViatico, GastoViatico

solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


@login_required
@solo_admin
def viaticos_panel(request):
    """Panel principal de viáticos."""
    qs = AsignacionViatico.objects.select_related('personal').all()

    # Filtros
    estado = request.GET.get('estado', '')
    periodo = request.GET.get('periodo', '')
    buscar = request.GET.get('q', '')

    if estado:
        qs = qs.filter(estado=estado)
    if periodo:
        try:
            y, m = periodo.split('-')
            qs = qs.filter(periodo__year=int(y), periodo__month=int(m))
        except (ValueError, IndexError):
            pass
    if buscar:
        qs = qs.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar) |
            Q(ubicacion__icontains=buscar)
        )

    # Stats
    hoy = date.today()
    mes_actual = qs.filter(periodo__year=hoy.year, periodo__month=hoy.month)
    total_asignado_mes = mes_actual.aggregate(
        t=Sum('monto_asignado')
    )['t'] or Decimal('0.00')
    total_rendido_mes = mes_actual.aggregate(
        t=Sum('monto_rendido')
    )['t'] or Decimal('0.00')

    # ── Analytics: gastos por concepto (año actual) ───────────────────────────
    gastos_por_concepto_json = '[]'
    try:
        gastos_concepto_qs = (
            GastoViatico.objects
            .filter(fecha_gasto__year=hoy.year, estado='APROBADO')
            .values('concepto__nombre')
            .annotate(total=Sum('monto'))
            .order_by('-total')[:6]
        )
        gastos_por_concepto_json = json.dumps([
            {'label': item['concepto__nombre'], 'value': float(item['total'] or 0)}
            for item in gastos_concepto_qs
        ])
    except Exception:
        pass

    # ── Analytics: top 5 beneficiarios (año actual) ───────────────────────────
    top_beneficiarios = []
    try:
        top_qs = (
            AsignacionViatico.objects
            .filter(periodo__year=hoy.year)
            .exclude(estado='CANCELADO')
            .values('personal__id', 'personal__apellidos_nombres')
            .annotate(
                total_asignado=Coalesce(Sum('monto_asignado'), Value(Decimal('0.00')), output_field=DecimalField()),
                total_adicional=Coalesce(Sum('monto_adicional'), Value(Decimal('0.00')), output_field=DecimalField()),
            )
            .order_by('-total_asignado')[:5]
        )
        for row in top_qs:
            top_beneficiarios.append({
                'nombre': row['personal__apellidos_nombres'],
                'total_monto': row['total_asignado'] + row['total_adicional'],
            })
    except Exception:
        pass

    # ── Analytics: tendencia mensual últimos 6 meses ──────────────────────────
    tendencia_mensual_json = '[]'
    try:
        tendencia_data = []
        for i in range(5, -1, -1):
            m = hoy.month - i
            y = hoy.year
            while m <= 0:
                m += 12
                y -= 1
            label = date(y, m, 1).strftime('%m/%y')
            total = (
                AsignacionViatico.objects
                .filter(periodo__year=y, periodo__month=m)
                .exclude(estado='CANCELADO')
                .aggregate(t=Sum('monto_asignado'))['t']
            ) or Decimal('0.00')
            tendencia_data.append({'label': label, 'value': float(total)})
        tendencia_mensual_json = json.dumps(tendencia_data)
    except Exception:
        pass

    # ── KPI strip ────────────────────────────────────────────────────────────
    kpi_viaticos_activos = 0
    try:
        kpi_viaticos_activos = AsignacionViatico.objects.exclude(
            estado__in=['CONCILIADO', 'CANCELADO']
        ).count()
    except Exception:
        kpi_viaticos_activos = 0

    kpi_monto_pendiente_rendicion = Decimal('0.00')
    try:
        kpi_monto_pendiente_rendicion = (
            AsignacionViatico.objects
            .filter(estado__in=['ENTREGADO', 'EN_RENDICION'])
            .aggregate(t=Sum('monto_asignado'))['t']
        ) or Decimal('0.00')
    except Exception:
        kpi_monto_pendiente_rendicion = Decimal('0.00')

    kpi_rendidos_mes = 0
    try:
        kpi_rendidos_mes = AsignacionViatico.objects.filter(
            estado='CONCILIADO',
            fecha_conciliacion__year=hoy.year,
            fecha_conciliacion__month=hoy.month,
        ).count()
    except Exception:
        kpi_rendidos_mes = 0

    kpi_promedio_dias_rendicion = 0
    try:
        conciliados = AsignacionViatico.objects.filter(
            estado='CONCILIADO',
            fecha_entrega__isnull=False,
            fecha_conciliacion__isnull=False,
        ).values_list('fecha_entrega', 'fecha_conciliacion')
        if conciliados:
            dias_lista = [
                (fc - fe).days
                for fe, fc in conciliados
                if fc >= fe
            ]
            kpi_promedio_dias_rendicion = round(
                sum(dias_lista) / len(dias_lista)
            ) if dias_lista else 0
    except Exception:
        kpi_promedio_dias_rendicion = 0

    context = {
        'titulo': 'Viáticos y CDT',
        'asignaciones': qs[:100],
        'total': qs.count(),
        'filtro_estado': estado,
        'filtro_periodo': periodo,
        'buscar': buscar,
        'stats': {
            'total_mes': mes_actual.count(),
            'total_asignado_mes': total_asignado_mes,
            'total_rendido_mes': total_rendido_mes,
            'pendientes_rendicion': qs.filter(estado='ENTREGADO').count(),
            'por_conciliar': qs.filter(estado='EN_RENDICION').count(),
        },
        'gastos_por_concepto_json': gastos_por_concepto_json,
        'top_beneficiarios': top_beneficiarios,
        'tendencia_mensual_json': tendencia_mensual_json,
        # KPI strip
        'kpi_viaticos_activos': kpi_viaticos_activos,
        'kpi_monto_pendiente_rendicion': kpi_monto_pendiente_rendicion,
        'kpi_rendidos_mes': kpi_rendidos_mes,
        'kpi_promedio_dias_rendicion': kpi_promedio_dias_rendicion,
    }
    return render(request, 'viaticos/panel.html', context)


@login_required
@solo_admin
def viatico_crear(request):
    """Crear nueva asignación de viáticos."""
    if request.method == 'POST':
        try:
            personal = get_object_or_404(Personal, pk=request.POST['personal_id'])
            periodo_str = request.POST['periodo']
            y, m = periodo_str.split('-')
            periodo = date(int(y), int(m), 1)
            monto = Decimal(request.POST['monto'])
            ubicacion = request.POST.get('ubicacion', '')
            dias_campo = int(request.POST.get('dias_campo', 0))
            observaciones = request.POST.get('observaciones', '')

            # Verificar que no exista duplicado
            if AsignacionViatico.objects.filter(personal=personal, periodo=periodo).exists():
                messages.error(request, f'{personal.apellidos_nombres} ya tiene viáticos para {periodo.strftime("%m/%Y")}')
                return redirect('viatico_crear')

            asignacion = AsignacionViatico.objects.create(
                personal=personal,
                periodo=periodo,
                monto_asignado=monto,
                ubicacion=ubicacion,
                dias_campo=dias_campo,
                observaciones=observaciones,
                estado='APROBADO',  # Auto-aprobado
                creado_por=request.user,
                aprobado_por=request.user,
            )

            from core.audit import log_create
            log_create(request, asignacion,
                       f'Viático {periodo.strftime("%m/%Y")} — S/ {monto} — {personal.apellidos_nombres}')

            messages.success(request, f'Viático registrado: S/ {monto} — {personal.apellidos_nombres}')
            return redirect('viatico_detalle', pk=asignacion.pk)
        except (ValueError, KeyError) as e:
            messages.error(request, f'Error en los datos: {e}')

    context = {
        'titulo': 'Nueva Asignación de Viáticos',
        'personal_list': Personal.objects.filter(estado='Activo').order_by('apellidos_nombres'),
    }
    return render(request, 'viaticos/crear.html', context)


@login_required
@solo_admin
def viatico_detalle(request, pk):
    """Detalle de una asignación con sus gastos."""
    asignacion = get_object_or_404(
        AsignacionViatico.objects.select_related('personal', 'creado_por', 'aprobado_por'),
        pk=pk
    )
    gastos = asignacion.gastos.select_related('concepto').all()
    conceptos = ConceptoViatico.objects.filter(activo=True)

    # Resumen por concepto
    resumen_conceptos = {}
    for g in gastos.filter(estado='APROBADO'):
        key = g.concepto.nombre
        if key not in resumen_conceptos:
            resumen_conceptos[key] = Decimal('0.00')
        resumen_conceptos[key] += g.monto

    context = {
        'titulo': f'Viático #{asignacion.pk}',
        'a': asignacion,
        'gastos': gastos,
        'conceptos': conceptos,
        'resumen_conceptos': resumen_conceptos,
    }
    return render(request, 'viaticos/detalle.html', context)


@login_required
@solo_admin
def viatico_entregar(request, pk):
    """Marcar asignación como entregada."""
    asignacion = get_object_or_404(AsignacionViatico, pk=pk)
    if request.method == 'POST' and asignacion.estado in ('BORRADOR', 'APROBADO'):
        asignacion.entregar()

        from core.audit import log_update
        log_update(request, asignacion, {'estado': {'old': 'APROBADO', 'new': 'ENTREGADO'}},
                   f'Viático entregado a {asignacion.personal.apellidos_nombres}')

        messages.success(request, 'Viático marcado como entregado.')
    return redirect('viatico_detalle', pk=pk)


@login_required
@solo_admin
def viatico_conciliar(request, pk):
    """Conciliar asignación con los gastos rendidos."""
    asignacion = get_object_or_404(AsignacionViatico, pk=pk)
    if request.method == 'POST' and asignacion.estado in ('ENTREGADO', 'EN_RENDICION'):
        asignacion.conciliar()

        from core.audit import log_update
        log_update(request, asignacion,
                   {'estado': {'old': 'EN_RENDICION', 'new': 'CONCILIADO'},
                    'monto_rendido': {'old': '0', 'new': str(asignacion.monto_rendido)}},
                   f'Viático conciliado — Rendido: S/ {asignacion.monto_rendido}')

        messages.success(request, f'Viático conciliado — Rendido: S/ {asignacion.monto_rendido}')
    return redirect('viatico_detalle', pk=pk)


@login_required
@solo_admin
def gasto_agregar(request, pk):
    """Agregar gasto a una asignación (AJAX)."""
    asignacion = get_object_or_404(AsignacionViatico, pk=pk)

    if request.method == 'POST':
        try:
            concepto = get_object_or_404(ConceptoViatico, pk=request.POST['concepto_id'])
            gasto = GastoViatico.objects.create(
                asignacion=asignacion,
                concepto=concepto,
                fecha_gasto=request.POST['fecha_gasto'],
                monto=Decimal(request.POST['monto']),
                descripcion=request.POST.get('descripcion', ''),
                tipo_comprobante=request.POST.get('tipo_comprobante', 'BOLETA'),
                numero_comprobante=request.POST.get('numero_comprobante', ''),
                ruc_proveedor=request.POST.get('ruc_proveedor', ''),
                estado='APROBADO',
            )

            # Recalcular rendido
            total = asignacion.gastos.filter(
                estado='APROBADO'
            ).aggregate(t=Sum('monto'))['t'] or Decimal('0.00')
            asignacion.monto_rendido = total
            if asignacion.estado == 'ENTREGADO':
                asignacion.estado = 'EN_RENDICION'
            asignacion.save(update_fields=['monto_rendido', 'estado'])

            return JsonResponse({
                'ok': True,
                'gasto_id': gasto.pk,
                'concepto': concepto.nombre,
                'monto': str(gasto.monto),
                'fecha': gasto.fecha_gasto.strftime('%d/%m/%Y'),
                'tipo': gasto.get_tipo_comprobante_display(),
                'total_rendido': str(total),
            })
        except (ValueError, KeyError) as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)


# ── Revisar gasto individual (AJAX) ──────────────────────────────────────────

@login_required
@solo_admin
def gasto_revisar(request, gasto_id):
    """
    Cambia el estado de un GastoViatico individual.
    Acción: APROBADO | RECHAZADO | OBSERVADO | PENDIENTE
    Recalcula monto_rendido de la asignación padre.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    gasto = get_object_or_404(GastoViatico.objects.select_related('asignacion'), pk=gasto_id)
    accion = request.POST.get('accion', '').upper()
    motivo = request.POST.get('motivo', '').strip()

    ESTADOS_VALIDOS = {'APROBADO', 'RECHAZADO', 'OBSERVADO', 'PENDIENTE'}
    if accion not in ESTADOS_VALIDOS:
        return JsonResponse({'ok': False, 'error': f'Acción inválida: {accion}'}, status=400)

    if accion in ('RECHAZADO', 'OBSERVADO') and not motivo:
        return JsonResponse({'ok': False, 'error': 'Se requiere motivo para rechazar u observar.'}, status=400)

    gasto.estado = accion
    gasto.motivo_rechazo = motivo if accion in ('RECHAZADO', 'OBSERVADO') else ''
    gasto.save(update_fields=['estado', 'motivo_rechazo'])

    # Recalcular monto_rendido (solo gastos APROBADOS)
    asignacion = gasto.asignacion
    total = asignacion.gastos.filter(estado='APROBADO').aggregate(
        t=Sum('monto')
    )['t'] or Decimal('0.00')
    asignacion.monto_rendido = total
    asignacion.save(update_fields=['monto_rendido'])

    return JsonResponse({
        'ok': True,
        'gasto_id': gasto.pk,
        'nuevo_estado': accion,
        'nuevo_estado_display': gasto.get_estado_display(),
        'total_rendido': str(total),
    })


@login_required
@solo_admin
def gasto_eliminar(request, gasto_id):
    """Elimina un gasto en estado PENDIENTE (no aprobados)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    gasto = get_object_or_404(GastoViatico.objects.select_related('asignacion'), pk=gasto_id)

    if gasto.estado == 'APROBADO':
        return JsonResponse({'ok': False, 'error': 'No se puede eliminar un gasto aprobado.'}, status=400)

    asignacion = gasto.asignacion
    gasto.delete()

    # Recalcular
    total = asignacion.gastos.filter(estado='APROBADO').aggregate(
        t=Sum('monto')
    )['t'] or Decimal('0.00')
    asignacion.monto_rendido = total
    asignacion.save(update_fields=['monto_rendido'])

    return JsonResponse({'ok': True, 'total_rendido': str(total)})


@login_required
@solo_admin
def viatico_anular(request, pk):
    """Cancela (anula) una asignación de viáticos."""
    asignacion = get_object_or_404(AsignacionViatico, pk=pk)
    if request.method == 'POST':
        if asignacion.estado == 'CONCILIADO':
            messages.error(request, 'No se puede anular un viático ya conciliado.')
        else:
            old = asignacion.estado
            asignacion.estado = 'CANCELADO'
            asignacion.save(update_fields=['estado'])

            from core.audit import log_update
            log_update(request, asignacion, {'estado': {'old': old, 'new': 'CANCELADO'}},
                       f'Viático anulado')
            messages.warning(request, 'Asignación de viáticos anulada.')

    return redirect('viatico_detalle', pk=pk)


@login_required
@solo_admin
def viaticos_exportar(request):
    """Exporta viáticos filtrados a CSV."""
    import csv
    from django.http import HttpResponse

    qs = AsignacionViatico.objects.select_related('personal', 'creado_por').all()

    # Aplicar mismos filtros del panel
    estado  = request.GET.get('estado', '')
    periodo = request.GET.get('periodo', '')
    buscar  = request.GET.get('q', '')

    if estado:
        qs = qs.filter(estado=estado)
    if periodo:
        try:
            y, m = periodo.split('-')
            qs = qs.filter(periodo__year=int(y), periodo__month=int(m))
        except (ValueError, IndexError):
            pass
    if buscar:
        qs = qs.filter(
            Q(personal__apellidos_nombres__icontains=buscar) |
            Q(personal__nro_doc__icontains=buscar) |
            Q(ubicacion__icontains=buscar)
        )

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="viaticos.csv"'
    writer = csv.writer(response)

    writer.writerow([
        'DNI', 'Trabajador', 'Período', 'Ubicación', 'Días Campo',
        'Monto Asignado', 'Monto Adicional', 'Total Asignado',
        'Monto Rendido', 'Saldo', 'Estado', 'Fecha Entrega',
        'Fecha Conciliación', 'Creado por',
    ])

    for a in qs:
        writer.writerow([
            a.personal.nro_doc,
            a.personal.apellidos_nombres,
            a.periodo.strftime('%m/%Y'),
            a.ubicacion or '',
            a.dias_campo,
            str(a.monto_asignado),
            str(a.monto_adicional),
            str(a.monto_total),
            str(a.monto_rendido),
            str(a.saldo),
            a.get_estado_display(),
            a.fecha_entrega.strftime('%d/%m/%Y') if a.fecha_entrega else '',
            a.fecha_conciliacion.strftime('%d/%m/%Y') if a.fecha_conciliacion else '',
            str(a.creado_por) if a.creado_por else '',
        ])

    return response


@login_required
@solo_admin
def exportar_viaticos_excel(request):
    """
    Exporta CDT del mes actual (o del período indicado) a Excel.
    Columnas: Empleado, Proyecto/Ubicación, Total Viáticos, Total Conciliado, Diferencia.
    """
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    hoy = date.today()

    # Período: parámetro GET o mes actual
    periodo_param = request.GET.get('periodo', '')
    if periodo_param:
        try:
            y, m = periodo_param.split('-')
            anio, mes = int(y), int(m)
        except (ValueError, IndexError):
            anio, mes = hoy.year, hoy.month
    else:
        anio, mes = hoy.year, hoy.month

    qs = (
        AsignacionViatico.objects
        .select_related('personal')
        .filter(periodo__year=anio, periodo__month=mes)
        .exclude(estado='CANCELADO')
        .order_by('personal__apellidos_nombres')
    )

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Viáticos {mes:02d}/{anio}'

    # Colores marca Harmoni
    COLOR_HEADER_BG = '0D2B27'   # teal oscuro sidebar
    COLOR_HEADER_FG = 'FFFFFF'
    COLOR_SUBTOTAL  = 'F0FDF4'
    COLOR_ALERT     = 'FEF2F2'
    COLOR_OK        = 'F0FDF4'

    thin = Side(border_style='thin', color='D1D5DB')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    titulo_cell = ws['A1']
    titulo_cell.value = f'Reporte de Viáticos CDT — {mes:02d}/{anio}'
    titulo_cell.font = Font(name='Calibri', bold=True, size=14, color=COLOR_HEADER_FG)
    titulo_cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    sub_cell = ws['A2']
    sub_cell.value = f'Generado: {hoy.strftime("%d/%m/%Y")}    |    Total registros: {qs.count()}'
    sub_cell.font = Font(name='Calibri', italic=True, size=9, color='6B7280')
    sub_cell.alignment = Alignment(horizontal='center')

    # ── Cabecera ──────────────────────────────────────────────────────────────
    headers = [
        ('N°', 5),
        ('Empleado', 35),
        ('Proyecto / Ubicación', 28),
        ('Total Viáticos (S/)', 20),
        ('Total Conciliado (S/)', 20),
        ('Diferencia (S/)', 18),
    ]

    fila_header = 4
    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=fila_header, column=col_idx, value=header_text)
        cell.font = Font(name='Calibri', bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill('solid', fgColor='134E4A')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borde
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[fila_header].height = 22

    # ── Datos ─────────────────────────────────────────────────────────────────
    FMT_NUM = '#,##0.00'
    fila_data = fila_header + 1
    total_asignado = Decimal('0.00')
    total_rendido = Decimal('0.00')
    total_diferencia = Decimal('0.00')

    for idx, a in enumerate(qs, start=1):
        diferencia = a.saldo  # saldo = rendido - total asignado
        es_diferencia_significativa = abs(diferencia) > Decimal('50.00')

        row_fill = None
        if es_diferencia_significativa:
            row_fill = PatternFill('solid', fgColor='FEF9C3' if diferencia > 0 else 'FEE2E2')

        # N°
        c = ws.cell(row=fila_data, column=1, value=idx)
        c.alignment = Alignment(horizontal='center')
        c.border = borde
        if row_fill:
            c.fill = row_fill

        # Empleado (nombre + DNI en la misma celda, dos líneas)
        c = ws.cell(row=fila_data, column=2,
                    value=f"{a.personal.apellidos_nombres}\n{a.personal.nro_doc}")
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = borde
        if row_fill:
            c.fill = row_fill

        # Ubicación/Proyecto
        c = ws.cell(row=fila_data, column=3, value=a.ubicacion or '—')
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = borde
        if row_fill:
            c.fill = row_fill

        # Total viáticos asignado
        c = ws.cell(row=fila_data, column=4, value=float(a.monto_total))
        c.number_format = FMT_NUM
        c.alignment = Alignment(horizontal='right')
        c.border = borde
        if row_fill:
            c.fill = row_fill

        # Total conciliado/rendido
        c = ws.cell(row=fila_data, column=5, value=float(a.monto_rendido))
        c.number_format = FMT_NUM
        c.alignment = Alignment(horizontal='right')
        c.border = borde
        if row_fill:
            c.fill = row_fill

        # Diferencia — rojo si significativa
        c = ws.cell(row=fila_data, column=6, value=float(diferencia))
        c.number_format = FMT_NUM
        c.alignment = Alignment(horizontal='right')
        c.border = borde
        if diferencia == 0:
            c.font = Font(color='16A34A', bold=True)
        elif es_diferencia_significativa:
            c.font = Font(color='DC2626', bold=True)
            c.fill = PatternFill('solid', fgColor='FEE2E2')
        else:
            c.font = Font(color='D97706')

        total_asignado += a.monto_total
        total_rendido += a.monto_rendido
        total_diferencia += diferencia
        fila_data += 1

    # ── Totales ───────────────────────────────────────────────────────────────
    fila_total = fila_data
    ws.cell(row=fila_total, column=1, value='').border = borde
    c_lbl = ws.cell(row=fila_total, column=2, value='TOTAL')
    c_lbl.font = Font(bold=True)
    c_lbl.fill = PatternFill('solid', fgColor='E2E8F0')
    c_lbl.border = borde

    ws.cell(row=fila_total, column=3, value='').fill = PatternFill('solid', fgColor='E2E8F0')
    ws.cell(row=fila_total, column=3).border = borde

    c_ta = ws.cell(row=fila_total, column=4, value=float(total_asignado))
    c_ta.number_format = FMT_NUM
    c_ta.font = Font(bold=True)
    c_ta.fill = PatternFill('solid', fgColor='E2E8F0')
    c_ta.alignment = Alignment(horizontal='right')
    c_ta.border = borde

    c_tr = ws.cell(row=fila_total, column=5, value=float(total_rendido))
    c_tr.number_format = FMT_NUM
    c_tr.font = Font(bold=True)
    c_tr.fill = PatternFill('solid', fgColor='E2E8F0')
    c_tr.alignment = Alignment(horizontal='right')
    c_tr.border = borde

    c_td = ws.cell(row=fila_total, column=6, value=float(total_diferencia))
    c_td.number_format = FMT_NUM
    c_td.font = Font(bold=True,
                     color='16A34A' if total_diferencia == 0 else ('DC2626' if abs(total_diferencia) > 50 else 'D97706'))
    c_td.fill = PatternFill('solid', fgColor='E2E8F0')
    c_td.alignment = Alignment(horizontal='right')
    c_td.border = borde

    # Freeze header
    ws.freeze_panes = ws.cell(row=fila_header + 1, column=1)

    # ── Respuesta HTTP ────────────────────────────────────────────────────────
    nombre_archivo = f'viaticos_CDT_{anio}_{mes:02d}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# ── Dashboard Viáticos ───────────────────────────────────────────────────────

@login_required
@solo_admin
def viaticos_dashboard(request):
    """
    Dashboard ejecutivo de viáticos CDT.
    KPIs, CDTs en riesgo, tabla activos y gráfico Chart.js por área/mes.
    """
    hoy = date.today()

    # ── KPIs globales ────────────────────────────────────────────────────────
    qs_activos = AsignacionViatico.objects.exclude(estado__in=['CONCILIADO', 'CANCELADO'])

    total_cdt_activos = qs_activos.count()
    monto_total_pendiente = qs_activos.aggregate(
        t=Sum('monto_asignado')
    )['t'] or Decimal('0.00')

    # Mes actual
    qs_mes = AsignacionViatico.objects.filter(
        periodo__year=hoy.year, periodo__month=hoy.month
    ).exclude(estado='CANCELADO')
    monto_total_mes = qs_mes.aggregate(t=Sum('monto_asignado'))['t'] or Decimal('0.00')

    # Por conciliar = en rendición con gastos registrados
    por_conciliar = AsignacionViatico.objects.filter(estado='EN_RENDICION').count()

    # ── CDTs activos con días transcurridos ──────────────────────────────────
    cdts_activos = (
        qs_activos
        .select_related('personal', 'personal__subarea', 'personal__subarea__area')
        .order_by('-periodo', 'personal__apellidos_nombres')
    )

    # Anotar días transcurridos y marcar en riesgo (>30 días sin conciliar)
    cdts_con_dias = []
    en_riesgo = []
    for cdt in cdts_activos:
        ref = cdt.fecha_entrega or cdt.periodo
        dias = (hoy - ref).days if ref else 0
        riesgo = dias > 30 and cdt.estado in ('ENTREGADO', 'EN_RENDICION')
        cdts_con_dias.append({
            'obj': cdt,
            'dias': dias,
            'en_riesgo': riesgo,
        })
        if riesgo:
            en_riesgo.append(cdt)

    # ── Gráfico: montos por área (últimos 6 meses) ───────────────────────────
    meses_labels = []
    meses_data = []
    for i in range(5, -1, -1):
        # Calcular mes i meses atrás
        m = hoy.month - i
        y = hoy.year
        while m <= 0:
            m += 12
            y -= 1
        label = date(y, m, 1).strftime('%b %Y')
        total = AsignacionViatico.objects.filter(
            periodo__year=y, periodo__month=m
        ).exclude(estado='CANCELADO').aggregate(t=Sum('monto_asignado'))['t'] or 0
        meses_labels.append(label)
        meses_data.append(float(total))

    # Montos por área (activos, agrupados por subarea.area.nombre)
    areas_data = {}
    for cdt in qs_activos.select_related('personal__subarea__area'):
        area_nombre = '—'
        if cdt.personal.subarea and cdt.personal.subarea.area:
            area_nombre = cdt.personal.subarea.area.nombre
        areas_data[area_nombre] = areas_data.get(area_nombre, Decimal('0.00')) + cdt.monto_asignado

    areas_labels = list(areas_data.keys())
    areas_montos = [float(v) for v in areas_data.values()]

    context = {
        'titulo': 'Dashboard Viáticos',
        'hoy': hoy,
        # KPIs
        'total_cdt_activos': total_cdt_activos,
        'monto_total_pendiente': monto_total_pendiente,
        'monto_total_mes': monto_total_mes,
        'por_conciliar': por_conciliar,
        # Tablas
        'cdts_con_dias': cdts_con_dias[:50],
        'en_riesgo': en_riesgo,
        'total_en_riesgo': len(en_riesgo),
        # Chart.js
        'meses_labels': meses_labels,
        'meses_data': meses_data,
        'areas_labels': areas_labels,
        'areas_montos': areas_montos,
    }
    return render(request, 'viaticos/dashboard.html', context)


# ── Reporte Excel de CDTs ─────────────────────────────────────────────────────

@login_required
@solo_admin
def viaticos_reporte_excel(request):
    """
    Exporta CDT del período indicado a Excel con detalle completo.
    GET ?mes=<1-12>&anio=<YYYY>
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    hoy = date.today()

    try:
        mes = int(request.GET.get('mes', hoy.month))
        anio = int(request.GET.get('anio', hoy.year))
        if not (1 <= mes <= 12):
            mes = hoy.month
    except (ValueError, TypeError):
        mes, anio = hoy.month, hoy.year

    qs = (
        AsignacionViatico.objects
        .select_related('personal', 'personal__subarea', 'personal__subarea__area', 'creado_por')
        .filter(periodo__year=anio, periodo__month=mes)
        .exclude(estado='CANCELADO')
        .order_by('personal__apellidos_nombres')
    )

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'CDT {mes:02d}-{anio}'

    COLOR_HEADER = '0D2B27'
    COLOR_ALT = 'F0FDFA'
    COLOR_WHITE = 'FFFFFF'
    COLOR_FG_HEADER = 'FFFFFF'

    thin = Side(border_style='thin', color='CBD5E1')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Título ────────────────────────────────────────────────────────────────
    num_cols = 10
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    t = ws['A1']
    t.value = f'Reporte de Viáticos CDT — {mes:02d}/{anio}'
    t.font = Font(name='Calibri', bold=True, size=14, color=COLOR_FG_HEADER)
    t.fill = PatternFill('solid', fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    s = ws['A2']
    s.value = f'Generado: {hoy.strftime("%d/%m/%Y")}   |   Registros: {qs.count()}'
    s.font = Font(name='Calibri', italic=True, size=9, color='6B7280')
    s.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # Espacio

    # ── Cabecera ──────────────────────────────────────────────────────────────
    cols = [
        ('N°',              5),
        ('Empleado',        32),
        ('Área',            20),
        ('Tipo/Concepto',   20),
        ('Ubicación',       22),
        ('Monto Adelanto',  17),
        ('Monto Gastado',   17),
        ('Diferencia',      15),
        ('Estado',          15),
        ('Período',         12),
    ]

    fila_hdr = 4
    for ci, (lbl, w) in enumerate(cols, 1):
        c = ws.cell(row=fila_hdr, column=ci, value=lbl)
        c.font = Font(name='Calibri', bold=True, color=COLOR_FG_HEADER)
        c.fill = PatternFill('solid', fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = borde
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[fila_hdr].height = 24

    # ── Datos ─────────────────────────────────────────────────────────────────
    FMT = '#,##0.00'
    fila = fila_hdr + 1
    totales = {'adelanto': Decimal('0'), 'gastado': Decimal('0'), 'diff': Decimal('0')}

    for idx, a in enumerate(qs, 1):
        fill_bg = PatternFill('solid', fgColor=COLOR_ALT if idx % 2 == 0 else COLOR_WHITE)
        area_nombre = '—'
        if a.personal.subarea and a.personal.subarea.area:
            area_nombre = a.personal.subarea.area.nombre

        diferencia = a.monto_rendido - a.monto_total

        def _cell(col, val, fmt=None, align='left', bold=False, color=None):
            c = ws.cell(row=fila, column=col, value=val)
            c.fill = fill_bg
            c.border = borde
            c.alignment = Alignment(horizontal=align, vertical='center')
            if fmt:
                c.number_format = fmt
            if bold or color:
                c.font = Font(name='Calibri', bold=bold,
                              color=color if color else '000000')
            return c

        _cell(1, idx, align='center')
        c2 = _cell(2, f"{a.personal.apellidos_nombres}\n{a.personal.nro_doc}")
        c2.alignment = Alignment(wrap_text=True, vertical='top')
        _cell(3, area_nombre)
        _cell(4, a.ubicacion or '—')
        _cell(5, a.ubicacion or '—')  # Tipo/Concepto = Ubicación por campo disponible
        _cell(6, float(a.monto_total), fmt=FMT, align='right')
        _cell(7, float(a.monto_rendido), fmt=FMT, align='right')

        # Diferencia — colorear
        c_diff = ws.cell(row=fila, column=8, value=float(diferencia))
        c_diff.fill = fill_bg
        c_diff.border = borde
        c_diff.number_format = FMT
        c_diff.alignment = Alignment(horizontal='right', vertical='center')
        if diferencia == 0:
            c_diff.font = Font(color='16A34A', bold=True)
        elif diferencia > 0:
            c_diff.font = Font(color='DC2626', bold=True)
        else:
            c_diff.font = Font(color='D97706')

        _cell(9, a.get_estado_display(), align='center')
        _cell(10, a.periodo.strftime('%m/%Y'), align='center')

        totales['adelanto'] += a.monto_total
        totales['gastado'] += a.monto_rendido
        totales['diff'] += diferencia
        fila += 1

    # ── Fila de totales ───────────────────────────────────────────────────────
    fill_total = PatternFill('solid', fgColor='E2E8F0')
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=fila, column=ci)
        c.fill = fill_total
        c.border = borde
    ws.cell(row=fila, column=2, value='TOTAL').font = Font(bold=True)

    for ci, key, val in [(6, 'adelanto', totales['adelanto']),
                         (7, 'gastado', totales['gastado']),
                         (8, 'diff', totales['diff'])]:
        c = ws.cell(row=fila, column=ci, value=float(val))
        c.number_format = FMT
        c.font = Font(bold=True,
                      color='16A34A' if val == 0 else ('DC2626' if val > 50 else 'D97706'))
        c.alignment = Alignment(horizontal='right')
        c.border = borde
        c.fill = fill_total

    ws.freeze_panes = ws.cell(row=fila_hdr + 1, column=1)

    nombre = f'viaticos_CDT_{anio}_{mes:02d}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


# ── Conciliación Masiva AJAX ──────────────────────────────────────────────────

@login_required
@solo_admin
def conciliar_masivo(request):
    """
    Concilia múltiples CDTs en una sola operación.
    POST JSON: {ids: [1,2,3], montos: {1: 1200.00, 2: 800.00}}
    Responde: {ok, count, total_conciliado}
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        # También acepta form-data
        data = request.POST

    ids = data.get('ids', [])
    montos = data.get('montos', {})

    if not ids:
        return JsonResponse({'ok': False, 'error': 'No se proporcionaron IDs.'}, status=400)

    qs = AsignacionViatico.objects.filter(
        pk__in=ids,
        estado__in=['ENTREGADO', 'EN_RENDICION']
    )

    count = 0
    total_conciliado = Decimal('0.00')

    for asignacion in qs:
        # Si se envió monto_real para este CDT, actualizarlo primero
        monto_real = montos.get(str(asignacion.pk)) or montos.get(asignacion.pk)
        if monto_real is not None:
            try:
                asignacion.monto_rendido = Decimal(str(monto_real))
                asignacion.save(update_fields=['monto_rendido'])
            except Exception:
                pass  # Mantener valor existente

        # Conciliar
        diff = asignacion.monto_rendido - asignacion.monto_total
        if diff > 0:
            asignacion.monto_reembolso = diff
            asignacion.monto_devuelto = Decimal('0.00')
        elif diff < 0:
            asignacion.monto_devuelto = abs(diff)
            asignacion.monto_reembolso = Decimal('0.00')
        else:
            asignacion.monto_devuelto = Decimal('0.00')
            asignacion.monto_reembolso = Decimal('0.00')

        asignacion.estado = 'CONCILIADO'
        asignacion.fecha_conciliacion = date.today()
        asignacion.save(update_fields=[
            'estado', 'fecha_conciliacion', 'monto_devuelto', 'monto_reembolso'
        ])

        try:
            from core.audit import log_update
            log_update(
                request, asignacion,
                {'estado': {'old': 'EN_RENDICION', 'new': 'CONCILIADO'}},
                f'Conciliación masiva — Rendido: S/ {asignacion.monto_rendido}'
            )
        except Exception:
            pass

        total_conciliado += asignacion.monto_rendido
        count += 1

    return JsonResponse({
        'ok': True,
        'count': count,
        'total_conciliado': str(total_conciliado),
    })


# ── Portal del trabajador ──
@login_required
def mis_viaticos(request):
    """Vista del portal: mis viáticos."""
    from portal.views import _get_empleado
    empleado = _get_empleado(request.user)

    asignaciones = []
    if empleado:
        asignaciones = AsignacionViatico.objects.filter(
            personal=empleado
        ).exclude(estado='CANCELADO').order_by('-periodo')

    context = {
        'titulo': 'Mis Viáticos',
        'empleado': empleado,
        'asignaciones': asignaciones,
    }
    return render(request, 'viaticos/mis_viaticos.html', context)
