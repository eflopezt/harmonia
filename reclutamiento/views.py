"""
Vistas del modulo de Reclutamiento y Seleccion.

Incluye vistas admin (solo_admin) y vistas publicas (portal empleo).
"""
import io
import json
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Count, Avg, F, ExpressionWrapper, DurationField
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.audit import log_create, log_update
from integraciones.models import LogPublicacionVacante
from integraciones.reclutamiento import (
    ComputrabajoExporter,
    BumeranExporter,
    LinkedInJobsPublisher,
    PortalPropio,
    TelegramJobPublisher,
    WhatsAppBusinessPublisher,
)
from personal.models import Area
from .models import (
    Vacante, EtapaPipeline, Postulacion,
    NotaPostulacion, EntrevistaPrograma,
)
from .forms import (
    VacanteForm, PostulacionAdminForm, PostulacionPublicaForm,
    EtapaPipelineForm, NotaPostulacionForm, EntrevistaProgramaForm,
)


solo_admin = user_passes_test(lambda u: u.is_superuser, login_url='login')


# ══════════════════════════════════════════════════════════════
# ADMIN — VACANTES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def vacantes_panel(request):
    """Panel principal de vacantes con estadisticas y filtros."""
    qs = Vacante.objects.select_related('area', 'responsable').all()

    # Filtros
    estado = request.GET.get('estado', '')
    area_id = request.GET.get('area', '')
    prioridad = request.GET.get('prioridad', '')
    buscar = request.GET.get('q', '')

    if estado:
        qs = qs.filter(estado=estado)
    if area_id:
        qs = qs.filter(area_id=area_id)
    if prioridad:
        qs = qs.filter(prioridad=prioridad)
    if buscar:
        qs = qs.filter(Q(titulo__icontains=buscar) | Q(descripcion__icontains=buscar))

    # Stats basicas
    todas = Vacante.objects.all()
    stats = {
        'abiertas': todas.filter(estado='PUBLICADA').count(),
        'en_proceso': todas.filter(estado='EN_PROCESO').count(),
        'cubiertas': todas.filter(estado='CUBIERTA').count(),
        'canceladas': todas.filter(estado='CANCELADA').count(),
        'total_postulaciones': Postulacion.objects.filter(estado='ACTIVA').count(),
    }

    # ── KPIs adicionales ──────────────────────────────────────
    hoy = date.today()

    # 1. Vacantes activas (PUBLICADA + EN_PROCESO)
    vacantes_activas_count = todas.filter(estado__in=['PUBLICADA', 'EN_PROCESO']).count()

    # 2. Candidatos del mes actual
    inicio_mes = hoy.replace(day=1)
    candidatos_mes = Postulacion.objects.filter(
        fecha_postulacion__date__gte=inicio_mes,
    ).count()

    # 3. Tasa de conversion del ultimo mes (contratados / postulantes) * 100
    inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
    postulantes_mes_ant = Postulacion.objects.filter(
        fecha_postulacion__date__gte=inicio_mes_anterior,
        fecha_postulacion__date__lt=inicio_mes,
    ).count()
    contratados_mes_ant = Postulacion.objects.filter(
        fecha_postulacion__date__gte=inicio_mes_anterior,
        fecha_postulacion__date__lt=inicio_mes,
        estado='CONTRATADA',
    ).count()
    if postulantes_mes_ant > 0:
        tasa_conversion = round((contratados_mes_ant / postulantes_mes_ant) * 100, 1)
    else:
        tasa_conversion = 0.0

    # 4. Vacantes por area (annotate)
    vacantes_por_area = (
        Vacante.objects.filter(estado__in=['PUBLICADA', 'EN_PROCESO'])
        .exclude(area__isnull=True)
        .values('area__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')[:8]
    )

    # 5. Tiempo promedio de cierre en dias (vacantes con fecha_publicacion y fecha de ultimo update en CUBIERTA)
    #    Usamos creado_en como proxy de apertura y actualizado_en como proxy de cierre.
    vacantes_cubiertas = todas.filter(
        estado='CUBIERTA',
        fecha_publicacion__isnull=False,
        actualizado_en__isnull=False,
    )
    if vacantes_cubiertas.exists():
        tiempos = [
            (v.actualizado_en.date() - v.fecha_publicacion).days
            for v in vacantes_cubiertas
            if v.actualizado_en.date() >= v.fecha_publicacion
        ]
        tiempo_promedio_cierre = round(sum(tiempos) / len(tiempos), 0) if tiempos else None
    else:
        tiempo_promedio_cierre = None
    # ──────────────────────────────────────────────────────────

    context = {
        'titulo': 'Vacantes',
        'vacantes': qs[:100],
        'total': qs.count(),
        'stats': stats,
        'filtro_estado': estado,
        'filtro_area': area_id,
        'filtro_prioridad': prioridad,
        'buscar': buscar,
        'areas': Area.objects.filter(activa=True),
        # KPIs
        'kpi_vacantes_activas': vacantes_activas_count,
        'kpi_candidatos_mes': candidatos_mes,
        'kpi_tasa_conversion': tasa_conversion,
        'kpi_vacantes_por_area': list(vacantes_por_area),
        'kpi_tiempo_promedio_cierre': tiempo_promedio_cierre,
    }
    return render(request, 'reclutamiento/vacantes_panel.html', context)


@login_required
@solo_admin
def vacante_crear(request):
    """Formulario para crear una nueva vacante."""
    if request.method == 'POST':
        form = VacanteForm(request.POST)
        if form.is_valid():
            vacante = form.save(commit=False)
            vacante.creado_por = request.user
            vacante.save()
            log_create(request, vacante)
            messages.success(request, f'Vacante "{vacante.titulo}" creada exitosamente.')
            return redirect('vacante_detalle', pk=vacante.pk)
    else:
        form = VacanteForm()

    return render(request, 'reclutamiento/vacante_crear.html', {
        'titulo': 'Nueva Vacante',
        'form': form,
    })


@login_required
@solo_admin
def vacante_detalle(request, pk):
    """Detalle de vacante con pipeline kanban."""
    vacante = get_object_or_404(
        Vacante.objects.select_related('area', 'responsable', 'creado_por'),
        pk=pk,
    )
    etapas = EtapaPipeline.objects.filter(activa=True)
    postulaciones = vacante.postulaciones.filter(estado='ACTIVA').select_related('etapa')

    # Calcular dias_en_etapa para cada postulacion.
    # Usamos la fecha de la ultima nota (las notas de cambio de etapa se crean automaticamente).
    # Si no hay notas, se usa fecha_postulacion como referencia.
    hoy_dt = timezone.now().date()
    post_ids = [p.pk for p in postulaciones]
    # Obtener la ultima nota por postulacion
    from django.db.models import Max
    ultimas_notas = dict(
        NotaPostulacion.objects.filter(postulacion_id__in=post_ids)
        .values('postulacion_id')
        .annotate(ultima=Max('fecha'))
        .values_list('postulacion_id', 'ultima')
    )

    def _dias_en_etapa(p):
        ultima = ultimas_notas.get(p.pk)
        if ultima:
            ref = ultima.date() if hasattr(ultima, 'date') else ultima
        else:
            ref = p.fecha_postulacion.date() if hasattr(p.fecha_postulacion, 'date') else p.fecha_postulacion
        return max((hoy_dt - ref).days, 0)

    for p in postulaciones:
        p.dias_en_etapa = _dias_en_etapa(p)

    # Organizar postulaciones por etapa
    pipeline = []
    for etapa in etapas:
        posts_en_etapa = [p for p in postulaciones if p.etapa_id == etapa.pk]
        pipeline.append({
            'etapa': etapa,
            'postulaciones': posts_en_etapa,
            'count': len(posts_en_etapa),
        })

    # Postulaciones descartadas
    descartadas = vacante.postulaciones.filter(estado='DESCARTADA').select_related('etapa')

    context = {
        'titulo': vacante.titulo,
        'vacante': vacante,
        'pipeline': pipeline,
        'etapas': etapas,
        'descartadas': descartadas,
        'total_activas': postulaciones.count(),
        'form_postulacion': PostulacionAdminForm(),
    }
    return render(request, 'reclutamiento/vacante_detalle.html', context)


@login_required
@solo_admin
def vacante_editar(request, pk):
    """Editar una vacante existente."""
    vacante = get_object_or_404(Vacante, pk=pk)

    if request.method == 'POST':
        form = VacanteForm(request.POST, instance=vacante)
        if form.is_valid():
            cambios = {}
            for field in form.changed_data:
                cambios[field] = {
                    'old': getattr(vacante, field),
                    'new': form.cleaned_data[field],
                }
            vacante = form.save()
            if cambios:
                log_update(request, vacante, cambios)
            messages.success(request, f'Vacante "{vacante.titulo}" actualizada.')
            return redirect('vacante_detalle', pk=vacante.pk)
    else:
        form = VacanteForm(instance=vacante)

    return render(request, 'reclutamiento/vacante_crear.html', {
        'titulo': f'Editar: {vacante.titulo}',
        'form': form,
        'vacante': vacante,
        'editando': True,
    })


# ══════════════════════════════════════════════════════════════
# ADMIN — POSTULACIONES
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def postulacion_crear(request, vacante_pk):
    """Agregar postulacion a una vacante (entrada admin)."""
    vacante = get_object_or_404(Vacante, pk=vacante_pk)

    if request.method == 'POST':
        form = PostulacionAdminForm(request.POST, request.FILES)
        if form.is_valid():
            postulacion = form.save(commit=False)
            postulacion.vacante = vacante
            # Asignar a la primera etapa activa
            primera_etapa = EtapaPipeline.objects.filter(activa=True).first()
            postulacion.etapa = primera_etapa
            postulacion.save()
            log_create(request, postulacion)
            messages.success(request, f'Postulacion de "{postulacion.nombre_completo}" registrada.')
            return redirect('vacante_detalle', pk=vacante.pk)
    else:
        form = PostulacionAdminForm()

    return render(request, 'reclutamiento/vacante_detalle.html', {
        'titulo': vacante.titulo,
        'vacante': vacante,
        'form_postulacion': form,
    })


@login_required
@solo_admin
def postulacion_detalle(request, pk):
    """Detalle de una postulacion con timeline, notas y entrevistas."""
    postulacion = get_object_or_404(
        Postulacion.objects.select_related('vacante', 'etapa'),
        pk=pk,
    )
    notas = postulacion.notas_detalle.select_related('autor').all()
    entrevistas = postulacion.entrevistas.select_related('entrevistador').all()
    etapas = EtapaPipeline.objects.filter(activa=True)

    context = {
        'titulo': postulacion.nombre_completo,
        'postulacion': postulacion,
        'notas': notas,
        'entrevistas': entrevistas,
        'etapas': etapas,
        'form_nota': NotaPostulacionForm(),
        'form_entrevista': EntrevistaProgramaForm(),
    }
    return render(request, 'reclutamiento/postulacion_detalle.html', context)


@login_required
@solo_admin
@require_POST
def postulacion_mover_etapa(request, pk):
    """Mover postulacion a otra etapa del pipeline (AJAX)."""
    postulacion = get_object_or_404(Postulacion, pk=pk)

    # Solo postulaciones activas pueden moverse entre etapas
    if postulacion.estado != 'ACTIVA':
        return JsonResponse(
            {'ok': False, 'error': 'Solo se puede mover postulaciones activas.'},
            status=400
        )

    etapa_id = request.POST.get('etapa_id')

    if not etapa_id:
        return JsonResponse({'ok': False, 'error': 'Etapa no especificada'}, status=400)

    etapa = get_object_or_404(EtapaPipeline, pk=etapa_id, activa=True)
    etapa_anterior = postulacion.etapa
    texto_etapa_anterior = str(etapa_anterior) if etapa_anterior else 'Sin etapa'

    postulacion.etapa = etapa
    postulacion.save(update_fields=['etapa'])

    log_update(request, postulacion, {
        'etapa': {'old': texto_etapa_anterior, 'new': str(etapa)},
    })

    # Crear nota automatica
    NotaPostulacion.objects.create(
        postulacion=postulacion,
        autor=request.user,
        texto=f'Movido de "{texto_etapa_anterior}" a "{etapa}"',
        tipo='NOTA',
    )

    return JsonResponse({
        'ok': True,
        'etapa_nombre': etapa.nombre,
        'etapa_color': etapa.color,
    })


@login_required
@solo_admin
@require_POST
def postulacion_descartar(request, pk):
    """Descartar una postulacion (AJAX)."""
    postulacion = get_object_or_404(Postulacion, pk=pk, estado='ACTIVA')
    motivo = request.POST.get('motivo', '')

    estado_anterior = postulacion.estado
    postulacion.estado = 'DESCARTADA'
    postulacion.save(update_fields=['estado'])

    log_update(request, postulacion, {
        'estado': {'old': estado_anterior, 'new': 'DESCARTADA'},
    })

    # Crear nota con motivo
    NotaPostulacion.objects.create(
        postulacion=postulacion,
        autor=request.user,
        texto=f'Candidato descartado. {motivo}'.strip(),
        tipo='NOTA',
    )

    return JsonResponse({'ok': True})


# ══════════════════════════════════════════════════════════════
# ADMIN — NOTAS Y ENTREVISTAS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
@require_POST
def nota_agregar(request, postulacion_pk):
    """Agregar nota a una postulacion (AJAX)."""
    postulacion = get_object_or_404(Postulacion, pk=postulacion_pk)
    form = NotaPostulacionForm(request.POST)

    if form.is_valid():
        nota = form.save(commit=False)
        nota.postulacion = postulacion
        nota.autor = request.user
        nota.save()
        return JsonResponse({
            'ok': True,
            'nota': {
                'texto': nota.texto,
                'tipo': nota.get_tipo_display(),
                'autor': request.user.get_full_name() or request.user.username,
                'fecha': nota.fecha.strftime('%d/%m/%Y %H:%M'),
            }
        })

    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


@login_required
@solo_admin
@require_POST
def entrevista_crear(request, postulacion_pk):
    """Programar entrevista para una postulacion (AJAX)."""
    postulacion = get_object_or_404(Postulacion, pk=postulacion_pk)
    form = EntrevistaProgramaForm(request.POST)

    if form.is_valid():
        entrevista = form.save(commit=False)
        entrevista.postulacion = postulacion
        entrevista.save()
        log_create(request, entrevista)
        return JsonResponse({
            'ok': True,
            'entrevista': {
                'tipo': entrevista.get_tipo_display(),
                'fecha': entrevista.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                'entrevistador': str(entrevista.entrevistador),
                'modalidad': entrevista.get_modalidad_display(),
            }
        })

    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


@login_required
@solo_admin
@require_POST
def entrevista_resultado(request, pk):
    """Registrar resultado de entrevista (AJAX)."""
    entrevista = get_object_or_404(EntrevistaPrograma, pk=pk)
    resultado = request.POST.get('resultado', '')
    calificacion = request.POST.get('calificacion')
    notas_post = request.POST.get('notas_post', '')

    if resultado not in dict(EntrevistaPrograma.RESULTADO_CHOICES):
        return JsonResponse({'ok': False, 'error': 'Resultado invalido'}, status=400)

    entrevista.resultado = resultado
    if calificacion:
        try:
            cal = int(calificacion)
            if 1 <= cal <= 10:
                entrevista.calificacion = cal
        except (ValueError, TypeError):
            pass
    entrevista.notas_post = notas_post
    entrevista.save()

    log_update(request, entrevista, {
        'resultado': {'old': 'PENDIENTE', 'new': resultado},
    })

    return JsonResponse({
        'ok': True,
        'resultado': entrevista.get_resultado_display(),
        'calificacion': entrevista.calificacion,
    })


# ══════════════════════════════════════════════════════════════
# ADMIN — PIPELINE PANEL (cross-vacante)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def pipeline_panel(request):
    """Vista kanban cross-vacante con todas las postulaciones activas."""
    etapas = EtapaPipeline.objects.filter(activa=True)
    vacante_id = request.GET.get('vacante', '')

    postulaciones = Postulacion.objects.filter(
        estado='ACTIVA',
    ).select_related('vacante', 'etapa')

    if vacante_id:
        postulaciones = postulaciones.filter(vacante_id=vacante_id)

    # Calcular dias_en_etapa para pipeline global
    hoy_dt_pl = timezone.now().date()
    pl_ids = [p.pk for p in postulaciones]
    from django.db.models import Max as _Max
    ultimas_notas_pl = dict(
        NotaPostulacion.objects.filter(postulacion_id__in=pl_ids)
        .values('postulacion_id')
        .annotate(ultima=_Max('fecha'))
        .values_list('postulacion_id', 'ultima')
    )
    for p in postulaciones:
        ultima = ultimas_notas_pl.get(p.pk)
        if ultima:
            ref = ultima.date() if hasattr(ultima, 'date') else ultima
        else:
            ref = p.fecha_postulacion.date() if hasattr(p.fecha_postulacion, 'date') else p.fecha_postulacion
        p.dias_en_etapa = max((hoy_dt_pl - ref).days, 0)

    # Organizar por etapa
    pipeline = []
    for etapa in etapas:
        posts = [p for p in postulaciones if p.etapa_id == etapa.pk]
        pipeline.append({
            'etapa': etapa,
            'postulaciones': posts,
            'count': len(posts),
        })

    vacantes_activas = Vacante.objects.filter(
        estado__in=['PUBLICADA', 'EN_PROCESO']
    ).order_by('titulo')

    context = {
        'titulo': 'Pipeline de Seleccion',
        'pipeline': pipeline,
        'etapas': etapas,
        'vacantes_activas': vacantes_activas,
        'filtro_vacante': vacante_id,
    }
    return render(request, 'reclutamiento/pipeline_panel.html', context)


# ══════════════════════════════════════════════════════════════
# ADMIN — EXPORTAR CANDIDATOS EXCEL
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def exportar_candidatos_excel(request, pk):
    """
    Exporta todas las postulaciones de una vacante a un archivo .xlsx.

    Columnas: Nombre, Email, Telefono, Etapa Actual, Puntaje (calificacion
    promedio de entrevistas), Fecha Aplicacion, Notas.
    Estilos Harmoni: header con fondo #0D2B27, texto blanco.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl no esta instalado. Ejecuta: pip install openpyxl')
        return redirect('vacante_detalle', pk=pk)

    vacante = get_object_or_404(Vacante.objects.select_related('area'), pk=pk)

    postulaciones = (
        Postulacion.objects
        .filter(vacante=vacante)
        .select_related('etapa')
        .prefetch_related('entrevistas')
        .order_by('estado', 'fecha_postulacion')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Candidatos'

    # ── Estilos Harmoni ──────────────────────────────────────
    COLOR_HEADER  = '0D2B27'
    COLOR_ACCENT  = '5EEAD4'
    COLOR_ALT_ROW = 'F0FDFA'

    header_fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    alt_fill  = PatternFill(fill_type='solid', fgColor=COLOR_ALT_ROW)
    body_font = Font(size=9)
    body_align = Alignment(vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # ── Fila titulo ──────────────────────────────────────────
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'Candidatos: {vacante.titulo}'
    title_cell.font = Font(bold=True, color='FFFFFF', size=12)
    title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # ── Fila subtitulo ───────────────────────────────────────
    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    area_nombre = vacante.area.nombre if vacante.area else 'Sin area'
    sub_cell.value = (
        f'Area: {area_nombre}  |  Estado: {vacante.get_estado_display()}  |  '
        f'Exportado: {date.today():%d/%m/%Y}'
    )
    sub_cell.font = Font(italic=True, color='64748B', size=8)
    sub_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 14

    # ── Encabezados ──────────────────────────────────────────
    headers = [
        'Nombre Completo',
        'Email',
        'Telefono',
        'Etapa Actual',
        'Puntaje Entrevista',
        'Fecha Aplicacion',
        'Estado',
        'Notas',
    ]
    col_widths = [30, 28, 14, 18, 16, 16, 12, 40]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 20

    # ── Filas de datos ───────────────────────────────────────
    for row_idx, post in enumerate(postulaciones, start=4):
        # Puntaje: promedio de calificaciones de entrevistas completadas
        calificaciones = [
            e.calificacion
            for e in post.entrevistas.all()
            if e.calificacion is not None
        ]
        puntaje = round(sum(calificaciones) / len(calificaciones), 1) if calificaciones else ''

        fecha_app = (
            post.fecha_postulacion.strftime('%d/%m/%Y %H:%M')
            if post.fecha_postulacion else ''
        )

        row_data = [
            post.nombre_completo,
            post.email,
            post.telefono,
            post.etapa.nombre if post.etapa else '—',
            puntaje,
            fecha_app,
            post.get_estado_display(),
            post.notas,
        ]

        fill = alt_fill if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 15

    # ── Freeze panes debajo de headers ───────────────────────
    ws.freeze_panes = 'A4'

    # ── Respuesta HTTP ───────────────────────────────────────
    nombre_archivo = (
        f"candidatos_{vacante.titulo[:30].replace(' ', '_')}_{date.today():%Y%m%d}.xlsx"
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ══════════════════════════════════════════════════════════════
# ADMIN — CONFIGURACION ETAPAS
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def etapas_config(request):
    """Configuracion de etapas del pipeline."""
    etapas = EtapaPipeline.objects.all()
    form = EtapaPipelineForm()

    context = {
        'titulo': 'Etapas del Pipeline',
        'etapas': etapas,
        'form': form,
    }
    return render(request, 'reclutamiento/etapas_config.html', context)


@login_required
@solo_admin
@require_POST
def etapa_crear(request):
    """Crear nueva etapa del pipeline (AJAX)."""
    form = EtapaPipelineForm(request.POST)
    if form.is_valid():
        etapa = form.save()
        log_create(request, etapa)
        return JsonResponse({
            'ok': True,
            'etapa': {
                'id': etapa.pk,
                'nombre': etapa.nombre,
                'codigo': etapa.codigo,
                'orden': etapa.orden,
                'color': etapa.color,
            }
        })
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


# ══════════════════════════════════════════════════════════════
# ADMIN — PUBLICAR EN PLATAFORMAS EXTERNAS
# ══════════════════════════════════════════════════════════════

# Plataformas disponibles para mostrar en el formulario
PLATAFORMAS_DISPONIBLES = [
    {
        'id':          'computrabajo',
        'nombre':      'Computrabajo Peru',
        'icono':       'bi bi-briefcase',
        'descripcion': 'Genera XML/JSON para importacion masiva en Computrabajo.com.pe',
        'tipo':        'EXPORT',                           # EXPORT = genera archivo, API = llama API
    },
    {
        'id':          'bumeran',
        'nombre':      'Bumeran Peru',
        'icono':       'bi bi-people',
        'descripcion': 'Genera XML/JSON para feed masivo en Bumeran.com.pe',
        'tipo':        'EXPORT',
    },
    {
        'id':          'linkedin',
        'nombre':      'LinkedIn Jobs',
        'icono':       'bi bi-linkedin',
        'descripcion': 'Publica via API OAuth2 (requiere access_token configurado)',
        'tipo':        'API',
    },
    {
        'id':          'telegram',
        'nombre':      'Telegram Bot',
        'icono':       'bi bi-telegram',
        'descripcion': 'Publica en un canal de Telegram via Bot API (requiere token y chat_id)',
        'tipo':        'API',
    },
    {
        'id':          'whatsapp',
        'nombre':      'WhatsApp Business',
        'icono':       'bi bi-whatsapp',
        'descripcion': 'Envia a número(s) via WhatsApp Business Cloud API (Meta Graph API)',
        'tipo':        'API',
    },
    {
        'id':          'portal',
        'nombre':      'Portal de Empleo Propio',
        'icono':       'bi bi-globe',
        'descripcion': 'Activa la vacante en el portal publico de la empresa',
        'tipo':        'INTERNO',
    },
]


@login_required
@solo_admin
def publicar_en_plataformas(request, pk):
    """
    Vista para publicar una vacante en una o varias plataformas externas.

    GET:  Muestra formulario con checkboxes para elegir plataformas.
    POST: Ejecuta los publicadores seleccionados y guarda logs.

    Los resultados (payloads XML/JSON para Computrabajo y Bumeran) se muestran
    en pantalla para que el usuario los descargue o copie.
    """
    vacante = get_object_or_404(
        Vacante.objects.select_related('area', 'responsable'),
        pk=pk,
    )

    # Logs historicos de publicacion para esta vacante
    logs = LogPublicacionVacante.objects.filter(vacante=vacante).order_by('-creado_en')[:20]

    if request.method == 'POST':
        plataformas_sel = request.POST.getlist('plataformas')

        if not plataformas_sel:
            messages.warning(request, 'Debes seleccionar al menos una plataforma.')
            return redirect('reclutamiento_publicar', pk=pk)

        resultados = []

        # Obtener base_url del request
        base_url = request.build_absolute_uri('/').rstrip('/')

        for plataforma_id in plataformas_sel:
            resultado = _ejecutar_publicador(request, vacante, plataforma_id, base_url)
            resultados.append(resultado)

            # Guardar log en BD
            _guardar_log_publicacion(vacante, resultado, request.user)

        # Mensajes de resumen
        exitosos = [r for r in resultados if r.get('ok')]
        fallidos  = [r for r in resultados if not r.get('ok')]

        if exitosos:
            nombres = ', '.join(r.get('plataforma', '') for r in exitosos)
            messages.success(request, f'Publicacion exitosa en: {nombres}.')

        if fallidos:
            for r in fallidos:
                messages.error(
                    request,
                    f'Error en {r.get("plataforma", "")}: {r.get("error", "Error desconocido")}',
                )

        return render(request, 'reclutamiento/publicar_plataformas.html', {
            'titulo':      f'Publicar: {vacante.titulo}',
            'vacante':     vacante,
            'plataformas': PLATAFORMAS_DISPONIBLES,
            'logs':        LogPublicacionVacante.objects.filter(vacante=vacante).order_by('-creado_en')[:20],
            'resultados':  resultados,
            'enviado':     True,
        })

    return render(request, 'reclutamiento/publicar_plataformas.html', {
        'titulo':      f'Publicar: {vacante.titulo}',
        'vacante':     vacante,
        'plataformas': PLATAFORMAS_DISPONIBLES,
        'logs':        logs,
        'resultados':  [],
        'enviado':     False,
    })


def _ejecutar_publicador(request, vacante, plataforma_id: str, base_url: str) -> dict:
    """
    Ejecuta el publicador correspondiente para una plataforma.

    Returns:
        dict estandarizado con ok, plataforma, mensaje, error, y campos extras segun plataforma.
    """
    if plataforma_id == 'computrabajo':
        exporter = ComputrabajoExporter()
        return exporter.publicar_vacante(vacante)

    elif plataforma_id == 'bumeran':
        exporter = BumeranExporter()
        return exporter.publicar_vacante(vacante)

    elif plataforma_id == 'linkedin':
        publisher = LinkedInJobsPublisher()

        # Obtener credenciales desde ConfiguracionSistema (si existe) o POST
        access_token    = request.POST.get('linkedin_access_token', '').strip()
        organization_id = request.POST.get('linkedin_organization_id', '').strip()

        # Intentar desde ConfiguracionSistema si el app existe
        if not access_token or not organization_id:
            try:
                from asistencia.models import ConfiguracionSistema
                config = ConfiguracionSistema.objects.first()
                if config:
                    if not access_token:
                        access_token    = getattr(config, 'linkedin_access_token', '') or ''
                    if not organization_id:
                        organization_id = getattr(config, 'linkedin_organization_id', '') or ''
            except Exception:
                pass

        if not access_token:
            # Devolver preview del payload sin publicar
            return publisher.generar_payload_preview(vacante, organization_id or 'ORG_ID')

        return publisher.publicar_vacante(vacante, access_token, organization_id)

    elif plataforma_id == 'telegram':
        publisher = TelegramJobPublisher()

        # Credenciales desde POST o ConfiguracionSistema
        bot_token = request.POST.get('telegram_bot_token', '').strip()
        chat_id   = request.POST.get('telegram_chat_id', '').strip()

        if not bot_token or not chat_id:
            try:
                from asistencia.models import ConfiguracionSistema
                config = ConfiguracionSistema.objects.first()
                if config:
                    if not bot_token:
                        bot_token = getattr(config, 'telegram_bot_token', '') or ''
                    if not chat_id:
                        chat_id = getattr(config, 'telegram_channel_id', '') or ''
            except Exception:
                pass

        portal_url = f'{base_url}/empleo/{vacante.pk}/postular/'

        if not bot_token or not chat_id:
            # Devolver preview del mensaje sin publicar
            return publisher.generar_preview(vacante, portal_url=portal_url)

        return publisher.publicar_vacante(
            vacante,
            bot_token=bot_token,
            chat_id=chat_id,
            portal_url=portal_url,
        )

    elif plataforma_id == 'whatsapp':
        publisher = WhatsAppBusinessPublisher()

        phone_number_id = request.POST.get('whatsapp_phone_number_id', '').strip()
        access_token    = request.POST.get('whatsapp_access_token', '').strip()
        to_numbers      = request.POST.get('whatsapp_to_number', '').strip()

        if not phone_number_id or not access_token or not to_numbers:
            try:
                from asistencia.models import ConfiguracionSistema
                config = ConfiguracionSistema.objects.first()
                if config:
                    if not phone_number_id:
                        phone_number_id = getattr(config, 'whatsapp_phone_number_id', '') or ''
                    if not access_token:
                        access_token = getattr(config, 'whatsapp_access_token', '') or ''
                    if not to_numbers:
                        to_numbers = getattr(config, 'whatsapp_to_number', '') or ''
            except Exception:
                pass

        portal_url = f'{base_url}/empleo/{vacante.pk}/postular/'

        if not phone_number_id or not access_token or not to_numbers:
            return publisher.generar_preview(vacante, portal_url=portal_url)

        return publisher.publicar_vacante(
            vacante,
            phone_number_id=phone_number_id,
            access_token=access_token,
            to_numbers=to_numbers,
            portal_url=portal_url,
        )

    elif plataforma_id == 'portal':
        portal = PortalPropio()
        return portal.publicar_vacante(vacante, base_url=base_url)

    else:
        return {
            'ok':        False,
            'plataforma': plataforma_id.upper(),
            'error':     f'Plataforma "{plataforma_id}" no reconocida.',
            'mensaje':   '',
        }


def _guardar_log_publicacion(vacante, resultado: dict, usuario) -> None:
    """Persiste el resultado de una publicacion en LogPublicacionVacante."""
    plataforma_id = resultado.get('plataforma', 'OTRO')

    # Mapear id interno al choice del modelo
    mapa_plataforma = {
        'COMPUTRABAJO': 'COMPUTRABAJO',
        'BUMERAN':      'BUMERAN',
        'LINKEDIN':     'LINKEDIN',
        'TELEGRAM':     'TELEGRAM',
        'WHATSAPP':     'WHATSAPP',
        'PORTAL':       'PORTAL',
    }
    plataforma_code = mapa_plataforma.get(plataforma_id.upper(), 'PORTAL')

    # Serializar respuesta_api de forma segura
    respuesta_api = ''
    if resultado.get('respuesta_api'):
        try:
            respuesta_api = json.dumps(resultado['respuesta_api'], ensure_ascii=False, indent=2)
        except (TypeError, ValueError):
            respuesta_api = str(resultado['respuesta_api'])

    LogPublicacionVacante.objects.create(
        vacante       = vacante,
        plataforma    = plataforma_code,
        estado        = 'OK' if resultado.get('ok') else 'ERROR',
        url_publicada = resultado.get('url_publicada', ''),
        respuesta_api = respuesta_api,
        mensaje       = resultado.get('mensaje', '') or resultado.get('error', ''),
        publicado_por = usuario,
    )


# ══════════════════════════════════════════════════════════════
# PUBLICO — PORTAL DE EMPLEO
# ══════════════════════════════════════════════════════════════

def portal_empleo(request):
    """Pagina publica de oportunidades laborales."""
    vacantes = Vacante.objects.filter(
        publica=True,
        estado__in=['PUBLICADA', 'EN_PROCESO'],
    ).select_related('area').order_by('-fecha_publicacion')

    area_id = request.GET.get('area', '')
    if area_id:
        vacantes = vacantes.filter(area_id=area_id)

    areas_con_vacantes = Area.objects.filter(
        vacantes__publica=True,
        vacantes__estado__in=['PUBLICADA', 'EN_PROCESO'],
    ).distinct().order_by('nombre')

    context = {
        'vacantes': vacantes,
        'areas': areas_con_vacantes,
        'filtro_area': area_id,
    }
    return render(request, 'reclutamiento/portal_empleo.html', context)


def portal_postular(request, pk):
    """Formulario publico para postularse a una vacante."""
    vacante = get_object_or_404(
        Vacante,
        pk=pk,
        publica=True,
        estado__in=['PUBLICADA', 'EN_PROCESO'],
    )

    enviado = False

    if request.method == 'POST':
        form = PostulacionPublicaForm(request.POST, request.FILES)
        if form.is_valid():
            postulacion = form.save(commit=False)
            postulacion.vacante = vacante
            postulacion.fuente = 'PORTAL'
            # Asignar a primera etapa
            primera_etapa = EtapaPipeline.objects.filter(activa=True).first()
            postulacion.etapa = primera_etapa
            postulacion.save()
            enviado = True
    else:
        form = PostulacionPublicaForm()

    context = {
        'vacante': vacante,
        'form': form,
        'enviado': enviado,
    }
    return render(request, 'reclutamiento/portal_postular.html', context)


# ══════════════════════════════════════════════════════════════
# SCORING DE CANDIDATOS
# ══════════════════════════════════════════════════════════════

# Orden jerarquico de niveles educativos para comparacion
_EDUCACION_ORDEN = {
    'NO_REQUERIDO': 0,
    'SECUNDARIA': 1,
    'TECNICO': 2,
    'UNIVERSITARIO': 3,
    'MAESTRIA': 4,
    'DOCTORADO': 5,
}


def _calcular_score(postulacion, vacante):
    """
    Calcula el score de una postulacion respecto a la vacante.

    Criterios:
      +30  Experiencia >= experiencia_minima de la vacante
      +25  Educacion >= educacion_minima de la vacante
      +20  Salario pretendido dentro del rango [salario_min, salario_max]
      +10  Tiene notas (considerado como carta de presentacion / informacion adicional)
      +15  Entrevista con calificacion >= 8 (escala 1-10)
      +8   Entrevista con calificacion >= 6 (escala 1-10)

    Retorna dict con score total y detalle por criterio.
    """
    score = 0
    detalle = {
        'experiencia': 0,
        'educacion': 0,
        'salario': 0,
        'carta': 0,
        'entrevista': 0,
    }

    # Experiencia
    if postulacion.experiencia_anos >= vacante.experiencia_minima:
        score += 30
        detalle['experiencia'] = 30

    # Educacion
    nivel_candidato = _EDUCACION_ORDEN.get(postulacion.educacion, 0)
    nivel_requerido = _EDUCACION_ORDEN.get(vacante.educacion_minima, 0)
    if nivel_candidato >= nivel_requerido:
        score += 25
        detalle['educacion'] = 25

    # Salario pretendido
    if postulacion.salario_pretendido is not None:
        sal = postulacion.salario_pretendido
        min_ok = (vacante.salario_min is None) or (sal >= vacante.salario_min)
        max_ok = (vacante.salario_max is None) or (sal <= vacante.salario_max)
        if min_ok and max_ok:
            score += 20
            detalle['salario'] = 20

    # Carta de presentacion / notas adicionales
    if postulacion.notas and postulacion.notas.strip():
        score += 10
        detalle['carta'] = 10

    # Mejor calificacion de entrevista (1-10)
    mejor_cal = (
        postulacion.entrevistas
        .filter(calificacion__isnull=False)
        .order_by('-calificacion')
        .values_list('calificacion', flat=True)
        .first()
    )
    if mejor_cal is not None:
        if mejor_cal >= 8:
            score += 15
            detalle['entrevista'] = 15
        elif mejor_cal >= 6:
            score += 8
            detalle['entrevista'] = 8

    return score, detalle


@login_required
@solo_admin
def scoring_candidatos(request, pk):
    """Scoring rules-based de todos los candidatos de una vacante."""
    vacante = get_object_or_404(
        Vacante.objects.select_related('area'),
        pk=pk,
    )
    postulaciones = (
        Postulacion.objects
        .filter(vacante=vacante)
        .select_related('etapa')
        .prefetch_related('entrevistas')
        .order_by('nombre_completo')
    )

    resultados = []
    for p in postulaciones:
        score, detalle = _calcular_score(p, vacante)
        # Color segun score
        if score >= 70:
            color = '#16a34a'
            nivel = 'success'
        elif score >= 50:
            color = '#f59e0b'
            nivel = 'warning'
        else:
            color = '#dc2626'
            nivel = 'danger'

        resultados.append({
            'postulacion': p,
            'score': score,
            'color': color,
            'nivel': nivel,
            'detalle': detalle,
        })

    # Ordenar por score descendente
    resultados.sort(key=lambda x: x['score'], reverse=True)

    etapas = EtapaPipeline.objects.filter(activa=True)

    context = {
        'titulo': f'Scoring — {vacante.titulo}',
        'vacante': vacante,
        'resultados': resultados,
        'etapas': etapas,
        'total': len(resultados),
    }
    return render(request, 'reclutamiento/scoring.html', context)


# ══════════════════════════════════════════════════════════════
# HISTORIAL DE POSTULACION
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def postulacion_historial(request, pk):
    """
    Historial completo de un candidato:
    - Todas sus postulaciones (mismo nombre/email)
    - Timeline de notas
    - Entrevistas
    - Documentos adjuntos
    """
    postulacion = get_object_or_404(
        Postulacion.objects.select_related('vacante', 'etapa', 'vacante__area'),
        pk=pk,
    )

    # Buscar otras postulaciones del mismo candidato (por email si disponible, sino por nombre)
    otras_postulaciones = Postulacion.objects.none()
    if postulacion.email:
        otras_postulaciones = (
            Postulacion.objects
            .filter(email=postulacion.email)
            .exclude(pk=pk)
            .select_related('vacante', 'etapa')
            .order_by('-fecha_postulacion')
        )
    else:
        otras_postulaciones = (
            Postulacion.objects
            .filter(nombre_completo__iexact=postulacion.nombre_completo)
            .exclude(pk=pk)
            .select_related('vacante', 'etapa')
            .order_by('-fecha_postulacion')
        )

    notas = postulacion.notas_detalle.select_related('autor').order_by('fecha')
    entrevistas = postulacion.entrevistas.select_related('entrevistador').order_by('fecha_hora')

    # Documentos: CV si existe
    documentos = []
    if postulacion.cv:
        documentos.append({
            'nombre': 'Curriculum Vitae',
            'url': postulacion.cv.url,
            'icono': 'fas fa-file-pdf',
        })

    context = {
        'titulo': f'Historial — {postulacion.nombre_completo}',
        'postulacion': postulacion,
        'otras_postulaciones': otras_postulaciones,
        'notas': notas,
        'entrevistas': entrevistas,
        'documentos': documentos,
        'total_postulaciones': otras_postulaciones.count() + 1,
    }
    return render(request, 'reclutamiento/postulacion_historial.html', context)


# ══════════════════════════════════════════════════════════════
# AGENDAR ENTREVISTA (standalone GET+POST)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def entrevista_agendar(request, pk):
    """
    Formulario standalone para agendar una entrevista a una postulacion.
    GET: muestra el formulario.
    POST: crea EntrevistaPrograma y redirige al detalle de la postulacion.
    """
    postulacion = get_object_or_404(
        Postulacion.objects.select_related('vacante', 'etapa'),
        pk=pk,
    )

    if request.method == 'POST':
        form = EntrevistaProgramaForm(request.POST)
        if form.is_valid():
            entrevista = form.save(commit=False)
            entrevista.postulacion = postulacion
            entrevista.save()
            log_create(request, entrevista)

            # Notificacion al entrevistador (best-effort)
            if entrevista.entrevistador and entrevista.entrevistador.email:
                try:
                    from comunicaciones.services import NotificacionService
                    NotificacionService.enviar_email(
                        destinatario=entrevista.entrevistador.email,
                        asunto=f'Entrevista agendada: {postulacion.nombre_completo}',
                        cuerpo=(
                            f'Se ha agendado una entrevista {entrevista.get_tipo_display()} '
                            f'({entrevista.get_modalidad_display()}) con {postulacion.nombre_completo} '
                            f'para el {entrevista.fecha_hora:%d/%m/%Y %H:%M}.\n\n'
                            f'Vacante: {postulacion.vacante.titulo}\n'
                            f'{"Enlace: " + entrevista.enlace_virtual if entrevista.enlace_virtual else ""}'
                        ),
                    )
                except Exception:
                    pass  # La notificacion falla silenciosamente

            messages.success(
                request,
                f'Entrevista agendada para {entrevista.fecha_hora:%d/%m/%Y %H:%M}.',
            )
            return redirect('postulacion_detalle', pk=postulacion.pk)
    else:
        form = EntrevistaProgramaForm()

    context = {
        'titulo': f'Agendar Entrevista — {postulacion.nombre_completo}',
        'postulacion': postulacion,
        'form': form,
    }
    return render(request, 'reclutamiento/entrevista_agendar.html', context)


# ══════════════════════════════════════════════════════════════
# MOVER ETAPA (AJAX, acepta etapa por nombre o pk)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
@require_POST
def mover_etapa(request, pk):
    """
    Mover postulacion a nueva etapa via AJAX.

    POST params:
        nueva_etapa  — nombre (str) o pk (int) de la etapa destino
        comentario   — comentario opcional para el historial
    """
    postulacion = get_object_or_404(Postulacion, pk=pk)
    nueva_etapa_raw = request.POST.get('nueva_etapa', '').strip()
    comentario = request.POST.get('comentario', '').strip()

    if not nueva_etapa_raw:
        return JsonResponse({'ok': False, 'error': 'Etapa no especificada'}, status=400)

    # Intentar lookup por pk primero, luego por nombre
    etapa = None
    if nueva_etapa_raw.isdigit():
        etapa = EtapaPipeline.objects.filter(pk=int(nueva_etapa_raw), activa=True).first()
    if etapa is None:
        etapa = EtapaPipeline.objects.filter(nombre__iexact=nueva_etapa_raw, activa=True).first()
    if etapa is None:
        etapa = EtapaPipeline.objects.filter(codigo__iexact=nueva_etapa_raw, activa=True).first()

    if etapa is None:
        return JsonResponse({'ok': False, 'error': f'Etapa "{nueva_etapa_raw}" no encontrada'}, status=404)

    etapa_anterior = postulacion.etapa
    postulacion.etapa = etapa
    postulacion.save(update_fields=['etapa'])

    log_update(request, postulacion, {
        'etapa': {'old': str(etapa_anterior), 'new': str(etapa)},
    })

    # Nota automatica
    texto_nota = f'Etapa cambiada de "{etapa_anterior}" a "{etapa}"'
    if comentario:
        texto_nota += f'. {comentario}'
    NotaPostulacion.objects.create(
        postulacion=postulacion,
        autor=request.user,
        texto=texto_nota,
        tipo='NOTA',
    )

    # Notificacion al candidato (best-effort)
    if postulacion.email:
        try:
            from comunicaciones.services import NotificacionService
            NotificacionService.enviar_email(
                destinatario=postulacion.email,
                asunto=f'Actualizacion de tu postulacion — {postulacion.vacante.titulo}',
                cuerpo=(
                    f'Estimado/a {postulacion.nombre_completo},\n\n'
                    f'Tu postulacion para el puesto "{postulacion.vacante.titulo}" '
                    f'ha avanzado a la etapa: {etapa.nombre}.\n\n'
                    f'Gracias por tu interes.'
                ),
            )
        except Exception:
            pass

    return JsonResponse({
        'ok': True,
        'etapa_id': etapa.pk,
        'etapa_nombre': etapa.nombre,
        'etapa_color': etapa.color,
    })


# ══════════════════════════════════════════════════════════════
# PUBLICAR OFERTA (accion rapida)
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
@require_POST
def publicar_oferta(request, pk):
    """
    Accion rapida para marcar una vacante como PUBLICADA
    y registrar logs para cada integracion activa.
    """
    vacante = get_object_or_404(Vacante, pk=pk)
    estado_anterior = vacante.estado

    vacante.estado = 'PUBLICADA'
    if not vacante.fecha_publicacion:
        from datetime import date as _date
        vacante.fecha_publicacion = _date.today()
    vacante.save(update_fields=['estado', 'fecha_publicacion'])

    log_update(request, vacante, {
        'estado': {'old': estado_anterior, 'new': 'PUBLICADA'},
    })

    # Registrar log de publicacion para cada plataforma configurada
    from integraciones.models import LogPublicacionVacante
    plataformas_notificadas = []

    # Portal propio siempre se notifica si publica=True
    if vacante.publica:
        LogPublicacionVacante.objects.create(
            vacante=vacante,
            plataforma='PORTAL',
            estado='OK',
            mensaje=f'Vacante publicada desde accion rapida por {request.user}',
            publicado_por=request.user,
        )
        plataformas_notificadas.append('Portal de Empleo')

    if plataformas_notificadas:
        msg = 'Vacante publicada. Plataformas notificadas: ' + ', '.join(plataformas_notificadas)
    else:
        msg = 'Vacante marcada como Publicada. Activa "Visible en Portal" para que aparezca en el portal de empleo.'

    messages.success(request, msg)
    return redirect('vacante_detalle', pk=vacante.pk)


# ══════════════════════════════════════════════════════════════
# DASHBOARD DE RECLUTAMIENTO
# ══════════════════════════════════════════════════════════════

@login_required
@solo_admin
def dashboard_reclutamiento(request):
    """
    Dashboard ejecutivo de reclutamiento con KPIs, funnel y actividad reciente.
    """
    from django.db.models import Min

    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_mes = hoy.replace(day=1)

    # ── KPI 1: Vacantes activas ───────────────────────────────
    vacantes_activas = Vacante.objects.filter(
        estado__in=['PUBLICADA', 'EN_PROCESO']
    ).count()

    # ── KPI 2: Total candidatos ───────────────────────────────
    total_candidatos = Postulacion.objects.filter(estado='ACTIVA').count()

    # ── KPI 3: Entrevistas esta semana ────────────────────────
    entrevistas_semana = EntrevistaPrograma.objects.filter(
        fecha_hora__date__gte=inicio_semana,
        fecha_hora__date__lte=hoy,
    ).count()

    # ── KPI 4: Ofertas publicadas este mes ────────────────────
    ofertas_mes = Vacante.objects.filter(
        fecha_publicacion__gte=inicio_mes,
    ).count()

    # ── Time-to-hire: dias desde creado_en hasta marcado CUBIERTA ──
    # Aproximacion: diferencia entre creado_en y actualizado_en en vacantes CUBIERTA
    vacantes_cubiertas = Vacante.objects.filter(
        estado='CUBIERTA',
        fecha_publicacion__isnull=False,
    )
    time_to_hire_avg = None
    if vacantes_cubiertas.exists():
        tiempos = []
        for v in vacantes_cubiertas:
            dias = (v.actualizado_en.date() - v.fecha_publicacion).days
            if dias >= 0:
                tiempos.append(dias)
        if tiempos:
            time_to_hire_avg = round(sum(tiempos) / len(tiempos), 0)

    # ── Funnel del pipeline ───────────────────────────────────
    etapas = EtapaPipeline.objects.filter(activa=True).order_by('orden')
    funnel = []
    for etapa in etapas:
        count = Postulacion.objects.filter(etapa=etapa, estado='ACTIVA').count()
        funnel.append({
            'etapa': etapa.nombre,
            'count': count,
            'color': etapa.color,
        })

    # Totales adicionales para funnel superior/inferior
    total_aplicaron = Postulacion.objects.count()
    total_contratados = Postulacion.objects.filter(estado='CONTRATADA').count()
    total_descartados = Postulacion.objects.filter(estado='DESCARTADA').count()

    # ── Top 3 vacantes mas aplicadas ─────────────────────────
    top_vacantes = (
        Vacante.objects
        .annotate(num_post=Count('postulaciones'))
        .filter(num_post__gt=0)
        .order_by('-num_post')[:5]
    )

    # ── Actividad reciente: ultimas 10 postulaciones ──────────
    recientes = (
        Postulacion.objects
        .select_related('vacante', 'etapa')
        .order_by('-fecha_postulacion')[:10]
    )

    # ── Tasa de conversión global ─────────────────────────────
    conversion_pct = round(
        (total_contratados / total_aplicaron * 100) if total_aplicaron > 0 else 0, 1
    )

    # ── Tendencia postulaciones — últimos 6 meses ─────────────
    postulaciones_6m_json = '[]'
    try:
        from datetime import date as _date
        from django.db.models.functions import TruncMonth
        meses = (
            Postulacion.objects
            .filter(fecha_postulacion__date__gte=hoy - timedelta(days=180))
            .annotate(mes=TruncMonth('fecha_postulacion'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        postulaciones_6m_json = json.dumps([
            {'label': m['mes'].strftime('%b %Y'), 'total': m['total']}
            for m in meses if m['mes']
        ])
    except Exception:
        pass

    # ── Entrevistas próximas (próximos 7 días) ────────────────
    entrevistas_proximas = []
    try:
        entrevistas_proximas = list(
            EntrevistaPrograma.objects
            .select_related('postulacion', 'postulacion__vacante', 'entrevistador')
            .filter(
                fecha_hora__date__gte=hoy,
                fecha_hora__date__lte=hoy + timedelta(days=7),
                estado__in=['PROGRAMADA', 'CONFIRMADA'],
            )
            .order_by('fecha_hora')[:8]
        )
    except Exception:
        pass

    # ── Vacantes por prioridad ────────────────────────────────
    vacantes_prioridad_json = '[]'
    try:
        prioridades = (
            Vacante.objects
            .filter(estado__in=['PUBLICADA', 'EN_PROCESO'])
            .values('prioridad')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        COLOR_PRIORIDAD = {
            'CRITICA': '#dc2626', 'ALTA': '#f59e0b',
            'MEDIA': '#0f766e', 'BAJA': '#94a3b8',
        }
        vacantes_prioridad_json = json.dumps([
            {'label': p['prioridad'], 'total': p['total'],
             'color': COLOR_PRIORIDAD.get(p['prioridad'], '#94a3b8')}
            for p in prioridades
        ])
    except Exception:
        pass

    context = {
        'titulo': 'Dashboard de Reclutamiento',
        # KPIs
        'kpi_vacantes_activas': vacantes_activas,
        'kpi_total_candidatos': total_candidatos,
        'kpi_entrevistas_semana': entrevistas_semana,
        'kpi_ofertas_mes': ofertas_mes,
        'kpi_time_to_hire': time_to_hire_avg,
        'kpi_conversion_pct': conversion_pct,
        # Funnel
        'funnel': funnel,
        'funnel_json': json.dumps([{'etapa': f['etapa'], 'count': f['count'], 'color': f['color']} for f in funnel]),
        'total_aplicaron': total_aplicaron,
        'total_contratados': total_contratados,
        'total_descartados': total_descartados,
        # Top vacantes
        'top_vacantes': top_vacantes,
        # Recientes
        'postulaciones_recientes': recientes,
        # Nuevos
        'postulaciones_6m_json': postulaciones_6m_json,
        'entrevistas_proximas': entrevistas_proximas,
        'vacantes_prioridad_json': vacantes_prioridad_json,
    }
    return render(request, 'reclutamiento/dashboard.html', context)


# ══════════════════════════════════════════════════════════════════════════════
# CONTRATAR CANDIDATO → onboarding automático
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def contratar_candidato(request, pk):
    """
    GET  → Modal/form para ingresar datos de contratación.
    POST → Crea Personal + ProcesoOnboarding automático.

    Flujo:
      1. Validar nro_doc único y fecha_alta
      2. Crear Personal desde datos de Postulacion + form
      3. Marcar Postulacion como CONTRATADA + vincular personal_creado
      4. Mover etapa del pipeline a "Contratado"
      5. Auto-crear ProcesoOnboarding desde primera PlantillaOnboarding activa
      6. Notificar vía NotificacionService
    """
    from .models import Postulacion, EtapaPipeline
    from personal.models import Personal

    postulacion = get_object_or_404(Postulacion, pk=pk, estado='ACTIVA')

    # Plantillas de onboarding disponibles
    from onboarding.models import PlantillaOnboarding
    plantillas_onb = PlantillaOnboarding.objects.filter(activa=True).order_by('nombre')

    if request.method == 'GET':
        return render(request, 'reclutamiento/contratar_modal.html', {
            'postulacion': postulacion,
            'plantillas_onb': plantillas_onb,
            'hoy': date.today(),
        })

    # ── POST: procesar contratación ───────────────────────────────────────────
    nro_doc    = request.POST.get('nro_doc', '').strip()
    fecha_alta = request.POST.get('fecha_alta', '').strip()
    tipo_trab  = request.POST.get('tipo_trab', 'Empleado')
    sueldo_raw = request.POST.get('sueldo_base', '0').strip()
    plantilla_id = request.POST.get('plantilla_id', '').strip()

    # Validaciones
    if not nro_doc:
        messages.error(request, 'El número de documento es obligatorio.')
        return redirect('contratar_candidato', pk=pk)

    if Personal.objects.filter(nro_doc=nro_doc).exists():
        messages.error(request, f'Ya existe un empleado con el documento {nro_doc}.')
        return redirect('contratar_candidato', pk=pk)

    if not fecha_alta:
        messages.error(request, 'La fecha de alta es obligatoria.')
        return redirect('contratar_candidato', pk=pk)

    try:
        from datetime import datetime
        fecha_alta_dt = datetime.strptime(fecha_alta, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Fecha de alta inválida.')
        return redirect('contratar_candidato', pk=pk)

    try:
        from decimal import Decimal
        sueldo = Decimal(sueldo_raw) if sueldo_raw else Decimal('0')
    except Exception:
        sueldo = Decimal('0')

    from django.db import transaction
    with transaction.atomic():
        # ── 1. Crear Personal ──────────────────────────────────────────────
        personal = Personal.objects.create(
            apellidos_nombres = postulacion.nombre_completo,
            nro_doc           = nro_doc,
            tipo_doc          = 'DNI',
            cargo             = postulacion.vacante.titulo[:150],
            tipo_trab         = tipo_trab,
            fecha_alta        = fecha_alta_dt,
            sueldo_base       = sueldo,
            correo_personal   = postulacion.email,
            estado            = 'Activo',
            # Área desde la vacante si existe
            **(
                {'subarea': postulacion.vacante.area.subareas.filter(activa=True).first()}
                if postulacion.vacante.area_id and
                   postulacion.vacante.area.subareas.filter(activa=True).exists()
                else {}
            ),
        )

        # ── 2. Vincular postulacion ────────────────────────────────────────
        postulacion.estado         = 'CONTRATADA'
        postulacion.personal_creado = personal
        postulacion.save(update_fields=['estado', 'personal_creado'])

        # ── 3. Mover etapa pipeline a "Contratado" ─────────────────────────
        etapa_contratado = EtapaPipeline.objects.filter(codigo='contratado').first()
        if etapa_contratado:
            postulacion.etapa = etapa_contratado
            postulacion.save(update_fields=['etapa'])

        # ── 4. Auto-crear ProcesoOnboarding ───────────────────────────────
        proceso_onb = None
        plantilla   = None
        if plantilla_id:
            plantilla = PlantillaOnboarding.objects.filter(pk=plantilla_id, activa=True).first()
        if not plantilla:
            plantilla = PlantillaOnboarding.objects.filter(activa=True).order_by('id').first()

        if plantilla:
            from onboarding.models import ProcesoOnboarding, PasoOnboarding
            proceso_onb = ProcesoOnboarding.objects.create(
                personal      = personal,
                plantilla     = plantilla,
                fecha_ingreso = fecha_alta_dt,
                fecha_inicio  = date.today(),
                estado        = 'EN_CURSO',
                iniciado_por  = request.user,
                notas         = f'Generado automáticamente al contratar desde reclutamiento. '
                                f'Postulación #{postulacion.pk} — Vacante: {postulacion.vacante.titulo}',
            )
            # Generar pasos desde plantilla
            for paso_tpl in plantilla.pasos.all().order_by('orden'):
                PasoOnboarding.objects.create(
                    proceso      = proceso_onb,
                    paso_plantilla = paso_tpl,
                    orden        = paso_tpl.orden,
                    titulo       = paso_tpl.titulo,
                    estado       = 'PENDIENTE',
                    fecha_limite = fecha_alta_dt + timedelta(days=paso_tpl.dias_plazo),
                )

        # ── 5. Notificar al personal recién contratado (in-app) ──────────
        try:
            from comunicaciones.services import NotificacionService
            cuerpo_notif = (
                f'<p>Bienvenido/a <strong>{personal.apellidos_nombres}</strong>. '
                f'Tu incorporación como <strong>{personal.cargo}</strong> ha sido registrada '
                f'con fecha de ingreso <strong>{fecha_alta_dt:%d/%m/%Y}</strong>.</p>'
                + (f'<p>Tu proceso de onboarding "<strong>{plantilla.nombre}</strong>" '
                   f'ha sido iniciado automáticamente.</p>' if plantilla else '')
            )
            NotificacionService.enviar(
                destinatario = personal,
                asunto       = f'¡Bienvenido/a, {personal.apellidos_nombres.split(",")[0].strip()}!',
                cuerpo       = cuerpo_notif,
                tipo         = 'IN_APP',
            )
        except Exception:
            pass  # La notificación no bloquea el flujo

    msg = f'✅ {personal.apellidos_nombres} contratado/a exitosamente.'
    if proceso_onb:
        msg += f' Onboarding "{plantilla.nombre}" iniciado ({plantilla.pasos.count()} pasos).'
    messages.success(request, msg)

    # Redirigir al registro de personal recién creado
    return redirect('personal_detail', pk=personal.pk)


# ══════════════════════════════════════════════════════════════
# API — GENERADOR DE DESCRIPCION CON IA
# ══════════════════════════════════════════════════════════════

@login_required
@require_POST
def api_generar_descripcion(request):
    """
    Genera descripción y requisitos del puesto usando IA.
    Similar a BUK AI "Crear" — generador de descripciones de puesto.
    """
    import json as _json
    try:
        data = _json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'JSON inválido'})

    titulo = data.get('titulo', '').strip()
    area = data.get('area', '').strip()

    if not titulo:
        return JsonResponse({'ok': False, 'error': 'Título requerido'})

    from asistencia.services.ai_service import get_service
    svc = get_service()
    if not svc:
        return JsonResponse({'ok': False, 'error': 'IA no configurada. Ir a Configuración > IA.'})

    system = (
        'Eres un experto en RRHH del sector construcción en Perú. '
        'Genera descripciones de puesto profesionales y requisitos para una empresa constructora. '
        'El formato debe ser claro y estructurado. '
        'Responde en español. No uses markdown, solo texto plano con viñetas (-).'
    )

    prompt = (
        f'Genera la DESCRIPCIÓN DEL PUESTO y los REQUISITOS para el cargo:\n\n'
        f'Puesto: {titulo}\n'
        f'Área: {area or "General"}\n'
        f'Empresa: Consorcio constructora (obra hospitalaria)\n\n'
        f'Formato de respuesta:\n'
        f'DESCRIPCION:\n[texto de 3-5 líneas describiendo responsabilidades principales]\n\n'
        f'REQUISITOS:\n[lista con - de 5-8 requisitos: educación, experiencia, conocimientos, habilidades]'
    )

    try:
        resultado = svc.generate(prompt, system=system)
        if not resultado:
            return JsonResponse({'ok': False, 'error': 'La IA no devolvió resultado.'})

        # Parsear respuesta
        desc = ''
        reqs = ''
        if 'DESCRIPCION:' in resultado.upper() and 'REQUISITOS:' in resultado.upper():
            parts = resultado.upper().split('REQUISITOS:')
            desc_part = resultado[:resultado.upper().index('REQUISITOS:')]
            reqs_part = resultado[resultado.upper().index('REQUISITOS:') + len('REQUISITOS:'):]
            desc = desc_part.replace('DESCRIPCION:', '').replace('DESCRIPCIÓN:', '').strip()
            reqs = reqs_part.strip()
        else:
            # Sin formato esperado — usar todo como descripción
            desc = resultado.strip()

        return JsonResponse({
            'ok': True,
            'descripcion': desc,
            'requisitos': reqs,
        })

    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Error IA: {str(e)[:200]}'})
